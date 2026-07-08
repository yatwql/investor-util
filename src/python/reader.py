from __future__ import annotations

import logging
import os
import zipfile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from src.python.models import Holding

logger = logging.getLogger("invest")

# 预期表头
_EXPECTED_HEADER = ["名称", "代码", "持仓份额", "每份成本"]


def list_xlsx_files(directory: str) -> list[str]:
    """列出指定目录下所有 .xlsx 文件，按修改时间降序排列。

    Args:
        directory: 目录路径

    Returns:
        xlsx 文件绝对路径列表
    """
    if not os.path.isdir(directory):
        return []

    files = [
        os.path.join(directory, f)
        for f in os.listdir(directory)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ]
    files.sort(key=os.path.getmtime, reverse=True)
    return files


def get_xlsx_info(filepath: str) -> dict:
    """快速读取 xlsx 文件的元信息（不解析全量持仓数据）。

    Args:
        filepath: .xlsx 文件路径

    Returns:
        {sheet_names: [str], accounts: int, total_rows: int}
        文件无法读取时返回含 error 字段的字典
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        try:
            sheets = wb.sheetnames
            total_rows = 0
            for s in sheets:
                ws = wb[s]
                total_rows += max(0, (ws.max_row or 1) - 1)  # 减去标题行
            return {
                "sheet_names": sheets,
                "accounts": len(sheets),
                "total_rows": total_rows,
            }
        finally:
            wb.close()
    except Exception as e:
        return {"error": str(e)}


def read_holdings(filepath: str) -> list[Holding]:
    """读取持仓 Excel 文件，返回持仓记录列表。

    每个工作表为一个独立账户，表头固定 4 列：
        名称、代码、持仓份额、每份成本

    Args:
        filepath: .xlsx 文件路径

    Returns:
        持仓记录列表

    Raises:
        FileNotFoundError: 文件不存在
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"持仓文件不存在: {filepath}")

    logger.info("正在读取持仓文件: %s", filepath)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile) as e:
        logger.error("文件格式错误 (非有效 xlsx): %s", e)
        raise ValueError(f"文件格式错误，请确认是有效的 Excel 文件: {e}") from e

    try:
        holdings = _parse_workbook(wb)
    finally:
        wb.close()

    if not holdings:
        logger.warning("未读取到任何持仓记录")

    return holdings


def _parse_workbook(wb) -> list[Holding]:
    """解析 workbook 中的所有工作表。"""
    holdings: list[Holding] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.debug("处理工作表: %s (行数=%s)", sheet_name, ws.max_row)

        if ws.max_row is None or ws.max_row < 2:
            logger.warning("工作表 '%s' 为空，跳过", sheet_name)
            continue

        # 校验表头
        header = [_safe_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not _match_header(header, _EXPECTED_HEADER):
            logger.warning("工作表 '%s' 表头不匹配: %s，跳过", sheet_name, header)
            continue

        # 逐行解析（跟踪实际 Excel 行号用于错误提示）
        sheet_holdings = 0
        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过全空行
            if all(cell is None for cell in row):
                continue
            name = _safe_str(row[0])
            if not name:
                continue

            code = _safe_str(row[1])
            shares = _safe_float(row[2], "份额", sheet_name, excel_row)
            cost_price = _safe_float(row[3], "成本", sheet_name, excel_row)
            if shares is None or cost_price is None:
                continue

            if shares <= 0 or cost_price < 0:
                logger.warning("工作表 '%s' 第 %d 行含无效数值 (份额=%s, 成本=%s), 跳过此行",
                               sheet_name, excel_row, shares, cost_price)
                continue

            try:
                holding = Holding(
                    account=sheet_name.strip(),
                    name=name,
                    code=code,
                    shares=shares,
                    cost_price=cost_price,
                )
                holdings.append(holding)
                sheet_holdings += 1
            except (ValueError, TypeError) as e:
                logger.warning("工作表 '%s' 第 %d 行解析失败: %s", sheet_name, excel_row, e)
                continue

        logger.info("工作表 '%s' 解析完成，共 %d 条持仓", sheet_name, sheet_holdings)

    return holdings


def _safe_str(value) -> str:
    """将单元格值转为字符串，None 返回空字符串。"""
    if value is None:
        return ""
    return str(value).strip()


def _safe_float(value, field: str, sheet: str, row: int) -> float | None:
    """将单元格值转为 float；None 或转换失败时记录警告并返回 None。"""
    if value is None:
        logger.warning("工作表 '%s' 第 %d 行 '%s' 为空，跳过该行", sheet, row, field)
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        logger.warning("工作表 '%s' 第 %d 行 '%s' 无法解析 (%s)，跳过该行", sheet, row, field, value)
        return None


def _match_header(actual: list[str], expected: list[str]) -> bool:
    """判断表头是否与期望匹配（允许比预期更多的列）。"""
    if len(actual) < len(expected):
        return False
    return all(a == e for a, e in zip(actual, expected))
