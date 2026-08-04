from __future__ import annotations

import logging
import os
import re
import zipfile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from src.python.core import holding_status
from src.python.core.models import DividendRecord, HoldingsFile, Holding, TradeRecord

logger = logging.getLogger("invest")

# 预期表头
_EXPECTED_HEADER = ["名称", "代码", "持仓份额", "每份成本"]

# 可选流水页签名与表头（持仓文件扩展——不触碰主表 4 列语义）
_TRADE_SHEET_NAME = "交易流水"
_DIVIDEND_SHEET_NAME = "分红流水"
# 交易流水列：日期/代码/操作/份额/价格 + 可选费用列（费用列缺失时按 0 处理）
_TRADE_HEADER = ["日期", "代码", "操作", "份额", "价格"]
_TRADE_HEADER_WITH_FEE = ["日期", "代码", "操作", "份额", "价格", "费用"]
# 分红流水列：日期/代码/每份分红
_DIVIDEND_HEADER = ["日期", "代码", "每份分红"]

# 操作方向归一化（买入/卖出 中英文同义）
_ACTION_BUY = {"买入", "买", "buy", "申购"}
_ACTION_SELL = {"卖出", "卖", "sell", "赎回"}

# 日期形态：YYYY-MM-DD 或 YYYY/MM/DD
_DATE_RE = re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$")


def list_xlsx_files(directory: str) -> list[str]:
    """列出指定目录下所有 .xlsx 文件，按修改时间降序排列。

    Args:
        directory: 目录路径

    Returns:
        xlsx 文件绝对路径列表
    """
    if not os.path.isdir(directory):
        return []

    files = [
        os.path.join(directory, f)
        for f in os.listdir(directory)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ]
    files.sort(key=os.path.getmtime, reverse=True)
    return files


def get_xlsx_info(filepath: str) -> dict:
    """快速读取 xlsx 文件的元信息（不解析全量持仓数据）。

    Args:
        filepath: .xlsx 文件路径

    Returns:
        {sheet_names: [str], accounts: int, total_rows: int}
        文件无法读取时返回含 error 字段的字典
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        try:
            sheets = wb.sheetnames
            total_rows = 0
            for s in sheets:
                ws = wb[s]
                total_rows += max(0, (ws.max_row or 1) - 1)  # 减去标题行
            return {
                "sheet_names": sheets,
                "accounts": len(sheets),
                "total_rows": total_rows,
            }
        finally:
            wb.close()
    except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException, OSError) as e:
        return {"error": str(e)}


def read_holdings(filepath: str) -> list[Holding]:
    """读取持仓 Excel 文件，返回持仓记录列表。

    每个工作表为一个独立账户，表头固定 4 列：
        名称、代码、持仓份额、每份成本

    Args:
        filepath: .xlsx 文件路径

    Returns:
        持仓记录列表

    Raises:
        FileNotFoundError: 文件不存在
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"持仓文件不存在: {filepath}")

    logger.info("正在读取持仓文件: %s", filepath)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile) as e:
        logger.error("文件格式错误 (非有效 xlsx): %s", e)
        raise ValueError(f"文件格式错误，请确认是有效的 Excel 文件: {e}") from e

    try:
        holdings = _parse_workbook(wb)
    finally:
        wb.close()

    if not holdings:
        logger.warning("未读取到任何持仓记录")

    return holdings


def _parse_workbook(wb) -> list[Holding]:
    """解析 workbook 中的所有工作表。"""
    holdings: list[Holding] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.debug("处理工作表: %s (行数=%s)", sheet_name, ws.max_row)

        if ws.max_row is None or ws.max_row < 2:
            logger.warning("工作表 '%s' 为空，跳过", sheet_name)
            continue

        # 校验表头
        header = [_safe_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not _match_header(header, _EXPECTED_HEADER):
            logger.warning("工作表 '%s' 表头不匹配: %s，跳过", sheet_name, header)
            continue

        # 逐行解析（跟踪实际 Excel 行号用于错误提示）
        sheet_holdings = 0
        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过全空行
            if all(cell is None for cell in row):
                continue
            name = _safe_str(row[0])
            if not name:
                continue

            code = _safe_str(row[1])
            shares = _safe_float(row[2], "份额", sheet_name, excel_row)
            cost_price = _safe_float(row[3], "成本", sheet_name, excel_row)
            if shares is None or cost_price is None:
                continue

            if shares <= 0 or cost_price < 0:
                logger.warning(
                    "工作表 '%s' 第 %d 行含无效数值 (份额=%s, 成本=%s), 跳过此行",
                    sheet_name,
                    excel_row,
                    shares,
                    cost_price,
                )
                continue

            try:
                holding = Holding(
                    account=sheet_name.strip(),
                    name=name,
                    code=code,
                    shares=shares,
                    cost_price=cost_price,
                    # 本地可判定的品种状态：代码格式异常在解析期即标注，
                    # 其余状态（净值缺失/可能退市/名称不匹配）需行情数据，后续标注
                    data_status=(
                        "bad_code_format"
                        if holding_status.classify_code_format(code) != holding_status.STATUS_OK
                        else ""
                    ),
                )
                holdings.append(holding)
                sheet_holdings += 1
            except (ValueError, TypeError) as e:
                logger.warning("工作表 '%s' 第 %d 行解析失败: %s", sheet_name, excel_row, e)
                continue

        logger.info("工作表 '%s' 解析完成，共 %d 条持仓", sheet_name, sheet_holdings)

    return holdings


def read_flow_sheets(filepath: str) -> tuple[list[TradeRecord], list[DividendRecord]]:
    """读取持仓 Excel 的可选「交易流水」「分红流水」页签。

    主表解析不受影响（与 read_holdings 独立）；无流水页签时返回空列表。

    Args:
        filepath: .xlsx 文件路径

    Returns:
        (transactions, dividends) 二元组：
          transactions — 交易流水记录列表（无页签/解析为空时为空列表）
          dividends    — 分红流水记录列表（无页签/解析为空时为空列表）

    Raises:
        FileNotFoundError: 文件不存在
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"持仓文件不存在: {filepath}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile) as e:
        logger.error("文件格式错误 (非有效 xlsx): %s", e)
        raise ValueError(f"文件格式错误，请确认是有效的 Excel 文件: {e}") from e

    try:
        transactions: list[TradeRecord] = []
        dividends: list[DividendRecord] = []
        for sheet_name in wb.sheetnames:
            if sheet_name.strip() == _TRADE_SHEET_NAME:
                transactions = _parse_trade_sheet(wb[sheet_name])
            elif sheet_name.strip() == _DIVIDEND_SHEET_NAME:
                dividends = _parse_dividend_sheet(wb[sheet_name])
    finally:
        wb.close()

    if transactions:
        logger.info("交易流水页签解析完成，共 %d 条记录", len(transactions))
    if dividends:
        logger.info("分红流水页签解析完成，共 %d 条记录", len(dividends))
    return transactions, dividends


def read_holdings_with_flows(filepath: str) -> HoldingsFile:
    """读取持仓 Excel 完整数据（主表 + 可选流水页签）。

    Args:
        filepath: .xlsx 文件路径

    Returns:
        HoldingsFile 容器：holdings（主表）+ transactions + dividends（可选流水）。
    """
    holdings = read_holdings(filepath)
    transactions, dividends = read_flow_sheets(filepath)
    return HoldingsFile(
        holdings=holdings,
        transactions=transactions,
        dividends=dividends,
    )


def _parse_trade_sheet(ws) -> list[TradeRecord]:
    """解析「交易流水」页签。

    表头须为 日期/代码/操作/份额/价格（费用列可选）。逐行容错：
    日期格式、数值、操作方向非法时跳过该行并告警，不阻塞其余行。
    """
    records: list[TradeRecord] = []
    if ws.max_row is None or ws.max_row < 2:
        logger.warning("交易流水页签为空，跳过")
        return records

    header = [_safe_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if header[:5] != _TRADE_HEADER:
        logger.warning("交易流水页签表头不匹配: %s，跳过（应为 %s）", header, _TRADE_HEADER)
        return records
    has_fee = len(header) >= 6 and header[5] == "费用"

    sheet_name = ws.title
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue
        code = _safe_str(row[1])
        if not code:
            continue
        date = _safe_str(row[0])
        if not _valid_date(date):
            logger.warning("交易流水页签第 %d 行日期无效 (%s)，跳过该行", excel_row, date)
            continue
        action = _normalize_action(_safe_str(row[2]))
        if action is None:
            logger.warning("交易流水页签第 %d 行操作无效 (%s)，跳过该行", excel_row, row[2])
            continue
        shares = _safe_float(row[3], "份额", _TRADE_SHEET_NAME, excel_row)
        price = _safe_float(row[4], "价格", _TRADE_SHEET_NAME, excel_row)
        if shares is None or price is None:
            continue
        if shares <= 0 or price < 0:
            logger.warning("交易流水页签第 %d 行含无效数值 (份额=%s, 价格=%s)，跳过该行", excel_row, shares, price)
            continue
        fee = 0.0
        if has_fee:
            fee = _safe_float(row[5], "费用", _TRADE_SHEET_NAME, excel_row) or 0.0
        records.append(
            TradeRecord(
                date=date,
                code=code,
                action=action,
                shares=shares,
                price=price,
                fee=max(0.0, fee),
                account=sheet_name.strip(),
            )
        )
    return records


def _parse_dividend_sheet(ws) -> list[DividendRecord]:
    """解析「分红流水」页签。

    表头须为 日期/代码/每份分红。逐行容错：日期格式、数值非法时跳过。
    """
    records: list[DividendRecord] = []
    if ws.max_row is None or ws.max_row < 2:
        logger.warning("分红流水页签为空，跳过")
        return records

    header = [_safe_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if header[:3] != _DIVIDEND_HEADER:
        logger.warning("分红流水页签表头不匹配: %s，跳过（应为 %s）", header, _DIVIDEND_HEADER)
        return records

    sheet_name = ws.title
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue
        code = _safe_str(row[1])
        if not code:
            continue
        date = _safe_str(row[0])
        if not _valid_date(date):
            logger.warning("分红流水页签第 %d 行日期无效 (%s)，跳过该行", excel_row, date)
            continue
        amount = _safe_float(row[2], "每份分红", _DIVIDEND_SHEET_NAME, excel_row)
        if amount is None or amount < 0:
            logger.warning("分红流水页签第 %d 行每份分红无效 (%s)，跳过该行", excel_row, row[2])
            continue
        records.append(
            DividendRecord(
                date=date,
                code=code,
                amount=amount,
                account=sheet_name.strip(),
            )
        )
    return records


def _valid_date(value: str) -> bool:
    """判断字符串是否为合法日期（YYYY-MM-DD 或 YYYY/MM/DD）。"""
    return bool(_DATE_RE.match(value))


def _normalize_action(value: str) -> str | None:
    """归一化操作方向为 buy/sell；无法识别时返回 None。"""
    v = value.strip().lower()
    if v in _ACTION_BUY:
        return "buy"
    if v in _ACTION_SELL:
        return "sell"
    return None


def _safe_str(value) -> str:
    """将单元格值转为字符串，None 返回空字符串。"""
    if value is None:
        return ""
    return str(value).strip()


def _safe_float(value, field: str, sheet: str, row: int) -> float | None:
    """将单元格值转为 float；None 或转换失败时记录警告并返回 None。"""
    if value is None:
        logger.warning("工作表 '%s' 第 %d 行 '%s' 为空，跳过该行", sheet, row, field)
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        logger.warning("工作表 '%s' 第 %d 行 '%s' 无法解析 (%s)，跳过该行", sheet, row, field, value)
        return None


def _match_header(actual: list[str], expected: list[str]) -> bool:
    """判断表头是否与期望匹配（允许比预期更多的列）。"""
    if len(actual) < len(expected):
        return False
    return all(a == e for a, e in zip(actual, expected))
