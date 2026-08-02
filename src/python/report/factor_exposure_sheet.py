"""因子暴露分析 Excel 写入模块 — 报告页签 17。

输出列：
  风格因子 | 暴露系数 β | t 值 | 显著（95%） | 风格归属占比
  基准对照：组合 β | 基准 β | 相对暴露
  备注：窗口 / 有效样本 / α / 因子相关性

数据不足或数据源故障时写入占位文本（available=False，§1.4.5 数据降级治理）。
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.python.core.registry import get_report_section_number, get_report_sheet_name
from src.python.report.data_status import STATUS_MESSAGES
from src.python.report.excel_writer import (
    _write_placeholder,
    auto_width,
    freeze_header,
    write_data_row,
    write_header_row,
    write_title_row,
)

logger = logging.getLogger("invest")

_NCOLS = 5
_HEADERS = ["风格因子", "暴露系数 β", "t 值", "显著（95%）", "风格归属占比"]

# 固定 MVP 因子展示顺序（与 FACTOR_INDICES 键一致，名字在编排层 data 中）
_FACTOR_ORDER = ["value", "growth", "quality"]

_FONT_GREEN = Font(color="009900")
_FONT_RED = Font(color="CC0000")


def _factor_name(factor: str, factor_names: dict | None) -> str:
    """因子 key → 中文名（编排层 data.factor_names 为单一数据源）。"""
    if factor_names and factor in factor_names:
        return str(factor_names[factor])
    return factor


def write_factor_exposure_sheet(
    ws: Worksheet,
    factor_exposure: dict[str, Any] | None,
    factor_names: dict | None = None,
) -> None:
    """写入因子暴露分析页签。

    Args:
        ws: openpyxl Worksheet 对象
        factor_exposure: C19 契约 dict；None 或 available=False 时写入占位。
        factor_names: 因子 key → 中文名映射（缺省回退 key 本身）。
    """
    _name = get_report_sheet_name("factor_exposure")
    write_title_row(ws, 1, f"{get_report_section_number('factor_exposure')}. {_name}", ncols=_NCOLS)

    if not factor_exposure or not factor_exposure.get("available"):
        row = write_header_row(ws, 2, _HEADERS)
        _write_placeholder(
            ws,
            STATUS_MESSAGES["factor_exposure_unavailable"],
            row=row + 1,
            max_cols=_NCOLS,
        )
        freeze_header(ws, row=2)
        auto_width(ws)
        logger.info("因子暴露分析：无数据，写入占位")
        return

    row = write_header_row(ws, 2, _HEADERS)
    betas = factor_exposure.get("betas", {})
    t_stats = factor_exposure.get("t_stats", {})
    significant = factor_exposure.get("significant", {})
    style_alloc = factor_exposure.get("style_allocation", {})

    for factor in _FACTOR_ORDER:
        if factor not in betas:
            continue
        _sig = bool(significant.get(factor))
        row_data = [
            _factor_name(factor, factor_names),
            betas[factor],
            t_stats.get(factor, "--") if t_stats.get(factor) is not None else "--",
            "✅ 显著" if _sig else "—",
            style_alloc.get(factor, 0.0),
        ]
        write_data_row(ws, row, row_data, formats=[None, "0.0000", "0.000", None, "0.00%"])
        if _sig:
            ws.cell(row=row, column=4).font = _FONT_GREEN
        row += 1

    # 基准对照（沪深300 同窗口回归）
    baseline_betas = factor_exposure.get("baseline_betas", {})
    if baseline_betas:
        row += 1
        row = write_title_row(ws, row, "基准对照（沪深300 同窗口回归）", ncols=_NCOLS)
        row = write_header_row(ws, row, ["风格因子", "组合 β", "基准 β", "相对暴露", ""])
        for factor in _FACTOR_ORDER:
            if factor not in betas or factor not in baseline_betas:
                continue
            rel = round(float(betas[factor]) - float(baseline_betas[factor]), 3)
            row_data = [
                _factor_name(factor, factor_names),
                betas[factor],
                baseline_betas[factor],
                rel,
                "",
            ]
            write_data_row(ws, row, row_data, formats=[None, "0.0000", "0.0000", "0.000", None])
            if rel > 0.1:
                ws.cell(row=row, column=4).font = _FONT_RED
            elif rel < -0.1:
                ws.cell(row=row, column=4).font = _FONT_GREEN
            row += 1

    # 备注区
    row += 1
    row = write_title_row(ws, row, "说明", ncols=_NCOLS)
    notes = [
        f"回归窗口：{factor_exposure.get('window', 0)} 个交易日；有效样本：{factor_exposure.get('sample_count', 0)} 期",
        f"α（截距）= {factor_exposure.get('alpha', 0.0)}；显著列为 95% 双尾 t 检验结果",
    ]
    stale = factor_exposure.get("stale_factors") or []
    if stale:
        notes.append(f"已剔除停更/不可用因子：{'、'.join(stale)}")
    corr = factor_exposure.get("factor_correlations") or {}
    if corr:
        corr_text = "；".join(f"{k}={v}" for k, v in corr.items())
        notes.append(f"因子间相关性：{corr_text}")
    corr_note = factor_exposure.get("correlation_note") or ""
    if corr_note:
        notes.append(corr_note)
    for n in notes:
        write_data_row(ws, row, [n, "", "", "", ""])
        row += 1

    freeze_header(ws, row=2)
    auto_width(ws, min_width=10, max_width=30)
    logger.info("因子暴露分析页签写入完成: %d 个因子", len(betas))
