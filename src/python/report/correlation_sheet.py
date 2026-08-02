"""持仓相关性矩阵 Excel 写入模块 — 报告页签 11。

输出列：
  下三角相关矩阵（行品种 × 列品种，红=正相关 / 蓝=负相关 / 白=不显著 / 灰=N/A）
  配对明细（按 |r| 降序）
  说明：窗口 / 有效样本 / 显著性 / 数据不足品种

数据不足或数据源故障时写入占位文本（available=False，§1.4.5 数据降级治理）。
配色与 HTML 模板保持一致（红=正相关 / 蓝=负相关 / 白=不显著 / 灰=数据不足）。
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.python.core.registry import get_report_section_number, get_report_sheet_name
from src.python.report.data_status import STATUS_MESSAGES
from src.python.report.excel_writer import (
    _write_placeholder,
    auto_width,
    freeze_header,
    write_data_row,
    write_header_row,
    write_title_row,
)

logger = logging.getLogger("invest")

# 显著性阈值（与 analysis.correlation.SIGNIFICANCE_LEVEL 一致）
_SIGNIFICANCE_LEVEL = 0.05

# 热力格填充：红=正相关（按强度分层）、蓝=负相关、白=不显著、灰=N/A、浅灰=对角线
_FILL_POS_STRONG = PatternFill("solid", fgColor="C00000")
_FILL_POS_MEDIUM = PatternFill("solid", fgColor="FF7F7F")
_FILL_POS_WEAK = PatternFill("solid", fgColor="FFE0E0")
_FILL_NEG_STRONG = PatternFill("solid", fgColor="1F4E79")
_FILL_NEG_MEDIUM = PatternFill("solid", fgColor="8DB4E2")
_FILL_NEG_WEAK = PatternFill("solid", fgColor="E0ECF7")
_FILL_INSIGNIFICANT = PatternFill("solid", fgColor="FFFFFF")
_FILL_NA = PatternFill("solid", fgColor="EEEEEE")
_FILL_DIAG = PatternFill("solid", fgColor="F5F5F5")

_FONT_STRONG = Font(color="FFFFFF", bold=True)
_FONT_BODY = Font(color="333333")
_FONT_GREY = Font(color="999999")
_FONT_LIGHT = Font(color="BBBBBB")
_FONT_POS_COLOR = Font(color="C00000", bold=True)
_FONT_NEG_COLOR = Font(color="1F4E79", bold=True)

_PAIRS_HEADERS = ["序号", "品种A", "品种B", "相关系数 r", "p 值", "样本数", "显著性"]
_PAIRS_NCOLS = len(_PAIRS_HEADERS)


def _cell_style(r: float | None, p: float | None) -> tuple[PatternFill, Font]:
    """返回相关性格的填充色 + 字体（与 HTML 模板配色一致）。

    Args:
        r: Pearson 相关系数；None 表示重叠样本不足（灰色 N/A）。
        p: 双侧 p 值；None 或 ≥0.05 判不显著（白色）。

    Returns:
        (fill, font) 用于 openpyxl 单元格样式。
    """
    if r is None:
        return _FILL_NA, _FONT_GREY
    if p is None or p >= _SIGNIFICANCE_LEVEL:
        return _FILL_INSIGNIFICANT, _FONT_GREY
    if r >= 0.5:
        return _FILL_POS_STRONG, _FONT_STRONG
    if r >= 0.3:
        return _FILL_POS_MEDIUM, _FONT_BODY
    if r > 0:
        return _FILL_POS_WEAK, _FONT_BODY
    if r <= -0.5:
        return _FILL_NEG_STRONG, _FONT_STRONG
    if r <= -0.3:
        return _FILL_NEG_MEDIUM, _FONT_BODY
    return _FILL_NEG_WEAK, _FONT_BODY


def write_correlation_sheet(
    ws: Worksheet,
    correlation_data: dict[str, Any] | None,
) -> None:
    """写入持仓相关性矩阵页签。

    Args:
        ws: openpyxl Worksheet 对象
        correlation_data: C19 契约 dict；None 或 available=False 时写入占位。
    """
    _name = get_report_sheet_name("correlation_analysis")
    codes = (correlation_data or {}).get("codes", []) if correlation_data else []
    _ncols = max(2, len(codes) + 2)
    write_title_row(ws, 1, f"{get_report_section_number('correlation_analysis')}. {_name}", ncols=_ncols)

    if not correlation_data or not correlation_data.get("available"):
        row = write_header_row(ws, 2, ["品种", "代码"] + list(codes))
        _write_placeholder(
            ws,
            STATUS_MESSAGES["correlation_unavailable"],
            row=row + 1,
            max_cols=_ncols,
        )
        freeze_header(ws, row=2)
        auto_width(ws)
        logger.info("持仓相关性矩阵：无数据，写入占位")
        return

    names = correlation_data.get("names", {})
    matrix = correlation_data.get("matrix", [])
    p_values = correlation_data.get("p_values", [])
    pairs = correlation_data.get("pairs", [])

    # ── 1. 矩阵表（下三角） ──
    row = write_header_row(
        ws, 2, [""] + [f"{names.get(c, c)} ({c})" for c in codes]
    )
    for i, ci in enumerate(codes):
        row += 1
        ws.cell(row=row, column=1, value=f"{names.get(ci, ci)} ({ci})")
        for j, cj in enumerate(codes):
            cell = ws.cell(row=row, column=2 + j)
            if j == i:
                cell.value = 1.00
                cell.number_format = "0.00"
                cell.fill = _FILL_DIAG
                cell.font = _FONT_LIGHT
            elif j < i:
                r_val = matrix[i][j]
                p_val = (
                    p_values[i][j]
                    if i < len(p_values) and j < len(p_values[i])
                    else None
                )
                if r_val is None:
                    cell.value = "N/A"
                    cell.fill, cell.font = _FILL_NA, _FONT_GREY
                else:
                    cell.value = round(float(r_val), 2)
                    cell.number_format = "0.00"
                    cell.fill, cell.font = _cell_style(float(r_val), p_val)
            # j > i（上三角）留空

    # ── 2. 配对明细（按 |r| 降序，与 HTML 一致） ──
    if pairs:
        row += 2
        row = write_title_row(ws, row, "配对明细（按 |r| 降序）", ncols=_ncols)
        row = write_header_row(ws, row, _PAIRS_HEADERS)
        for idx, p in enumerate(pairs, start=1):
            _sig = bool(p.get("significant"))
            row_data = [
                idx,
                f"{p.get('name_a', p.get('code_a', ''))} ({p.get('code_a', '')})",
                f"{p.get('name_b', p.get('code_b', ''))} ({p.get('code_b', '')})",
                p.get("pearson", 0.0),
                p.get("p_value", 1.0),
                p.get("samples", 0),
                "✅ 显著" if _sig else "— 不显著",
            ]
            write_data_row(ws, row, row_data, formats=[None, None, None, "0.00", "0.0000", None, None])
            _corr_cell = ws.cell(row=row, column=4)
            _r = float(p.get("pearson", 0.0))
            if _r >= 0.3:
                _corr_cell.font = _FONT_POS_COLOR
            elif _r <= -0.3:
                _corr_cell.font = _FONT_NEG_COLOR
            else:
                _corr_cell.font = _FONT_GREY
            row += 1

    # ── 3. 说明区 ──
    row += 2
    row = write_title_row(ws, row, "说明", ncols=_ncols)
    notes = [
        f"计算窗口：最近 {correlation_data.get('window', 0)} 个交易日重叠样本；有效样本：{correlation_data.get('sample_count', 0)} 期",
        "相关系数 r = 两品种日收益率的 Pearson 相关系数；显著列为 95% 双尾 t 检验结果（p < 0.05）",
        "红=正相关（同向波动，伪分散风险），蓝=负相关（反向对冲），白=不显著，灰=重叠样本不足",
    ]
    insufficient = correlation_data.get("insufficient_codes") or []
    if insufficient:
        notes.append(f"下列品种重叠样本不足窗口期，相关性格标为 N/A（灰色）：{'、'.join(insufficient)}")
    for n in notes:
        write_data_row(ws, row, [n] + [""] * (_ncols - 1))
        row += 1

    freeze_header(ws, row=2)
    auto_width(ws, min_width=10, max_width=30)
    logger.info("持仓相关性矩阵页签写入完成: %d 个品种, %d 对配对", len(codes), len(pairs))
