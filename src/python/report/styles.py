"""Excel 样式常量与辅助函数。

盈亏配色规则：
  - 正数 → 红色字体（A 股红色代表上涨）
  - 负数 → 绿色字体（A 股绿色代表下跌）
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── 颜色 ──────────────────────────────────────────────────

RED_FONT = Font(color="CC0000")  # 正数 / 优秀 → 红色
GREEN_FONT = Font(color="009900")  # 负数 / 偏差 → 绿色
DARK_GREEN_FONT = Font(color="006400")  # 较差 → 深绿色
BLUE_FONT = Font(color="0066CC")  # 稳定 → 蓝色
NORMAL_FONT = Font(size=10)
CONTENT_FONT = Font(size=11, color="000000")  # LLM 内容区正文
BOLD_FONT = Font(size=10, bold=True)
TITLE_FONT = Font(size=14, bold=True, color="FFFFFF")

# ── 填充 ──────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
TITLE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# ── 对齐 ──────────────────────────────────────────────────

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# ── 边线 ──────────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── 数字格式 ──────────────────────────────────────────────

FMT_MONEY = "#,##0.00"  # 金额：千分位带两位小数
FMT_SHARES = "#,##0.00"  # 份额
FMT_PERCENT = "0.00%"  # 百分比
FMT_PRICE = "#,##0.0000"  # 价格保留四位小数

# ── 对特定值应用颜色 ─────────────────────────────────────


def profit_font(value: float) -> Font:
    """根据盈亏值返回对应颜色的字体。

    Args:
        value: 盈亏金额或收益率（正数红色，负数绿色）

    Returns:
        红色或绿色字体
    """
    if value > 0:
        return RED_FONT
    if value < 0:
        return GREEN_FONT
    return NORMAL_FONT
