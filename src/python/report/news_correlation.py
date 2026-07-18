"""财经新闻热点与持仓关联分析模块 — 报告增补页签。

从新浪财经/东方财富/财联社三个源获取最新财经新闻，
与持仓名称/代码以及穿透 TOP10 资产名称进行关键词匹配，
按关联度排序输出 TOP N，在 Excel 中以单独页签呈现。

输出列：
  序号 | 新闻标题 | 摘要 | 来源 | 发布时间 | 关联关键词

可选 LLM 增强：为每条新闻做关联度判定，增加"LLM 关联分析"列。
"""

from __future__ import annotations

import logging
import re
from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.python.models import Holding
from src.python.registry import get_llm_module_name
from src.python.report.data_status import STATUS_MESSAGES
from src.python.report.excel_writer import (
    _write_placeholder,
    auto_width,
    freeze_header,
    write_data_row,
    write_header_row,
    write_title_row,
)

logger = logging.getLogger("invest")

_NCOLS = 6
_BASE_HEADERS = ["序号", "新闻标题", "摘要", "来源", "发布时间", "关联关键词"]

# ── 关键词富化 ────────────────────────────────────────────────


# ── 关键词构建辅助函数 ──────────────────────────────────────────

_SUFFIXES = [
    "ETF",
    "联接",
    "A",
    "C",
    "(QDII)",
    "基金",
    "混合",
    "指数",
    "开放",
    "式",
    "发起",
    "LOF",
]


def _extract_terms(name: str) -> list[str]:
    """从名称中提取中文词（去掉后缀后）。"""
    clean = name
    for suffix in _SUFFIXES:
        clean = clean.replace(suffix, "")
    return re.findall(r"[一-鿿]{2,}", clean)


def _index_holdings(lookup: dict, holdings: list[Holding]) -> None:
    """将持仓名称/代码加入关键词查找表。"""
    for h in holdings:
        code = h.code.strip()
        name = h.name.strip()
        if code and code not in lookup:
            lookup[code] = {"type": "holding", "name": name, "code": code}
        for t in _extract_terms(name):
            if t not in lookup:
                lookup[t] = {"type": "holding", "name": name, "code": code}
        if "ETF" in name:
            core = name.replace("ETF", "").strip()
            for t in re.findall(r"[一-鿿]{2,}", core):
                if t not in lookup:
                    lookup[t] = {"type": "holding", "name": name, "code": code}


def _index_penetrated_assets(lookup: dict, penetrated_assets: list[dict]) -> None:
    """将穿透资产名称/代码加入关键词查找表。"""
    for asset in penetrated_assets:
        asset_name = (asset.get("name") or "").strip()
        for ac in asset.get("codes") or []:
            ac_stripped = ac.strip()
            if ac_stripped and ac_stripped not in lookup:
                lookup[ac_stripped] = {"type": "penetration", "name": asset_name, "code": ac_stripped}
        if asset_name:
            for t in _extract_terms(asset_name):
                if t not in lookup:
                    lookup[t] = {"type": "penetration", "name": asset_name, "code": t}


def _index_industry_concepts(lookup: dict, industry_data: dict[str, dict]) -> None:
    """将行业分类/概念板块加入关键词查找表。"""
    for code, idata in industry_data.items():
        industry_name = (idata.get("industry") or "").strip()
        if industry_name and industry_name not in lookup:
            lookup[industry_name] = {
                "type": "concept",
                "name": industry_name,
                "code": code,
                "source": "industry",
            }
        for cname in idata.get("concepts", []):
            cname = cname.strip()
            if cname and cname not in lookup:
                lookup[cname] = {
                    "type": "concept",
                    "name": cname,
                    "code": code,
                    "source": "concept",
                }


def _enrich_with_industry_data(lookup: dict, industry_data: dict[str, dict]) -> None:
    """为持仓/穿透条目附加行业/概念信息。"""
    for entry in lookup.values():
        if entry.get("type") in ("holding", "penetration"):
            code = entry.get("code", "")
            idata = industry_data.get(code)
            if idata:
                if idata.get("industry"):
                    entry["industry"] = idata["industry"]
                if idata.get("concepts"):
                    entry["concepts_list"] = idata["concepts"]


def _build_keyword_lookup(
    holdings: list[Holding],
    penetrated_assets: list[dict] | None = None,
    industry_data: dict[str, dict] | None = None,
) -> dict[str, dict]:
    """构建关键词→来源正向查找表。

    对每条持仓的名称片段和代码、穿透资产的名称和代码建立索引，
    可选加入行业分类和概念板块关键词（type="concept" 类型），
    用于将 matched_keywords 还原为具体的持仓/穿透/概念/行业标签。

    Args:
        holdings: 持仓列表
        penetrated_assets: 穿透 TOP10 资产列表
        industry_data: 行业/概念数据字典 {code: {industry, concepts, ...}}

    Returns:
        {keyword: {"type": "holding"|"penetration"|"concept", "name": str, "code": str}}
        同一关键词若同时匹配持仓和穿透，持仓优先（先到先得）。
    """
    lookup: dict[str, dict] = {}
    _index_holdings(lookup, holdings)
    if penetrated_assets:
        _index_penetrated_assets(lookup, penetrated_assets)
    if industry_data:
        _index_industry_concepts(lookup, industry_data)
        _enrich_with_industry_data(lookup, industry_data)
    return lookup


def _format_industry_tags(entry: dict) -> str:
    """为持仓/穿透条目生成行业/概念标签后缀。

    Args:
        entry: lookup 条目（可能含 industry / concepts_list 字段）

    Returns:
        标签字符串，如 " [电力·水电]"；无信息时返回 ""
    """
    tags: list[str] = []
    if entry.get("industry"):
        tags.append(entry["industry"])
    concepts_list = entry.get("concepts_list", [])
    if concepts_list:
        tags.extend(concepts_list[:2])  # 最多取前 2 个概念
    if tags:
        return f" [{' · '.join(tags)}]"
    return ""


def _enrich_keywords_for_item(
    item: dict[str, Any],
    keyword_lookup: dict[str, dict],
) -> list[dict]:
    """将单条新闻的 matched_keywords 富化为 enriched_keywords。

    返回列表：
        [{"display": "长江电力(600900)", "type": "holding"}, ...]

    去重规则：同一持仓只出现一次，同一穿透资产同理。
    """
    matched = item.get("matched_keywords", [])
    if not matched:
        return []

    seen: set[tuple[str, str]] = set()
    enriched: list[dict] = []

    for kw in matched:
        entry = keyword_lookup.get(kw)
        if entry and entry["type"] == "holding":
            dedup_key = entry["code"]
            if ("holding", dedup_key) not in seen:
                seen.add(("holding", dedup_key))
                enriched.append(
                    {
                        "display": f"{entry['name']}({entry['code']}){_format_industry_tags(entry)}",
                        "type": "holding",
                    }
                )
        elif entry and entry["type"] == "penetration":
            dedup_key = entry.get("code", entry["name"])
            if ("penetration", dedup_key) not in seen:
                seen.add(("penetration", dedup_key))
                enriched.append(
                    {
                        "display": f"{entry['name']}[穿透]{_format_industry_tags(entry)}",
                        "type": "penetration",
                    }
                )
        elif entry and entry["type"] == "concept":
            dedup_key = entry.get("name", kw)
            if ("concept", dedup_key) not in seen:
                seen.add(("concept", dedup_key))
                enriched.append(
                    {
                        "display": f"{entry['name']}[概念]",
                        "type": "concept",
                    }
                )
        else:
            if ("industry", kw) not in seen:
                seen.add(("industry", kw))
                enriched.append(
                    {
                        "display": kw,
                        "type": "industry",
                    }
                )

    type_order = {"holding": 0, "penetration": 1, "concept": 2, "industry": 3}
    enriched.sort(key=lambda x: type_order.get(x["type"], 99))

    return enriched


def _format_enriched_keywords(enriched: list[dict]) -> str:
    """将 enriched_keywords 列表格式化为单行显示字符串。"""
    if not enriched:
        return ""
    return ", ".join(item["display"] for item in enriched)


# ── 核心函数 ──────────────────────────────────────────────────


def _expand_industry_keywords(
    holdings: list[Holding],
    penetrated_assets: list[dict] | None,
    keywords: list[str],
) -> tuple[list[str], dict[str, dict], set[str]]:
    """扩展关键词：获取行业/概念数据，将行业名和概念名追加为关键词。

    Returns:
        (expanded_keywords, industry_data, lightweight_kw)
        lightweight_kw 为扩展中的行业/概念关键词集合（轻量级关键词），
        用于下游判定是否需要至少 2 个匹配才视为关联。
        行业数据获取失败时返回 (keywords, {}, set())。
    """
    industry_data: dict[str, dict] = {}
    try:
        all_codes: set[str] = set()
        for h in holdings:
            if h.code and h.code.strip():
                all_codes.add(h.code.strip())
        if penetrated_assets:
            for asset in penetrated_assets:
                for ac in asset.get("codes") or []:
                    if ac and ac.strip():
                        all_codes.add(ac.strip())

        if all_codes:
            from src.python.fetcher.industry import batch_fetch_industry_data as _batch_industry

            industry_data = _batch_industry(list(all_codes))
            if industry_data:
                extra_kw: list[str] = []
                for idata in industry_data.values():
                    if idata.get("industry"):
                        extra_kw.append(idata["industry"])
                    extra_kw.extend(cname.strip() for cname in idata.get("concepts", []) if cname.strip())
                if extra_kw:
                    all_kw = list(set(keywords + extra_kw))
                    all_kw.sort(key=lambda x: (-len(x), x))
                    lightweight_kw = set(extra_kw) - set(keywords)
                    new_count = len(lightweight_kw)
                    logger.info("行业/概念关键词扩展: 行业/概念 %d 个 → 共 %d 个", new_count, len(all_kw))
                    return all_kw, industry_data, lightweight_kw
    except Exception as e:
        logger.warning("行业/概念数据获取失败（非关键错误，继续）: %s", e)
    return keywords, industry_data, set()


def _extract_active_sources(news_items: list[dict]) -> list[str]:
    """提取成功访问的数据源标签列表。"""
    active: list[str] = []
    seen: list[str] = []
    for item in news_items:
        label = item.get("_source", "")
        if label and label not in seen:
            seen.append(label)
            active.append(label)
    return active


def _news_source_cb(label: str, count: int, status: str) -> None:
    """回调：各新闻源获取完成后在 TUI 输出状态。"""
    if status == "OK":
        logger.info("新闻源 %s: %d 条", label, count)
    else:
        logger.info("新闻源 %s: %s", label, status)


def _apply_llm_enhancement(
    news_items: list[dict[str, Any]],
    holdings: list[Holding],
    penetrated_assets: list[dict] | None,
    industry_data: dict[str, dict],
    meta: dict,
) -> dict:
    """可选 LLM 增强：对新闻逐条判定关联度。

    注意：此模块不提供 output_brief 配置项。其他 LLM 模块支持精简模式，
    但新闻模块输出严格 JSON 供程序解析，精简会破坏 JSON 结构，故不支持。

    Returns:
        更新后的 meta 字典
    """
    from src.python.config import get_llm_config

    llm_config = get_llm_config()
    enabled_llm = llm_config.get("enabled_llm") if llm_config else None
    llm_enabled = enabled_llm.get("news_correlation", False) if isinstance(enabled_llm, dict) else False

    if not llm_config or not llm_enabled:
        return meta

    # 多链模式：api_key 在 chain 条目内，不在顶级 llm_config
    provider_list = llm_config.get("_provider_list")
    if not provider_list:
        api_key = (llm_config.get("api_key") or "").strip()
        if not api_key:
            logger.warning("enabled_llm.news_correlation 已开启但未配置 api_key，降级为传统关键词匹配分析")
            return meta

    meta["llm_enabled"] = True
    from src.python.llm import run_news_correlation_safe
    from src.python.llm.pricing import estimate_cost

    try:
        news_items[:], cached, token_usage = run_news_correlation_safe(
            news_items,
            holdings,
            penetrated_assets=penetrated_assets,
            industry_data=industry_data,
        )
    except Exception:
        logger.warning("LLM 新闻关联分析失败，降级为传统关键词匹配", exc_info=True)
        cached = False
        token_usage = {}
    meta["llm_cached"] = cached
    meta["token_usage"] = token_usage
    meta["thinking_enabled"] = llm_config.get("thinking_enabled_news_correlation", False)
    if token_usage and token_usage.get("model"):
        meta["cost_estimation"] = estimate_cost(
            token_usage.get("model", ""),
            token_usage.get("input_tokens", 0),
            token_usage.get("output_tokens", 0),
        )
    else:
        meta["cost_estimation"] = "-"
    if cached:
        logger.info(
            "%s（缓存）: 富化 %d 条",
            get_llm_module_name("news_correlation"),
            sum(1 for n in news_items if n.get("llm_analysis")),
        )
    return meta


def _enrich_news_keywords(
    news_items: list[dict[str, Any]],
    holdings: list[Holding],
    penetrated_assets: list[dict] | None,
    industry_data: dict[str, dict],
) -> None:
    """为每条新闻做关键词富化（标注来源为持仓/穿透/概念）。"""
    lookup = _build_keyword_lookup(holdings, penetrated_assets, industry_data=industry_data)
    for item in news_items:
        enriched = _enrich_keywords_for_item(item, lookup)
        item["enriched_keywords"] = enriched if enriched else []


def build_news_data(
    holdings: list[Holding],
    top_n: int = 100,
    penetrated_assets: list[dict] | None = None,
) -> tuple[list[dict[str, Any]], dict]:
    """获取新闻数据并与持仓关联。

    从多个财经新闻源获取最新新闻，
    与持仓名称/代码及穿透 TOP10 资产进行关键词匹配，
    按关联度排序返回 TOP N。

    若 llm_settings.json 中 enabled_llm.news_correlation 为 true，自动启用 LLM 二次分析，
    对新闻逐条判定关联度并给出原因分析，结果写入 llm_analysis 字段。

    Args:
        holdings: 持仓列表
        top_n: 最多返回的关联新闻条数
        penetrated_assets: 穿透 TOP10 资产列表（可选），
            每项含 name 和 codes 字段。传入后新闻关键词
            会额外覆盖穿透到的底层资产。

    Returns:
        (news_data, meta)
        news_data: [{title, intro, url, ctime, media_name, matched_keywords, llm_analysis?}, ...]
        meta: {
            "token_usage": {...},        # LLM token 消耗（启用时）
            "llm_cached": bool,          # LLM 结果是否来自缓存
            "llm_enabled": bool,         # 是否启用了 LLM 分析
        }
        获取失败时 news_data 为 []。
    """
    from src.python.fetcher.news import aggregate_news, build_holding_keywords

    keywords = build_holding_keywords(holdings, penetrated_assets=penetrated_assets)
    logger.info("%s关键词（含穿透）: %s", get_llm_module_name("news_correlation"), keywords)

    keywords, industry_data, lightweight_kw = _expand_industry_keywords(holdings, penetrated_assets, keywords)
    # per_source 取大值保证召回覆盖面，避免去重后候选不足
    per_source = max(500, top_n * 2)
    news_items = aggregate_news(
        keywords,
        top_n=top_n,
        per_source=per_source,
        progress_callback=_news_source_cb,
        lightweight_keywords=lightweight_kw,
    )

    active_sources = _extract_active_sources(news_items)
    meta: dict = {
        "token_usage": {},
        "llm_cached": False,
        "llm_enabled": False,
        "active_sources": active_sources,
        "source_status": {},
    }

    if not news_items:
        # 即使 news_items 为空，也能从 aggregate_news 获取各源状态
        try:
            from src.python.fetcher.news import get_last_source_status as _glss

            meta["source_status"] = _glss()
        except Exception:
            logger.warning("获取新闻源状态失败（news_aggregator.get_last_source_status），不影响核心新闻数据")
        logger.warning("新闻获取失败")
        return news_items, meta

    logger.info(
        "%s完成: 获取 %d 条, 匹配 %d 条",
        get_llm_module_name("news_correlation"),
        len(news_items),
        sum(1 for n in news_items if n.get("matched_keywords")),
    )

    meta = _apply_llm_enhancement(news_items, holdings, penetrated_assets, industry_data, meta)
    _enrich_news_keywords(news_items, holdings, penetrated_assets, industry_data)

    # 补充各源状态（在 aggregate_news 之后获取）
    try:
        from src.python.providers.news_aggregator import get_last_source_status as _glss

        meta["source_status"] = _glss()
    except Exception:
        logger.warning("补充新闻源状态失败（news_aggregator.get_last_source_status），不影响新闻匹配结果")

    return news_items, meta


def _build_news_footer(news_data: list[dict], llm_meta: dict | None, llm_count: int = 0) -> str:
    """构建新闻页签底部说明文本。"""
    if llm_meta and llm_meta.get("llm_enabled"):
        if llm_meta.get("llm_cached"):
            hint = f"共获取 {len(news_data)} 条关联新闻。本次使用LLM缓存，未直接使用LLM服务能力"
            if llm_meta.get("thinking_enabled", False):
                hint += " | Extended Thinking"
            return hint
        parts = [f"共获取 {len(news_data)} 条关联新闻"]
        if llm_count:
            parts.append(f"（其中 LLM 关联分析 {llm_count} 条）")
        parts.append("，关键词匹配基于持仓名称和代码")
        return "".join(parts)
    result = (
        f"共获取 {len(news_data)} 条关联新闻。"
        "基于持仓名称和代码进行关键词匹配。"
        "本次未使用LLM服务能力增强支持，使用传统关键字匹配技术"
    )
    # 追加部分源失败信息
    _source_status = (llm_meta or {}).get("source_status", {})
    _failed_sources = [s["label"] for s in _source_status.values() if not s["success"]]
    if _failed_sources:
        result += f" 以下新闻源不可用：{'、'.join(_failed_sources)}"
    return result


def _write_news_token_footer(ws: Worksheet, row: int, llm_meta: dict | None) -> int:
    """写入 Token 用量行（LLM 启用且非缓存时）。返回写入后的行号。"""
    if not (llm_meta and llm_meta.get("llm_enabled") and not llm_meta.get("llm_cached")):
        return row
    token_usage = llm_meta.get("token_usage") or {}
    if token_usage.get("total_tokens", 0) <= 0:
        return row
    row += 1
    parts = [
        f"模型：{token_usage.get('model', '')}",
        f"Token 用量：输入 {token_usage.get('input_tokens', 0):,} / 输出 {token_usage.get('output_tokens', 0):,} = {token_usage.get('total_tokens', 0):,}",
    ]
    cost_est = llm_meta.get("cost_estimation", "-")
    if cost_est and cost_est != "-":
        parts.append(f"估算费用：{cost_est}")
    if llm_meta.get("thinking_enabled", False):
        parts.append("Extended Thinking")
    write_data_row(ws, row, [" | ".join(parts)])
    return row


def _set_news_column_widths(ws: Worksheet, has_llm: bool) -> None:
    """设置新闻页签列宽。"""
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    if has_llm:
        llm_col = _NCOLS + 1
        ws.column_dimensions[get_column_letter(llm_col)].width = 30


# ── LLM 关联分析列着色 ────────────────────────────────────

_BULLISH_RED = "CC0000"
_BEARISH_GREEN = "009900"


def _colorize_llm_cell(ws: Worksheet, row: int, col: int, text: str) -> None:
    """根据 llm_analysis 文本中的利好/利空标记为单元格字体着色。

    [利好] → 红色 (#CC0000)，[利空] → 绿色 (#009900)。
    中性或无标记 → 不染色（保持默认黑色）。
    """
    if "[利好]" in text:
        ws.cell(row=row, column=col).font = Font(color=_BULLISH_RED)
    elif "[利空]" in text:
        ws.cell(row=row, column=col).font = Font(color=_BEARISH_GREEN)


def write_news_sheet(
    ws: Worksheet,
    news_data: list[dict[str, Any]],
    llm_meta: dict | None = None,
) -> None:
    """写入财经新闻热点与持仓关联分析页签。

    这是 Excel 的增补页签（仅在 N 选项时生成）。
    若有 LLM 分析数据，自动增加 "LLM 关联分析" 列。

    Args:
        ws: 目标工作表
        news_data: build_news_data() 返回的数据
        llm_meta: LLM 元数据，含 token_usage / llm_cached / llm_enabled
    """
    has_llm = any(isinstance(item, dict) and item.get("llm_analysis") for item in news_data)
    llm_count = sum(1 for n in news_data if n.get("llm_analysis"))
    ncols = _NCOLS + (1 if has_llm else 0)
    headers = _BASE_HEADERS + (["LLM 关联分析"] if has_llm else [])

    row = write_title_row(ws, 1, get_llm_module_name("news_correlation"), ncols)
    row = write_header_row(ws, row, headers)

    if not news_data:
        # 全源失败 → 写占位
        _source_status = (llm_meta or {}).get("source_status", {})
        _all_failed = _source_status and all(not s["success"] for s in _source_status.values())
        if _all_failed:
            _write_placeholder(ws, STATUS_MESSAGES["news_all_failed"], row=row, max_cols=ncols)
            logger.warning("%s：所有新闻源均获取失败，写入占位", get_llm_module_name("news_correlation"))
        else:
            write_data_row(ws, row, ["暂无关联新闻"])
            logger.info("%s：无数据", get_llm_module_name("news_correlation"))
        freeze_header(ws, 2)
        auto_width(ws)
        return

    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    llm_col = _NCOLS + 1 if has_llm else None
    for idx, item in enumerate(news_data, 1):
        enriched = item.get("enriched_keywords", [])
        keywords_str = _format_enriched_keywords(enriched) if enriched else ", ".join(item.get("matched_keywords", []))
        vals = [
            idx,
            item.get("title", ""),
            item.get("intro", ""),
            item.get("media_name", ""),
            item.get("ctime", ""),
            keywords_str,
        ]
        if has_llm:
            llm_text = item.get("llm_analysis", "")
            vals.append(llm_text)
        write_data_row(ws, row, vals)
        ws.cell(row=row, column=2).alignment = wrap_left
        ws.cell(row=row, column=3).alignment = wrap_left
        if has_llm and llm_text:
            _colorize_llm_cell(ws, row, llm_col, llm_text)
        row += 1

    row += 1
    write_data_row(ws, row, [_build_news_footer(news_data, llm_meta, llm_count)])
    row = _write_news_token_footer(ws, row, llm_meta)

    freeze_header(ws, 2)
    auto_width(ws)
    _set_news_column_widths(ws, has_llm)
    llm_info = f"，其中 LLM 分析 {llm_count} 条" if llm_count else ""
    logger.info("%s写入完成%s，共 %d 条", get_llm_module_name("news_correlation"), llm_info, len(news_data))
