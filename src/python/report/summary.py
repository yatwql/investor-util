"""投资分析汇总模块。

显示当前日期、持仓概况（分类统计+价格更新状态）、盈亏汇总、市场指数。"""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.python.cache import get_cache_age_by_data_type, get_ttl
from src.python.core.registry import get_report_sheet_name
from src.python.report.data_status import (
    STATUS_MESSAGES,
    DataStatus,
    DataStatusItem,
    get_tracker,
)
from src.python.report.excel_writer import (
    _write_data_status_foot,
    auto_width,
    freeze_header,
    write_data_row,
    write_header_row,
    write_title_row,
)
from src.python.report.market_value import get_last_trading_day
from src.python.report.styles import profit_font

logger = logging.getLogger("invest")

# 指数涨跌颜色
_INDEX_UP_FONT = Font(size=10, bold=True, color="CC0000")  # 涨→红
_INDEX_DOWN_FONT = Font(size=10, bold=True, color="009900")  # 跌→绿

# 单元格对齐
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# 模块级降级阈值控制器（单例工厂共享，统一管理）
_tracker = get_tracker()

_NCOLS = 8
_HEADERS = ["指标", "数值"]

# 样式
_SECTION_FONT = Font(size=11, bold=True, color="2E75B6")  # 章节标题：深蓝
_BLUE_FONT = Font(size=10, bold=True, color="2E75B6")  # 更新完成：蓝色
_RED_FONT = Font(size=10, bold=True, color="CC0000")  # 未完成：红色
_NORMAL_FONT = Font(size=10)
_STAT_TIME_FONT = Font(size=10, bold=True)  # 统计时间值：加粗
_TRADING_DAY_FONT = Font(size=12, bold=True, color="2E75B6")  # 所属交易日值：加粗+加大+蓝色


def _write_section(ws, row: int, label: str) -> int:
    """写入章节标题行（如 【持仓概况】），占 2 列合并。"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NCOLS)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = _SECTION_FONT
    cell.alignment = _CENTER_ALIGN
    return row + 1


def _write_kv_row(ws, row: int, key: str, value: Any) -> int:
    """写入一个指标=数值行（标准 2 列）。"""
    write_data_row(ws, row, [key, value])
    return row + 1


def _write_kv_row_colored(ws, row: int, key: str, value: Any, font: Font) -> int:
    """写入带颜色的指标行。"""
    write_data_row(ws, row, [key, value])
    for col in (1, 2):
        ws.cell(row=row, column=col).font = font
    return row + 1


def _write_index_row(ws, row: int, name: str, price: float, change_pct: float) -> int:
    """写入指数行，涨跌幅自动着色（涨红跌绿）。

    Args:
        ws: 工作表
        row: 当前行号
        name: 指数中文名
        price: 当前点数
        change_pct: 涨跌幅百分数（如 0.35 表示 +0.35%）

    Returns:
        下一行号
    """
    sign = "+" if change_pct >= 0 else ""
    label = f"  {name}"
    value = f"{price:.2f}  ({sign}{change_pct:.2f}%)"
    write_data_row(ws, row, [label, value])
    font = _INDEX_UP_FONT if change_pct >= 0 else _INDEX_DOWN_FONT
    ws.cell(row=row, column=2).font = font
    return row + 1


def _write_blanks(_ws, row: int, n: int = 1) -> int:
    """写入 n 行空白。"""
    return row + n


def _write_basic_info(ws: Worksheet, row: int, now: datetime | None = None) -> int:
    """写入基本信息和所属交易日。

    Args:
        ws: 工作表
        row: 当前行号
        now: 当前时间（允许外部注入便于测试）
    """
    now = now or datetime.now()
    row = _write_kv_row(ws, row, "统计时间", now.strftime("%Y-%m-%d %H:%M:%S"))
    # 统计时间值加粗
    ws.cell(row=row - 1, column=2).font = _STAT_TIME_FONT
    row = _write_kv_row(ws, row, "所属交易日", get_last_trading_day())
    # 所属交易日值加粗+加大+蓝色
    ws.cell(row=row - 1, column=2).font = _TRADING_DAY_FONT
    row = _write_blanks(ws, row)
    return row


def _write_holdings_overview(
    ws: Worksheet,
    row: int,
    categories: dict[str, list] | None,
    update_status: tuple[int, int, bool] | None,
) -> int:
    """写入持仓概况分类计数和价格更新状态。"""
    row = _write_section(ws, row, "【持仓概况】")
    total_count = 0
    if categories:
        for cat_label in ("场内股票", "场内ETF", "国内场外", "QDII"):
            count = len(categories.get(cat_label, []))
            total_count += count
            row = _write_kv_row(ws, row, f"  {cat_label}", count)
        row = _write_kv_row(ws, row, "持仓总数", total_count)
    else:
        row = _write_kv_row(ws, row, "持仓总数", "--")

    if update_status:
        updated, total, all_done = update_status
        if total > 0:
            status_text = f"{updated}/{total}  (全部已更新)" if all_done else f"{updated}/{total}  (尚有缺失)"
            status_font = _BLUE_FONT if all_done else _RED_FONT
        else:
            status_text = "--"
            status_font = _NORMAL_FONT
        row = _write_kv_row_colored(ws, row, "价格更新状态", status_text, status_font)
    else:
        row = _write_kv_row(ws, row, "价格更新状态", "--")
    row = _write_blanks(ws, row)
    return row


def _write_profit_summary(
    ws: Worksheet,
    row: int,
    total_mv: float,
    total_cost: float,
    total_profit: float,
    today_profit: float,
    fund_flow_data: dict | None = None,
) -> int:
    """写入盈亏汇总（数值以原始小数/金额写入，由 Excel 数字格式控制显示）。

    Args:
        fund_flow_data: 成本流水数据契约。非 None 时在汇总末尾追加
            「资金加权收益率 (XIRR)」行（无可用现金流时写占位文本）；
            None 时保持既有输出（report_submodules.cost_lots 关闭）。
    """
    # 检测行情数据是否全部不可用 — 有持仓成本但市值全零
    _data_unavailable = total_mv == 0 and total_cost > 0
    if _data_unavailable:
        _WARN_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        _WARN_FONT = Font(size=10, bold=True, color="CC0000")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NCOLS)
        cell = ws.cell(
            row=row,
            column=1,
            value="⚠ 行情数据全部不可用（非交易时段/网络异常/API限速），以下市值为 0，请于交易时段重新生成",
        )
        cell.font = _WARN_FONT
        cell.fill = _WARN_FILL
        row += 1
        row = _write_blanks(ws, row)

    profit_rate = (total_profit / total_cost) if total_cost > 0 else 0.0  # 小数，0.00% 格式
    denominator = total_cost + total_profit - today_profit
    today_rate = (today_profit / denominator) if denominator > 0 else 0.0  # 小数，0.00% 格式

    from src.python.report.styles import FMT_MONEY, FMT_PERCENT

    row = _write_section(ws, row, "【盈亏汇总】")
    summary_data: list[tuple[str, float, str]] = [
        ("总市值 (元)", total_mv, FMT_MONEY),
        ("总成本 (元)", total_cost, FMT_MONEY),
        ("总盈亏 (元)", total_profit, FMT_MONEY),
        ("总收益率", profit_rate, FMT_PERCENT),
        ("本日盈亏 (元)", today_profit, FMT_MONEY),
        ("本日收益率", today_rate, FMT_PERCENT),
    ]
    for label, val, fmt in summary_data:
        write_data_row(ws, row, [label, val])
        ws.cell(row=row, column=2).number_format = fmt
        if "盈亏" in label and isinstance(val, (int, float)) or "收益率" in label and isinstance(val, (int, float)):
            ws.cell(row=row, column=2).font = profit_font(val)
        row += 1

    # 资金加权收益率 (XIRR)：成本流水子模块开启时展示（无可用现金流写占位）
    if fund_flow_data is not None:
        xirr = fund_flow_data.get("xirr")
        if xirr and xirr.get("rate") is not None:
            rate = xirr["rate"]
            write_data_row(ws, row, ["资金加权收益率 (XIRR)", rate])
            ws.cell(row=row, column=2).number_format = FMT_PERCENT
            ws.cell(row=row, column=2).font = profit_font(rate)
        else:
            write_data_row(ws, row, ["资金加权收益率 (XIRR)", "未录入流水/无法计算"])
        row += 1

    row = _write_blanks(ws, row)
    return row


def _write_a_share_indices(ws: Worksheet, row: int, a_indices: dict[str, dict[str, Any]] | None) -> int:
    """写入 A 股指数（本日 + 上日）。"""
    if not a_indices:
        return _write_kv_row(ws, row, "── A股指数 ──", "暂无数据")

    row = _write_kv_row(ws, row, "── A股指数（本日）──", "")
    a_list = [
        ("sh000001", "上证指数"),
        ("sz399001", "深证成指"),
        ("sh000300", "沪深300"),
        ("sh000688", "科创板50"),
        ("sz399006", "创业板指"),
    ]
    for code, cname in a_list:
        idx = a_indices.get(code)
        if idx and idx.get("price", 0) > 0:
            row = _write_index_row(ws, row, cname, idx["price"], idx.get("change_pct", 0))
        else:
            row = _write_kv_row(ws, row, f"  {cname}", "--")

    row = _write_kv_row(ws, row, "── A股指数（上日）──", "")
    for code, cname in a_list:
        idx = a_indices.get(code)
        yclose = idx.get("yesterday_close", 0) if idx else 0
        if yclose > 0:
            row = _write_kv_row(ws, row, f"  {cname}", f"{yclose:.2f}")
        else:
            row = _write_kv_row(ws, row, f"  {cname}", "--")
    return row


def _write_us_indices(ws: Worksheet, row: int, us_indices: dict[str, dict[str, Any]] | None) -> int:
    """写入美股指数（最新 + 上日）。"""
    if not us_indices:
        return _write_kv_row(ws, row, "── 美股指数 ──", "暂无数据")

    row = _write_kv_row(ws, row, "── 美股指数（最新）──", "")
    us_list = [
        ("gb_dji", "道琼斯"),
        ("gb_ixic", "纳斯达克"),
        ("gb_inx", "标普500"),
    ]
    for code, cname in us_list:
        idx = us_indices.get(code)
        if idx and idx.get("price", 0) > 0:
            row = _write_index_row(ws, row, cname, idx["price"], idx.get("change_pct", 0))
        else:
            row = _write_kv_row(ws, row, f"  {cname}", "--")

    row = _write_kv_row(ws, row, "── 美股指数（上日）──", "")
    for code, cname in us_list:
        idx = us_indices.get(code)
        yclose = idx.get("yesterday_close", 0) if idx else 0
        if yclose > 0:
            row = _write_kv_row(ws, row, f"  {cname}", f"{yclose:.2f}")
        else:
            row = _write_kv_row(ws, row, f"  {cname}", "--")
    return row


def _write_market_temperature(ws: Worksheet, row: int, temperature: dict | None) -> int:
    """写入市场温度刻度行（三因子合成温度计，无仓位指令）。

    Args:
        ws: 目标工作表
        row: 起始行
        temperature: 市场温度数据契约（market_temperature_data）。None/不可用时
            写占位文本 + 免责声明；可用时写分数刻度 + 三因子明细。

    Returns:
        写入后的下一行行号。
    """
    row = _write_section(ws, row, "【市场温度】")
    disclaimer = (temperature or {}).get(
        "disclaimer"
    ) or "市场温度为价格分位、均线偏离与波动率三因子合成的信号，仅供参考，不构成任何仓位建议"
    if not temperature or not temperature.get("available"):
        write_data_row(ws, row, ["市场温度", "--（数据不足，暂不显示）"])
        row += 1
        row = _write_kv_row(ws, row, "注", disclaimer)
        return row

    score = temperature.get("score")
    tier = temperature.get("tier") or "合理"
    index_name = temperature.get("index_name") or "沪深300"
    pct = temperature.get("price_percentile")
    dev = temperature.get("ma_deviation")
    vol = temperature.get("volatility")
    if score is not None:
        write_data_row(ws, row, ["市场温度", f"{score:.0f} / 100（{tier}）"])
    else:
        write_data_row(ws, row, ["市场温度", f"--（{tier}）"])
    row += 1
    if all(v is not None for v in (pct, dev, vol)):
        # 分位为 0~100，均线偏离/波动率为小数比例（0.032=3.2%），转百分数展示
        row = _write_kv_row(
            ws,
            row,
            f"三因子（{index_name}）",
            f"价格分位 {pct:.1f}% · 20日均线偏离 {dev * 100:+.1f}% · 年化波动率 {vol * 100:.1f}%",
        )
    else:
        row = _write_kv_row(ws, row, f"三因子（{index_name}）", "因子数据不完整")
    row = _write_kv_row(ws, row, "注", disclaimer)
    return row


def build_index_data_status(
    a_indices: dict[str, dict[str, Any]] | None,
    us_indices: dict[str, dict[str, Any]] | None,
) -> DataStatus:
    """检查指数数据来源，构建数据源状态字典。

    腾讯为 A 股主链路 → 新浪备用为降级。
    新浪为美股主链路 → 腾讯备用为降级。
    过期缓存同样视为降级。

    Returns:
        数据源状态字典（可能为空 = 全部正常）
    """
    status: DataStatus = {}

    # A 股指数（T2）
    if a_indices:
        has_degraded = any(idx.get("_source") in ("sina", "stale_cache") for idx in a_indices.values())
        if has_degraded:
            cache_age = get_cache_age_by_data_type("index", "sh000001")
            ttl = get_ttl("index")
            degraded, _, _ = _tracker.record(
                "index_a",
                "T2",
                success=False,
                failure_type="unreachable",
                cache_age_hours=cache_age / 3600 if cache_age else None,
                cache_ttl_hours=ttl / 3600 if ttl else 24,
            )
            if degraded:
                status["index_a"] = DataStatusItem(
                    available=False,
                    tier="T2",
                    message=STATUS_MESSAGES["index_degraded"],
                )

    # 美股指数（T2）
    if us_indices:
        has_degraded = any(idx.get("_source") in ("tencent", "stale_cache") for idx in us_indices.values())
        if has_degraded:
            cache_age = get_cache_age_by_data_type("index", "gb_dji")
            ttl = get_ttl("index")
            degraded, _, _ = _tracker.record(
                "index_us",
                "T2",
                success=False,
                failure_type="unreachable",
                cache_age_hours=cache_age / 3600 if cache_age else None,
                cache_ttl_hours=ttl / 3600 if ttl else 24,
            )
            if degraded:
                status["index_us"] = DataStatusItem(
                    available=False,
                    tier="T2",
                    message=STATUS_MESSAGES["index_degraded"],
                )

    return status


def write_summary_sheet(
    ws: Worksheet,
    total_mv: float,
    total_cost: float,
    total_profit: float,
    today_profit: float,
    categories: dict[str, list] | None = None,
    update_status: tuple[int, int, bool] | None = None,
    a_indices: dict[str, dict[str, Any]] | None = None,
    us_indices: dict[str, dict[str, Any]] | None = None,
    fund_flow_data: dict | None = None,
    market_temperature_data: dict | None = None,
) -> None:
    """写入投资分析汇总。

    Args:
        ws: 目标工作表
        total_mv: 总市值
        total_cost: 总成本
        total_profit: 总盈亏
        today_profit: 本日总盈亏
        categories: 分类结果 {类型: [Holdings]}，来自 classify_holdings()
        update_status: (已更新数, 总数, 是否全部更新)，来自 price_update_status()
        a_indices: A 股指数 {代码: {name, price, yesterday_close, change_pct}}
        us_indices: 美股指数 {代码: {name, price, yesterday_close, change_pct}}
        fund_flow_data: 成本流水数据契约（非 None 时盈亏汇总追加 XIRR 行；
            None 时保持既有输出，report_submodules.cost_lots 关闭）
        market_temperature_data: 市场温度数据契约（非 None 时在「市场指数」后
            追加「市场温度」刻度行；None 时保持既有输出，report_submodules.market_temperature 关闭）
    """
    row = write_title_row(ws, 1, get_report_sheet_name("summary"), _NCOLS)
    row = write_header_row(ws, row, _HEADERS)

    row = _write_basic_info(ws, row)
    row = _write_holdings_overview(ws, row, categories, update_status)
    row = _write_profit_summary(ws, row, total_mv, total_cost, total_profit, today_profit, fund_flow_data)
    row = _write_blanks(ws, row)

    # ── 市场指数 ──
    row = _write_section(ws, row, "【市场指数】")
    row = _write_a_share_indices(ws, row, a_indices)
    row = _write_blanks(ws, row)
    row = _write_us_indices(ws, row, us_indices)

    # 市场温度刻度行（report_submodules.market_temperature 子模块，非 None 才渲染）
    if market_temperature_data is not None:
        row = _write_blanks(ws, row)
        row = _write_market_temperature(ws, row, market_temperature_data)

    # 指数数据源状态
    data_status = build_index_data_status(a_indices, us_indices)
    _write_data_status_foot(ws, data_status, start_row=row)
    freeze_header(ws, 2)
    auto_width(ws)
    logger.info("投资分析汇总写入完成，共 %d 行", row)


# ── re-export: write_llm_usage_sheet ──
from src.python.report.summary_llm_usage import write_llm_usage_sheet  # noqa: F401, E402
