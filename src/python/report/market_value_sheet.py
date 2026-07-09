"""行情市值核算 — Excel 写入层。

职责：将 market_value.DetailRow 数据写入 Excel 工作表。
仅依赖 DetailRow 类型，不包含行情获取/计算逻辑。

依赖方向：
  market_value.py ← market_value_sheet.py（仅 DetailRow 类型导入）
  excel_generator.py → market_value.py + market_value_sheet.py（编排器分别导入）
"""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.python.code_utils import is_qdii_extended
from src.python.registry import get_report_sheet_name
from src.python.report.excel_writer import (
    auto_width,
    freeze_header,
    write_data_row,
    write_header_row,
    write_subtotal_row,
    write_title_row,
    write_total_row,
)
from src.python.report.market_value import DetailRow
from src.python.report.styles import BLUE_FONT, FMT_MONEY, FMT_PERCENT, FMT_PRICE, FMT_SHARES, profit_font

import logging
logger = logging.getLogger("invest")

__all__ = [
    "write_market_value_sheet",
    "_detail_to_row_values",
    "_apply_profit_colors",
    "_apply_price_type_colors",
]

# 15 列表头（R-197 拆分后归 sheet 层维护）
_HEADERS = [
    "账户", "名称", "代码", "最新价", "净值日期", "昨日价",
    "取价方式", "溢价率", "份额", "市值", "成本",
    "盈亏", "收益率", "本日盈亏", "取价渠道",
]
_NCOLS = len(_HEADERS)

_PRICE_TYPE_COL = 7
_NAME_COL = 2


def _detail_to_row_values(d: DetailRow) -> list[Any]:
    """将 DetailRow 转为 Excel 行值列表。"""
    return [
        d.account, d.name, d.code, d.price, d.nav_date,
        d.yesterday_close, d.price_type, d.premium,
        d.shares, d.market_value, d.cost, d.profit,
        d.profit_rate, d.today_profit, d.source,
    ]


def _num_formats() -> list[str | None]:
    """每列的 Excel 数字格式。"""
    return [
        "",           # 1  账户
        "",           # 2  名称
        "",           # 3  代码
        FMT_PRICE,    # 4  最新价
        "",           # 5  净值日期
        FMT_PRICE,    # 6  昨日价
        "",           # 7  取价方式
        "",           # 8  溢价率
        FMT_SHARES,   # 9  份额
        FMT_MONEY,    # 10 市值
        FMT_MONEY,    # 11 成本
        FMT_MONEY,    # 12 盈亏
        FMT_PERCENT,  # 13 收益率
        FMT_MONEY,    # 14 本日盈亏
        "",           # 15 取价渠道
    ]


def _apply_profit_colors(ws, start_row: int, end_row: int,
                         profit_col: int, rate_col: int, today_col: int) -> None:
    """对盈亏列（12）、收益率列（13）、本日盈亏列（14）着色。"""
    for r in range(start_row, end_row + 1):
        for col in (profit_col, today_col):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                cell.font = profit_font(cell.value)
        rate_cell = ws.cell(row=r, column=rate_col)
        if isinstance(rate_cell.value, float):
            rate_cell.font = profit_font(rate_cell.value)


def _apply_price_type_colors(ws, start_row: int, end_row: int) -> None:
    """对取价方式列（第 7 列）着色：蓝色代表价格来源可靠/时效性高。"""
    for r in range(start_row, end_row + 1):
        cell = ws.cell(row=r, column=_PRICE_TYPE_COL)
        val = str(cell.value) if cell.value else ""
        if val in ("场内收盘价(T)", "场内午市收盘(T)", "官方净值(T)"):
            cell.font = BLUE_FONT
        elif val == "官方净值(T-1)":
            name_cell = ws.cell(row=r, column=_NAME_COL)
            name = str(name_cell.value) if name_cell.value else ""
            if is_qdii_extended(name):
                cell.font = BLUE_FONT


def _write_account_groupings(
    ws, details: list[DetailRow], data_start: int,
) -> tuple[float, float, float, float, int]:
    """按账户分组写入明细行和小计，返回汇总数据及最终行号。

    Returns:
        (grand_mv, grand_cost, grand_profit, grand_today, final_row)
    """
    accounts: dict[str, list[DetailRow]] = {}
    for d in details:
        accounts.setdefault(d.account, []).append(d)

    row = data_start
    grand_mv = grand_cost = grand_profit = grand_today = 0.0

    for acc_name, acc_details in accounts.items():
        for d in acc_details:
            vals = _detail_to_row_values(d)
            write_data_row(ws, row, vals, _num_formats())
            row += 1

        acc_mv = sum(d.market_value for d in acc_details)
        acc_cost = sum(d.cost for d in acc_details)
        acc_profit = sum(d.profit for d in acc_details)
        acc_today = sum(d.today_profit for d in acc_details)
        acc_rate = acc_profit / acc_cost if acc_cost > 0 else 0.0

        subtotal_vals = [
            f"{acc_name} 小计", "", "", "", "", "", "", "",
            sum(d.shares for d in acc_details),
            acc_mv, acc_cost, acc_profit, acc_rate, acc_today, "",
        ]
        write_subtotal_row(ws, row, f"{acc_name} 小计",
                           subtotal_vals[1:], _NCOLS, _num_formats())
        row += 1

        grand_mv += acc_mv
        grand_cost += acc_cost
        grand_profit += acc_profit
        grand_today += acc_today

    return grand_mv, grand_cost, grand_profit, grand_today, row


def write_market_value_sheet(ws: Worksheet, holdings: list,
                             today_str: str = "",
                             details: list[DetailRow] | None = None) -> tuple[float, float, float, float, list[DetailRow]]:
    """写入市值核算明细表，返回汇总数据供汇总页签使用。

    Args:
        ws: 目标工作表
        holdings: 持仓列表
        today_str: 日期字符串（YYYY-MM-DD），默认当天
        details: 预计算明细行（必须传入，由编排器预计算）

    Returns:
        (总市值, 总成本, 总盈亏, 本日总盈亏, 明细行列表)
    """
    row = write_title_row(ws, 1, get_report_sheet_name('market_value'), _NCOLS)
    row = write_header_row(ws, row, _HEADERS)
    data_start = row

    # 若所有行情数据全零，写一行醒目提示
    _all_zero = all(d.price == 0 for d in details) if details else False
    if _all_zero:
        _WARN_FONT = Font(size=10, bold=True, color="CC0000")
        ws.cell(row=row, column=1).font = _WARN_FONT
        ws.cell(row=row, column=1, value="⚠ 行情数据全部不可用（非交易时段/网络异常），以下市值/盈亏均为占位 —")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        row += 1

    # 按账户分组写入明细 + 小计
    grand_mv, grand_cost, grand_profit, grand_today, row = _write_account_groupings(
        ws, details or [], data_start)

    # 总计
    grand_rate = grand_profit / grand_cost if grand_cost > 0 else 0.0
    total_vals = [
        "总计", "", "", "", "", "", "", "",
        sum(d.shares for d in details) if details else 0,
        grand_mv, grand_cost, grand_profit, grand_rate, grand_today, "",
    ]
    write_total_row(ws, row, "总计", total_vals[1:], _NCOLS, _num_formats())

    # 对盈亏列着色
    _apply_profit_colors(ws, data_start, row, profit_col=12, rate_col=13, today_col=14)

    # 对取价方式列着色
    _apply_price_type_colors(ws, data_start, row)

    freeze_header(ws, 2)
    auto_width(ws)

    logger.info("%s写入完成，共 %d 个账户，%d 条持仓",
                get_report_sheet_name('market_value'),
                len(set(d.account for d in (details or []))), len(details or []))

    _details = details or []
    return grand_mv, grand_cost, grand_profit, grand_today, _details
