"""调仓 What-if 模拟报告输出。

编排双产物输出：
  - Excel 调仓模拟工作簿（调仓摘要 / 分类配置对比 / 持仓变动明细
    + 指定生效日时的「时序回测」页签）
  - HTML 双栏对比页（含资产配置对比环形图 + 回测折线图，复用 Chart.js 本地 bundle）

报告按主报告归档惯例输出到 output_dir（与主报告分离）：
  - 最新版固定名 `调仓模拟.xlsx` / `调仓模拟.html`（每次覆盖为最新对比）
  - 归档版 `YYYYMMDD/调仓模拟-YYYYMMDD-HHMMSS.xlsx` / `.html`（日期子目录）
并复制 Chart.js 前端资产到同目录（离线自包含，R21 约束）；
超过 180 天的归档目录自动清理。
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Any

from src.python.report.excel_writer import _cleanup_old_archives, _ensure_reports_dir
from src.python.report.html_writer import _copy_js_assets
from src.python.report.whatif_sheet import (
    write_whatif_backtest_sheet,
    write_whatif_category_sheet,
    write_whatif_changes_sheet,
    write_whatif_summary_sheet,
)

logger = logging.getLogger("invest")


def write_whatif_excel(whatif_data: dict[str, Any], output_dir: str = "reports") -> str:
    """输出调仓模拟 Excel 工作簿（最新版固定名 + 日期目录归档版），返回最新文件路径。

    归档格式对齐主报告：`调仓模拟.xlsx`（最新版，覆盖）+ `YYYYMMDD/调仓模拟-YYYYMMDD-HHMMSS.xlsx`（归档版）。

    Args:
        whatif_data: C19 契约 dict
        output_dir: 输出目录

    Returns:
        最新版 Excel 绝对路径
    """
    from openpyxl import Workbook

    _ensure_reports_dir(output_dir)
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "调仓摘要"
    write_whatif_summary_sheet(ws_sum, whatif_data)
    ws_cat = wb.create_sheet("分类配置对比")
    write_whatif_category_sheet(ws_cat, whatif_data)
    ws_chg = wb.create_sheet("持仓变动明细")
    write_whatif_changes_sheet(ws_chg, whatif_data)
    ws_bt = wb.create_sheet("时序回测")
    write_whatif_backtest_sheet(ws_bt, whatif_data)

    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%H%M%S")
    latest = os.path.join(output_dir, "调仓模拟.xlsx")
    archive = os.path.join(output_dir, date_str, f"调仓模拟-{date_str}-{time_str}.xlsx")
    try:
        wb.save(latest)
    except PermissionError:
        logger.error("文件被占用: %s", latest)
        raise
    try:
        wb.save(archive)
    except (PermissionError, OSError) as e:
        logger.warning("存档 Excel 写入失败（非关键）: %s", e)
    _cleanup_old_archives(output_dir)
    logger.info("调仓模拟 Excel 已保存: %s", latest)
    return os.path.abspath(latest)


def _trim_whatif_chart_data(whatif_data: dict[str, Any] | None) -> dict[str, Any] | None:
    """What-if 图表数据专用裁剪（避免整包 tojson，R9 数据最小化）。

    whatif_data（C19 契约）含 summary/changes/stats/base/candidate 等表格字段，
    双环图只需 categories（图表 JS 读取 whatif.categories）。保留 available 便于
    JS 侧可用性判断；数据不足（None/available=False）返回 None（模板不输出数据段）。
    """
    if not whatif_data or not whatif_data.get("available"):
        return None
    return {"available": True, "categories": whatif_data.get("categories") or []}


def _trim_whatif_backtest_chart_data(whatif_data: dict[str, Any] | None) -> dict[str, Any] | None:
    """时序回测图表数据专用裁剪（R9 数据最小化）。

    只透传 series 字段（labels/base/candidate/base_drawdown/candidate_drawdown），
    避免把 metrics/reason 等表格字段整包 tojson 到前端。回测缺失/不可用时返回 None。
    """
    bt = (whatif_data or {}).get("backtest") if whatif_data else None
    if not bt or not bt.get("available"):
        return None
    series = bt.get("series")
    if not series or not series.get("labels"):
        return None
    return {
        "available": True,
        "effective_date": bt.get("effective_date"),
        "series": {
            "labels": series.get("labels"),
            "base": series.get("base"),
            "candidate": series.get("candidate"),
            "base_drawdown": series.get("base_drawdown"),
            "candidate_drawdown": series.get("candidate_drawdown"),
        },
    }


def render_whatif_html(whatif_data: dict[str, Any], now_str: str) -> str:
    """渲染 whatif_template.html，返回完整 HTML 字符串。

    Args:
        whatif_data: C19 契约 dict
        now_str: 展示用时间字符串

    Returns:
        HTML 字符串
    """
    from src.python.report.html_jinja_env import _ENV

    return _ENV.get_template("whatif_template.html").render(
        whatif_data=whatif_data,
        now=now_str,
        whatif_chart_data=_trim_whatif_chart_data(whatif_data),
        whatif_backtest_chart_data=_trim_whatif_backtest_chart_data(whatif_data),
    )


def write_whatif_html(whatif_data: dict[str, Any], output_dir: str = "reports") -> str:
    """输出调仓模拟 HTML 页面（最新版固定名 + 日期目录归档版，含 Chart.js 资产复制），返回最新文件路径。

    归档格式对齐主报告：`调仓模拟.html`（最新版，覆盖）+ `YYYYMMDD/调仓模拟-YYYYMMDD-HHMMSS.html`（归档版）。

    Args:
        whatif_data: C19 契约 dict
        output_dir: 输出目录

    Returns:
        最新版 HTML 绝对路径
    """
    _ensure_reports_dir(output_dir)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    html = render_whatif_html(whatif_data, now_str)
    _copy_js_assets(output_dir)

    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%H%M%S")
    latest = os.path.join(output_dir, "调仓模拟.html")
    with open(latest, "w", encoding="utf-8") as f:
        f.write(html)
    archive_dir = os.path.join(output_dir, date_str)
    os.makedirs(archive_dir, exist_ok=True)
    archive = os.path.join(archive_dir, f"调仓模拟-{date_str}-{time_str}.html")
    with open(archive, "w", encoding="utf-8") as f:
        f.write(html)
    _cleanup_old_archives(output_dir)
    logger.info("调仓模拟 HTML 已保存: %s", latest)
    return os.path.abspath(latest)


def write_whatif_report(
    whatif_data: dict[str, Any],
    output_dir: str = "reports",
    reporter=None,
) -> dict[str, str]:
    """同时输出 Excel + HTML 调仓模拟报告。

    Args:
        whatif_data: C19 契约 dict
        output_dir: 输出目录
        reporter: 进度输出（CliProgressReporter），None 时静默

    Returns:
        {"excel": 最新 Excel 绝对路径, "html": 最新 HTML 绝对路径}
    """
    if reporter is not None:
        reporter.info("正在输出调仓 What-if 模拟报告...")
    excel_path = write_whatif_excel(whatif_data, output_dir)
    html_path = write_whatif_html(whatif_data, output_dir)
    if reporter is not None:
        reporter.ok(f"调仓模拟报告生成完成: Excel {excel_path} / HTML {html_path}")
    return {"excel": excel_path, "html": html_path}
