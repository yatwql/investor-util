"""调仓 What-if 模拟 Excel 页签写入模块。

输出 3 个固定页签 + 1 个条件页签：
  1. 调仓摘要 — 基准/目标文件 + 变动统计 + 汇总指标对比（含箭头）
  2. 分类配置对比 — 资产大类成本权重基准 vs 目标（成本口径）
  3. 持仓变动明细 — 新增/清仓/加仓/减仓/不变，行底色按变动类型标注
  4. 时序回测（条件）— 指定调仓生效日时追加；未指定或数据不足时写占位文本

数据不足（whatif_data=None 或 available=False）时摘要页写占位文本
（§1.4.5 数据降级治理）。
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.python.report.data_status import STATUS_MESSAGES
from src.python.report.excel_writer import (
    _write_placeholder,
    auto_width,
    freeze_header,
    write_data_row,
    write_header_row,
    write_title_row,
)
from src.python.report.styles import FMT_MONEY, FMT_SHARES

logger = logging.getLogger("invest")

# 变动类型行底色（新增→绿 / 清仓→红 / 加仓→黄 / 减仓→蓝 / 不变→浅灰）
_ACTION_FILL: dict[str, PatternFill] = {
    "新增": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "清仓": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "加仓": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "减仓": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "不变": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

# 变动类型 → 字体色（新增绿 / 清仓红 / 其余普通）
_ACTION_FONT: dict[str, Font] = {
    "新增": Font(color="006100"),
    "清仓": Font(color="9C0006"),
}

# summary 行单位 → 数字格式
_UNIT_FORMATS: dict[str, str | None] = {
    "money": FMT_MONEY,
    "shares": FMT_SHARES,
    "count": None,
    "hhi": "0.000000",
}

# 回测指标行单位 → 数字格式（pct 后拼 % 号）
_BT_UNIT_FORMATS: dict[str, str | None] = {
    "pct": '0.00"%"',
    "ratio": "0.00",
}


def _file_label(whatif_data: dict[str, Any]) -> str:
    """基准 → 目标 文件展示串。"""
    base_file = whatif_data.get("base_file") or "基准持仓"
    cand_file = whatif_data.get("candidate_file") or "目标持仓"
    return f"{base_file}  →  {cand_file}"


def write_whatif_summary_sheet(ws: Worksheet, whatif_data: dict[str, Any] | None) -> None:
    """写入「调仓摘要」页签。

    Args:
        ws: openpyxl Worksheet 对象
        whatif_data: 数据契约 dict；None 或 available=False 时写占位。
    """
    _ncols = 8
    write_title_row(ws, 1, "调仓 What-if 模拟", ncols=_ncols)

    if not whatif_data or not whatif_data.get("available"):
        _write_placeholder(
            ws, STATUS_MESSAGES.get("whatif_unavailable", "调仓对比数据暂不可用"), row=3, max_cols=_ncols
        )
        freeze_header(ws, row=2)
        auto_width(ws)
        logger.info("调仓 What-if：数据不足，写入占位")
        return

    row = 2
    # ── 文件对比 + 变动统计 ──
    row = write_data_row(ws, row, [f"对比：{_file_label(whatif_data)}"] + [""] * (_ncols - 1))
    stats = whatif_data.get("stats", {})
    row = write_data_row(
        ws,
        row,
        [
            f"变动统计：新增 {stats.get('added', 0)} | 清仓 {stats.get('removed', 0)} | "
            f"加仓 {stats.get('increased', 0)} | 减仓 {stats.get('decreased', 0)} | "
            f"不变 {stats.get('unchanged', 0)}"
        ]
        + [""] * (_ncols - 1),
    )

    # ── 汇总指标对比 ──
    row += 1
    row = write_title_row(ws, row, "汇总指标对比（成本口径）", ncols=4)
    row = write_header_row(ws, row, ["指标", "基准", "目标", "变化"])
    for s in whatif_data.get("summary", []):
        _fmt = _UNIT_FORMATS.get(s.get("unit", ""))
        row = write_data_row(
            ws,
            row,
            [s.get("label", ""), s.get("base", 0), s.get("candidate", 0), s.get("delta", 0)],
            formats=[None, _fmt, _fmt, _fmt],
        )
        # 箭头列在变化量列内拼上方向（HTML 侧用字符，Excel 侧用文字）
        ws.cell(row=row - 1, column=4).value = f"{s.get('delta', 0)} {s.get('arrow', '')}"

    # ── 口径说明 ──
    row += 1
    row = write_title_row(ws, row, "说明", ncols=_ncols)
    notes = [
        "口径：权重基于成本 = 份额 × 每份成本（成本口径截面比较，本地计算、零网络请求）",
        "时序回测：指定调仓生效日时 opt-in 联网取生效日后行情，追加第 4 页签「时序回测」",
        "假设推演：回测为模拟推演，不构成收益承诺",
    ]
    for n in notes:
        row = write_data_row(ws, row, [n] + [""] * (_ncols - 1))

    freeze_header(ws, row=2)
    auto_width(ws, min_width=10, max_width=40)
    logger.info("调仓摘要页签写入完成: %s", _file_label(whatif_data))


def write_whatif_category_sheet(ws: Worksheet, whatif_data: dict[str, Any] | None) -> None:
    """写入「分类配置对比」页签（成本口径权重 %）。"""
    _ncols = 6
    write_title_row(ws, 1, "资产配置对比（成本口径）", ncols=_ncols)

    if not whatif_data or not whatif_data.get("available"):
        _write_placeholder(
            ws, STATUS_MESSAGES.get("whatif_unavailable", "调仓对比数据暂不可用"), row=3, max_cols=_ncols
        )
        freeze_header(ws, row=2)
        auto_width(ws)
        return

    row = 2
    row = write_header_row(ws, row, ["资产大类", "基准权重(%)", "目标权重(%)", "变化(百分点)"])
    for c in whatif_data.get("categories", []):
        row = write_data_row(
            ws,
            row,
            [
                c.get("label", c.get("key", "")),
                c.get("base_weight", 0.0),
                c.get("cand_weight", 0.0),
                c.get("delta_pct", 0.0),
            ],
            formats=[None, "0.00", "0.00", "+0.00;-0.00;0.00"],
        )
    freeze_header(ws, row=2)
    auto_width(ws, min_width=10, max_width=30)
    logger.info("分类配置对比页签写入完成: %d 大类", len(whatif_data.get("categories", [])))


def write_whatif_changes_sheet(ws: Worksheet, whatif_data: dict[str, Any] | None) -> None:
    """写入「持仓变动明细」页签，行底色按变动类型标注。"""
    _ncols = 12
    write_title_row(ws, 1, "持仓变动明细", ncols=_ncols)

    if not whatif_data or not whatif_data.get("available"):
        _write_placeholder(
            ws, STATUS_MESSAGES.get("whatif_unavailable", "调仓对比数据暂不可用"), row=3, max_cols=_ncols
        )
        freeze_header(ws, row=2)
        auto_width(ws)
        return

    row = 2
    row = write_header_row(
        ws,
        row,
        [
            "变动",
            "名称",
            "代码",
            "基准份额",
            "目标份额",
            "份额变化",
            "基准成本(元)",
            "目标成本(元)",
            "成本变化(元)",
            "基准权重(%)",
            "目标权重(%)",
            "权重变化(百分点)",
        ],
    )
    for c in whatif_data.get("changes", []):
        action = c.get("action", "不变")
        _row = write_data_row(
            ws,
            row,
            [
                action,
                c.get("name", ""),
                c.get("code", ""),
                c.get("base_shares", 0.0),
                c.get("cand_shares", 0.0),
                c.get("shares_diff", 0.0),
                c.get("base_cost", 0.0),
                c.get("cand_cost", 0.0),
                c.get("cost_diff", 0.0),
                c.get("base_weight", 0.0),
                c.get("cand_weight", 0.0),
                c.get("weight_delta_pct", 0.0),
            ],
            formats=[
                None,
                None,
                None,
                FMT_SHARES,
                FMT_SHARES,
                FMT_SHARES,
                FMT_MONEY,
                FMT_MONEY,
                FMT_MONEY,
                "0.00",
                "0.00",
                "+0.00;-0.00;0.00",
            ],
        )
        # 行底色 + 变动类型字体
        fill = _ACTION_FILL.get(action)
        if fill is not None:
            for col in range(1, _ncols + 1):
                ws.cell(row=_row - 1, column=col).fill = fill
        font = _ACTION_FONT.get(action)
        if font is not None:
            ws.cell(row=_row - 1, column=1).font = font
        row = _row

    freeze_header(ws, row=2)
    auto_width(ws, min_width=8, max_width=26)
    logger.info("持仓变动明细页签写入完成: %d 条", len(whatif_data.get("changes", [])))


def write_whatif_backtest_sheet(ws: Worksheet, whatif_data: dict[str, Any] | None) -> None:
    """写入「时序回测」页签（指定调仓生效日时才有数据）。

    结构：生效日信息行 + 回测指标对比（5 行，unit 映射 _BT_UNIT_FORMATS，
    变化列拼 arrow）+ 归一化净值与回撤序列数据表 + 口径/局限说明。
    未指定生效日 / backtest 缺失 / 回测不可用 → 写占位文本（不阻塞主报告）。

    Args:
        ws: openpyxl Worksheet 对象
        whatif_data: 数据契约 dict（含可选 backtest 键）
    """
    _ncols = 6
    write_title_row(ws, 1, "时序回测（指定生效日后行情推演）", ncols=_ncols)

    bt = (whatif_data or {}).get("backtest") if whatif_data else None
    if not bt or not bt.get("available"):
        _write_placeholder(
            ws,
            (bt or {}).get("reason") or STATUS_MESSAGES.get("whatif_backtest_unavailable", "时序回测不可用"),
            row=3,
            max_cols=_ncols,
        )
        freeze_header(ws, row=2)
        auto_width(ws)
        return

    row = 2
    eff = bt.get("effective_date") or "—"
    row = write_data_row(
        ws,
        row,
        [f"调仓生效日：{eff}；基准/目标组合在生效日后的 as-if 净值均归一化到 100 基点"] + [""] * (_ncols - 1),
    )

    # ── 指标对比 ──
    row += 1
    row = write_title_row(ws, row, "回测指标对比（生效日后，假设推演）", ncols=4)
    row = write_header_row(ws, row, ["指标", "基准", "目标", "变化"])
    for m in bt.get("metrics", []):
        _fmt = _BT_UNIT_FORMATS.get(m.get("unit", ""))
        row = write_data_row(
            ws,
            row,
            [
                m.get("label", ""),
                m.get("base"),
                m.get("candidate"),
                m.get("delta"),
            ],
            formats=[None, _fmt, _fmt, _fmt],
        )
        # 变化列拼方向箭头（与摘要页一致）
        delta_v = m.get("delta")
        ws.cell(row=row - 1, column=4).value = (
            f"{delta_v} {m.get('arrow', '')}" if delta_v is not None else f"-- {m.get('arrow', '')}"
        )

    # ── 归一化净值与回撤序列 ──
    row += 1
    row = write_title_row(ws, row, "归一化净值与回撤序列（基准/目标）", ncols=_ncols)
    series = bt.get("series") or {}
    labels = series.get("labels") or []
    if labels:
        row = write_header_row(
            ws, row, ["日期", "基准净值(100基点)", "目标净值(100基点)", "基准回撤(%)", "目标回撤(%)"]
        )
        for i, d in enumerate(labels):
            row = write_data_row(
                ws,
                row,
                [
                    d,
                    series.get("base", [])[i],
                    series.get("candidate", [])[i],
                    series.get("base_drawdown", [])[i],
                    series.get("candidate_drawdown", [])[i],
                ],
                formats=[None, "0.0000", "0.0000", "0.0000", "0.0000"],
            )
    else:
        row = write_data_row(ws, row, ["无有效净值序列"] + [""] * (_ncols - 1))

    # ── 说明 ──
    row += 1
    row = write_title_row(ws, row, "说明", ncols=_ncols)
    notes = [
        "口径：as-if 市值 = 份额 × 每日价格；基准/目标组合各自归一化到 100 基点后比较",
        "局限：假设推演、不构成收益承诺；未指定生效日时无本页签",
        "数据来源：生效日后行情历史（opt-in 联网获取）；数据不足时回测降级不阻塞主报告",
    ]
    for n in notes:
        row = write_data_row(ws, row, [n] + [""] * (_ncols - 1))

    freeze_header(ws, row=2)
    auto_width(ws, min_width=10, max_width=40)
    logger.info("时序回测页签写入完成: %s", eff)
