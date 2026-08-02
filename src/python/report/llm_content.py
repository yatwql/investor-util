"""LLM 内容输出模块 — 全球政经局势 / 智囊团深度复盘 / 持仓体检报告 / 穿透深度分析。

由调用方预生成 LLM 内容后传入本模块写入 Excel 页签。
"""

from __future__ import annotations

import logging
import re
from math import ceil
from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from src.python.llm import (
    FAIL_REASON_API_ERROR,
    FAIL_REASON_CIRCUIT_OPEN,
    FAIL_REASON_DISABLED,
    FAIL_REASON_NETWORK_ERROR,
    FAIL_REASON_NOT_CONFIGURED,
    FAIL_REASON_TIMEOUT,
)
from src.python.llm.prompts import LLM_MODULE_FAILURE
from src.python.core.registry import get_llm_module_name, get_report_section_order
from src.python.report.excel_writer import (
    freeze_header,
    write_title_row,
)
from src.python.report.llm_module_info import get_llm_module_failure_reason
from src.python.report.styles import CONTENT_FONT, GREEN_FONT

logger = logging.getLogger("invest")

_CONTENT_NCOLS = 2
_COL_WIDTH = 80  # 固定列宽（字符宽度）
_CHARS_PER_LINE = 40  # 每行约 40 中文字符
_ROW_HEIGHT_MIN = 30  # 最小行高 (pt)
_ROW_HEIGHT_PER_LINE = 15  # 每行增加高度 (pt)

# 匹配 LLM 生成的底部标识行：<p style="color:#888;font-size:12px">...</p>
_FOOTER_RE = re.compile(r'<p style="color:#888;font-size:12px">([^<]*)</p>')


def _extract_footer_text(html: str) -> str:
    """从 LLM 生成的 HTML 内容中提取底部标识行的纯文本。

    匹配最后出现的 <p style="color:#888;font-size:12px">…</p> 标签内容，
    该标签由 _generate_llm_content() 在 HTML 尾部追加，包含模型名、
    Token 用量、估算费用及 Extended Thinking 标识。

    Returns:
        标签内的纯文本，未找到时返回空字符串
    """
    if not html:
        return ""
    matches = _FOOTER_RE.findall(html)
    if matches:
        return matches[-1].strip()
    return ""


# 块级结束标签/自闭合 → 换行，保证剥离后的纯文本保留段落/列表结构
_BLOCK_END_TO_NEWLINE_RE = re.compile(
    r"(</p>|</li>|</h[1-6]>|</div>|</ul>|</ol>|<br\s*/?>|<hr\s*/?>)",
    re.IGNORECASE,
)


def _strip_html(text: str) -> str:
    """移除文本中的 HTML 标签，保留纯文本内容。

    块级结束标签（``</p>``、``</li>`` 等）与自闭合块（``<br>``/``<hr>``）
    先替换为换行，使剥离后的多行块（如 ``<ul>`` 列表、校验摘要明细行）
    保留段落结构，避免标签间文本粘连成一行；连续换行折叠为单个，
    消除块级结束标签与既有换行叠加产生的空行。
    """
    if not text:
        return ""
    text = _BLOCK_END_TO_NEWLINE_RE.sub("\n", text)
    text = re.sub(r"\n+", "\n", text)
    return re.sub(r"<[^>]+>", "", text).strip()


# 块级 HTML 标签（含自闭合 <br>/<hr>），用于将 LLM 内容分段写入 Excel 单元格
_BLOCK_TAG_RE = re.compile(
    r"(<p\b[^>]*>.*?</p>|<li\b[^>]*>.*?</li>|<div\b[^>]*>.*?</div>|"
    r"<h[1-6]\b[^>]*>.*?</h[1-6]>|<ul\b[^>]*>.*?</ul>|<ol\b[^>]*>.*?</ol>|<br\s*/?>|<hr\s*/?>)",
    re.IGNORECASE | re.DOTALL,
)


def _split_html_blocks(html: str) -> list[str]:
    """按块级 HTML 元素将内容切分为多个块，供逐块写入 Excel 单元格。

    真实 LLM 内容由 ``markdown_to_html`` 以无换行方式拼接 ``<p>`` 标签，
    按 ``\\n\\n`` 分段会把整模块坍缩进一个单元格，且 footer、事实校验
    摘要无法独立成行。本函数以块级标签为切分点；游离文本
    （如校验摘要尾部的 ``<span>`` 附属行）并入前一块，纯文本输入则回退
    按 ``\\n\\n`` 分段以兼容历史输入。

    Returns:
        按出现顺序排列的 HTML 块片段列表（含标签）。
    """
    if not html:
        return []
    parts = _BLOCK_TAG_RE.split(html)
    blocks: list[str] = []
    pending = ""
    for part in parts:
        if not part or not part.strip():
            continue
        if _BLOCK_TAG_RE.fullmatch(part.strip()):
            # 块级元素本身：先收拢 pending 游离文本，再入块
            if pending:
                blocks.append(pending)
                pending = ""
            blocks.append(part.strip())
        else:
            # 游离文本：并入前一块（通常为校验摘要的 <span> 附属行），
            # 与前块以 \n 分隔，避免剥离后粘连成一行导致明细样式丢失。
            # part 自身换行先 strip，防与前块拼接产生 \n\n 被二次分段。
            if blocks:
                sep = "" if blocks[-1].endswith("\n") else "\n"
                blocks[-1] = blocks[-1] + sep + part.strip()
            else:
                pending += part
    if pending:
        blocks.append(pending)
    # 纯文本段（无块级标签）回退按 \n\n 分段，兼容历史输入
    result: list[str] = []
    for block in blocks:
        sub = [seg.strip() for seg in block.split("\n\n") if seg.strip()]
        result.extend(sub)
    return result


def _calc_row_height(text: str) -> int:
    """根据文本长度估算行高。

    按中文字符宽度估算：列宽 _COL_WIDTH 约容纳 _CHARS_PER_LINE 个中文。
    含 ``\\n`` 的多行块（如 ``<ul>`` 列表剥离后的列表项、摘要明细行）
    逐行估算宽度并求和，避免行高不足导致内容截断。
    """
    if not text:
        return _ROW_HEIGHT_MIN
    width_rows = sum(ceil(len(line) / _CHARS_PER_LINE) for line in text.split("\n"))
    lines = max(width_rows, 1)
    return max(lines * _ROW_HEIGHT_PER_LINE, _ROW_HEIGHT_MIN)


_CONTENT_ALIGN = Alignment(
    wrap_text=True,
    vertical="top",
    horizontal="left",
)

_MODEL_NAME_FONT = Font(
    color="888888",
    size=9,
    italic=True,
)

_THINKING_FONT = Font(
    color="888888",
    size=9,
    italic=True,
)

# 匹配事实校验摘要行（run_fact_check 追加到内容尾部）
_FACT_CHECK_PASS_RE = re.compile(r"✓.*事实校验")  # 通过：✓ 事实校验通过
_FACT_CHECK_FAIL_RE = re.compile(r"事实校验：.*提示|^⚠ ")  # 告警：事实校验：N/M 项通过，K 项提示 / ⚠ 品种代码

# 事实校验字体（与 HTML 报告的 #4a4 / #a40 对应）
_FACT_CHECK_PASS_FONT = GREEN_FONT
_FACT_CHECK_WARN_FONT = Font(color="CC6600", size=11)  # 琥珀色
# 摘要明细行灰色小字（对应 HTML 中 #888「已修正明细」/ #999「建议提及」行）
_FACT_CHECK_DETAIL_FONT = Font(color="888888", size=9)


def _get_module_key_map(section_order: list[dict] | None = None) -> dict[str, str]:
    """从 section_order 构建 Excel 页签标题 → 模块键映射。

    格式：{"12.全球政经局势": "global_macro", ...}
    序号跟随 section_order 配置，news_correlation 无独立 LLM 分析章节，排除在外。
    注：从 section_order 动态构建，不缓存（最多 14 项，< 0.01ms）。
    """
    result: dict[str, str] = {}
    order = section_order or get_report_section_order()
    for sec in order:
        mk = sec["key"]
        if mk != "news_correlation" and mk != "llm_usage":
            title = f"{sec['number']}.{sec['name']}"
            result[title] = mk
    return result


_PLACEHOLDER_BY_REASON: dict[str, str] = {
    FAIL_REASON_NOT_CONFIGURED: "本节内容待生成 — LLM 未配置（请配置 data/config/llm_providers.json 或 llm_key.json）",
    FAIL_REASON_API_ERROR: "本节内容待生成 — LLM API 调用失败（请检查 API Key 和网络连接后重新生成）",
    FAIL_REASON_TIMEOUT: "本节内容待生成 — LLM API 请求超时（可尝试在 llm_settings.json 中增大 timeout 配置）",
    FAIL_REASON_NETWORK_ERROR: "本节内容待生成 — LLM API 网络连接失败（请检查网络后重新生成）",
    FAIL_REASON_CIRCUIT_OPEN: "本节内容待生成 — LLM API 暂时不可用（熔断冷却中，请稍后重试）",
}


def _get_placeholder(title: str, section_order: list[dict] | None = None) -> str:
    """根据页签标题查找对应的失败原因占位文本。"""
    mk = _get_module_key_map(section_order).get(title)
    if mk:
        reason = get_llm_module_failure_reason(LLM_MODULE_FAILURE, mk)
        if reason in _PLACEHOLDER_BY_REASON:
            return _PLACEHOLDER_BY_REASON[reason]
        if reason:
            return "本节内容待生成 — LLM 生成失败"
    return "本节内容待生成 — 请配置 LLM API Key（data/config/llm_key.json）"


def _write_fact_check_block(ws: Worksheet, start_row: int, block_text: str) -> int:
    """写事实校验摘要块：首行绿/琥珀，明细行灰色小字。

    摘要块内部以 ``\\n`` 分隔首行（通过/告警概述）与逐条明细，
    拆为多行分别着色，避免整块坍缩进一个单元格后明细样式丢失。

    Args:
        ws: 目标工作表
        start_row: 起始行号
        block_text: 摘要块纯文本（含 ``\\n`` 分隔的明细行）

    Returns:
        写完最后一个明细行后的下一行号。
    """
    lines = [ln.strip() for ln in block_text.split("\n") if ln.strip()]
    is_warn = _FACT_CHECK_FAIL_RE.search(block_text)
    first_font = _FACT_CHECK_WARN_FONT if is_warn else _FACT_CHECK_PASS_FONT
    row = start_row
    for i, line in enumerate(lines):
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = first_font if i == 0 else _FACT_CHECK_DETAIL_FONT
        cell.alignment = _CONTENT_ALIGN
        ws.row_dimensions[row].height = _calc_row_height(line)
        row += 1
    return row


def _write_content_sheet(
    ws: Worksheet,
    title: str,
    content: str | None,
    section_order: list[dict] | None = None,
    debate_mode_label: str | None = None,
) -> None:
    """写入一个 LLM 分析章节。

    结构：
      - 第 1 行：标题（合并 A1:B1，居中标题样式）
      - 第 2 行起：每段落占一行 + 一行空行间距
      - 底部标识行从 HTML 内容末尾的 ``<p style="color:#888;font-size:12px">`` 标签提取

    Args:
        ws: 目标工作表
        title: 页签标题行文本
        content: LLM 返回的 HTML 文本（已剥离标签），为 None 时写入占位符
        section_order: 可选，用于占位符查找的 section_order
    """
    # 标题行
    row = write_title_row(ws, 1, title, _CONTENT_NCOLS)

    # 固定列宽
    ws.column_dimensions["A"].width = _COL_WIDTH

    # 辩论模式标注（灰色小字注释）
    if debate_mode_label:
        row += 1
        _note_cell = ws.cell(row=row, column=1, value=f"本报告为实验模式输出（{debate_mode_label}），结果仅供参考")
        _note_cell.font = Font(size=9, color="999999")
        _note_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

    if content:
        # 先提取底部 LLM footer（在剥离 HTML 前进行）
        footer_text = _extract_footer_text(content)

        # 按块级 HTML 元素分段（真实内容由 markdown_to_html 无换行拼接 <p> 标签，
        # 按 "\n\n" 分段会坍缩整模块进一个单元格、footer/摘要无法独立成行）
        blocks = _split_html_blocks(content)

        for block in blocks:
            block_text = _strip_html(block)
            if not block_text:
                continue
            # footer 块：正文跳过，末尾统一写一次，避免重复
            if footer_text and block_text == footer_text:
                continue
            # 事实校验摘要块：首行绿/琥珀，明细行灰色小字
            if _FACT_CHECK_PASS_RE.search(block_text) or _FACT_CHECK_FAIL_RE.search(block_text):
                row = _write_fact_check_block(ws, row, block_text)
                row += 1  # 摘要块后空行
                continue
            # 普通段落
            cell = ws.cell(row=row, column=1, value=block_text)
            cell.font = CONTENT_FONT
            cell.alignment = _CONTENT_ALIGN
            ws.row_dimensions[row].height = _calc_row_height(block_text)
            row += 2

        # 底部 LLM 标识行（从 HTML 中提取，与 HTML 报告格式保持一致）
        if footer_text:
            cell = ws.cell(row=row, column=1, value=footer_text)
            cell.font = _MODEL_NAME_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 20
            row += 1
    else:
        placeholder = _get_placeholder(title, section_order)
        cell = ws.cell(row=row, column=1, value=placeholder)
        cell.font = CONTENT_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = _ROW_HEIGHT_MIN
        row += 1

    # 冻结标题行
    freeze_header(ws, 1)


def write_llm_sheets(
    sheets: dict[str, Any],
    llm_content: tuple[str | None, str | None, str | None, str | None],
    section_order: list[dict] | None = None,
    debate_mode_label: str | None = None,
) -> tuple[str, str, str, str]:
    """写入 LLM 分析章节（全球政经局势 & 智囊团深度复盘 & 持仓体检报告 & 穿透深度分析）。

    调用方预创建页签并通过 sheets 字典传入，本函数仅负责写入 Excel 内容。
    底部标识行（模型名/Token用量/缓存/Extended Thinking）已嵌入 HTML 内容中，
    由 ``_extract_footer_text()`` 提取后追加到页签尾部。

    Args:
        sheets: {key: ws} 字典，页签已由 _create_sheets 按序预创建
        llm_content: (global_macro_html, expert_review_html, health_check_html, penetration_deep_html) 预生成内容
        section_order: 可选，用于 LLM 内部标题行序号跟随用户配置
        debate_mode_label: 辩论模式标签，非 None 时在 expert_review 页签追加实验模式注释

    Returns:
        (global_macro_text, expert_review_text, health_check_text, penetration_deep_text) 纯文本四元组，供 TUI 展示
    """
    _reverse = {v: k for k, v in _get_module_key_map(section_order).items()}

    _module_keys = ["global_macro", "expert_review", "health_check", "penetration_deep"]
    _disabled = tuple(
        get_llm_module_failure_reason(LLM_MODULE_FAILURE, mk) == FAIL_REASON_DISABLED for mk in _module_keys
    )

    content7, content8, content9, contentA = llm_content

    _module_contents = [
        ("global_macro", content7),
        ("expert_review", content8),
        ("health_check", content9),
        ("penetration_deep", contentA),
    ]
    for i, (mk, content) in enumerate(_module_contents):
        if _disabled[i]:
            logger.info("LLM 分析章节跳过（已禁用）: %s", get_llm_module_name(mk))
            continue
        ws = sheets.get(mk)
        if ws is None:
            logger.warning("LLM 分析章节页签 %s 未由 _create_sheets 创建，跳过", mk)
            continue
        _write_content_sheet(
            ws,
            _reverse.get(mk, get_llm_module_name(mk)),
            content,
            section_order,
            debate_mode_label=(debate_mode_label if mk == "expert_review" else None),
        )

    logger.info("LLM 分析章节写入完成（含%s）", get_llm_module_name("penetration_deep"))

    # 返回纯文本内容，供 TUI 展示（禁用模块返回空字符串）
    def _text_or_empty(c: str | None, disabled: bool) -> str:
        """禁用或空内容返回 ""，否则返回剥离 HTML 的纯文本。"""
        if disabled or not c:
            return ""
        return _strip_html(c)

    return (
        _text_or_empty(content7, _disabled[0]),
        _text_or_empty(content8, _disabled[1]),
        _text_or_empty(content9, _disabled[2]),
        _text_or_empty(contentA, _disabled[3]),
    )
