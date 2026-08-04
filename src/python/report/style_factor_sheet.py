"""风格与因子分析 Excel 写入模块 — 报告页签 9（一章三区块）。

「风格与因子分析」一章分「基金风格表 + 风格因子回归」两区块
（sheet key `style_factor`），另含可选「行业 Beta」子表：

  一、基金风格表 —— 基金名称/当前风格/漂移等级等（数据不足写占位）
  二、风格因子回归 —— 价值/成长/质量 3 因子 OLS 暴露 + 基准对照（数据不足写占位）
  三、行业 Beta 子表 —— 穿透行业暴露占比 + 各行业指数 β/相关性
      （由编排层 C19 `style_factor_data.industry_beta` 提供；None/available=False 时该区块不渲染）

任一区块数据不足时该区块独立降级（§1.4.5），互不影响；
三区块均无数据时整页写占位。
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.python.core.registry import get_report_section_number, get_report_sheet_name
from src.python.report.data_status import STATUS_MESSAGES
from src.python.report.excel_writer import (
    _write_placeholder,
    auto_width,
    freeze_header,
    write_data_row,
    write_header_row,
    write_title_row,
)

logger = logging.getLogger("invest")

# ── 风格表区块 ────────────────────────────────────────────────

_STYLE_NCOLS = 8
_STYLE_HEADERS = [
    "基金名称",
    "基金代码",
    "当前风格",
    "上期风格",
    "漂移等级",
    "漂移评分",
    "备注",
    "标识",
]

_DRIFT_FONTS: dict[str, Font] = {
    "严重": Font(color="CC0000"),
    "中度": Font(color="FF8C00"),
    "轻度": Font(color="DAA520"),
}

# ── 因子回归区块 ──────────────────────────────────────────────

_FACTOR_NCOLS = 5
_FACTOR_HEADERS = ["风格因子", "暴露系数 β", "t 值", "显著（95%）", "风格归属占比"]
# 固定 MVP 因子展示顺序（与 FACTOR_INDICES 键一致，名字在编排层 data 中）
_FACTOR_ORDER = ["value", "growth", "quality"]

_FONT_GREEN = Font(color="009900")
_FONT_RED = Font(color="CC0000")

# ── 行业 Beta 子表 ────────────────────────────────────────────

_IND_NCOLS = 7
_IND_HEADERS = ["行业", "暴露占比", "指数代码", "β", "t 值", "显著（95%）", "相关性 r"]


# ── 区块工具 ─────────────────────────────────────────────────


def _drift_font(level: str) -> Font:
    return _DRIFT_FONTS.get(level, Font())


def _style_remark(is_estimated: bool, is_first: bool) -> str:
    """生成风格表备注列文本。"""
    if is_first:
        return "基准确立中"
    return "估算风格" if is_estimated else ""


def _factor_name(factor: str, factor_names: dict | None) -> str:
    """因子 key → 中文名（编排层 data.factor_names 为单一数据源）。"""
    if factor_names and factor in factor_names:
        return str(factor_names[factor])
    return factor


def _compute_ncols(
    style_data: list[dict] | None,
    factor_exposure: dict | None,
    industry_beta: dict | None,
) -> int:
    """计算标题栏跨列数：取三区块所需列数的最大值。"""
    return max(_STYLE_NCOLS, _FACTOR_NCOLS, _IND_NCOLS)


# ── 区块一：基金风格表 ────────────────────────────────────────


def _write_style_block(
    ws: Worksheet,
    row: int,
    style_data: list[dict] | None,
    ncols: int,
) -> int:
    """写入一、基金风格表区块，返回下一行起始行号。"""
    write_title_row(ws, row, "一、基金风格表", ncols=ncols)
    row += 1
    row = write_header_row(ws, row, _STYLE_HEADERS)
    row += 1

    style_data = style_data or []
    if not style_data:
        row = _write_placeholder(ws, STATUS_MESSAGES["style_unavailable"], row=row, max_cols=ncols)
        logger.info("风格与因子分析·风格表：无数据，写入占位")
        return row

    for item in style_data:
        drift_level = item.get("drift_level", "")
        row_font = _drift_font(drift_level)
        is_first = item.get("is_first_check", False)
        row_data = [
            item.get("name", ""),
            item.get("code", ""),
            item.get("current_style", "--"),
            item.get("prev_style", "--"),
            drift_level,
            item.get("drift_score", "--") if item.get("drift_score") is not None else "--",
            _style_remark(item.get("is_estimated", False), is_first),
            "📋 基线" if is_first else "✅",
        ]
        write_data_row(ws, row, row_data)
        for col in range(1, _STYLE_NCOLS + 1):
            ws.cell(row=row, column=col).font = row_font
        row += 1

    row += 1  # 区块间隔空行
    return row


# ── 区块二：风格因子回归 ──────────────────────────────────────


def _write_factor_block(
    ws: Worksheet,
    row: int,
    factor_exposure: dict | None,
    factor_names: dict | None,
    ncols: int,
) -> int:
    """写入二、风格因子回归区块，返回下一行起始行号。"""
    write_title_row(ws, row, "二、风格因子回归", ncols=ncols)
    row += 1

    if not factor_exposure or not factor_exposure.get("available"):
        row = write_header_row(ws, row, _FACTOR_HEADERS)
        row = _write_placeholder(ws, STATUS_MESSAGES["factor_exposure_unavailable"], row=row, max_cols=ncols)
        logger.info("风格与因子分析·因子回归：无数据，写入占位")
        return row

    row = write_header_row(ws, row, _FACTOR_HEADERS)
    betas = factor_exposure.get("betas", {})
    t_stats = factor_exposure.get("t_stats", {})
    significant = factor_exposure.get("significant", {})
    style_alloc = factor_exposure.get("style_allocation", {})

    for factor in _FACTOR_ORDER:
        if factor not in betas:
            continue
        _sig = bool(significant.get(factor))
        row_data = [
            _factor_name(factor, factor_names),
            betas[factor],
            t_stats.get(factor, "--") if t_stats.get(factor) is not None else "--",
            "✅ 显著" if _sig else "—",
            style_alloc.get(factor, 0.0),
        ]
        write_data_row(ws, row, row_data, formats=[None, "0.0000", "0.000", None, "0.00%"])
        if _sig:
            ws.cell(row=row, column=4).font = _FONT_GREEN
        row += 1

    # 基准对照（沪深300 同窗口回归）
    baseline_betas = factor_exposure.get("baseline_betas", {})
    if baseline_betas:
        row += 1
        row = write_title_row(ws, row, "基准对照（沪深300 同窗口回归）", ncols=ncols)
        row = write_header_row(ws, row, ["风格因子", "组合 β", "基准 β", "相对暴露", ""])
        for factor in _FACTOR_ORDER:
            if factor not in betas or factor not in baseline_betas:
                continue
            rel = round(float(betas[factor]) - float(baseline_betas[factor]), 3)
            row_data = [
                _factor_name(factor, factor_names),
                betas[factor],
                baseline_betas[factor],
                rel,
                "",
            ]
            write_data_row(ws, row, row_data, formats=[None, "0.0000", "0.0000", "0.000", None])
            if rel > 0.1:
                ws.cell(row=row, column=4).font = _FONT_RED
            elif rel < -0.1:
                ws.cell(row=row, column=4).font = _FONT_GREEN
            row += 1

    # 说明区
    row += 1
    row = write_title_row(ws, row, "说明", ncols=ncols)
    notes = [
        f"回归窗口：{factor_exposure.get('window', 0)} 个交易日；有效样本：{factor_exposure.get('sample_count', 0)} 期",
        f"α（截距）= {factor_exposure.get('alpha', 0.0)}；显著列为 95% 双尾 t 检验结果",
    ]
    stale = factor_exposure.get("stale_factors") or []
    if stale:
        notes.append(f"已剔除停更/不可用因子：{'、'.join(stale)}")
    corr = factor_exposure.get("factor_correlations") or {}
    if corr:
        corr_text = "；".join(f"{k}={v}" for k, v in corr.items())
        notes.append(f"因子间相关性：{corr_text}")
    corr_note = factor_exposure.get("correlation_note") or ""
    if corr_note:
        notes.append(corr_note)
    for n in notes:
        write_data_row(ws, row, [n] + [""] * (ncols - 1))
        row += 1

    row += 1  # 区块间隔空行
    return row


# ── 区块三：行业 Beta 子表 ────────────────────────────────────


def _write_industry_beta_block(
    ws: Worksheet,
    row: int,
    industry_beta: dict | None,
    ncols: int,
) -> int:
    """写入三、行业 Beta 子表区块，返回下一行起始行号。

    编排层 C19 `style_factor_data.industry_beta` 子键：
      - None（开关 report_submodules.industry_beta 关闭）→ 区块不渲染；
      - available=False（push2 行业分类 / 指数 K 线不足）→ 标题 + "数据不足"占位，
        Beta 子表不渲染，不阻塞该章其余内容（§1.4.5）。
    """
    if industry_beta is None:
        logger.info("风格与因子分析·行业 Beta：开关关闭，区块不渲染")
        return row

    write_title_row(ws, row, "三、行业 Beta（组合对各行业指数敏感性）", ncols=ncols)
    row += 1

    if not industry_beta.get("available"):
        row = _write_placeholder(ws, STATUS_MESSAGES["industry_beta_unavailable"], row=row, max_cols=ncols)
        logger.info("风格与因子分析·行业 Beta：数据不足，写入占位")
        return row

    row = write_header_row(ws, row, _IND_HEADERS)
    row += 1

    exposure = industry_beta.get("exposure", {})
    betas = industry_beta.get("betas", {})
    t_stats = industry_beta.get("t_stats", {})
    significant = industry_beta.get("significant", {})
    correlations = industry_beta.get("correlations", {})

    # 有 Beta 的行业优先，再补仅有暴露占比的行业（无指数映射）
    industries: list[str] = list(betas.keys()) + [i for i in exposure if i not in betas]

    for ind in industries:
        _sig = bool(significant.get(ind))
        if ind in betas:
            row_data = [
                ind,
                exposure.get(ind, "--") if exposure.get(ind) is not None else "--",
                industry_beta.get("index_codes", {}).get(ind, "--"),
                betas[ind],
                t_stats.get(ind, "--") if t_stats.get(ind) is not None else "--",
                "✅ 显著" if _sig else "—",
                correlations.get(ind, "--") if correlations.get(ind) is not None else "--",
            ]
            write_data_row(
                ws,
                row,
                row_data,
                formats=[None, "0.00%", None, "0.0000", "0.000", None, "0.0000"],
            )
            if _sig:
                ws.cell(row=row, column=6).font = _FONT_GREEN
        else:
            # 有暴露但无指数映射的行业：仅显示占比
            row_data = [ind, exposure.get(ind, "--"), "—", "—", "—", "—", "—"]
            write_data_row(ws, row, row_data, formats=[None, "0.00%", None, None, None, None, None])
        row += 1

    # 说明区
    row += 1
    row = write_title_row(ws, row, "说明", ncols=ncols)
    notes = [
        f"回归窗口：最近 {industry_beta.get('window', 0)} 个交易日；有效样本：{industry_beta.get('sample_count', 0)} 期",
        "行业暴露占比按持仓市值加权；行业指数为中证行业指数（近似代理）",
        "β = 组合日收益对行业指数日收益的一元 OLS 系数（复用因子回归机制）；显著列为 95% 双尾 t 检验",
        "仅参与暴露占比、无指数映射的行业（—）不参与 Beta 回归",
    ]
    for n in notes:
        write_data_row(ws, row, [n] + [""] * (ncols - 1))
        row += 1

    row += 1  # 区块间隔空行
    return row


# ── 页签入口 ─────────────────────────────────────────────────


def write_style_factor_sheet(
    ws: Worksheet,
    style_data: list[dict] | None = None,
    factor_exposure: dict[str, Any] | None = None,
    factor_names: dict | None = None,
    industry_beta: dict[str, Any] | None = None,
) -> None:
    """写入风格与因子分析页签（一章三区块：风格表 + 因子回归 + 行业 Beta）。

    Args:
        ws: openpyxl Worksheet 对象
        style_data: analyze_style_for_all_funds 的 results 列表（渲染期派生，不进 C19）。
        factor_exposure: C19 `style_factor_data` dict（因子回归区块数据源）；
            None 或 available=False 时因子回归区块写占位。
        factor_names: 因子 key → 中文名映射（缺省回退 key 本身）。
        industry_beta: C19 `style_factor_data.industry_beta` 子键（行业 Beta 区块数据源）；
            None 或 available=False 时该区块不渲染（开关关/数据不可用）。
    """
    _name = get_report_sheet_name("style_factor")
    ncols = _compute_ncols(style_data, factor_exposure, industry_beta)
    write_title_row(ws, 1, f"{get_report_section_number('style_factor')}. {_name}", ncols=ncols)

    row = 2
    row = _write_style_block(ws, row, style_data, ncols)
    row = _write_factor_block(ws, row, factor_exposure, factor_names, ncols)
    row = _write_industry_beta_block(ws, row, industry_beta, ncols)

    freeze_header(ws, row=2)
    auto_width(ws, min_width=10, max_width=30)
    logger.info("风格与因子分析页签写入完成")
