"""持仓关系矩阵 Excel 写入模块 — 持仓重合度 + 持仓相关性合一页签（一章两区块）。

输出内容：
  一、持仓重合度矩阵（基金×基金对称矩阵 + 配对明细）—— 数据不足时写占位
  二、持仓相关性矩阵（品种×品种下三角矩阵 + 配对明细 + 说明）—— 数据不足时写占位

任一区块数据不足时该区块独立降级（§1.4.5），互不影响；
两区块均无数据时整页写占位。

配色说明：
  重合度区块热力图：>50% 红 / 30-50% 橙 / 15-30% 黄 / >0 绿；
  相关性区块与 HTML 模板保持一致（红=正相关 / 蓝=负相关 / 白=不显著 / 灰=N/A）。
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.python.core.registry import get_report_section_number, get_report_sheet_name
from src.python.report.data_status import STATUS_MESSAGES
from src.python.report.excel_writer import (
    _write_placeholder,
    auto_width,
    freeze_header,
    write_data_row,
    write_header_row,
    write_title_row,
)

logger = logging.getLogger("invest")

# ── 重合度区块热力图颜色 ───────────────────────────────────────

_FILL_RED = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
_FILL_ORANGE = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
_FILL_GREEN = PatternFill(start_color="AAFFAA", end_color="AAFFAA", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

# ── 相关性区块热力格填充（与 HTML 模板配色一致） ──────────────

_SIGNIFICANCE_LEVEL = 0.05
_FILL_POS_STRONG = PatternFill("solid", fgColor="C00000")
_FILL_POS_MEDIUM = PatternFill("solid", fgColor="FF7F7F")
_FILL_POS_WEAK = PatternFill("solid", fgColor="FFE0E0")
_FILL_NEG_STRONG = PatternFill("solid", fgColor="1F4E79")
_FILL_NEG_MEDIUM = PatternFill("solid", fgColor="8DB4E2")
_FILL_NEG_WEAK = PatternFill("solid", fgColor="E0ECF7")
_FILL_INSIGNIFICANT = PatternFill("solid", fgColor="FFFFFF")
_FILL_NA = PatternFill("solid", fgColor="EEEEEE")
_FILL_DIAG = PatternFill("solid", fgColor="F5F5F5")

_FONT_STRONG = Font(color="FFFFFF", bold=True)
_FONT_BODY = Font(color="333333")
_FONT_GREY = Font(color="999999")
_FONT_LIGHT = Font(color="BBBBBB")
_FONT_POS_COLOR = Font(color="C00000", bold=True)
_FONT_NEG_COLOR = Font(color="1F4E79", bold=True)

# ── 表头常量 ───────────────────────────────────────────────────

_OVERLAP_PAIR_HEADERS = ["序号", "基金A", "基金B", "重合度", "共同标的数", "共同标的"]
_OVERLAP_PAIR_NCOLS = len(_OVERLAP_PAIR_HEADERS)

_PAIRS_HEADERS = ["序号", "品种A", "品种B", "相关系数 r", "p 值", "样本数", "显著性"]
_PAIRS_NCOLS = len(_PAIRS_HEADERS)


# ── 重合度区块工具 ───────────────────────────────────────────


def _heat_fill(pct: float) -> PatternFill | None:
    """根据比例返回重合度热力图填充色。"""
    if pct >= 0.50:
        return _FILL_RED
    if pct >= 0.30:
        return _FILL_ORANGE
    if pct >= 0.15:
        return _FILL_YELLOW
    if pct > 0:
        return _FILL_GREEN
    return None


def _format_pct(value: float) -> str:
    """格式化百分比（0.35 → 35.00%）。"""
    return f"{value * 100:.2f}%"


# ── 相关性区块工具 ───────────────────────────────────────────


def _cell_style(r: float | None, p: float | None) -> tuple[PatternFill, Font]:
    """返回相关性格的填充色 + 字体（与 HTML 模板配色一致）。

    Args:
        r: Pearson 相关系数；None 表示重叠样本不足（灰色 N/A）。
        p: 双侧 p 值；None 或 ≥0.05 判不显著（白色）。

    Returns:
        (fill, font) 用于 openpyxl 单元格样式。
    """
    if r is None:
        return _FILL_NA, _FONT_GREY
    if p is None or p >= _SIGNIFICANCE_LEVEL:
        return _FILL_INSIGNIFICANT, _FONT_GREY
    if r >= 0.5:
        return _FILL_POS_STRONG, _FONT_STRONG
    if r >= 0.3:
        return _FILL_POS_MEDIUM, _FONT_BODY
    if r > 0:
        return _FILL_POS_WEAK, _FONT_BODY
    if r <= -0.5:
        return _FILL_NEG_STRONG, _FONT_STRONG
    if r <= -0.3:
        return _FILL_NEG_MEDIUM, _FONT_BODY
    return _FILL_NEG_WEAK, _FONT_BODY


# ── 区块渲染 ─────────────────────────────────────────────────


def _compute_ncols(
    overlap_result: dict[str, Any] | None,
    correlation_data: dict[str, Any] | None,
) -> int:
    """计算标题栏跨列数：取两区块所需列数的最大值。"""
    n_funds = len((overlap_result or {}).get("funds", [])) if overlap_result else 0
    n_codes = len((correlation_data or {}).get("codes", [])) if correlation_data else 0
    return max(n_funds + 2, n_codes + 2, _PAIRS_NCOLS, _OVERLAP_PAIR_NCOLS)


def _write_overlap_block(
    ws: Worksheet,
    row: int,
    overlap_result: dict[str, Any] | None,
    fund_names: dict[str, str] | None,
    ncols: int,
) -> int:
    """写入一、持仓重合度矩阵区块，返回下一行起始行号。"""
    write_title_row(ws, row, "一、持仓重合度矩阵", ncols=ncols)
    row += 1

    overlap_result = overlap_result or {}
    funds = overlap_result.get("funds", [])
    matrix = overlap_result.get("matrix", [])
    pairs = overlap_result.get("pairs", [])
    n = len(funds)

    if n < 2:
        row = _write_placeholder(ws, STATUS_MESSAGES["overlap_unavailable"], row=row, max_cols=ncols)
        logger.info("持仓重合度区块：基金数 < 2（%d），写入占位", n)
        return row

    # ── 表头：基金A / 基金B / 各基金代码列 ──
    headers = ["", ""] + [fund_names.get(c, c) if fund_names else c for c in funds]
    write_header_row(ws, row, headers)
    for col_idx in range(3, n + 3):
        ws.cell(row=row, column=col_idx).fill = _FILL_HEADER
    row += 1

    # ── 矩阵体 ──
    for i in range(n):
        code_i = funds[i]
        label_i = fund_names.get(code_i, code_i) if fund_names else code_i
        row_data = [label_i, f"({code_i})"] + [_format_pct(matrix[i][j]) for j in range(n)]
        write_data_row(ws, row, row_data)
        for j in range(n):
            fill = _heat_fill(matrix[i][j])
            if fill:
                ws.cell(row=row, column=j + 3).fill = fill
        row += 1

    row += 1  # 空行

    # ── 配对明细表 ──
    write_title_row(ws, row, "配对明细（按重合度降序）", ncols=_OVERLAP_PAIR_NCOLS)
    row += 1
    write_header_row(ws, row, _OVERLAP_PAIR_HEADERS)
    row += 1

    for idx, pair in enumerate(pairs, 1):
        a_name = fund_names.get(pair["fund_a"], pair["fund_a"]) if fund_names else pair["fund_a"]
        b_name = fund_names.get(pair["fund_b"], pair["fund_b"]) if fund_names else pair["fund_b"]
        common_stocks_str = (
            "、".join(s.get("name", s.get("code", "")) for s in pair.get("common_stocks", []))
            if pair.get("common_stocks")
            else "—"
        )
        overlap_pct = max(pair["jaccard"], 0.01 if pair["common_count"] > 0 else 0)
        write_data_row(
            ws,
            row,
            [
                idx,
                f"{a_name}({pair['fund_a']})",
                f"{b_name}({pair['fund_b']})",
                _format_pct(overlap_pct),
                pair.get("common_count", 0),
                common_stocks_str,
            ],
        )
        row += 1

    row += 1  # 区块间隔空行
    return row


def _write_correlation_block(
    ws: Worksheet,
    row: int,
    correlation_data: dict[str, Any] | None,
    ncols: int,
) -> int:
    """写入二、持仓相关性矩阵区块，返回下一行起始行号。"""
    write_title_row(ws, row, "二、持仓相关性矩阵", ncols=ncols)
    row += 1

    if not correlation_data or not correlation_data.get("available"):
        codes = (correlation_data or {}).get("codes", []) if correlation_data else []
        row = write_header_row(ws, row, ["品种", "代码"] + list(codes))
        row = _write_placeholder(ws, STATUS_MESSAGES["correlation_unavailable"], row=row, max_cols=ncols)
        logger.info("持仓相关性区块：无数据，写入占位")
        return row

    codes = correlation_data.get("codes", [])
    names = correlation_data.get("names", {})
    matrix = correlation_data.get("matrix", [])
    p_values = correlation_data.get("p_values", [])
    pairs = correlation_data.get("pairs", [])

    # ── 1. 矩阵表（下三角） ──
    row = write_header_row(ws, row, [""] + [f"{names.get(c, c)} ({c})" for c in codes])
    for i, ci in enumerate(codes):
        row += 1
        ws.cell(row=row, column=1, value=f"{names.get(ci, ci)} ({ci})")
        for j, cj in enumerate(codes):
            cell = ws.cell(row=row, column=2 + j)
            if j == i:
                cell.value = 1.00
                cell.number_format = "0.00"
                cell.fill = _FILL_DIAG
                cell.font = _FONT_LIGHT
            elif j < i:
                r_val = matrix[i][j]
                p_val = p_values[i][j] if i < len(p_values) and j < len(p_values[i]) else None
                if r_val is None:
                    cell.value = "N/A"
                    cell.fill, cell.font = _FILL_NA, _FONT_GREY
                else:
                    cell.value = round(float(r_val), 2)
                    cell.number_format = "0.00"
                    cell.fill, cell.font = _cell_style(float(r_val), p_val)
            # j > i（上三角）留空

    # ── 2. 配对明细（按 |r| 降序，与 HTML 一致） ──
    if pairs:
        row += 2
        row = write_title_row(ws, row, "配对明细（按 |r| 降序）", ncols=ncols)
        row = write_header_row(ws, row, _PAIRS_HEADERS)
        for idx, p in enumerate(pairs, start=1):
            _sig = bool(p.get("significant"))
            row_data = [
                idx,
                f"{p.get('name_a', p.get('code_a', ''))} ({p.get('code_a', '')})",
                f"{p.get('name_b', p.get('code_b', ''))} ({p.get('code_b', '')})",
                p.get("pearson", 0.0),
                p.get("p_value", 1.0),
                p.get("samples", 0),
                "✅ 显著" if _sig else "— 不显著",
            ]
            write_data_row(ws, row, row_data, formats=[None, None, None, "0.00", "0.0000", None, None])
            _corr_cell = ws.cell(row=row, column=4)
            _r = float(p.get("pearson", 0.0))
            if _r >= 0.3:
                _corr_cell.font = _FONT_POS_COLOR
            elif _r <= -0.3:
                _corr_cell.font = _FONT_NEG_COLOR
            else:
                _corr_cell.font = _FONT_GREY
            row += 1

    # ── 3. 说明区 ──
    row += 2
    row = write_title_row(ws, row, "说明", ncols=ncols)
    notes = [
        f"计算窗口：最近 {correlation_data.get('window', 0)} 个交易日重叠样本；有效样本：{correlation_data.get('sample_count', 0)} 期",
        "相关系数 r = 两品种日收益率的 Pearson 相关系数；显著列为 95% 双尾 t 检验结果（p < 0.05）",
        "红=正相关（同向波动，伪分散风险），蓝=负相关（反向对冲），白=不显著，灰=重叠样本不足",
    ]
    insufficient = correlation_data.get("insufficient_codes") or []
    if insufficient:
        notes.append(f"下列品种重叠样本不足窗口期，相关性格标为 N/A（灰色）：{'、'.join(insufficient)}")
    for note in notes:
        write_data_row(ws, row, [note] + [""] * (ncols - 1))
        row += 1

    return row


# ── 页签入口 ─────────────────────────────────────────────────


def write_position_relationship_sheet(
    ws: Worksheet,
    overlap_result: dict[str, Any] | None = None,
    fund_names: dict[str, str] | None = None,
    correlation_data: dict[str, Any] | None = None,
) -> None:
    """写入持仓关系矩阵页签（一章两区块：持仓重合度 + 持仓相关性）。

    Args:
        ws: openpyxl Worksheet 对象
        overlap_result: compute_overlap_matrix() 的返回结果；None 或基金数 < 2 时重合度区块写占位。
        fund_names: {fund_code: fund_name} 覆盖默认名称显示（重合度区块）。
        correlation_data: `position_relationship_data` 契约 dict（相关性区块数据源）；
            None 或 available=False 时相关性区块写占位。
    """
    _name = get_report_sheet_name("position_relationship")
    ncols = _compute_ncols(overlap_result, correlation_data)
    write_title_row(ws, 1, f"{get_report_section_number('position_relationship')}. {_name}", ncols=ncols)

    row = 2
    row = _write_overlap_block(ws, row, overlap_result, fund_names, ncols)
    row = _write_correlation_block(ws, row, correlation_data, ncols)

    freeze_header(ws, row=2)
    auto_width(ws, min_width=10, max_width=30)
    logger.info("持仓关系矩阵页签写入完成")
