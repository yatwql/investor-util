"""LLM API 用量页签写入模块。

职责：LLM 用量汇总/模块明细/状态着色/缓存统计/列宽设置。
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.python.cache import get_cache_hit_rate
from src.python.report.excel_writer import write_title_row

logger = logging.getLogger("invest")


def _init_llm_usage_sheet(ws: Any) -> int:
    """初始化 LLM 用量页签内容，返回当前行号。"""
    title = ws.title
    row = write_title_row(ws, 1, title, 10)
    row += 1
    _SUB_FONT = Font(size=9, color="666666")
    ws.cell(
        row=row, column=1, value="以下展示本次 LLM 全量生成的 API 调用统计和模块明细，帮助了解 Token 消耗和费用构成。"
    )
    ws.cell(row=row, column=1).font = _SUB_FONT
    row += 2
    return row


def _write_llm_summary_section(
    ws: Any,
    row: int,
    session_usage: dict[str, Any] | None,
    llm_endpoint: str = "",
    debate_mode_label: str | None = None,
) -> int:
    """写入 LLM 用量汇总数据区，返回下一行号。"""
    if not session_usage or not session_usage.get("has_usage"):
        return row

    _SECTION_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    _SECTION_FONT = Font(size=10, bold=True, color="1A1A1A")
    _KV_KEY_FONT = Font(size=10, bold=True, color="2E75B6")
    _KV_VAL_FONT = Font(size=10)

    for ci in range(1, 3):
        ws.cell(row=row, column=ci).fill = _SECTION_FILL
    ws.cell(row=row, column=1, value="▎汇总数据").font = _SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    pairs = [
        ("API 调用次数", f"{session_usage.get('call_count', 0)} 次"),
        ("模型", session_usage.get("model_display", "未指定")),
        ("Endpoint", llm_endpoint or "—"),
        ("输入 Token", f"{session_usage.get('input_tokens', 0):,}"),
        ("输出 Token", f"{session_usage.get('output_tokens', 0):,}"),
        ("总 Token", f"{session_usage.get('total_tokens', 0):,}"),
    ]
    _cache_hit = session_usage.get("cache_hit_tokens", 0)
    if _cache_hit:
        pairs.append(("缓存命中 Token", f"{_cache_hit:,}"))
    pairs.append(("累计费用", session_usage.get("cost_display", "—")))

    for key, val in pairs:
        ws.cell(row=row, column=1, value=key).font = _KV_KEY_FONT
        ws.cell(row=row, column=2, value=val).font = _KV_VAL_FONT
        row += 1
    if debate_mode_label:
        ws.cell(row=row, column=1, value="实验模式").font = _KV_KEY_FONT
        ws.cell(row=row, column=2, value=debate_mode_label).font = _KV_VAL_FONT
        row += 1
    return row + 1


def _write_module_table_header(ws: Any, row: int, headers: list[str]) -> int:
    """写入「各模块明细」区域标题 + 列头，返回下一行号。"""
    ncols = len(headers)
    _SECTION_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    _SECTION_FONT = Font(size=10, bold=True, color="1A1A1A")
    _HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _THIN_BORDER = Border(bottom=Side(style="thin", color="d0d0d0"))

    for ci in range(1, ncols + 1):
        ws.cell(row=row, column=ci).fill = _SECTION_FILL
    ws.cell(row=row, column=1, value="▎各模块明细").font = _SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    row += 1

    header_font = Font(size=10, bold=True, color="333333")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
    return row + 1


def _write_module_data_rows(ws: Any, row: int, module_info: list[dict]) -> int:
    """写入各模块明细行，返回下一行号。"""
    _KV_VAL_FONT = Font(size=10)
    _STATUS_COLORS = {
        "disabled": "9ca3af",
        "failed": "c0392b",
        "cached": "2e86c1",
        "success": "27ae60",
    }
    _THIN_BORDER = Border(bottom=Side(style="thin", color="d0d0d0"))
    ncols = 10

    for mi in module_info:
        if not mi.get("status_label"):
            continue

        ws.cell(row=row, column=1, value=mi.get("name", "")).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        _sc = _STATUS_COLORS.get(mi.get("status", ""), "999999")
        ws.cell(row=row, column=2, value=mi.get("status_label", "")).font = Font(size=10, color=_sc)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=row, column=3, value=mi.get("model") or "—").font = _KV_VAL_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center")

        _token_fields = ["total_tokens", "input_tokens", "output_tokens", "cache_hit_tokens"]
        right_align = Alignment(horizontal="right", vertical="center")
        for ci, tf in enumerate(_token_fields, 4):
            val = mi.get(tf, 0)
            if val:
                ws.cell(row=row, column=ci, value=f"{val:,}").font = _KV_VAL_FONT
            else:
                ws.cell(row=row, column=ci, value="—").font = Font(size=9, color="cccccc")
            ws.cell(row=row, column=ci).alignment = right_align

        _cost = mi.get("cost", 0.0)
        _status_val = mi.get("status", "")
        if _cost > 0:
            from src.python.llm.pricing import CURRENCY_SYMBOLS, PRICING_CURRENCY

            _sym = CURRENCY_SYMBOLS.get(PRICING_CURRENCY, "¥")
            ws.cell(row=row, column=8, value=f"{_sym}{_cost:.4f}").font = _KV_VAL_FONT
            ws.cell(row=row, column=8).alignment = right_align
        elif _status_val == "cached":
            ws.cell(row=row, column=8, value="已计入原调用").font = Font(size=9, color="999999")
            ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws.cell(row=row, column=8, value="—").font = Font(size=9, color="cccccc")
            ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="center")

        _cached = mi.get("cached", False)
        if _cached:
            ws.cell(row=row, column=9, value="✓").font = Font(size=10, color="2e86c1")
        else:
            ws.cell(row=row, column=9, value="—").font = Font(size=9, color="cccccc")
        ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")

        _thinking = mi.get("thinking", False)
        if _thinking:
            ws.cell(row=row, column=10, value="✓").font = Font(size=10, color="8e44ad")
        else:
            ws.cell(row=row, column=10, value="—").font = Font(size=9, color="cccccc")
        ws.cell(row=row, column=10).alignment = Alignment(horizontal="center", vertical="center")

        for ci in range(1, ncols + 1):
            ws.cell(row=row, column=ci).border = _THIN_BORDER
        row += 1
    return row


def _write_legend(ws: Any, row: int) -> None:
    """写入底部状态颜色图例。"""
    row += 1
    ws.cell(row=row, column=1, value="状态标色说明：").font = Font(size=9, bold=True, color="999999")
    row += 1
    legends = [
        ("成功  ", "27ae60", "API 调用成功，生成正常"),
        ("缓存  ", "2e86c1", "结果来自 LLM 缓存，未发起 API 请求"),
        ("已禁用  ", "9ca3af", "该模块在配置中未启用"),
        ("已失败  ", "c0392b", "API 调用失败/超时/熔断/未配置"),
    ]
    for lbl, clr, desc in legends:
        ws.cell(row=row, column=1, value=lbl).font = Font(size=9, color=clr)
        ws.cell(row=row, column=2, value=desc).font = Font(size=8, color="999999")
        row += 1
    row += 1
    ws.cell(
        row=row,
        column=1,
        value="注：API 调用次数统计实际发起的 API 请求总数（含截断后自动重试等），"
        "各模块明细仅列最终结果；两者不一致属正常现象。",
    ).font = Font(size=8, color="999999")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)


def _write_cache_stats_section(ws: Any) -> None:
    """写入数据缓存命中率统计区（仅在有过缓存请求时显示）。"""
    stats = get_cache_hit_rate()
    total = stats.get("total", 0)
    if not total:
        return

    _SECTION_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    _SECTION_FONT = Font(size=10, bold=True, color="1A1A1A")
    _KV_KEY_FONT = Font(size=10, bold=True, color="2E75B6")
    _KV_VAL_FONT = Font(size=10)

    row = ws.max_row + 2

    for ci in range(1, 3):
        ws.cell(row=row, column=ci).fill = _SECTION_FILL
    ws.cell(row=row, column=1, value="▎数据缓存系统").font = _SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    rate = stats.get("rate", 0.0)
    rate_display = f"{rate * 100:.1f}%" if rate > 0 else "—"
    pairs = [
        ("缓存命中", f"{stats.get('hits', 0):,} 次"),
        ("缓存未命中", f"{stats.get('misses', 0):,} 次"),
        ("总请求", f"{total:,} 次"),
        ("命中率", rate_display),
    ]
    for key, val in pairs:
        ws.cell(row=row, column=1, value=key).font = _KV_KEY_FONT
        ws.cell(row=row, column=2, value=val).font = _KV_VAL_FONT
        row += 1


def _set_column_widths(ws: Any, widths: list[int]) -> None:
    """设置列宽并冻结标题行。"""
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A1"


def write_llm_usage_sheet(
    ws: Any,
    llm_session_usage: dict[str, Any] | None,
    llm_module_info: list[dict[str, Any]] | None,
    llm_endpoint: str = "",
    debate_mode_label: str | None = None,
) -> None:
    """写入 'LLM API 用量' 页签内容。

    Args:
        ws: 预创建的 LLM API 用量工作表
        llm_session_usage: format_session_usage() 返回值
        llm_module_info: 合并后的模块明细列表
        llm_endpoint: 全局 LLM endpoint
        debate_mode_label: 辩论模式标签，非 None 时在汇总区显示实验模式行
    """
    if not llm_module_info:
        return

    _HEADERS = [
        "模块",
        "状态",
        "模型",
        "总 Token 用量",
        "输入 Token",
        "输出 Token",
        "缓存命中 Token",
        "费用",
        "LLM 缓存",
        "Thinking",
    ]

    row = _init_llm_usage_sheet(ws)
    row = _write_llm_summary_section(
        ws, row, llm_session_usage, llm_endpoint=llm_endpoint, debate_mode_label=debate_mode_label
    )

    row = _write_module_table_header(ws, row, _HEADERS)
    row = _write_module_data_rows(ws, row, llm_module_info)
    _write_legend(ws, row)
    _write_cache_stats_section(ws)
    _set_column_widths(ws, [20, 16, 26, 16, 14, 14, 18, 16, 12, 12])

    logger.info("LLM API 用量页签写入完成")
