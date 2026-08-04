"""数据质量仪表盘页签写入 —「数据源可用性矩阵」章节改造。

`report_submodules.data_quality` 开关（默认关）开启时，该页签以
「数据质量仪表盘」标题展示三个区块：

  区块 A 源健康   — 数据源可用性矩阵（现状保留，来自 DegradationTracker 聚合）
  区块 B 品种覆盖 — 逐品种数据状态清单（`position_status` 契约，
                     由品种覆盖诊断组装）
  区块 C 可信度   — 逐品种数据新鲜度分类 + 单日 ±20% 异常跳变检测
                     （`data_freshness` 契约，由数据新鲜度诊断组装）

开关关闭时，该章保持旧「数据源可用性矩阵」样式（见 `excel_generator`
的开关分支）。本模块仅在开关开启时被调用。
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.styles import Font

from src.python.report.excel_writer import auto_width, write_data_row, write_header_row, write_title_row

logger = logging.getLogger("invest")

# 品种覆盖区块降级占位（未获取行情数据时提示）
_COVERAGE_PLACEHOLDER = "未获取行情数据，品种覆盖无法判定"

_FONT_RED = Font(color="CC0000")
_FONT_GREEN = Font(color="009900")
_FONT_ORANGE = Font(color="E67E22")


# ── 数据规范化 ──────────────────────────────────────────────


def build_coverage_block(position_status: dict | None) -> dict:
    """规范化品种覆盖区块数据，供 Excel/HTML 渲染。

    Args:
        position_status: `position_status` 契约 dict（build_coverage_summary
            输出），为 None 或 available=False 表示品种覆盖数据不可用。

    Returns:
        {"has_data": bool, "items": list[dict], "abnormal_count": int,
         "summary": str}；不可用时 has_data=False、items 为空。
    """
    if not position_status or not position_status.get("available"):
        return {"has_data": False, "items": [], "abnormal_count": 0, "summary": ""}
    items = position_status.get("items") or []
    return {
        "has_data": True,
        "items": items,
        "abnormal_count": int(position_status.get("abnormal_count") or 0),
        "summary": position_status.get("summary") or "",
    }


# ── Excel 页签写入 ──────────────────────────────────────────


def write_data_quality_sheet(
    ws,
    matrix: list[dict],
    position_status: dict | None,
    data_freshness: dict | None = None,
) -> int:
    """写入数据质量仪表盘页签（源健康 + 品种覆盖 + 可信度三区块）。

    Args:
        ws: openpyxl worksheet
        matrix: build_data_source_matrix() 输出的源健康矩阵行列表
        position_status: `position_status` 契约 dict，None 表示无行情数据
        data_freshness: `data_freshness` 契约 dict（新鲜度 + 单日跳变），
            None 表示可信度数据不可用

    Returns:
        写入结束行号
    """
    ncols = 5
    row = write_title_row(ws, 1, "数据质量仪表盘", ncols)
    row = _write_source_health_block(ws, row, matrix, ncols)
    row = _write_coverage_block(ws, row, position_status, ncols)
    row = _write_freshness_block(ws, row, data_freshness, ncols)
    auto_width(ws)
    logger.info("数据质量仪表盘页签已写入")
    return row


def _write_source_health_block(ws, row: int, matrix: list[dict], ncols: int) -> int:
    """写入区块 A：源健康（数据源可用性矩阵，现状保留）。

    Returns:
        区块结束行号
    """
    row = write_title_row(ws, row, "源健康（数据源可用性）", ncols)
    row = write_header_row(ws, row, ["数据源", "状态", "详情", "成功", "失败/降级"])
    for m in matrix:
        if m["status"] == "ok":
            status_label = "✅ 正常"
            _font = _FONT_GREEN
        elif m["status"] == "degraded":
            status_label = "⚠️ 降级"
            _font = _FONT_ORANGE
        else:
            status_label = "❌ 失败"
            _font = _FONT_RED
        row = write_data_row(
            ws,
            row,
            [m["name"], status_label, m["detail"], m["ok"], f"{m['degraded']}/{m['failed']}"],
        )
        if m["status"] != "ok":
            for col in range(1, 6):
                ws.cell(row=row - 1, column=col).font = _font

    _degraded_list = [dg for m in matrix for dg in m.get("degraded_list", [])]
    if _degraded_list:
        row += 1
        row = write_title_row(ws, row, "降级明细", ncols)
        for m in matrix:
            for dg in m.get("degraded_list", []):
                row = write_data_row(ws, row, [m["name"], dg, "", "", ""])

    _failures = [sf for m in matrix for sf in m.get("sample_failures", [])]
    if _failures:
        row += 1
        row = write_title_row(ws, row, "失败明细", ncols)
        for m in matrix:
            for sf in m.get("sample_failures", []):
                row = write_data_row(ws, row, [m["name"], sf, "", "", ""])
    return row


def _write_coverage_block(ws, row: int, position_status: dict | None, ncols: int) -> int:
    """写入区块 B：品种覆盖（逐品种数据状态清单）。

    position_status 不可用（None / available=False）时写降级占位。

    Returns:
        区块结束行号
    """
    row += 1
    row = write_title_row(ws, row, "品种覆盖（逐品种数据状态）", ncols)
    block = build_coverage_block(position_status)
    if not block["has_data"]:
        row = write_data_row(ws, row, [_COVERAGE_PLACEHOLDER, "", "", "", ""])
        return row

    row = write_header_row(ws, row, ["代码", "名称", "账户", "状态", "原因"])
    for item in block["items"]:
        abnormal = item["status"] != "ok"
        row = write_data_row(
            ws,
            row,
            [item["code"], item["name"], item["account"], item["status_label"], item["reason"]],
        )
        if abnormal:
            for col in range(1, 6):
                ws.cell(row=row - 1, column=col).font = _FONT_RED
    return row


# 可信度区块降级占位（未获取可信度数据时提示）
_FRESHNESS_PLACEHOLDER = "未获取可信度数据，新鲜度/跳变无法判定"


def _write_freshness_block(ws, row: int, data_freshness: dict | None, ncols: int) -> int:
    """写入区块 C：可信度（逐品种新鲜度分类 + 单日跳变）。

    data_freshness 不可用（None / available=False）时写降级占位。

    Returns:
        区块结束行号
    """
    row += 1
    row = write_title_row(ws, row, "可信度（数据新鲜度 + 单日跳变）", ncols)
    if not data_freshness or not data_freshness.get("available"):
        row = write_data_row(ws, row, [_FRESHNESS_PLACEHOLDER, "", "", "", ""])
        return row

    row = write_header_row(ws, row, ["代码", "名称", "账户", "新鲜度", "单日变化/跳变"])
    for item in data_freshness.get("items") or []:
        freshness_label = item.get("freshness_label", "") or ""
        jump_label = item.get("jump_label", "") or ""
        change_pct = item.get("change_pct", 0.0) or 0.0
        change_text = f"{change_pct:+.2f}%"
        detail_text = jump_label if jump_label else change_text
        abnormal = bool(item.get("jump")) or item.get("freshness") in ("stale", "degraded")
        row = write_data_row(
            ws,
            row,
            [item.get("code", ""), item.get("name", ""), item.get("account", ""), freshness_label, detail_text],
        )
        if abnormal:
            for col in range(1, 6):
                ws.cell(row=row - 1, column=col).font = _FONT_RED
    return row
