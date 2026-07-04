"""持仓重合度矩阵 Excel 写入模块 — 报告页签 14。

输出内容：
  - 基金 × 基金对称矩阵，热力图着色
  - 矩阵下方：按重叠度降序排列的 TOP 配对 + 共同标的明细
  - 仅 >= 2 只基金时渲染矩阵

Excel 着色规则：
  >50%  红色背景（高重叠预警）
  30-50% 橙色背景（中等重叠）
  15-30% 黄色背景（低重叠）
  <15%  默认背景
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.python.report.excel_writer import auto_width, write_data_row, write_header_row, write_title_row

logger = logging.getLogger("invest")

_NCOLS_BASE = 4  # 序号 + 基金A + 基金B + 重合度
_NCOLS_DETAIL = 6  # 带共同标的矩阵：序号 + 基金A + 基金B + 重合度 + 共同数 + 共同标的

# ── 热力图颜色 ────────────────────────────────────────────────

_FILL_RED = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
_FILL_ORANGE = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
_FILL_GREEN = PatternFill(start_color="AAFFAA", end_color="AAFFAA", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")


def _heat_fill(pct: float) -> PatternFill | None:
    """根据比例返回热力图填充色。"""
    if pct >= 0.50:
        return _FILL_RED
    elif pct >= 0.30:
        return _FILL_ORANGE
    elif pct >= 0.15:
        return _FILL_YELLOW
    elif pct > 0:
        return _FILL_GREEN
    return None


def _format_pct(value: float) -> str:
    """格式化百分比（0.35 → 35.00%）。"""
    return f"{value * 100:.2f}%"


def write_overlap_matrix_sheet(
    ws: Worksheet,
    overlap_result: dict[str, Any],
    fund_names: dict[str, str] | None = None,
) -> None:
    """写入持仓重合度矩阵页签。

    Args:
        ws: openpyxl Worksheet 对象
        overlap_result: compute_overlap_matrix() 的返回结果
        fund_names: {fund_code: fund_name} 覆盖默认名称显示
    """
    funds = overlap_result.get("funds", [])
    matrix = overlap_result.get("matrix", [])
    pairs = overlap_result.get("pairs", [])
    n = len(funds)

    # ── 标题 ──
    write_title_row(ws, 1, "14. 持仓重合度矩阵", ncols=n + 2)

    if n < 2:
        write_data_row(ws, 2, ["无可比较的基金（至少需要 2 只基金）"])
        auto_width(ws)
        logger.info("持仓重合度矩阵：基金数 < 2，仅输出提示")
        return

    row = 2  # 当前行号

    # ── 表头：基金A / 基金B / 各基金代码列 ──
    headers = ["", ""] + [fund_names.get(c, c) if fund_names else c for c in funds]
    write_header_row(ws, row, headers)
    # 给基金名称列着色
    for col_idx in range(3, n + 3):
        ws.cell(row=row, column=col_idx).fill = _FILL_HEADER
    row += 1

    # ── 矩阵体 ──
    for i in range(n):
        code_i = funds[i]
        label_i = fund_names.get(code_i, code_i) if fund_names else code_i
        row_data = [label_i, f"({code_i})"] + [_format_pct(matrix[i][j]) for j in range(n)]
        write_data_row(ws, row, row_data)
        # 热力图着色
        for j in range(n):
            fill = _heat_fill(matrix[i][j])
            if fill:
                ws.cell(row=row, column=j + 3).fill = fill
        row += 1

    row += 1  # 空行

    # ── 配对明细表 ──
    write_title_row(ws, row, "配对明细（按重合度降序）", ncols=_NCOLS_BASE)
    row += 1

    pair_headers = ["序号", "基金A", "基金B", "重合度", "共同标的数", "共同标的"]
    write_header_row(ws, row, pair_headers)
    row += 1

    for idx, pair in enumerate(pairs, 1):
        a_name = fund_names.get(pair["fund_a"], pair["fund_a"]) if fund_names else pair["fund_a"]
        b_name = fund_names.get(pair["fund_b"], pair["fund_b"]) if fund_names else pair["fund_b"]
        common_stocks_str = "、".join(
            s.get("name", s.get("code", "")) for s in pair.get("common_stocks", [])
        ) if pair.get("common_stocks") else "—"

        overlap_pct = max(pair["jaccard"], 0.01 if pair["common_count"] > 0 else 0)

        row_data = [
            idx,
            f"{a_name}({pair['fund_a']})",
            f"{b_name}({pair['fund_b']})",
            _format_pct(overlap_pct),
            pair.get("common_count", 0),
            common_stocks_str,
        ]
        write_data_row(ws, row, row_data)
        row += 1

    auto_width(ws)
    logger.info("持仓重合度矩阵页签写入完成: %d 只基金, %d 对", n, len(pairs))
