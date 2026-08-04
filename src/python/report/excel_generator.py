"""Excel 报告生成核心函数。

通过 ProgressReporter 接口输出进度，不依赖 TUI。
"""

from __future__ import annotations

from typing import Any

from src.python.core.logger import setup_logger
from src.python.core.registry import get_report_section_order
from src.python.report.excel_fund_deep_analysis import write_fund_deep_analysis_sheets
from src.python.report.excel_content_sheets import write_content_sheets
from src.python.report.excel_llm_usage import write_llm_section_and_usage
from src.python.report.excel_market_data import resolve_indices, resolve_market_data
from src.python.report.excel_module_loader import load_report_modules
from src.python.report.excel_news_warning import write_news_sheet
from src.python.report.excel_sheet_factory import create_sheets
from src.python.report.progress import ProgressReporter, SilentProgressReporter, Timer

logger = setup_logger()


# ── 组合历史走势与回撤页签写入（合并章：走势表 + 回撤矩阵 + 危机区间标注） ──────


def _write_portfolio_history_drawdown_sheet(
    sheets: dict[str, Any],
    history_data: dict | None,
    crisis_annotation: dict[str, Any] | None = None,
    tail_risk: dict[str, Any] | None = None,
) -> None:
    """写入组合历史走势与回撤合并页签（一章两区块 + 危机区间标注 + 尾部风险）。

    Args:
        sheets: 页签字典，读取 `portfolio_history_drawdown` 键。
        history_data: `history_data` C19 契约 dict；不可用时整页写占位。
        crisis_annotation: `crisis_annotation_data` C19 契约 dict（危机区间标注）；
            None 时危机区块写占位。
        tail_risk: `tail_risk_data` C19 契约 dict（尾部风险统计）；
            None 时尾部指标行写占位。
    """
    ws = sheets.get("portfolio_history_drawdown")
    if ws is None:
        return

    data_ok = history_data and history_data.get("status") != "unavailable" and history_data.get("bars")
    effective = history_data if data_ok else None

    from src.python.report.portfolio_history_drawdown_sheet import write_portfolio_history_drawdown_sheet

    write_portfolio_history_drawdown_sheet(ws, effective, crisis_annotation, tail_risk)


def _write_data_source_matrix_sheet(ws, prog) -> None:
    """写入数据源可用性矩阵页签（旧样式，数据质量仪表盘开关关闭时使用）。

    Args:
        ws: data_source_status 页签 worksheet
        prog: 进度上报接口
    """
    prog.info("正在写入数据源可用性矩阵...")
    try:
        from src.python.report.data_source_matrix import build_data_source_matrix
        from src.python.report.excel_writer import (
            auto_width,
            write_data_row,
            write_header_row,
            write_title_row,
        )
        from openpyxl.styles import Font

        _FONT_RED = Font(color="CC0000")
        _FONT_GREEN = Font(color="009900")
        _FONT_ORANGE = Font(color="E67E22")

        matrix = build_data_source_matrix()
        if matrix:
            ncols = 5
            row = write_title_row(ws, 1, "数据源可用性矩阵", ncols)
            row = write_header_row(
                ws,
                row,
                ["数据源", "状态", "详情", "成功", "失败/降级"],
            )
            for m in matrix:
                if m["status"] == "ok":
                    status_label = "✅ 正常"
                    _font = _FONT_GREEN
                elif m["status"] == "degraded":
                    status_label = "⚠️ 降级"
                    _font = _FONT_ORANGE
                else:
                    status_label = "❌ 失败"
                    _font = _FONT_RED
                row = write_data_row(
                    ws,
                    row,
                    [m["name"], status_label, m["detail"], m["ok"], f"{m['degraded']}/{m['failed']}"],
                )
                if m["status"] != "ok":
                    for col in range(1, 6):
                        ws.cell(row=row - 1, column=col).font = _font

            has_degraded = any(m["degraded_list"] for m in matrix)
            if has_degraded:
                row += 1
                row = write_title_row(ws, row, "降级明细", ncols)
                for m in matrix:
                    for dg in m.get("degraded_list", []):
                        row = write_data_row(ws, row, [m["name"], dg, "", "", ""])

            has_failures = any(m["sample_failures"] for m in matrix)
            if has_failures:
                row += 1
                row = write_title_row(ws, row, "失败明细", ncols)
                for m in matrix:
                    for sf in m.get("sample_failures", []):
                        row = write_data_row(ws, row, [m["name"], sf, "", "", ""])
            auto_width(ws)
            logger.info("数据源可用性矩阵页签已写入")
        else:
            logger.debug("[excel] 数据源矩阵为空，跳过页签写入")
    except Exception:
        logger.debug("[excel] 数据源可用性矩阵页签写入失败（非关键）", exc_info=True)


def generate_excel_report(
    holdings: list,
    include_news: bool = False,
    output_dir: str = "reports",
    news_top_count: int = 100,
    include_llm: bool = False,
    llm_content: tuple[str | None, str | None, str | None, str | None] | None = None,
    details: list | None = None,
    a_indices: dict[str, dict[str, Any]] | None = None,
    us_indices: dict[str, dict[str, Any]] | None = None,
    news_data: list | None = None,
    news_llm_meta: dict | None = None,
    enable_fund_deep_analysis: bool = False,  # board 层：基金深度分析是否开启
    enable_news: bool = True,  # board 层：市场新闻是否开启（配置值）
    enable_llm: bool = True,  # board 层：LLM 分析章节是否开启
    enable_history: bool = True,  # board 层：历史走势章节是否开启
    enable_portfolio_evolution: bool = True,  # board 层：组合演进章节是否开启
    enable_action: bool = False,  # board 层：行动建议章节是否开启（默认关）
    enable_data_quality: bool = False,  # 子模块：数据质量仪表盘（report_submodules.data_quality）
    progress: ProgressReporter | None = None,
    section_order: list[dict] | None = None,
    pipeline_data: dict | None = None,  # 组合历史走势：环比对比数据（drives delta columns）
    history_data: dict | None = None,  # 组合历史走势数据（含基准指数）
    debate_info: dict | None = None,
    enable_cost_lots: bool = False,  # 子模块：成本流水（成本分档 + XIRR + 分红累计，report_submodules.cost_lots）
    transactions: list | None = None,  # 交易流水记录（「交易流水」页签，无则 None）
    dividends: list | None = None,  # 分红流水记录（「分红流水」页签，无则 None）
) -> None:
    """生成 Excel 报告的核心逻辑。

    Args:
        holdings: 持仓列表
        include_news: 是否包含新闻页签（data 层：菜单类型决定）
        output_dir: 输出目录
        news_top_count: 新闻条数上限
        include_llm: 是否包含 LLM 分析章节
        llm_content: 预生成的 LLM 内容元组
        details: 预获取的市值核算明细
        a_indices: A 股指数数据
        us_indices: 美股指数数据
        news_data: 预获取的新闻数据
        news_llm_meta: 新闻 LLM 元数据
        enable_fund_deep_analysis: board 层 — 基金深度分析是否开启
        enable_news: board 层 — 市场新闻是否开启（配置值）
        enable_llm: board 层 — LLM 分析章节是否开启
        enable_history: board 层 — 历史走势章节是否开启
        enable_portfolio_evolution: board 层 — 组合演进章节是否开启
        enable_data_quality: 子模块 — 数据质量仪表盘（源健康+品种覆盖），
            默认 False（向后兼容，该章保持旧「数据源可用性矩阵」样式）
        progress: 进度报告接口（默认 SilentProgressReporter，不输出）
        section_order: 可选的自定义报告模块顺序，来自 get_report_section_order(config)
        pipeline_data: 组合历史走势环比对比数据（含 diff 等），注入 summary 页签生成 δ 列对比摘要
        history_data: 组合历史走势数据（含基准指数），来自 PortfolioHistoryCalculator。
                      未提供或 status=unavailable 时页签显示占位文本。
        enable_cost_lots: 子模块 — 成本流水（成本分档 + XIRR + 分红累计）。
            默认 False（向后兼容，「投资分析汇总」/「市值核算明细表」/「持仓分类表」保持既有输出）
        transactions: 交易流水记录（「交易流水」页签），成本分档/FIFO 批次与 XIRR 现金流用
        dividends: 分红流水记录（「分红流水」页签），分红累计与 XIRR 现金流用
    """
    prog = progress if progress is not None else SilentProgressReporter()

    modules = load_report_modules(prog)
    if not modules:
        return  # excel_writer 缺失，无法继续

    create_workbook = modules["create_workbook"]
    save_workbook = modules["save_workbook"]

    # ── 创建工作簿，按需创建全部页签 ──
    wb = create_workbook()
    wb.remove(wb.active)
    order = section_order or get_report_section_order()  # 内部名 order，避免影子覆盖参数

    # 构造 data 层可用性字典
    data_availability: dict[str, bool] = {}
    if include_news:
        data_availability["news_data_available"] = True
    if include_llm:
        data_availability["llm_data_available"] = True

    sheets = create_sheets(
        wb,
        order,
        enable_fund_deep_analysis=enable_fund_deep_analysis,
        enable_news=enable_news,
        enable_history=enable_history,
        enable_portfolio_evolution=enable_portfolio_evolution,
        enable_action=enable_action,
        enable_llm=enable_llm,
        data_availability=data_availability,
    )

    # ── 行情市值 + 指数 ──
    # 成本流水子模块：开关开启时由 resolve_market_data 组装 C19 fund_flow_data
    # （成本分档 + XIRR + 分红累计，基于交易/分红流水 + 行情明细价格）
    data = resolve_market_data(
        holdings,
        details,
        modules,
        sheets["market_value"],
        prog,
        enable_cost_lots=enable_cost_lots,
        transactions=transactions,
        dividends=dividends,
    )
    a_idx, us_idx = resolve_indices(a_indices, us_indices, modules, prog)

    # ── 各页签写入 ──
    pen_result = write_content_sheets(sheets, holdings, data, a_idx, us_idx, modules, prog, enable_cost_lots=enable_cost_lots)
    write_news_sheet(sheets, holdings, pen_result, include_news, news_data, news_llm_meta, news_top_count, prog)
    # 风格与因子分析：C19 数据在编排层注入 pipeline_data（style_factor_data 主键），
    # 此处透传页签写入（一章三区块：风格表 + 因子回归 + 行业 Beta 子表）
    write_fund_deep_analysis_sheets(
        sheets,
        holdings,
        enable_fund_deep_analysis,
        data,
        modules,
        prog,
        style_factor_data=(pipeline_data or {}).get("style_factor_data"),
        position_relationship_data=(pipeline_data or {}).get("position_relationship_data"),
    )
    # 辩论模式标签（从 debate_info 提取或从 feature flag 检测）
    from src.python.report._debate_utils import detect_debate_mode

    _debate_mode_label, _debate_mode_combination = detect_debate_mode(debate_info)

    if _debate_mode_label and _debate_mode_combination:
        _debate_mode_label = f"{_debate_mode_label} · {_debate_mode_combination}"
    write_llm_section_and_usage(
        sheets, include_llm, llm_content, prog, section_order=order, debate_mode_label=_debate_mode_label
    )

    # ── 组合历史走势与回撤页签（合并章：走势表 + 回撤矩阵 + 危机区间标注） ──
    if enable_history:
        ws_hd = sheets.get("portfolio_history_drawdown")
        if ws_hd is not None:
            prog.info("正在写入组合历史走势与回撤页签...")
            try:
                _write_portfolio_history_drawdown_sheet(
                    sheets,
                    history_data,
                    (pipeline_data or {}).get("crisis_annotation_data"),
                    (pipeline_data or {}).get("tail_risk_data"),
                )
            except Exception:
                logger.debug("[excel] 组合历史走势与回撤页签写入失败（非关键）", exc_info=True)

    # ── 组合演进页签（多快照趋势，C19 evolution_data） ──
    ws_evo = sheets.get("portfolio_evolution")
    if ws_evo is not None:
        prog.info("正在写入组合演进页签...")
        try:
            from src.python.report.evolution_sheet import write_evolution_sheet

            write_evolution_sheet(
                ws_evo,
                (pipeline_data or {}).get("evolution_data"),
                snapshot_diff_data=(pipeline_data or {}).get("snapshot_diff_data"),
            )
        except Exception:
            logger.debug("[excel] 组合演进页签写入失败（非关键）", exc_info=True)

    # ── 行动建议页签（行动板块，C19 action_data） ──
    ws_action = sheets.get("action")
    if ws_action is not None:
        prog.info("正在写入行动建议页签...")
        try:
            from src.python.report.action_sheet import write_action_sheet

            write_action_sheet(ws_action, (pipeline_data or {}).get("action_data"))
        except Exception:
            logger.debug("[excel] 行动建议页签写入失败（非关键）", exc_info=True)

    # ── 数据质量仪表盘 / 数据源可用性矩阵页签 ──
    ws_ds = sheets.get("data_source_status")
    if ws_ds is not None:
        if enable_data_quality:
            # 子模块开关开启：该章改造为「数据质量仪表盘」（源健康 + 品种覆盖 + 可信度）
            prog.info("正在写入数据质量仪表盘...")
            try:
                from src.python.report.data_quality_sheet import write_data_quality_sheet
                from src.python.report.data_source_matrix import build_data_source_matrix

                write_data_quality_sheet(
                    ws_ds,
                    build_data_source_matrix(),
                    (pipeline_data or {}).get("position_status"),
                    (pipeline_data or {}).get("data_freshness"),
                )
            except Exception:
                logger.debug("[excel] 数据质量仪表盘页签写入失败（非关键）", exc_info=True)
        else:
            # 开关关闭：该章保持旧「数据源可用性矩阵」样式（向后兼容）
            _write_data_source_matrix_sheet(ws_ds, prog)

    # ── 组合历史走势：环比对比摘要（写入 summary 页签底部） ──
    if pipeline_data and pipeline_data.get("diff") and "summary" in sheets:
        prog.info("正在写入环比对比摘要...")
        try:
            _diff = pipeline_data["diff"]
            _ws_sum = sheets["summary"]
            # 找到最后一行的行号
            _last_row = _ws_sum.max_row + 2
            from openpyxl.styles import Font

            _section_font = Font(size=12, bold=True, color="2E75B6")
            _bold_font = Font(bold=True)
            # 标题行
            _ws_sum.cell(row=_last_row, column=1, value="【环比对比】").font = _section_font
            _last_row += 1
            # 对比数据
            if _diff.get("days_since_last_report") is not None:
                _ws_sum.cell(row=_last_row, column=1, value="距上次报告")
                _ws_sum.cell(row=_last_row, column=2, value=f"{_diff['days_since_last_report']} 天")
                _last_row += 1
            if _diff.get("total_value_diff") is not None:
                _ws_sum.cell(row=_last_row, column=1, value="总市值变化").font = _bold_font
                _cell = _ws_sum.cell(row=_last_row, column=2, value=_diff["total_value_diff"])
                _cell.number_format = "#,##0.00"
                _cell.font = Font(color="CC0000" if _diff["total_value_diff"] >= 0 else "009900")
                _last_row += 1
            if _diff.get("total_value_diff_pct") is not None:
                _ws_sum.cell(row=_last_row, column=1, value="总市值变化率")
                _cell = _ws_sum.cell(row=_last_row, column=2, value=round(_diff["total_value_diff_pct"] / 100, 4))
                _cell.number_format = "0.00%"
                _last_row += 1
            if _diff.get("total_pnl_diff") is not None:
                _ws_sum.cell(row=_last_row, column=1, value="总盈亏变化").font = _bold_font
                _cell = _ws_sum.cell(row=_last_row, column=2, value=_diff["total_pnl_diff"])
                _cell.number_format = "#,##0.00"
                _cell.font = Font(color="CC0000" if _diff["total_pnl_diff"] >= 0 else "009900")
                _last_row += 1
            # 持仓变动概要
            for _label, _key in [
                ("新增持仓", "added"),
                ("清仓标的", "removed"),
                ("增持标的", "increased"),
                ("减持标的", "decreased"),
            ]:
                _items = _diff.get(_key, [])
                if _items:
                    _ws_sum.cell(row=_last_row, column=1, value=_label).font = _bold_font
                    _names = ", ".join(f"{i.get('name', '')}({i.get('code', '')})" for i in _items[:5])
                    if len(_items) > 5:
                        _names += f" 等{len(_items)}只"
                    _ws_sum.cell(row=_last_row, column=2, value=_names)
                    _last_row += 1
            logger.info("环比对比摘要已写入 summary 页签")
        except Exception:
            logger.debug("[F delta] Excel 环比对比摘要写入失败（非关键）", exc_info=True)

    # ── 保存 ──
    with Timer("保存 Excel/HTML 文件"):
        # 在每个页签底部写入隐私声明脚注
        for _ws_name, _ws in sheets.items():
            if _ws is not None:
                try:
                    from src.python.report.excel_writer import write_privacy_footer

                    write_privacy_footer(_ws, ncols=5)
                except Exception:
                    logger.debug("[privacy] 页签 %s 写入隐私脚注失败（非关键）", _ws_name, exc_info=True)
        prog.info("正在保存 Excel 报告...")
        path = save_workbook(wb, output_dir=output_dir)
        logger.info("Excel 报告已生成: %s", path)
        logger.info(
            "总市值: %.2f元, 总成本: %.2f元, 总盈亏: %.2f元, 本日盈亏: %.2f元",
            data["total_mv"],
            data["total_cost"],
            data["total_profit"],
            data["today_profit"],
        )
        prog.ok(f"Excel 报告已保存: {path}")
