"""行动建议 Excel 写入模块 — 20 章「行动建议」页签。

决策闭环的核心产出页签，展示四个行动子块：
  1. 再平衡信号 — 单品占比超警戒线（复用 simple_rebalance 计算）
  2. 交易纪律   — 止盈/止损/回撤触发（analysis/trade_discipline 计算）
  3. 调仓建议   — 可行化调仓清单（analysis/rebalance_advisor 计算，份额取整/费用/现金）
  4. 收益归因   — 品种收益贡献占比（后续轮次填充）

数据源为 C19 `action_data` 契约（`analysis/action_advisor.build_action_data` 组装、
orchestrator 注入 pipeline_data）。子块为空时写「暂无」占位——收益归因框架先行，
后续轮次逐步填充，报告结构保持稳定。

数据不可用（available=False）时写入占位文本（§1.4.5 数据降级治理）。
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.python.core.registry import get_report_section_number, get_report_sheet_name
from src.python.report.excel_writer import (
    auto_width,
    write_data_row,
    write_header_row,
    write_title_row,
)

logger = logging.getLogger("invest")

_FONT_ACCENT = Font(size=12, bold=True, color="2E75B6")
_FONT_SUB_BLOCK = Font(bold=True, color="404040")
_FONT_WARN = Font(color="CC0000")

# 空子块占位（框架先行，后续轮次填充时被真实数据替换）
_PLACEHOLDER_EMPTY = "暂无触发"
_PLACEHOLDER_UNAVAILABLE = "无持仓数据，行动建议无法生成"


def _write_sub_block(
    ws: Worksheet,
    row: int,
    title: str,
    items: list[dict[str, Any]],
    headers: list[str],
    ncols: int,
    row_values,
    placeholder: str = _PLACEHOLDER_EMPTY,
) -> int:
    """写入一个行动子块（标题 + 表头 + 数据行，空时写占位）。

    Args:
        ws: worksheet
        row: 起始行号
        title: 子块标题
        items: 子块数据行列表
        headers: 表头
        ncols: 表格列数
        row_values: 数据行取值函数（item → list）
        placeholder: 空数据占位文本

    Returns:
        子块结束行号
    """
    row += 1
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _FONT_SUB_BLOCK
    row += 1
    if not items:
        row = write_data_row(ws, row, [placeholder] + [""] * (ncols - 1))
        return row
    row = write_header_row(ws, row, headers[:ncols])
    for item in items:
        row = write_data_row(ws, row, row_values(item))
    return row


def write_action_sheet(ws: Worksheet, action_data: dict[str, Any] | None) -> None:
    """写入行动建议页签。

    Args:
        ws: openpyxl Worksheet 对象
        action_data: C19 `action_data` 契约 dict；None 或 available=False 时写占位。
    """
    _name = get_report_sheet_name("action")
    _ncols = 5
    write_title_row(ws, 1, f"{get_report_section_number('action')}. {_name}", ncols=_ncols)
    row = 2

    if not action_data or not action_data.get("available"):
        row = write_data_row(ws, row, [_PLACEHOLDER_UNAVAILABLE, "", "", "", ""])
        auto_width(ws)
        logger.info("行动建议：无持仓数据，写入占位")
        return

    # 行动摘要行
    _summary = (action_data.get("summary") or "").strip()
    if _summary:
        ws.cell(row=row, column=1, value=_summary).font = _FONT_ACCENT
        row += 1

    # 子块 1：再平衡信号（单品超限）
    row = _write_sub_block(
        ws,
        row,
        "再平衡信号（单品占比超警戒线）",
        action_data.get("rebalance_signals") or [],
        ["代码", "名称", "占比", "警戒线", "建议动作"],
        _ncols,
        lambda s: [
            s.get("code", ""),
            s.get("name", ""),
            f"{s.get('weight', 0) * 100:.1f}%",
            f"{s.get('threshold', 0) * 100:.0f}%",
            s.get("action", ""),
        ],
        placeholder="组合内无品种超警戒线",
    )

    # 子块 2：交易纪律（框架，后续轮次填充）
    row = _write_sub_block(
        ws,
        row,
        "交易纪律（止盈/止损/回撤）",
        action_data.get("discipline_signals") or [],
        ["代码", "名称", "规则", "当前值", "触发状态"],
        _ncols,
        lambda s: [
            s.get("code", ""),
            s.get("name", ""),
            s.get("rule", ""),
            s.get("value", ""),
            s.get("status_label", ""),
        ],
    )

    # 子块 3：调仓建议（可行化清单：份额取整一手/费用估算/现金缓冲）
    row = _write_sub_block(
        ws,
        row,
        "调仓建议清单",
        action_data.get("rebalance_advice") or [],
        ["代码", "名称", "操作", "份额", "金额", "预估费用", "调仓后现金"],
        7,
        lambda s: [
            s.get("code", ""),
            s.get("name", ""),
            s.get("operation", ""),
            s.get("shares", ""),
            s.get("amount", ""),
            s.get("fee", ""),
            s.get("cash_after", ""),
        ],
    )

    # 子块 4：收益归因（框架，后续轮次填充）
    _attr = action_data.get("attribution")
    row += 1
    ws.cell(row=row, column=1, value="收益归因（品种贡献占比）").font = _FONT_SUB_BLOCK
    row += 1
    if not _attr or not _attr.get("available"):
        row = write_data_row(ws, row, ["待生成", "", "", "", ""])
    else:
        row = write_header_row(ws, row, ["来源", "品种", "贡献占比", "盈亏金额", ""])
        for src in ("盈利来源", "亏损来源"):
            for item in _attr.get(src) or []:
                row = write_data_row(
                    ws,
                    row,
                    [src, item.get("name", ""), item.get("contribution_pp", ""), item.get("profit", ""), ""],
                )

    auto_width(ws)
    logger.info("行动建议页签已写入")
