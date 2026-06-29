"""市值核算模块 — 报告第 2 页。

15 列明细表，含分账户小计和总计。盈亏数值正数红色、负数绿色。"""

from __future__ import annotations

import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Any, List

from openpyxl.worksheet.worksheet import Worksheet

from src.python import cache
from src.python.fetcher import fetch_market_data
from src.python.models import Holding
from src.python.report.excel_writer import auto_width, freeze_header, write_data_row, write_header_row, write_subtotal_row, \
    write_title_row, write_total_row
from src.python.report.styles import BLUE_FONT, FMT_MONEY, FMT_PERCENT, FMT_PRICE, FMT_SHARES, NORMAL_FONT, profit_font

logger = logging.getLogger("invest")

# 15 列表头
_HEADERS = [
    "账户", "名称", "代码", "最新价", "净值日期", "昨日价",
    "取价方式", "溢价率", "份额", "市值", "成本",
    "盈亏", "收益率", "本日盈亏", "取价渠道",
]
_NCOLS = len(_HEADERS)

# 常见管理费率，用于估算 QDII 溢价（简化处理，不考虑实时溢价）
_FUND_PREMIUM_PLACEHOLDER = "--"


@dataclass
class DetailRow:
    """单条持仓估值明细（15 列对应字段）。"""
    account: str = ""
    name: str = ""
    code: str = ""
    price: float = 0.0
    nav_date: str = ""
    yesterday_close: float = 0.0
    price_type: str = ""
    premium: str = _FUND_PREMIUM_PLACEHOLDER
    shares: float = 0.0
    market_value: float = 0.0
    cost: float = 0.0
    profit: float = 0.0
    profit_rate: float = 0.0
    today_profit: float = 0.0
    source: str = ""
    source_api: str = ""


def _is_qdii(name: str) -> bool:
    return "QDII" in name.upper()


def _date_within_days(date_str: str, today_str: str, max_days: int) -> bool:
    """检查日期是否在指定天数内。"""
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d")
        t = datetime.strptime(today_str, "%Y-%m-%d")
        return 0 <= (t - d).days <= max_days
    except (ValueError, TypeError):
        return False


def _is_etf(name: str) -> bool:
    return "ETF" in name.upper()


def classify_holdings(holdings: List[Holding]) -> dict[str, list]:
    """按类型分类持仓。

    判断逻辑（按优先级）:
        1. 名称含 QDII → QDII
        2. 账户名为场外渠道（含基金/支付宝/微信/银行等）→ 国内场外
        3. 名称含 ETF 或代码 5/1 开头 → 场内ETF
        4. 代码 6/0/3 开头 → 场内股票
        5. 其余 → 国内场外

    用账户名做第一道区分是因为：
    - 6 位代码可能既是股票/ETF 也是场外基金（如 002943、530021）
    - 证券账户里的 5xxxxx 是 ETF，基金账户里的 5xxxxx 是场外基金
    先判断账户属性可以避免代码重叠导致的误分类。

    Returns:
        {type: [Holding, ...]}，type 取值:
        "场内股票", "场内ETF", "国内场外", "QDII"
    """
    categories: dict[str, list] = {
        "场内股票": [],
        "场内ETF": [],
        "国内场外": [],
        "QDII": [],
    }

    # 场外渠道关键词（账户名中包含则视为场外基金账户）
    _FUND_ACCOUNT_KEYWORDS = ("基金", "支付宝", "微信", "银行")

    for h in holdings:
        code = h.code.strip()
        name = h.name.strip()
        account = h.account.strip()

        # 1) QDII（名称含 QDII）
        if _is_qdii(name):
            categories["QDII"].append(h)
        # 2) 场外渠道 → 国内场外（基金账户不会持有场内品种）
        elif any(kw in account for kw in _FUND_ACCOUNT_KEYWORDS):
            categories["国内场外"].append(h)
        # 3) 场内 ETF（名称含 ETF，或代码 5/1 开头的场内品种）
        elif _is_etf(name) or code.startswith(("5", "1")):
            categories["场内ETF"].append(h)
        # 4) A 股股票（代码 6/0/3 开头）
        elif code.startswith(("6", "0", "3")):
            categories["场内股票"].append(h)
        # 5) 其余归入场外
        else:
            categories["国内场外"].append(h)

    return categories


def price_update_status(details: List[DetailRow], trading_day: str) -> tuple[int, int, bool]:
    """检查今日价格更新状态。

    判断逻辑：
    - 场内资产（tencent）：nav_date == trading_day 视为已更新
    - QDII（eastmoney + QDII）：nav_date 在 trading_day 3 天内视为已更新
    - 国内场外（eastmoney + 非 QDII）：nav_date == trading_day 或
      nav_date == prev_trading_day 视为已更新

    Args:
        details: 明细行列表
        trading_day: 最近交易日 YYYY-MM-DD

    Returns:
        (已更新数量, 总数, 是否全部更新)
    """
    total = len(details)
    updated = 0
    prev_td = get_prev_trading_day(trading_day)
    for d in details:
        if d.source_api == "tencent":
            # 场内资产：净值日期等于交易日即视为已更新
            if d.nav_date == trading_day:
                updated += 1
        elif d.source_api == "eastmoney" and _is_qdii(d.name):
            # QDII：净值日期在 3 天内即视为已更新
            if d.nav_date and _date_within_days(d.nav_date, trading_day, 3):
                updated += 1
        elif d.source_api == "eastmoney":
            # 国内场外：净值日期等于交易日或前一个交易日即视为已更新
            if d.nav_date == trading_day or (prev_td and d.nav_date == prev_td):
                updated += 1
    return updated, total, updated >= total


def is_market_open() -> bool:
    """检查当前是否为 A 股交易时间。

    A 股交易时间：周一至周五 9:30-11:30, 13:00-15:00
    周末及非交易时段返回 False（即取价方式应为"收盘价"）。

    Returns:
        True 表示正处交易时段，False 表示已收市或休市
    """
    now = datetime.now()
    # 周末不开市
    if now.weekday() >= 5:  # 5=Saturday, 6=Sunday
        return False
    hour = now.hour
    minute = now.minute
    time_decimal = hour + minute / 60.0
    # 9:30-11:30 或 13:00-15:00
    return (9.5 <= time_decimal <= 11.5) or (13.0 <= time_decimal <= 15.0)


# ── 交易日历（节假日感知） ───────────────────────────────
_TRADING_CALENDAR_CACHE_KEY = "trading_calendar"
_TRADING_CALENDAR_TTL = 86400 * 7  # 7 天刷新一次


def _get_trading_calendar() -> set[str]:
    """获取 A 股交易日历（YYYY-MM-DD 字符串集合）。

    通过 akshare 获取全年交易日数据并缓存。若获取失败，返回空集合，
    由调用方（get_last_trading_day）回退到简易周度判断。

    Returns:
        交易日日期字符串集合
    """
    cached = cache.get(_TRADING_CALENDAR_CACHE_KEY, _TRADING_CALENDAR_TTL)
    if cached is not None and isinstance(cached, list):
        logger.debug("交易日历缓存命中")
        return set(cached)

    try:
        import akshare as ak

        df = ak.tool_trade_date_hist_sina()
        dates: set[str] = set(df["trade_date"].dropna().astype(str).tolist())
        if dates:
            cache.set(_TRADING_CALENDAR_CACHE_KEY, sorted(dates))
            logger.info("交易日历已更新（%d 个交易日）", len(dates))
            return dates
    except Exception as exc:
        logger.warning("获取交易日历失败: %s，使用简易节假日判断回退", exc)

    return set()


def _is_trading_day(date: datetime) -> bool:
    """判断给定日期是否为 A 股交易日。

    优先使用 akshare 日历，失败时回退到简易判断（非周六日即为交易日）。

    Args:
        date: 待判断的日期

    Returns:
        True 表示为交易日
    """
    calendar = _get_trading_calendar()
    date_str = date.strftime("%Y-%m-%d")
    if calendar:
        return date_str in calendar
    # 回退：仅排除周六日
    return date.weekday() < 5


def get_last_trading_day() -> str:
    """获取最近一个交易日（YYYY-MM-DD），含节假日感知。

    判断逻辑：
    1. 使用 akshare 交易日历判定节假日（端午、中秋、国庆等）
    2. A 股盘前（< 9:30）退回上一交易日
    3. 盘中/盘后（≥ 9:30）且当天为交易日 → 返回当天
    4. 非交易日则向前查找最近一个交易日

    Returns:
        YYYY-MM-DD 格式的交易日字符串
    """
    now = datetime.now()
    # 若盘前（< 9:30），基准日设为昨天
    if now.hour < 9 or (now.hour == 9 and now.minute < 30):
        check = now - timedelta(days=1)
    else:
        check = now

    # 从基准日起向前查找最近一个交易日
    for _ in range(14):  # 最多回溯 14 天（覆盖长假）
        if _is_trading_day(check):
            return check.strftime("%Y-%m-%d")
        check -= timedelta(days=1)

    # 极端回退（不应到达）
    return now.strftime("%Y-%m-%d")


def get_prev_trading_day(trading_day: str = "") -> str:
    """获取指定交易日的前一个交易日，含节假日感知。

    使用 akshare 交易日历向前查找，找不到时回退到简易周度判断。

    Args:
        trading_day: YYYY-MM-DD 格式的交易日，默认取最近交易日

    Returns:
        前一个交易日 YYYY-MM-DD
    """
    if not trading_day:
        trading_day = get_last_trading_day()
    try:
        dt = datetime.strptime(trading_day, "%Y-%m-%d")
        # 从 trading_day - 1 起向前查找最近一个交易日
        check = dt - timedelta(days=1)
        for _ in range(14):
            if _is_trading_day(check):
                return check.strftime("%Y-%m-%d")
            check -= timedelta(days=1)
        return (dt - timedelta(days=1)).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return ""


def _determine_price_type(source_api: str, nav_date: str, trading_day: str) -> str:
    """判断取价方式。

    "T" 指"所属交易日"（而非自然日），由 get_last_trading_day() 确定。

    Tencent（场内股票/ETF）：
      - 交易时段 → "场内实时价"
      - 已收市，nav_date == T → "场内收盘价(T)"
      - 已收市，nav_date == T-1 → "场内收盘价(T-1)"

    East Money（场外基金）：
      - nav_date == T → "官方净值(T)"
      - nav_date == T-1 → "官方净值(T-1)"
      - nav_date 为 2~5 天前 → "官方净值(T-N)"
      - nav_date 为 6 天以上 → "官方净值(YYYY-MM-DD)"

    Args:
        source_api: "tencent"（场内）或 "eastmoney"（场外）
        nav_date: 净值日期（YYYY-MM-DD，可为空）
        trading_day: 所属交易日（YYYY-MM-DD）
        is_qdii: 是否为 QDII 基金

    Returns:
        取价方式标签字符串
    """
    if source_api == "tencent":
        if is_market_open():
            return "场内实时价"
        # 已收市，用交易日作为 T
        if not nav_date:
            return "场内收盘价(--)"
        prev_td = get_prev_trading_day(trading_day)
        if nav_date == trading_day:
            return "场内收盘价(T)"
        elif prev_td and nav_date == prev_td:
            return "场内收盘价(T-1)"
        return f"场内收盘价({nav_date})"

    # 场外基金
    if not nav_date:
        return "官方净值(--)"

    prev_td = get_prev_trading_day(trading_day)
    if nav_date == trading_day:
        return "官方净值(T)"
    elif prev_td and nav_date == prev_td:
        return "官方净值(T-1)"

    # 计算天数差（用自然日大致估算，基金净值可能跳过周末）
    try:
        nav_dt = datetime.strptime(nav_date, "%Y-%m-%d")
        td_dt = datetime.strptime(trading_day, "%Y-%m-%d")
        days_diff = (td_dt - nav_dt).days
        if 2 <= days_diff <= 5:
            return f"官方净值(T-{days_diff})"
        elif days_diff > 5:
            return f"官方净值({nav_date})"
        # days_diff < 0? 不太可能
        return "官方净值(T)"
    except ValueError:
        return f"官方净值({nav_date})"


def _generate_details(holdings: List[Holding], today_str: str) -> List[DetailRow]:
    """获取所有持仓的行情数据并生成明细行（并行 HTTP 请求）。"""
    details: List[DetailRow] = []
    today_str = today_str or datetime.now().strftime("%Y-%m-%d")

    # 并行发起行情请求（缓存命中时秒回，冷启动加速最高 8×）
    future_map = {}
    with ThreadPoolExecutor(max_workers=8) as executor:
        for h in holdings:
            future = executor.submit(fetch_market_data, h.code, h.name)
            future_map[future] = h

        for future in as_completed(future_map):
            h = future_map[future]
            try:
                mkt = future.result()
            except Exception:
                logger.warning("获取行情异常: %s (%s)", h.name, h.code)
                mkt = None

            if mkt is None:
                logger.warning("无法获取行情数据: %s (%s)", h.name, h.code)
                price = 0.0
                yclose = 0.0
                nav_date = ""
                source = "--"
                source_api = ""
                price_type = "--"
            else:
                price = mkt.get("price", 0.0) or 0.0
                yclose = mkt.get("yesterday_close", 0.0) or 0.0
                nav_date = mkt.get("price_date", "")  # 统一使用 price_date
                source = mkt.get("source", "--")
                source_api = mkt.get("source_api", "")
                price_type = _determine_price_type(source_api, nav_date, get_last_trading_day())

            cost = round(h.cost_price * h.shares, 2)
            mv = round(price * h.shares, 2)
            profit = round(mv - cost, 2)
            profit_rate = profit / cost if cost > 0 else 0.0

            # 本日盈亏
            if source_api == "tencent":
                # 场内：(最新价 - 昨收盘) × 份额
                today_profit = round((price - yclose) * h.shares, 2)
            elif nav_date:
                # 场外：nav_date 是净值的所属交易日
                # 用最近交易日/前一交易日做对比，判断是否为最新可用数据
                trading_day = get_last_trading_day()
                prev_td = get_prev_trading_day(trading_day)
                if nav_date == trading_day or (prev_td and nav_date == prev_td):
                    today_profit = round((price - yclose) * h.shares, 2)
                else:
                    today_profit = 0.0
            else:
                today_profit = 0.0

            detail = DetailRow(
                account=h.account.strip(),
                name=h.name,
                code=h.code,
                price=price,
                nav_date=nav_date,
                yesterday_close=yclose,
                price_type=price_type,
                premium=_FUND_PREMIUM_PLACEHOLDER,
                shares=h.shares,
                market_value=mv,
                cost=cost,
                profit=profit,
                profit_rate=profit_rate,
                today_profit=today_profit,
                source=source,
                source_api=source_api,
            )
            details.append(detail)

    logger.info("市值核算明细数据生成完成，共 %d 条", len(details))
    return details


def _detail_to_row_values(d: DetailRow) -> List[Any]:
    """将 DetailRow 转为 Excel 行值列表。"""
    return [
        d.account,
        d.name,
        d.code,
        d.price,
        d.nav_date,
        d.yesterday_close,
        d.price_type,
        d.premium,
        d.shares,
        d.market_value,
        d.cost,
        d.profit,
        d.profit_rate,
        d.today_profit,
        d.source,
    ]


def _num_formats() -> List[str]:
    """每列的 Excel 数字格式。"""
    return [
        "",           # 1  账户
        "",           # 2  名称
        "",           # 3  代码
        FMT_PRICE,    # 4  最新价
        "",           # 5  净值日期
        FMT_PRICE,    # 6  昨日价
        "",           # 7  取价方式
        "",           # 8  溢价率
        FMT_SHARES,   # 9  份额
        FMT_MONEY,    # 10 市值
        FMT_MONEY,    # 11 成本
        FMT_MONEY,    # 12 盈亏
        FMT_PERCENT,  # 13 收益率
        FMT_MONEY,    # 14 本日盈亏
        "",           # 15 取价渠道
    ]


def _apply_profit_colors(ws, start_row: int, end_row: int,
                         profit_col: int, rate_col: int, today_col: int) -> None:
    """对盈亏列（12）、收益率列（13）、本日盈亏列（14）着色。"""
    for r in range(start_row, end_row + 1):
        for col in (profit_col, today_col):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                cell.font = profit_font(cell.value)
        # 收益率特殊处理（字符串含 % 号）
        rate_cell = ws.cell(row=r, column=rate_col)
        if isinstance(rate_cell.value, float):
            rate_cell.font = profit_font(rate_cell.value)


_PRICE_TYPE_COL = 7     # 取价方式列
_NAME_COL = 2           # 名称列（用于识别 QDII）


def _apply_price_type_colors(ws, start_row: int, end_row: int) -> None:
    """对取价方式列（第 7 列）着色：蓝色代表价格来源可靠/时效性高。

    着色规则：
      - "场内收盘价(T)" → 蓝色（最新场内收盘数据）
      - "官方净值(T)" → 蓝色（最新官方净值）
      - QDII 基金的 "官方净值(T-1)" → 蓝色（QDII 因时差延迟一天属正常）

    Args:
        ws: 目标工作表
        start_row: 起始行号（含）
        end_row: 结束行号（含）
    """
    for r in range(start_row, end_row + 1):
        cell = ws.cell(row=r, column=_PRICE_TYPE_COL)
        val = str(cell.value) if cell.value else ""

        if val in ("场内收盘价(T)", "官方净值(T)"):
            cell.font = BLUE_FONT
        elif val == "官方净值(T-1)":
            name_cell = ws.cell(row=r, column=_NAME_COL)
            name = str(name_cell.value) if name_cell.value else ""
            if _is_qdii(name):
                cell.font = BLUE_FONT


def write_market_value_sheet(ws: Worksheet, holdings: List[Holding],
                             today_str: str = "",
                             details: List[DetailRow] | None = None) -> tuple[float, float, float, float, List[DetailRow]]:
    """写入市值核算页签，返回汇总数据供汇总页签使用。

    Args:
        ws: 目标工作表
        holdings: 持仓列表
        today_str: 日期字符串（YYYY-MM-DD），默认当天
        details: 可选预计算明细行，传入时跳过内部行情获取。

    Returns:
        (总市值, 总成本, 总盈亏, 本日总盈亏, 明细行列表)
    """
    ws.title = "2. 市值核算"
    if details is None:
        details = _generate_details(holdings, today_str)

    row = write_title_row(ws, 1, "市值核算明细表", _NCOLS)
    row = write_header_row(ws, row, _HEADERS)
    data_start = row

    # 按账户分组，组内输出
    accounts: dict[str, List[DetailRow]] = {}
    for d in details:
        accounts.setdefault(d.account, []).append(d)

    grand_mv = grand_cost = grand_profit = grand_today = 0.0

    for acc_name, acc_details in accounts.items():
        # 该账户的明细行
        for d in acc_details:
            vals = _detail_to_row_values(d)
            write_data_row(ws, row, vals, _num_formats())
            row += 1

        # 小计
        acc_mv = sum(d.market_value for d in acc_details)
        acc_cost = sum(d.cost for d in acc_details)
        acc_profit = sum(d.profit for d in acc_details)
        acc_today = sum(d.today_profit for d in acc_details)
        acc_rate = acc_profit / acc_cost if acc_cost > 0 else 0.0

        subtotal_vals = [
            f"{acc_name} 小计",
            "", "", "", "", "", "", "",
            sum(d.shares for d in acc_details),
            acc_mv, acc_cost, acc_profit, acc_rate, acc_today, "",
        ]
        write_subtotal_row(ws, row, f"{acc_name} 小计",
                           subtotal_vals[1:], _NCOLS, _num_formats())
        row += 1

        grand_mv += acc_mv
        grand_cost += acc_cost
        grand_profit += acc_profit
        grand_today += acc_today

    # 总计
    grand_rate = grand_profit / grand_cost if grand_cost > 0 else 0.0
    total_vals = [
        "总计", "", "", "", "", "", "", "",
        sum(d.shares for d in details),
        grand_mv, grand_cost, grand_profit, grand_rate, grand_today, "",
    ]
    write_total_row(ws, row, "总计", total_vals[1:], _NCOLS, _num_formats())

    # 对盈亏列着色（数据行 + 小计 + 总计）
    _apply_profit_colors(ws, data_start, row, profit_col=12, rate_col=13, today_col=14)

    # 对取价方式列着色（蓝色标识最新可靠数据来源）
    _apply_price_type_colors(ws, data_start, row)

    freeze_header(ws, 2)
    auto_width(ws)

    logger.info("市值核算页签写入完成，共 %d 个账户，%d 条持仓",
                len(accounts), len(details))

    return grand_mv, grand_cost, grand_profit, grand_today, details
