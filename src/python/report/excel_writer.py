"""Excel 输出引擎 — 封装 openpyxl 的通用写表操作。"""

from __future__ import annotations

import logging
import os
import re
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.python.report.styles import (
    BOLD_FONT,
    CENTER_ALIGN,
    HEADER_FILL,
    NORMAL_FONT,
    SUBTOTAL_FILL,
    THIN_BORDER,
    TITLE_FILL,
    TITLE_FONT,
    TOTAL_FILL,
)

logger = logging.getLogger("invest")

def _ensure_reports_dir(output_dir: str) -> None:
    """创建 output_dir/YYYYMMDD/ 目录。

    Raises:
        PermissionError: 目录无写入权限
        OSError: 目录创建失败
    """
    date_str = datetime.now().strftime("%Y%m%d")
    date_dir = os.path.join(output_dir, date_str)
    os.makedirs(date_dir, exist_ok=True)
    # 验证 output_dir 可写
    test_file = os.path.join(output_dir, ".write_test")
    try:
        open(test_file, "a").close()
        os.remove(test_file)
    except (PermissionError, OSError) as e:
        raise PermissionError(
            f"输出目录 '{output_dir}' 无写入权限"
        ) from e
    # 验证存档子目录可写
    archive_test_file = os.path.join(date_dir, ".write_test")
    try:
        open(archive_test_file, "a").close()
        os.remove(archive_test_file)
    except (PermissionError, OSError) as e:
        raise PermissionError(
            f"存档子目录 '{date_dir}' 无写入权限"
        ) from e


def _latest_path(output_dir: str) -> str:
    """最新 Excel 报告路径。"""
    return os.path.join(output_dir, "个人投资分析报告.xlsx")


def _archive_path(output_dir: str) -> str:
    """存档 Excel 报告路径（带时间戳）。"""
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%H%M%S")
    fname = f"个人投资分析报告-{date_str}-{time_str}.xlsx"
    return os.path.join(output_dir, date_str, fname)


def save_workbook(wb: Workbook, output_dir: str = "reports") -> str:
    """保存 workbook 到最新路径和存档路径，返回最新文件路径。

    - 覆盖 {output_dir}/个人投资分析报告.xlsx（最新版）
    - 同时保存 {output_dir}/YYYYMMDD/个人投资分析报告-YYYYMMDD-HHmmss.xlsx（存档版）

    Args:
        wb: 已写入数据的 Workbook 对象
        output_dir: 报告输出根目录，默认 "reports"

    Returns:
        最新文件绝对路径

    Raises:
        PermissionError: 输出目录无写入权限
        OSError: 文件写入失败
    """
    _ensure_reports_dir(output_dir)
    latest = _latest_path(output_dir)
    archive = _archive_path(output_dir)

    try:
        wb.save(latest)
        logger.info("最新报告已保存: %s", latest)
    except PermissionError:
        print()
        print("  [ERR] 报告文件被占用，无法保存")
        print("     请关闭已打开的 Excel 文件后重试")
        print(f"     文件路径: {os.path.abspath(latest)}")
        logger.error("文件被占用: %s", latest)
        raise

    try:
        wb.save(archive)
        logger.info("存档报告已保存: %s", archive)
    except (PermissionError, OSError) as e:
        logger.warning("存档报告写入失败: %s", e)
        print("  [!] 存档报告写入失败（文件可能被占用），最新版已保存")

    return os.path.abspath(latest)


def create_workbook() -> Workbook:
    """创建并返回一个新的 Workbook。"""
    wb = Workbook()
    return wb


def write_title_row(ws, row: int, text: str, ncols: int) -> int:
    """写入标题行（合并单元格居中）。

    Args:
        ws: 工作表
        row: 行号（1-based）
        text: 标题文本
        ncols: 跨列数

    Returns:
        下一行起始行号
    """
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = TITLE_FILL
    return row + 1


def write_header_row(ws, row: int, headers: list[str]) -> int:
    """写入表头行。

    Args:
        ws: 工作表
        row: 行号
        headers: 列名列表

    Returns:
        下一行起始行号
    """
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    return row + 1


def write_data_row(ws, row: int, values: list[Any], formats: list[str | None] | None = None) -> int:
    """写一行数据。

    Args:
        ws: 工作表
        row: 行号
        values: 值列表
        formats: 可选的格式列表（None 表示默认文本格式）

    Returns:
        下一行起始行号
    """
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        if formats and col - 1 < len(formats) and formats[col - 1]:
            cell.number_format = formats[col - 1]
    return row + 1


def write_subtotal_row(ws, row: int, label: str, values: List[Any], ncols: int,
                       formats: list[str | None] | None = None) -> int:
    """写小计行（首列标签，后续各列居中对齐，黄底）。

    Args:
        ws: 工作表
        row: 行号
        label: 小计标签（如 "证券账户 小计"）
        values: 从第 2 列开始的数值列表（长度 = ncols - 1）
        ncols: 总列数

    Returns:
        下一行起始行号
    """
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = BOLD_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = SUBTOTAL_FILL
    cell.border = THIN_BORDER

    for col in range(2, ncols + 1):
        idx = col - 2
        v = values[idx] if idx < len(values) else None
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = SUBTOTAL_FILL
        cell.border = THIN_BORDER
        # +1 因为 values 从列 2 开始，而 formats 从列 1 起索引
        fmt_idx = idx + 1
        if formats and fmt_idx < len(formats) and formats[fmt_idx]:
            cell.number_format = formats[fmt_idx]
    return row + 1


def write_total_row(ws, row: int, label: str, values: List[Any], ncols: int,
                    formats: Optional[List[Optional[str]]] = None) -> int:
    """写总计行（绿底加粗，与上方的分隔线）。"""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = BOLD_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER

    for col in range(2, ncols + 1):
        idx = col - 2
        v = values[idx] if idx < len(values) else None
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        # +1 因为 values 从列 2 开始，而 formats 从列 1 起索引
        fmt_idx = idx + 1
        if formats and fmt_idx < len(formats) and formats[fmt_idx]:
            cell.number_format = formats[fmt_idx]
    return row + 1


def auto_width(ws, min_width: int = 8, max_width: int = 30) -> None:
    """自动调整列宽。

    Args:
        ws: 工作表
        min_width: 最小列宽
        max_width: 最大列宽
    """
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                val = str(cell.value)
                # 中文字符算 2 个宽度
                char_len = sum(2 if re.match(r"[一-鿿]", c) else 1 for c in val)
                max_len = max(max_len, char_len)
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def freeze_header(ws, row: int = 1) -> None:
    """冻结指定行以上的区域（默认首行表头）。"""
    ws.freeze_panes = f"A{row + 1}"


