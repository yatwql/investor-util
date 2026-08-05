"""Excel 页签 / HTML 章节 / 注册表三方一致性契约。

核心契约：两端可见章节的「名称 + 顺序 + 集合」均以
registry.get_report_section_order() 为单一数据源，llm_usage 恒强制末位。
锁死两条防漂移保证：
  1. 同一逻辑输入（board 开关 + 数据就绪状态）下，Excel create_sheets 与
     HTML _compute_section_visibility 产出相同的可见章节集合与顺序；
  2. Excel 页签标题与注册表章节名严格一致（HTML 侧标题一致性由
     test_html_report_structure.py 的导航标题测试覆盖）。
"""

from __future__ import annotations

import unittest

import pytest

from src.python.core.registry import get_report_section_order

pytestmark = [pytest.mark.integration, pytest.mark.integration_contract]


@pytest.mark.integration
@pytest.mark.integration_contract
class TestReportChapterConsistency(unittest.TestCase):
    """Excel 页签 / HTML 章节 / 注册表三方一致性契约。"""

    # ── 双端可见性镜像工具 ──────────────────────────────────────

    def _excel_visible(
        self,
        order,
        *,
        fund_deep=True,
        news=True,
        history=True,
        evolution=True,
        action=False,
        llm=True,
        data_availability=None,
    ):
        """镜像 generate_excel_report 的页签构造，返回 (key 列表, 标题列表)。

        data_availability 缺省为 None → 按真实流仅注入 news/llm 两个标志；
        基金深度等 data_flag 缺省 → should_create_sheet 默认 True
        （建页签，内容由下游写占位）。
        """
        from openpyxl import Workbook

        from src.python.report.excel_sheet_factory import create_sheets

        wb = Workbook()
        wb.remove(wb.active)
        if data_availability is None:
            data_availability = {}
            if news:
                data_availability["news_data_available"] = True
            if llm:
                data_availability["llm_data_available"] = True
        sheets = create_sheets(
            wb,
            order,
            enable_fund_deep_analysis=fund_deep,
            enable_news=news,
            enable_history=history,
            enable_portfolio_evolution=evolution,
            enable_action=action,
            enable_llm=llm,
            data_availability=data_availability,
        )
        return list(sheets.keys()), [ws.title for ws in sheets.values()]

    def _html_visible_keys(
        self, order, *, fund_deep=True, news=True, history=True, evolution=True, action=False, llm=True
    ):
        """镜像 write_html_report 的可见性计算，返回按连续编号的有序 key 列表。

        翻译规则（与真实流水线一致）：
          - 基金深度开启时 _render_* 恒返回占位 dict（非 None）→ data_flags True；
          - evolution/style/position 数据契约 dict 编排层恒注入（非 None）；
          - llm_enabled_flag = enable_llm（真实 full 流经 build_fallback_llm_content
            兜底后内容恒非空，llm_enabled_flag 恒与 board 开关一致）。
        """
        from src.python.report.html_writer import _compute_section_visibility

        placeholder = {} if fund_deep else None
        visible_numbers, _visible_dict, _ = _compute_section_visibility(
            order,
            manager_analysis=placeholder,
            overlap_matrix=placeholder,
            concentration_analysis=placeholder,
            style_analysis=placeholder,
            include_news=news,
            llm_enabled_flag=llm,
            enable_news=news,
            enable_fund_deep_analysis=fund_deep,
            enable_history=history,
            enable_portfolio_evolution=evolution,
            enable_action=action,
            enable_llm=llm,
            style_factor_data=placeholder,
            position_relationship_data=placeholder,
            evolution_data={} if evolution else None,
        )
        ordered = sorted(visible_numbers.items(), key=lambda kv: kv[1])
        return [k for k, _ in ordered]

    # ── 1. 注册表完整性 ─────────────────────────────────────────

    def test_registry_order_integrity(self):
        """注册表全量章节：序号连续唯一、llm_usage 强制末位、章节名唯一非空。"""
        order = get_report_section_order()
        keys = [s["key"] for s in order]
        names = [s["name"] for s in order]
        numbers = [s["number"] for s in order]

        self.assertEqual(len(set(keys)), len(keys), "章节 key 必须唯一")
        self.assertEqual(len(set(names)), len(names), "章节名必须唯一")
        self.assertEqual(
            numbers,
            list(range(1, len(order) + 1)),
            "章节序号应连续且与注册表条目数一致",
        )
        self.assertEqual(keys[-1], "llm_usage", "llm_usage 必须强制末位")
        self.assertTrue(all(s["name"].strip() for s in order), "章节名不得为空")

    # ── 2. Excel 端与注册表一致性 ───────────────────────────────

    def test_excel_sheets_match_registry_order_all_enabled(self):
        """全开时 Excel 页签顺序 == 注册表顺序，标题 == 「序号.章节名」，llm_usage 末位。"""
        order = get_report_section_order()
        excel_keys, excel_titles = self._excel_visible(order, action=True)

        registry_keys = [s["key"] for s in order]
        self.assertEqual(excel_keys, registry_keys, "Excel 页签顺序必须等于注册表顺序")
        self.assertEqual(len(excel_keys), len(order), "Excel 页签数必须等于注册表条目数")
        self.assertEqual(excel_keys[-1], "llm_usage", "Excel 端 llm_usage 必须末位")

        expected_titles = [f"{i}.{s['name']}" for i, s in enumerate(order, start=1)]
        self.assertEqual(excel_titles, expected_titles, "Excel 页签标题应为「序号.章节名」")

    # ── 3. HTML 端与注册表一致性 ───────────────────────────────

    def test_html_visible_numbers_match_registry_all_data_present(self):
        """数据全就绪时 HTML 连续编号与注册表顺序一致，llm_usage 位于末位。"""
        from src.python.report.html_writer import _compute_section_visibility

        order = get_report_section_order()
        placeholder = {}
        visible_numbers, visible_dict, _ = _compute_section_visibility(
            order,
            placeholder,
            placeholder,
            placeholder,
            placeholder,
            include_news=True,
            llm_enabled_flag=True,
            enable_fund_deep_analysis=True,
            enable_action=True,  # 行动建议默认关，全开场景显式开启
            style_factor_data={},
            position_relationship_data={},
            evolution_data={},
        )

        self.assertEqual(
            set(visible_dict.keys()),
            {s["key"] for s in order},
            "HTML 可见性字典必须覆盖注册表全部章节",
        )
        self.assertTrue(all(visible_dict.values()), "全开时所有章节应可见")

        ordered = sorted(visible_numbers.items(), key=lambda kv: kv[1])
        self.assertEqual([k for k, _ in ordered], [s["key"] for s in order], "HTML 章节顺序必须等于注册表顺序")
        self.assertEqual(
            list(visible_numbers.values()),
            list(range(1, len(order) + 1)),
            "HTML 连续编号应等于注册表序号",
        )
        self.assertEqual(visible_numbers["llm_usage"], len(order), "llm_usage 编号必须为末位序号")

    # ── 4. 双端可见章节集合与顺序交叉一致 ───────────────────────

    def test_dual_side_visible_sections_agree(self):
        """同一逻辑输入下，Excel 与 HTML 可见章节集合/顺序逐场景一致。"""
        order = get_report_section_order()
        scenarios = [
            # (场景名, board 开关 kwargs)；action 默认关，全开场景显式开启
            ("全开", {"action": True}),
            ("基金深度关闭", {"fund_deep": False, "action": True}),
            ("新闻关闭", {"news": False, "action": True}),
            ("LLM 关闭", {"llm": False, "action": True}),
            ("组合演进关闭", {"evolution": False, "action": True}),
            ("历史关闭", {"history": False, "action": True}),
            ("行动建议关闭(默认)", {"action": False}),
            ("多开关组合", {"fund_deep": False, "news": False, "llm": False, "action": False}),
        ]

        for name, kw in scenarios:
            excel_keys, _ = self._excel_visible(order, **kw)
            html_keys = self._html_visible_keys(order, **kw)

            self.assertEqual(
                excel_keys,
                html_keys,
                f"场景「{name}」两端可见章节集合/顺序不一致: Excel={excel_keys} HTML={html_keys}",
            )
            for key_list in (excel_keys, html_keys):
                if "llm_usage" in key_list:
                    self.assertEqual(key_list[-1], "llm_usage", f"场景「{name}」llm_usage 未强制末位")

    def test_dual_side_agree_on_explicit_data_gaps(self):
        """显式数据缺口：同一 data_flag 两端同时报告未就绪 → 两端隐藏相同章节。

        锁定 data_flag 命名契约：注册表 data_flag 键在两端的查表行为一致。
        该场景为假设性输入（真实 generate_excel_report 只传 news/llm 标志），
        用于锁定 should_create_sheet 缺省 True 与 HTML data_flags 缺省 False
        在「显式 False」下的一致性。
        """
        from src.python.report.html_writer import _compute_section_visibility

        order = get_report_section_order()
        gap_avail = {
            "manager_data": False,
            "concentration_data": False,
            "style_factor_data": False,
            "position_relationship_data": False,
            "evolution_data": False,
            "news_data_available": True,
            "llm_data_available": True,
        }
        excel_keys, _ = self._excel_visible(order, fund_deep=True, evolution=True, data_availability=gap_avail)

        visible_numbers, _visible_dict, _ = _compute_section_visibility(
            order,
            manager_analysis=None,
            overlap_matrix=None,
            concentration_analysis=None,
            style_analysis=None,
            include_news=True,
            llm_enabled_flag=True,
            enable_fund_deep_analysis=True,
            style_factor_data=None,
            position_relationship_data=None,
            evolution_data=None,
        )
        ordered = sorted(visible_numbers.items(), key=lambda kv: kv[1])
        html_keys = [k for k, _ in ordered]

        self.assertEqual(excel_keys, html_keys, "显式数据缺口下两端可见集合应一致")
        for hidden in (
            "fund_manager",
            "fund_concentration",
            "style_factor",
            "position_relationship",
            "portfolio_evolution",
        ):
            self.assertNotIn(hidden, excel_keys, f"Excel 不应创建缺数据页签 {hidden}")
            self.assertNotIn(hidden, html_keys, f"HTML 不应显示缺数据章节 {hidden}")
