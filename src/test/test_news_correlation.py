"""财经新闻关联模块单元测试 — 异常场景与边界测试。

测试目标：
  - build_news_data — 空持仓/API 失败降级（mock API）
  - write_news_sheet — 空数据/正常数据/LLM 分析列渲染
  - LLM 增强路径 — 配置开启时调用 enhance_news_correlation

运行：
  cd D:/codebase/zoo/investor-util
  python -m unittest src.test_news_correlation -v
"""

from __future__ import annotations

import unittest
from unittest.mock import MagicMock, patch

from src.python.models import Holding
from src.python.report import news_correlation as nc


class TestBuildNewsData(unittest.TestCase):
    """build_news_data 异常场景测试（mock 网络层 + LLM 配置）。"""

    @patch("src.python.config.get_llm_config")
    @patch("src.python.providers.news_aggregator.aggregate_news")
    def test_empty_holdings_returns_empty_list(
        self, mock_aggregate: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """空持仓 → aggregate_news 不会被调或返回空列表。"""
        mock_aggregate.return_value = []
        mock_llm_cfg.return_value = None
        result, meta = nc.build_news_data([])
        self.assertEqual(result, [])
        self.assertFalse(meta.get("llm_enabled"))

    @patch("src.python.config.get_llm_config")
    @patch("src.python.providers.news_aggregator.aggregate_news")
    def test_api_failure_returns_empty(
        self, mock_aggregate: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """API 失败 → 返回空列表。"""
        mock_aggregate.return_value = []
        mock_llm_cfg.return_value = None
        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)
        self.assertEqual(result, [])
        self.assertFalse(meta.get("llm_enabled"))

    @patch("src.python.config.get_llm_config")
    @patch("src.python.providers.news_aggregator.aggregate_news")
    def test_api_returns_data(
        self, mock_aggregate: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """API 返回数据 → 正确传递。"""
        mock_aggregate.return_value = [
            {"title": "新闻标题", "intro": "简介", "url": "http://example.com",
             "ctime": "2026-06-27", "media_name": "新浪财经", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": False}
        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)
        self.assertEqual(len(result), 1)
        self.assertIn("matched_keywords", result[0])
        self.assertFalse(meta.get("llm_enabled"))


class TestBuildNewsDataWithLLM(unittest.TestCase):
    """build_news_data 开启 LLM 增强时的测试。"""

    @patch("src.python.config.get_llm_config")
    @patch("src.python.llm_client.enhance_news_correlation")
    @patch("src.python.providers.news_aggregator.aggregate_news")
    def test_llm_enabled_calls_enhance(
        self, mock_aggregate: MagicMock, mock_enhance: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """llm_news_analysis=true → 调用 enhance_news_correlation。"""
        mock_aggregate.return_value = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": True, "api_key": "sk-test"}
        enriched = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪",
             "matched_keywords": ["长江电力"],
             "llm_analysis": "[高] 直接涉及电力政策"},
        ]
        mock_enhance.return_value = (enriched, False, {"total_tokens": 300})

        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)

        mock_enhance.assert_called_once()
        self.assertTrue(meta.get("llm_enabled"))
        self.assertFalse(meta.get("llm_cached"))
        self.assertEqual(meta.get("token_usage", {}).get("total_tokens"), 300)
        self.assertEqual(len(result), 1)
        self.assertIn("llm_analysis", result[0])

    @patch("src.python.config.get_llm_config")
    @patch("src.python.llm_client.enhance_news_correlation")
    @patch("src.python.providers.news_aggregator.aggregate_news")
    def test_llm_cache_hit(
        self, mock_aggregate: MagicMock, mock_enhance: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """LLM 缓存命中 → meta 中 llm_cached 为 True。"""
        mock_aggregate.return_value = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": True, "api_key": "sk-test"}
        enriched = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪",
             "matched_keywords": ["长江电力"],
             "llm_analysis": "[高] 直接涉及电力政策"},
        ]
        mock_enhance.return_value = (enriched, True, {})  # cached

        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)

        self.assertTrue(meta.get("llm_cached"))

    @patch("src.python.config.get_llm_config")
    @patch("src.python.llm_client.enhance_news_correlation")
    @patch("src.python.providers.news_aggregator.aggregate_news")
    def test_llm_disabled_when_no_api_key(
        self, mock_aggregate: MagicMock, mock_enhance: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """llm_news_analysis=true 但无 api_key → 降级为传统分析，llm_enabled=False。"""
        mock_aggregate.return_value = [
            {"title": "新闻标题", "intro": "简介", "url": "http://example.com",
             "ctime": "2026-06-27", "media_name": "新浪财经", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": True, "api_key": ""}
        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)
        mock_enhance.assert_not_called()
        self.assertFalse(meta.get("llm_enabled"))
        self.assertEqual(len(result), 1)
        self.assertNotIn("llm_analysis", result[0])

    @patch("src.python.config.get_llm_config")
    @patch("src.python.llm_client.enhance_news_correlation")
    @patch("src.python.providers.news_aggregator.aggregate_news")
    def test_llm_enhance_failure_falls_back_gracefully(
        self, mock_aggregate: MagicMock, mock_enhance: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """LLM 增强失败 → 返回原始新闻数据 + llm_enabled=True 但无分析。"""
        mock_aggregate.return_value = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": True, "api_key": "sk-test"}
        mock_enhance.side_effect = Exception("LLM 服务异常")

        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)

        self.assertEqual(len(result), 1)
        self.assertNotIn("llm_analysis", result[0])
        self.assertTrue(meta.get("llm_enabled"))
        # 失败后 token_usage 为空
        self.assertEqual(meta.get("token_usage", {}), {})

    @patch("src.python.config.get_llm_config")
    @patch("src.python.providers.news_aggregator.aggregate_news")
    def test_llm_disabled_returns_no_llm_meta(
        self, mock_aggregate: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """llm_news_analysis=false → llm_enabled=False。"""
        mock_aggregate.return_value = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": False}

        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)

        self.assertEqual(len(result), 1)
        self.assertNotIn("llm_analysis", result[0])
        self.assertFalse(meta.get("llm_enabled"))


class TestWriteNewsSheet(unittest.TestCase):
    """write_news_sheet 边界测试。"""

    def setUp(self) -> None:
        from openpyxl import Workbook
        self.wb = Workbook()
        self.ws = self.wb.active

    def test_empty_data(self) -> None:
        """空数据 → 写入"暂无关联新闻"占位。"""
        nc.write_news_sheet(self.ws, [])
        any_text = False
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and "暂无" in str(cell.value):
                    any_text = True
        self.assertTrue(any_text)

    def test_with_data(self) -> None:
        """有数据 → 正确写入标题、序号、关键词（6 列，无 LLM）。"""
        data = [
            {"title": "新闻A", "intro": "简介A", "url": "http://a.com",
             "ctime": "2026-06-27", "media_name": "新浪", "matched_keywords": ["茅台"]},
            {"title": "新闻B", "intro": "简介B", "url": "http://b.com",
             "ctime": "2026-06-26", "media_name": "新浪", "matched_keywords": ["五粮液", "白酒"]},
        ]
        nc.write_news_sheet(self.ws, data)
        headers = [self.ws.cell(row=2, column=c).value for c in range(1, 7)]
        self.assertIn("序号", headers)
        self.assertEqual(self.ws.cell(row=3, column=1).value, 1)

    def test_partial_missing_fields(self) -> None:
        """数据缺少可选字段 → 不崩溃。"""
        data = [
            {"title": "仅标题", "matched_keywords": ["茅台"]},
        ]
        try:
            nc.write_news_sheet(self.ws, data)
        except Exception as e:
            self.fail(f"write_news_sheet with partial data raised: {e}")

    def test_with_llm_analysis_column(self) -> None:
        """数据含 llm_analysis → 自动增加第 7 列。"""
        data = [
            {"title": "新闻A", "intro": "简介A", "url": "http://a.com",
             "ctime": "2026-06-27", "media_name": "新浪",
             "matched_keywords": ["茅台"], "llm_analysis": "[高] 白酒利好"},
        ]
        nc.write_news_sheet(self.ws, data)
        col7 = self.ws.cell(row=2, column=7).value
        self.assertEqual(col7, "LLM 关联分析")
        self.assertEqual(self.ws.cell(row=3, column=7).value, "[高] 白酒利好")

    def test_llm_enabled_with_token_usage(self) -> None:
        """LLM 启用 + 非缓存 → 底部显示 token 消耗。"""
        data = [
            {"title": "新闻A", "intro": "简介A", "url": "http://a.com",
             "ctime": "2026-06-27", "media_name": "新浪",
             "matched_keywords": ["茅台"], "llm_analysis": "[高] 白酒利好"},
        ]
        llm_meta = {
            "llm_enabled": True,
            "llm_cached": False,
            "token_usage": {"input_tokens": 2000, "output_tokens": 500, "total_tokens": 2500},
        }
        nc.write_news_sheet(self.ws, data, llm_meta=llm_meta)

        found_token = False
        for row in self.ws.iter_rows():
            for cell in row:
                val = str(cell.value) if cell.value else ""
                if "Token" in val and "2,500" in val:
                    found_token = True
        self.assertTrue(found_token, "Token 消耗行应出现在页签底部")

    def test_llm_cache_hit_footnote(self) -> None:
        """LLM 缓存命中 → 底部写"使用了LLM缓存"。"""
        data = [{"title": "新闻A", "intro": "简介", "url": "http://a.com",
                 "ctime": "2026-06-27", "media_name": "新浪",
                 "matched_keywords": ["茅台"], "llm_analysis": "[高] 利好"}]
        llm_meta = {"llm_enabled": True, "llm_cached": True, "token_usage": {}}
        nc.write_news_sheet(self.ws, data, llm_meta=llm_meta)

        found = False
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and "LLM缓存" in str(cell.value):
                    found = True
        self.assertTrue(found, "缓存命中时应显示'使用了LLM缓存'")

    def test_llm_disabled_footnote(self) -> None:
        """LLM 未启用 → 底部写"未使用LLM服务能力增强支持"。"""
        data = [{"title": "新闻A", "intro": "简介", "url": "http://a.com",
                 "ctime": "2026-06-27", "media_name": "新浪", "matched_keywords": ["茅台"]}]
        llm_meta = {"llm_enabled": False, "llm_cached": False, "token_usage": {}}
        nc.write_news_sheet(self.ws, data, llm_meta=llm_meta)

        found = False
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and "未使用LLM服务能力" in str(cell.value):
                    found = True
        self.assertTrue(found, "LLM 未启用时应显示'未使用LLM服务能力增强支持'")

    def test_llm_analysis_is_empty_when_all_irrelevant(self) -> None:
        """所有新闻 LLM 判定为无关 → 不显示 LLM 分析列。"""
        data = [
            {"title": "无关新闻", "intro": "无关内容", "url": "http://irr.com",
             "ctime": "2026-06-28", "media_name": "新浪",
             "matched_keywords": [], "llm_analysis": ""},
        ]
        nc.write_news_sheet(self.ws, data)
        col7 = self.ws.cell(row=2, column=7).value
        self.assertIsNone(col7)


# ============================================================
#  关键词富化单元测试
# ============================================================


class TestBuildKeywordLookup(unittest.TestCase):
    """_build_keyword_lookup 正向查找表构建测试。"""

    def test_holding_name_and_code(self):
        """持仓名称和代码 → lookup 中包含两条记录。"""
        holdings = [
            Holding("证券", "长江电力", "600900", 100, 50.0),
        ]
        lookup = nc._build_keyword_lookup(holdings)
        self.assertIn("600900", lookup)
        self.assertEqual(lookup["600900"]["type"], "holding")
        self.assertEqual(lookup["600900"]["name"], "长江电力")
        # 中文名称片段
        self.assertIn("长江", lookup)
        self.assertEqual(lookup["长江"]["type"], "holding")

    def test_penetration_asset(self):
        """穿透资产 → lookup 包含穿透资产条目。"""
        holdings: list = []
        penetrated = [
            {"name": "腾讯控股", "codes": ["00700"]},
        ]
        lookup = nc._build_keyword_lookup(holdings, penetrated)
        self.assertIn("00700", lookup)
        self.assertEqual(lookup["00700"]["type"], "penetration")
        self.assertIn("腾讯", lookup)
        self.assertEqual(lookup["腾讯"]["type"], "penetration")

    def test_holding_overrides_penetration(self):
        """同一关键词同时匹配持仓和穿透 → 持仓优先。"""
        holdings = [
            Holding("证券", "腾讯控股", "00700", 100, 50.0),
        ]
        penetrated = [
            {"name": "腾讯控股", "codes": ["00700"]},
        ]
        lookup = nc._build_keyword_lookup(holdings, penetrated)
        self.assertEqual(lookup["00700"]["type"], "holding")

    def test_empty_holdings(self):
        """空持仓 → 返回空字典。"""
        self.assertEqual(nc._build_keyword_lookup([]), {})

    def test_industry_data_adds_concepts(self):
        """industry_data → lookup 包含概念类型条目。"""
        industry_data = {
            "600900": {"industry": "电力设备", "concepts": ["CPO光模块", "人工智能"]},
        }
        lookup = nc._build_keyword_lookup([], industry_data=industry_data)
        self.assertIn("电力设备", lookup)
        self.assertEqual(lookup["电力设备"]["type"], "concept")
        self.assertEqual(lookup["电力设备"]["source"], "industry")
        self.assertIn("CPO光模块", lookup)
        self.assertEqual(lookup["CPO光模块"]["type"], "concept")
        self.assertEqual(lookup["CPO光模块"]["source"], "concept")

    def test_holding_overrides_industry(self):
        """同一关键词同时匹配持仓和行业 → 持仓优先。"""
        holdings = [Holding("证券", "电力设备", "000xxx", 100, 10.0)]
        industry_data = {
            "600900": {"industry": "电力设备", "concepts": []},
        }
        lookup = nc._build_keyword_lookup(holdings, industry_data=industry_data)
        # "电力设备" 来自持仓名称 "电力设备"，应该保留为 holding 类型
        self.assertIn("电力设备", lookup)
        self.assertEqual(lookup["电力设备"]["type"], "holding")


class TestEnrichKeywordsForItem(unittest.TestCase):
    """_enrich_keywords_for_item 富化逻辑测试。"""

    def setUp(self):
        self.lookup = nc._build_keyword_lookup([
            Holding("证券", "长江电力", "600900", 100, 50.0),
        ], [
            {"name": "腾讯控股", "codes": ["00700"]},
        ])

    def test_holding_match(self):
        """关键词匹配持仓名称 → display = "名称(代码)", type = holding。"""
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["长江"]}, self.lookup,
        )
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["display"], "长江电力(600900)")
        self.assertEqual(result[0]["type"], "holding")

    def test_holding_code_match(self):
        """关键词匹配持仓代码 → 与名称去重后合并。"""
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["600900", "长江"]}, self.lookup,
        )
        self.assertEqual(len(result), 1)  # 去重
        self.assertEqual(result[0]["type"], "holding")

    def test_penetration_match(self):
        """关键词匹配穿透资产 → display = "名称[穿透]", type = penetration。"""
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["00700"]}, self.lookup,
        )
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["display"], "腾讯控股[穿透]")
        self.assertEqual(result[0]["type"], "penetration")

    def test_industry_match(self):
        """关键词未匹配持仓/穿透 → type = industry。"""
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["电力行业"]}, self.lookup,
        )
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["display"], "电力行业")
        self.assertEqual(result[0]["type"], "industry")

    def test_mixed_types_sorted(self):
        """混合关键词 → 排序: holding → penetration → industry。"""
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["长江", "00700", "行业词汇"]}, self.lookup,
        )
        self.assertGreaterEqual(len(result), 3)
        types = [r["type"] for r in result]
        self.assertEqual(types[0], "holding")
        self.assertEqual(types[1], "penetration")
        # industry 在最后
        self.assertEqual(types[-1], "industry")

    def test_empty_keywords_returns_empty_list(self):
        """空 matched_keywords → []。"""
        self.assertEqual(nc._enrich_keywords_for_item({}, self.lookup), [])
        self.assertEqual(
            nc._enrich_keywords_for_item({"matched_keywords": []}, self.lookup), [],
        )

    def test_concept_type_enriched(self):
        """concept 类型关键词 → 显示为 XXX[概念] 格式。"""
        concept_lookup = {"CPO光模块": {"type": "concept", "name": "CPO光模块", "code": "600900", "source": "concept"}}
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["CPO光模块"]}, concept_lookup,
        )
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["display"], "CPO光模块[概念]")
        self.assertEqual(result[0]["type"], "concept")

    def test_industry_name_as_concept(self):
        """行业名称作为 concept 类型 → 显示为 行业名[概念]。"""
        concept_lookup = {"电力设备": {"type": "concept", "name": "电力设备", "code": "600900", "source": "industry"}}
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["电力设备"]}, concept_lookup,
        )
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["display"], "电力设备[概念]")
        self.assertEqual(result[0]["type"], "concept")

    def test_mixed_types_with_concept_sorted(self):
        """含 concept 的混合关键词 → 排序: holding → penetration → concept → industry。"""
        concept_lookup = {
            "长江": {"type": "holding", "name": "长江电力", "code": "600900"},
            "00700": {"type": "penetration", "name": "腾讯控股", "code": "00700"},
            "CPO光模块": {"type": "concept", "name": "CPO光模块", "code": "600900", "source": "concept"},
            "普通词汇": {},
        }
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["长江", "00700", "CPO光模块", "普通词汇"]}, concept_lookup,
        )
        types = [r["type"] for r in result]
        self.assertEqual(types[0], "holding")
        self.assertEqual(types[1], "penetration")
        self.assertEqual(types[2], "concept")
        self.assertEqual(types[3], "industry")


class TestFormatEnrichedKeywords(unittest.TestCase):
    """_format_enriched_keywords 格式化测试。"""

    def test_single_item(self):
        """单条 → 直接返回 display 字符串。"""
        result = nc._format_enriched_keywords([
            {"display": "长江电力(600900)", "type": "holding"},
        ])
        self.assertEqual(result, "长江电力(600900)")

    def test_multiple_items(self):
        """多条 → 逗号分隔。"""
        result = nc._format_enriched_keywords([
            {"display": "长江电力(600900)", "type": "holding"},
            {"display": "电力行业", "type": "industry"},
        ])
        self.assertEqual(result, "长江电力(600900), 电力行业")

    def test_empty_list(self):
        """空列表 → ""。"""
        self.assertEqual(nc._format_enriched_keywords([]), "")


# ============================================================
#  write_news_sheet 格式增强测试
# ============================================================


class TestWriteNewsSheetFormatting(unittest.TestCase):
    """write_news_sheet 的 wrap_text / column_width 断言。"""

    def setUp(self) -> None:
        from openpyxl import Workbook
        self.wb = Workbook()
        self.ws = self.wb.active
        self.data = [
            {"title": "测试新闻标题", "intro": "这是一段较长的新闻摘要内容用于测试。",
             "url": "http://a.com", "ctime": "2026-06-28", "media_name": "新浪",
             "matched_keywords": ["长江电力"],
             "enriched_keywords": [{"display": "长江电力(600900)", "type": "holding"}]},
        ]

    def test_column_b_width_after_auto_width(self):
        """B 列宽度 = 40（覆盖 auto_width 的 max_width=30）。"""
        nc.write_news_sheet(self.ws, self.data)
        self.assertEqual(self.ws.column_dimensions["B"].width, 40)

    def test_column_c_width_after_auto_width(self):
        """C 列宽度 = 50。"""
        nc.write_news_sheet(self.ws, self.data)
        self.assertEqual(self.ws.column_dimensions["C"].width, 50)

    def test_data_row_b_alignment_wrap_text(self):
        """B 列数据行 alignment 包含 wrap_text。"""
        nc.write_news_sheet(self.ws, self.data)
        cell = self.ws.cell(row=3, column=2)
        self.assertTrue(cell.alignment.wrap_text)
        self.assertEqual(cell.alignment.horizontal, "left")

    def test_data_row_c_alignment_wrap_text(self):
        """C 列数据行 alignment 包含 wrap_text。"""
        nc.write_news_sheet(self.ws, self.data)
        cell = self.ws.cell(row=3, column=3)
        self.assertTrue(cell.alignment.wrap_text)
        self.assertEqual(cell.alignment.horizontal, "left")

    def test_empty_data_does_not_set_alignment(self):
        """空数据 → B/C alignment 不会被异常撑爆。"""
        from openpyxl import Workbook
        ws2 = Workbook().active
        try:
            nc.write_news_sheet(ws2, [])
        except Exception as e:
            self.fail(f"空数据写入 failed: {e}")

    def test_enriched_keywords_in_excel(self):
        """enriched_keywords 存在 → excel 关键词列显示富化文本。"""
        nc.write_news_sheet(self.ws, self.data)
        cell_val = self.ws.cell(row=3, column=6).value
        self.assertIn("长江电力(600900)", cell_val)


# ═══════════════════════════════════════════════════════════
#  _format_industry_tags
# ═══════════════════════════════════════════════════════════


class TestFormatIndustryTags(unittest.TestCase):
    """_format_industry_tags 行业/概念标签后缀测试。"""

    def test_holding_with_industry_only(self):
        """持仓条目仅有行业 → 返回 [行业名]。"""
        entry = {"industry": "电力设备"}
        result = nc._format_industry_tags(entry)
        self.assertEqual(result, " [电力设备]")

    def test_holding_with_concepts_only(self):
        """持仓条目仅有概念 → 返回 [概念名]。"""
        entry = {"concepts_list": ["锂电池", "储能"]}
        result = nc._format_industry_tags(entry)
        self.assertEqual(result, " [锂电池 · 储能]")

    def test_holding_with_both(self):
        """持仓条目有行业和概念 → 返回 [行业 · 概念]。
        行业在前，概念在后。"""
        entry = {"industry": "电力设备", "concepts_list": ["锂电池", "储能", "新能源"]}
        result = nc._format_industry_tags(entry)
        # 概念只取前 2 个
        self.assertEqual(result, " [电力设备 · 锂电池 · 储能]")

    def test_holding_no_tags(self):
        """持仓条目无行业/概念 → 返回空字符串。"""
        entry = {"type": "holding", "name": "长江电力", "code": "600900"}
        result = nc._format_industry_tags(entry)
        self.assertEqual(result, "")

    def test_empty_concepts_list(self):
        """concepts_list 为空列表 → 不报错。"""
        entry = {"industry": "电力设备", "concepts_list": []}
        result = nc._format_industry_tags(entry)
        self.assertEqual(result, " [电力设备]")

    def test_missing_fields(self):
        """缺少 industry 和 concepts_list → 返回空字符串。"""
        entry = {}
        result = nc._format_industry_tags(entry)
        self.assertEqual(result, "")


# ═══════════════════════════════════════════════════════════
#  富化显示 — 带行业/概念标签
# ═══════════════════════════════════════════════════════════


class TestEnrichKeywordsWithIndustryTags(unittest.TestCase):
    """验证 enriched_keywords 在 industry_data 传入时附加行业/概念标签。"""

    def _make_lookup(self, industry_data: dict | None = None) -> dict:
        """构建带 industry_data 的 lookup。"""
        from src.python.models import Holding
        holdings = [
            Holding("证券", "长江电力", "600900", 100, 10.0),
        ]
        return nc._build_keyword_lookup(holdings, industry_data=industry_data)

    def test_holding_display_with_industry_tag(self):
        """持仓匹配时 display 包含行业标签。"""
        industry_data = {
            "600900": {"industry": "电力", "concepts": ["水电", "大盘蓝筹"]},
        }
        lookup = self._make_lookup(industry_data)
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["600900"]}, lookup,
        )
        self.assertTrue(len(result) >= 1)
        display = result[0]["display"]
        self.assertIn("长江电力(600900)", display)
        self.assertIn("电力", display, "应包含行业名称")
        self.assertIn("水电", display, "应包含概念名称")

    def test_holding_display_without_industry_data(self):
        """无 industry_data → display 不加标签。"""
        lookup = self._make_lookup(industry_data=None)
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["600900"]}, lookup,
        )
        display = result[0]["display"]
        self.assertEqual(display, "长江电力(600900)")

    def test_penetration_display_with_concept(self):
        """穿透匹配时 display 包含概念标签。"""
        pen_assets = [
            {"name": "腾讯控股", "codes": ["00700"]},
        ]
        industry_data = {
            "00700": {"industry": "互联网科技", "concepts": ["社交", "云计算"]},
        }
        lookup = nc._build_keyword_lookup([], penetrated_assets=pen_assets,
                                           industry_data=industry_data)
        result = nc._enrich_keywords_for_item(
            {"matched_keywords": ["00700"]}, lookup,
        )
        display = result[0]["display"]
        self.assertIn("腾讯控股[穿透]", display)
        self.assertIn("互联网科技", display, "应包含行业名称")


if __name__ == "__main__":
    unittest.main()
