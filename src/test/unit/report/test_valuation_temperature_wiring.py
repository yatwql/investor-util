"""估值分位 + 市场温度：报告层接线单元测试。

覆盖 HTML 展示构建器 / Excel 穿透估值列 / Excel 汇总温度行 /
编排层 compute_valuation_data 与 compute_market_temperature_data 的
开关关闭降级（None）与可用契约。

纯计算层（valuation_percentile / market_temperature）已有独立测试，
本文件只覆盖报告层接线，不重复计算逻辑。
"""

from __future__ import annotations

import unittest
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

pytestmark = [pytest.mark.unit, pytest.mark.unit_report]


class TestBuildTemperatureDisplay(unittest.TestCase):
    """_build_temperature_display：市场温度契约 → HTML 展示映射。"""

    def test_none_when_switch_off(self):
        from src.python.report.html_writer import _build_temperature_display

        assert _build_temperature_display(None) is None

    def test_placeholder_when_unavailable(self):
        from src.python.report.html_writer import _build_temperature_display

        data = {"available": False, "status": "insufficient"}
        d = _build_temperature_display(data)
        assert d is not None
        assert d["available"] is False
        assert d["score"] is None
        assert d["components"] is None
        assert d["index_name"] == "沪深300"

    def test_available_full_components(self):
        from src.python.report.html_writer import _build_temperature_display

        data = {
            "available": True,
            "price_percentile": 42.5,
            "ma_deviation": 0.035,
            "volatility": 0.18,
            "score": 55,
            "tier": "合理",
            "index_name": "沪深300",
        }
        d = _build_temperature_display(data)
        assert d["available"] is True
        assert d["score"] == 55
        assert d["tier"] == "合理"
        assert d["components"] == {
            "price_percentile": "42.5%",
            "ma_deviation": "+3.5%",
            "volatility": "18.0%",
        }
        assert d["index_name"] == "沪深300"

    def test_available_missing_factor(self):
        from src.python.report.html_writer import _build_temperature_display

        data = {"available": True, "score": 50, "tier": "合理"}
        d = _build_temperature_display(data)
        assert d["available"] is True
        assert d["components"] is None

    def test_available_tier_fallback(self):
        from src.python.report.html_writer import _build_temperature_display

        d = _build_temperature_display({"available": True})
        assert d["tier"] == "合理"


class TestAttachValuationToPenetration(unittest.TestCase):
    """_attach_valuation_to_penetration：穿透 TOP10 附加估值分位文本。"""

    def test_none_valuation_returns_original(self):
        """开关关闭（None）时原样返回，不附加列。"""
        from src.python.report.html_writer import _attach_valuation_to_penetration

        pen = {"top10": [{"rank": 1, "codes": ["600001"]}]}
        assert _attach_valuation_to_penetration(pen, None) is pen

    def test_none_penetration(self):
        from src.python.report.html_writer import _attach_valuation_to_penetration

        assert _attach_valuation_to_penetration(None, {"by_code": {}}) is None

    def test_attach_keeps_original_unmodified(self):
        """返回新 dict，原 penetration 对象不被修改（不可变）。"""
        from src.python.report.html_writer import _attach_valuation_to_penetration

        original = {
            "top10": [{"rank": 1, "name": "A", "codes": ["600001"]}],
            "summary": {"total_funds": 1},
        }
        val = {
            "by_code": {
                "600001": {
                    "pe": 12.3,
                    "pb": 1.56,
                    "price_percentile": 45.0,
                    "tier": "合理",
                    "percentile_available": True,
                }
            }
        }
        result = _attach_valuation_to_penetration(original, val)
        assert "valuation_text" not in original["top10"][0]
        assert result is not original
        assert result["top10"][0]["valuation_text"] == "PE 12.3 · PB 1.56 · 分位 45%（合理）"
        assert result["summary"] == original["summary"]

    def test_no_match_code(self):
        from src.python.report.html_writer import _attach_valuation_to_penetration

        pen = {"top10": [{"rank": 1, "codes": ["600999"]}]}
        val = {"by_code": {"600001": {"pe": 12.3}}}
        result = _attach_valuation_to_penetration(pen, val)
        assert result["top10"][0]["valuation_text"] == "--"


class TestGetValuationText(unittest.TestCase):
    """_get_valuation_text：穿透估值分位列文本（Excel）。"""

    def test_none_returns_dash(self):
        from src.python.report.penetration_sheet import _get_valuation_text

        assert _get_valuation_text(None, ["600001"]) == "--"

    def test_full_text(self):
        from src.python.report.penetration_sheet import _get_valuation_text

        val = {
            "by_code": {
                "600001": {
                    "pe": 12.3,
                    "pb": 1.56,
                    "price_percentile": 45.0,
                    "tier": "合理",
                    "percentile_available": True,
                }
            }
        }
        assert _get_valuation_text(val, ["600001"]) == "PE 12.3 · PB 1.56 · 分位 45%（合理）"

    def test_second_code_fallback(self):
        """首个代码无数据 → 依次查找后续代码。"""
        from src.python.report.penetration_sheet import _get_valuation_text

        val = {"by_code": {"600001": {"pe": 12.3}}}
        assert _get_valuation_text(val, ["600000", "600001"]) == "PE 12.3"

    def test_no_match(self):
        from src.python.report.penetration_sheet import _get_valuation_text

        val = {"by_code": {"600001": {"pe": 12.3}}}
        assert _get_valuation_text(val, ["600999"]) == "--"

    def test_percentile_only_without_pe_pb(self):
        from src.python.report.penetration_sheet import _get_valuation_text

        val = {"by_code": {"600001": {"price_percentile": 30.0, "tier": "低估", "percentile_available": True}}}
        assert _get_valuation_text(val, ["600001"]) == "分位 30%（低估）"


class TestWriteMarketTemperature(unittest.TestCase):
    """_write_market_temperature：Excel 汇总章市场温度刻度行。"""

    def _make_ws(self):
        wb = Workbook()
        return wb.active

    def test_unavailable_placeholder(self):
        from src.python.report.summary import _write_market_temperature

        ws = self._make_ws()
        row = _write_market_temperature(ws, 5, {"available": False, "status": "insufficient"})
        assert ws.cell(row=5, column=1).value == "【市场温度】"
        assert ws.cell(row=6, column=1).value == "市场温度"
        assert ws.cell(row=6, column=2).value == "--（数据不足，暂不显示）"
        assert ws.cell(row=7, column=1).value == "注"
        assert row == 8

    def test_none_temperature_placeholder(self):
        from src.python.report.summary import _write_market_temperature

        ws = self._make_ws()
        row = _write_market_temperature(ws, 5, None)
        assert ws.cell(row=6, column=2).value == "--（数据不足，暂不显示）"
        assert row == 8

    def test_available_rows(self):
        from src.python.report.summary import _write_market_temperature

        ws = self._make_ws()
        data = {
            "available": True,
            "score": 62,
            "tier": "偏暖",
            "price_percentile": 68.0,
            "ma_deviation": 0.032,
            "volatility": 0.15,
            "index_name": "沪深300",
        }
        row = _write_market_temperature(ws, 5, data)
        assert ws.cell(row=5, column=1).value == "【市场温度】"
        assert ws.cell(row=6, column=1).value == "市场温度"
        assert ws.cell(row=6, column=2).value == "62 / 100（偏暖）"
        assert ws.cell(row=7, column=1).value == "三因子（沪深300）"
        assert ws.cell(row=7, column=2).value == "价格分位 68.0% · 20日均线偏离 +3.2% · 年化波动率 15.0%"
        assert ws.cell(row=8, column=1).value == "注"
        assert row == 9


class TestComputeValuationData(unittest.TestCase):
    """compute_valuation_data：编排估值分位数据契约。"""

    def test_switch_off_returns_none(self):
        from src.python.report.orchestrator import compute_valuation_data

        result = compute_valuation_data([], [], {}, MagicMock())
        assert result is None

    def test_switch_on_contract(self):
        from src.python.report.orchestrator import compute_valuation_data

        config = {"report_submodules": {"valuation_percentile": True}}
        detail = MagicMock()
        detail.code = "600001"
        detail.name = "测试股票"
        fetched = {
            "pe": 12.3,
            "pb": 1.5,
            "price_percentile": 45.0,
            "tier": "合理",
            "sample_count": 200,
            "percentile_available": True,
        }
        with patch("src.python.report.orchestrator._fetch_valuation_for_code", return_value=fetched):
            result = compute_valuation_data([], [detail], config, MagicMock())
        assert result["available"] is True
        assert result["status"] == "ok"
        assert result["by_code"]["600001"]["pe"] == 12.3

    def test_switch_on_source_failed(self):
        """PE/PB 与 K 线均不可得 → available=False 占位（§1.4.5）。"""
        from src.python.report.orchestrator import compute_valuation_data

        config = {"report_submodules": {"valuation_percentile": True}}
        detail = MagicMock()
        detail.code = "600001"
        detail.name = "测试股票"
        with patch("src.python.report.orchestrator._fetch_valuation_for_code", return_value=None):
            result = compute_valuation_data([], [detail], config, MagicMock())
        assert result["available"] is False
        assert result["status"] == "source_failed"


class TestComputeMarketTemperatureData(unittest.TestCase):
    """compute_market_temperature_data：编排市场温度数据契约。"""

    def test_switch_off_returns_none(self):
        from src.python.report.orchestrator import compute_market_temperature_data

        assert compute_market_temperature_data({}, MagicMock()) is None

    def test_switch_on_available(self):
        from src.python.report.orchestrator import compute_market_temperature_data

        config = {"report_submodules": {"market_temperature": True}}
        # 需 ≥ MIN_SAMPLES(60) 根 K 线；恒平序列 → 分位 100、偏离/波动率 0，合成可得
        bars = [{"date": f"2024-01-{i:02d}", "close": 100.0} for i in range(1, 91)]
        with patch("src.python.fetcher.index.fetch_index_history", return_value=bars):
            result = compute_market_temperature_data(config, MagicMock())
        assert result["available"] is True
        assert result["status"] == "ok"
        assert result["score"] is not None
        assert result["index_code"] == "sh000300"
        assert result["index_name"] == "沪深300"

    def test_switch_on_insufficient(self):
        from src.python.report.orchestrator import compute_market_temperature_data

        config = {"report_submodules": {"market_temperature": True}}
        with patch("src.python.fetcher.index.fetch_index_history", return_value=[]):
            result = compute_market_temperature_data(config, MagicMock())
        assert result["available"] is False
        assert result["status"] == "insufficient"

    def test_switch_on_exception(self):
        """编排异常 → source_failed 占位。"""
        from src.python.report.orchestrator import compute_market_temperature_data

        config = {"report_submodules": {"market_temperature": True}}
        with patch(
            "src.python.fetcher.index.fetch_index_history",
            side_effect=RuntimeError("boom"),
        ):
            result = compute_market_temperature_data(config, MagicMock())
        assert result["available"] is False
        assert result["status"] == "source_failed"
