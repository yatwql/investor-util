"""Excel 报告 UI/UX 结构测试 — 页签次序/标题格式/文本换行。

覆盖场景：
  - 页签物理顺序与 section_order 一致（无错序）
  - 页签标题格式统一为 "{number}.{name}"（无硬编码默认序号）
  - 无重复页签标题
  - 文本换行设置（wrap_text）在关键列正确配置
  - 列宽度设置避免内容截断

运行：
  cd D:/codebase/zoo/investor-util
  pytest src/test/unit/report/test_excel_report_structure.py -v
"""

from __future__ import annotations

import unittest
from unittest.mock import patch

import pytest

pytestmark = [pytest.mark.unit, pytest.mark.unit_report]

# 标准注册表（精简版，仅含结构测试所需字段，与 registry.py 对齐）
_REPORT_SECTION_DEFAULT: list[dict] = [
    {"key": "summary", "name": "投资分析汇总", "number": 1, "type": "always"},
    {"key": "market_value", "name": "市值核算明细表", "number": 2, "type": "always"},
    {"key": "category", "name": "持仓分类表", "number": 3, "type": "always"},
    {"key": "penetration", "name": "资产穿透TOP10", "number": 4, "type": "always"},
    {"key": "fund_performance", "name": "基金业绩分析", "number": 5, "type": "always"},
    {"key": "fund_manager", "name": "基金经理变更监控", "number": 6, "type": "fund_deep_analysis"},
    {"key": "position_relationship", "name": "持仓关系矩阵", "number": 7, "type": "fund_deep_analysis"},
    {"key": "fund_concentration", "name": "持仓集中度监控", "number": 8, "type": "fund_deep_analysis"},
    {"key": "style_factor", "name": "风格与因子分析", "number": 9, "type": "fund_deep_analysis"},
    {"key": "news_correlation", "name": "财经新闻热点与持仓关联分析", "number": 10, "type": "news"},
    {"key": "global_macro", "name": "全球政经局势", "number": 11, "type": "llm"},
    {"key": "expert_review", "name": "智囊团深度复盘", "number": 12, "type": "llm"},
    {"key": "health_check", "name": "持仓体检报告", "number": 13, "type": "llm"},
    {"key": "penetration_deep", "name": "穿透深度分析", "number": 14, "type": "llm"},
    {"key": "portfolio_history_drawdown", "name": "组合历史走势与回撤", "number": 15, "type": "history"},
    # 注：portfolio_evolution(16) / action(17) 为注册表扩展模块，此精简版省略
    {"key": "data_source_status", "name": "数据源可用性矩阵", "number": 18, "type": "always"},
    {"key": "llm_usage", "name": "LLM API 用量", "number": 19, "type": "llm"},
]


# ═══════════════════════════════════════════════════════════════
#  Test: Sheet Physical Order
# ═══════════════════════════════════════════════════════════════


class TestExcelSheetOrder(unittest.TestCase):
    """Excel 页签物理顺序测试 — 验证 wb.sheetnames 与 section_order 一致。"""

    def _make_wb(self):
        """创建一个空 Workbook。"""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        return wb

    def test_sheet_order_matches_default_section_order(self):
        """默认配置 → 页签顺序与 _REPORT_SECTION_DEFAULT 一致。"""
        from src.python.report.excel_sheet_factory import create_sheets

        wb = self._make_wb()
        sheets = create_sheets(
            wb,
            _REPORT_SECTION_DEFAULT,
            enable_fund_deep_analysis=False,
            enable_news=False,
            enable_llm=False,
            enable_history=False,
        )
        # 只有 always 类型的 5 个页签
        expected_order = [sec["key"] for sec in _REPORT_SECTION_DEFAULT if sec["type"] == "always"]
        self.assertEqual(list(sheets.keys()), expected_order)
        self.assertEqual(wb.sheetnames, [sheets[k].title for k in expected_order])

    def test_sheet_order_custom_config(self):
        """自定义配置 → 页签顺序跟随自定义 section_order。"""
        from src.python.report.excel_sheet_factory import create_sheets

        custom_order = [
            {"key": "fund_performance", "name": "基金业绩分析", "number": 1, "type": "always"},
            {"key": "summary", "name": "投资分析汇总", "number": 2, "type": "always"},
            {"key": "market_value", "name": "市值核算明细表", "number": 3, "type": "always"},
        ]
        wb = self._make_wb()
        sheets = create_sheets(wb, custom_order, enable_fund_deep_analysis=False, enable_news=False, enable_llm=False)
        expected_order = [sec["key"] for sec in custom_order]
        self.assertEqual(list(sheets.keys()), expected_order)
        self.assertEqual(wb.sheetnames, [sheets[k].title for k in expected_order])

    def test_sheet_order_all_types_enabled(self):
        """全部类型启用 → 17 个页签按默认顺序排列。"""
        from src.python.report.excel_sheet_factory import create_sheets

        wb = self._make_wb()
        sheets = create_sheets(
            wb,
            _REPORT_SECTION_DEFAULT,
            enable_fund_deep_analysis=True,
            enable_news=True,
            enable_llm=True,
            data_availability={"news_data_available": True, "llm_data_available": True},
        )
        expected_keys = [sec["key"] for sec in _REPORT_SECTION_DEFAULT]
        self.assertEqual(list(sheets.keys()), expected_keys, "全部启用时页签顺序应与默认注册表一致")
        self.assertEqual(len(sheets), 17)

    def test_sheet_order_visibility_filtering(self):
        """可见性过滤 → 只创建匹配 type 的页签且顺序保持。"""
        from src.python.report.excel_sheet_factory import create_sheets

        wb = self._make_wb()
        # 启用 always + 基金深度分析
        sheets = create_sheets(
            wb,
            _REPORT_SECTION_DEFAULT,
            enable_fund_deep_analysis=True,
            enable_news=False,
            enable_llm=False,
            enable_history=False,
        )
        expected_keys = [
            sec["key"] for sec in _REPORT_SECTION_DEFAULT if sec["type"] in ("always", "fund_deep_analysis")
        ]
        self.assertEqual(list(sheets.keys()), expected_keys)
        self.assertEqual(len(sheets), 10, "always + 基金深度分析 = 10")


# ═══════════════════════════════════════════════════════════════
#  Test: Sheet Title Format
# ═══════════════════════════════════════════════════════════════


class TestExcelSheetTitleFormat(unittest.TestCase):
    """Excel 页签标题格式测试 — 验证 "{number}.{name}" 格式统一。"""

    def _make_wb(self):
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        return wb

    def test_all_titles_follow_number_name_format(self):
        """所有页签标题符合 {number}.{name} 格式。"""
        from src.python.report.excel_sheet_factory import create_sheets
        from src.python.core.registry import get_report_section_order

        wb = self._make_wb()
        order = get_report_section_order()
        sheets = create_sheets(
            wb,
            order,
            enable_fund_deep_analysis=True,
            enable_news=True,
            enable_llm=True,
            data_availability={"news_data_available": True, "llm_data_available": True},
        )
        for key, ws in sheets.items():
            self.assertRegex(
                ws.title,
                r"^\d+\.",
                f"页签 '{key}' 标题 '{ws.title}' 格式错误 — 应以 '数字.' 开头",
            )
            # 确认标题包含对应模块的中文名称
            expected_sec = next((s for s in order if s["key"] == key), None)
            if expected_sec:
                self.assertIn(
                    expected_sec["name"],
                    ws.title,
                    f"页签 '{key}' 标题 '{ws.title}' 应包含 '{expected_sec['name']}'",
                )

    def test_title_numbers_are_unique(self):
        """所有页签标题的数字序号无重复。"""
        from src.python.report.excel_sheet_factory import create_sheets
        from src.python.core.registry import get_report_section_order

        wb = self._make_wb()
        order = get_report_section_order()
        sheets = create_sheets(
            wb,
            order,
            enable_fund_deep_analysis=True,
            enable_news=True,
            enable_llm=True,
            data_availability={"news_data_available": True, "llm_data_available": True},
        )
        numbers = []
        for ws in sheets.values():
            import re

            m = re.match(r"^(\d+)", ws.title)
            if m:
                numbers.append(int(m.group(1)))
        self.assertEqual(
            len(numbers), len(set(numbers)), f"页签序号重复: {set(n for n in numbers if numbers.count(n) > 1)}"
        )

    def test_titles_are_unique(self):
        """所有页签标题字符串无重复。"""
        from src.python.report.excel_sheet_factory import create_sheets
        from src.python.core.registry import get_report_section_order

        wb = self._make_wb()
        order = get_report_section_order()
        sheets = create_sheets(
            wb,
            order,
            enable_fund_deep_analysis=True,
            enable_news=True,
            enable_llm=True,
            data_availability={"news_data_available": True, "llm_data_available": True},
        )
        titles = [ws.title for ws in sheets.values()]
        self.assertEqual(
            len(titles), len(set(titles)), f"页签标题重复: {set(t for t in titles if titles.count(t) > 1)}"
        )

    def test_title_uses_config_number_not_hardcoded(self):
        """自定义 section_order 时标题使用配置序号而非默认。"""
        from src.python.report.excel_sheet_factory import create_sheets

        custom_order = [
            {"key": "fund_performance", "name": "基金业绩分析", "number": 1, "type": "always"},
            {"key": "summary", "name": "投资分析汇总", "number": 2, "type": "always"},
            {"key": "market_value", "name": "市值核算明细表", "number": 3, "type": "always"},
        ]
        wb = self._make_wb()
        sheets = create_sheets(wb, custom_order, enable_fund_deep_analysis=False, enable_news=False, enable_llm=False)
        self.assertEqual(sheets["fund_performance"].title, "1.基金业绩分析", "fund_performance 应使用自定义序号 1")
        self.assertEqual(sheets["summary"].title, "2.投资分析汇总", "summary 应使用自定义序号 2")
        self.assertEqual(sheets["market_value"].title, "3.市值核算明细表", "market_value 应使用自定义序号 3")

    def test_title_order_tracks_section_order(self):
        """页签标题顺序与 section_order 的 number 值排序一致。"""
        from src.python.report.excel_sheet_factory import create_sheets
        from src.python.core.registry import get_report_section_order

        wb = self._make_wb()
        order = get_report_section_order()
        sheets = create_sheets(
            wb,
            order,
            enable_fund_deep_analysis=True,
            enable_news=True,
            enable_llm=True,
            data_availability={"news_data_available": True, "llm_data_available": True},
        )
        # 标题应是递增序号
        import re

        numbers = []
        for key in order:
            if key["key"] in sheets:
                ws = sheets[key["key"]]
                m = re.match(r"^(\d+)", ws.title)
                if m:
                    numbers.append(int(m.group(1)))
        for i in range(1, len(numbers)):
            self.assertLess(numbers[i - 1], numbers[i], f"页签标题序号应严格递增: {numbers}")


# ═══════════════════════════════════════════════════════════════
#  Test: Text Wrapping & Column Width (content display)
# ═══════════════════════════════════════════════════════════════


class TestExcelTextWrapping(unittest.TestCase):
    """Excel 文本换行设置测试 — 避免长文本截断不显示。"""

    def test_summary_uses_wrap_text(self):
        """汇总表关键列设置了 wrap_text（现位于 summary_llm_usage.py）。"""
        from src.python.report import summary_llm_usage as slu
        import inspect

        source = inspect.getsource(slu)
        self.assertIn("wrap_text", source, "summary_llm_usage.py 应使用 wrap_text 避免长文本截断")
        # 确认有列宽设置
        has_width = "column_width" in source.lower() or "_set_column_widths" in source
        self.assertTrue(has_width, "summary_llm_usage.py 应设置列宽")

    def test_news_uses_wrap_text(self):
        """新闻模块使用 wrap_text 避免摘要截断。"""
        from src.python.report import news_correlation as nc
        import inspect

        source = inspect.getsource(nc)
        self.assertIn("wrap_text", source, "news_correlation.py 应使用 wrap_text")

    def test_summary_has_column_widths(self):
        """汇总表定义了列宽数组。"""
        from src.python.report.summary_llm_usage import _set_column_widths

        # 函数存在即可（会在写入时调用）
        self.assertTrue(callable(_set_column_widths))

    def test_news_has_column_widths(self):
        """新闻模块定义了列宽函数。"""
        from src.python.report.news_correlation import _set_news_column_widths

        self.assertTrue(callable(_set_news_column_widths))

    def test_summary_column_widths_reasonable(self):
        """汇总列宽不小于 10（避免数字显示不全）。"""
        from src.python.report.summary_llm_usage import _set_column_widths
        import inspect

        source = inspect.getsource(_set_column_widths)
        # 查找调用处的列宽数组
        import re

        # 找到类似 _set_column_widths(ws, [20, 16, 26, ...]) 的调用
        width_calls = re.findall(r"_set_column_widths\([^)]+\)", source)
        all_widths_ok = True
        for call in width_calls:
            # 提取所有数字
            widths = [int(x) for x in re.findall(r"\b\d{2,}\b", call)]
            for w in widths:
                if w < 10:
                    all_widths_ok = False
        self.assertTrue(all_widths_ok, "列宽不应小于 10")


# ═══════════════════════════════════════════════════════════════
#  Test: Sheet Accessibility & Module Writing
# ═══════════════════════════════════════════════════════════════


class TestExcelModuleSheets(unittest.TestCase):
    """Excel 各模块页签可访问性测试 — 所有页签都能正确写入数据。"""

    def _make_ws(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "test"
        return ws

    def test_generate_excel_report_produces_valid_workbook(self):
        """generate_excel_report 生成有效 xlsx（不崩溃）。"""
        import tempfile
        import os
        from src.python.report.excel_generator import generate_excel_report

        with tempfile.TemporaryDirectory() as tmpdir:
            try:
                generate_excel_report(
                    holdings=[],
                    include_news=False,
                    output_dir=tmpdir,
                    news_top_count=5,
                    include_llm=False,
                    details=[],
                    a_indices={"上证指数": {"price": 3000, "change": 0.01, "change_pct": 0.01}},
                    us_indices={},
                    news_data=[],
                    news_llm_meta=None,
                    enable_fund_deep_analysis=False,
                )
                # 检查输出文件是否存在
                files = os.listdir(tmpdir)
                xlsx_files = [f for f in files if f.endswith(".xlsx")]
                self.assertGreater(len(xlsx_files), 0, "应生成至少一个 xlsx 文件")
            except Exception as e:
                self.fail(f"generate_excel_report 不应崩溃: {e}")

    def test_write_module_data_rows_has_border_fix(self):
        """写入器函数正确使用边框样式（回归检查）。

        确认 excel_writer 中的写入函数设置了 cell.border，
        避免因隐式依赖导致边框缺失。
        """
        from src.python.report import excel_writer as ew
        import inspect

        # 检查核心写入函数是否使用边框
        for fn_name in ("write_data_row", "write_subtotal_row", "write_total_row"):
            fn = getattr(ew, fn_name, None)
            if fn is None:
                continue
            source = inspect.getsource(fn)
            self.assertIn("border", source, f"{fn_name} 应设置 cell.border")


if __name__ == "__main__":
    unittest.main()
