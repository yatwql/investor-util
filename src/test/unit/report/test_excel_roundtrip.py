"""Excel 文件往返测试 — 保存的 xlsx 可重新打开。

运行：
  cd D:/codebase/zoo/investor-util
  pytest src/test/unit/report/test_excel_roundtrip.py -v
"""

from __future__ import annotations

import os
import tempfile
import unittest

import pytest

from src.python.report.excel_writer import create_workbook, save_workbook

pytestmark = [pytest.mark.unit, pytest.mark.unit_report]


class TestWorkbookSaveRoundtrip(unittest.TestCase):
    """Workbook 保存/读取往返测试。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()

    def tearDown(self):
        self.tmp.cleanup()

    def test_save_and_reopen(self):
        """保存的 xlsx 可被 openpyxl 重新打开。"""
        from openpyxl import load_workbook

        wb = create_workbook()
        ws = wb.active
        ws.title = "测试页签"
        ws["A1"] = "测试内容"

        path = save_workbook(wb, output_dir=self.tmp.name)
        self.assertTrue(os.path.exists(path))

        # 重新打开
        loaded = load_workbook(path, read_only=True)
        self.assertIn("测试页签", loaded.sheetnames)
        self.assertEqual(loaded["测试页签"]["A1"].value, "测试内容")
        loaded.close()

    def test_multiple_sheets_saved(self):
        """多页签 workbook 正确保存。"""
        wb = create_workbook()
        wb.remove(wb.active)

        names = []
        for i in range(4):
            ws = wb.create_sheet()
            name = f"页签{i + 1}"
            ws.title = name
            ws["A1"] = name
            names.append(name)

        path = save_workbook(wb, output_dir=self.tmp.name)
        self.assertTrue(os.path.exists(path))

        from openpyxl import load_workbook
        loaded = load_workbook(path, read_only=True)
        self.assertEqual(len(loaded.sheetnames), 4)
        for name in names:
            self.assertIn(name, loaded.sheetnames)
        loaded.close()


if __name__ == "__main__":
    unittest.main()
