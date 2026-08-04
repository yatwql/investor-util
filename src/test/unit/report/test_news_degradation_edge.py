"""新闻全源/部分失败占位 + source_status — 边缘测试。

测试目标：
  - write_news_sheet: news_data=[] + source_status 全失败 → _write_placeholder
  - write_news_sheet: news_data=[] + source_status 部分失败 → "暂无关联新闻"
  - write_news_sheet: news_data=[] + source_status 全正常 → "暂无关联新闻"
  - write_news_sheet: news_data=[] + 无 source_status → "暂无关联新闻"
  - write_news_sheet: news_data 非空 + 部分失败 → 底部注明不可用源
  - write_news_sheet: news_data 非空 + 全正常 → 底部不出现不可用源

运行：
  pytest src/test/unit/report/test_news_degradation_edge.py -v
"""

from __future__ import annotations

import unittest

import openpyxl
import pytest

from src.python.report.data_status import STATUS_MESSAGES

pytestmark = [pytest.mark.unit, pytest.mark.unit_report, pytest.mark.edge]


def _make_source_status(successes: dict[str, bool]) -> dict:
    """生成 source_status 测试数据。"""
    labels = {
        "sina": "新浪", "eastmoney": "东方财富", "cls": "财联社",
        "wallstreetcn": "华尔街见闻", "akshare": "akshare",
    }
    return {
        k: {"label": labels.get(k, k), "success": v}
        for k, v in successes.items()
    }


class TestWriteNewsSheetDegradation(unittest.TestCase):
    """write_news_sheet 降级占位测试"""

    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_all_sources_failed_writes_placeholder(self):
        """news_data=[] + source_status 全失败 → 写占位文本。"""
        llm_meta = {
            "source_status": _make_source_status(
                {"sina": False, "eastmoney": False, "cls": False}
            ),
            "llm_enabled": False,
        }
        from src.python.report.news_correlation import write_news_sheet
        write_news_sheet(self.ws, [], llm_meta)
        cell_val = self.ws.cell(row=3, column=1).value
        self.assertIn("新闻数据暂不可用", str(cell_val or ""))

    def test_partial_sources_failed_shows_empty_note(self):
        """news_data=[] + source_status 部分失败 → 显示"暂无关联新闻"。"""
        llm_meta = {
            "source_status": _make_source_status(
                {"sina": False, "eastmoney": True, "cls": True}
            ),
        }
        from src.python.report.news_correlation import write_news_sheet
        write_news_sheet(self.ws, [], llm_meta)
        cell_val = self.ws.cell(row=3, column=1).value
        self.assertIn("暂无关联新闻", str(cell_val or ""))

    def test_all_sources_ok_shows_empty_note(self):
        """news_data=[] + source_status 全成功 → 显示"暂无关联新闻"。"""
        llm_meta = {
            "source_status": _make_source_status(
                {"sina": True, "eastmoney": True}
            ),
        }
        from src.python.report.news_correlation import write_news_sheet
        write_news_sheet(self.ws, [], llm_meta)
        cell_val = self.ws.cell(row=3, column=1).value
        self.assertIn("暂无关联新闻", str(cell_val or ""))

    def test_no_source_status_shows_empty_note(self):
        """news_data=[] + source_status 缺失 → 显示"暂无关联新闻"。"""
        from src.python.report.news_correlation import write_news_sheet
        write_news_sheet(self.ws, [], {})
        cell_val = self.ws.cell(row=3, column=1).value
        self.assertIn("暂无关联新闻", str(cell_val or ""))

    def test_footer_lists_failed_sources(self):
        """news_data 非空 + source_status 有失败 → 底部列出不可用源。"""
        llm_meta = {
            "source_status": _make_source_status(
                {"sina": False, "eastmoney": True, "cls": False}
            ),
            "llm_enabled": False,
        }
        news_data = [
            {"title": "测试新闻", "intro": "简介", "url": "http://example.com",
             "ctime": "2026-07-07", "media_name": "测试", "matched_keywords": ["keyword"]},
        ]
        from src.python.report.news_correlation import write_news_sheet
        write_news_sheet(self.ws, news_data, llm_meta)
        for r in range(self.ws.max_row, 0, -1):
            val = self.ws.cell(row=r, column=1).value
            if val and "以下新闻源不可用" in str(val):
                self.assertIn("新浪", str(val))
                self.assertIn("财联社", str(val))
                return
        self.fail("底部未找到'以下新闻源不可用'")

    def test_footer_no_failure_when_all_ok(self):
        """news_data 非空 + source_status 全成功 → 底部无失败信息。"""
        llm_meta = {
            "source_status": _make_source_status(
                {"sina": True, "eastmoney": True}
            ),
            "llm_enabled": False,
        }
        news_data = [
            {"title": "测试新闻", "intro": "简介", "url": "http://example.com",
             "ctime": "2026-07-07", "media_name": "测试", "matched_keywords": ["keyword"]},
        ]
        from src.python.report.news_correlation import write_news_sheet
        write_news_sheet(self.ws, news_data, llm_meta)
        for r in range(self.ws.max_row, 0, -1):
            val = self.ws.cell(row=r, column=1).value
            if val and "以下新闻源不可用" in str(val):
                self.fail("所有源正常时不应出现'以下新闻源不可用'")

    def test_placeholder_uses_status_message_constant(self):
        """全源失败占位文本来自 STATUS_MESSAGES。"""
        llm_meta = {
            "source_status": _make_source_status(
                {"sina": False, "eastmoney": False}
            ),
            "llm_enabled": False,
        }
        from src.python.report.news_correlation import write_news_sheet
        write_news_sheet(self.ws, [], llm_meta)
        self.assertEqual(
            self.ws.cell(row=3, column=1).value,
            STATUS_MESSAGES["news_all_failed"],
        )
