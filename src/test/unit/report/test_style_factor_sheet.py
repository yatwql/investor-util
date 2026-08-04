"""风格与因子分析页签 Excel 渲染测试（一章三区块）。

覆盖：
  - 两区块同章渲染（一、基金风格表 + 二、风格因子回归）
  - 风格表数据渲染（基金名称/风格/漂移等级）
  - 因子回归数据渲染（因子名/β/t 值/显著/基准对照）
  - 行业 Beta 子表渲染（暴露占比 + β/显著/相关性）
  - 行业 Beta 无数据/开关关 → 区块不渲染
  - 各区块独立降级（风格表空 / 因子空 → 占位，不影响另一区块）
  - 全空 → 整页占位
  - 无指数映射行业仅显示暴露占比（— 占位）

运行：
  python -m pytest src/test/unit/report/test_style_factor_sheet.py -v
"""

from __future__ import annotations

import unittest

import pytest

pytestmark = [pytest.mark.unit, pytest.mark.unit_report]


def _style_results() -> list[dict]:
    """基金风格表数据（1 条基准确立中 + 1 条中度漂移）。"""
    return [
        {
            "name": "易方达中小盘",
            "code": "110011",
            "current_style": "大盘成长",
            "prev_style": "--",
            "drift_level": "基准确立中",
            "drift_score": 0,
            "is_first_check": True,
            "is_estimated": False,
        },
        {
            "name": "某成长混合",
            "code": "002001",
            "current_style": "小盘成长",
            "prev_style": "大盘成长",
            "drift_level": "中度",
            "drift_score": 2,
            "is_first_check": False,
            "is_estimated": True,
        },
    ]


def _factor_exposure(**extra) -> dict:
    """style_factor_data dict（因子回归区块 mock）。"""
    d = {
        "available": True,
        "status": "ok",
        "betas": {"value": 0.8, "growth": 1.2, "quality": -0.3},
        "t_stats": {"value": 4.5, "growth": 6.1, "quality": -1.2},
        "significant": {"value": True, "growth": True, "quality": False},
        "style_allocation": {"value": 0.35, "growth": 0.52, "quality": 0.13},
        "baseline_betas": {"value": 0.5, "growth": 0.4, "quality": 0.1},
        "factor_correlations": {"价值-成长": 0.3},
        "correlation_note": "",
        "alpha": 0.0001,
        "window": 60,
        "sample_count": 60,
        "stale_factors": [],
    }
    d.update(extra)
    return d


_FACTOR_NAMES = {"value": "价值", "growth": "成长", "quality": "质量"}


def _industry_beta(**extra) -> dict:
    """style_factor_data.industry_beta 子键 mock。"""
    d = {
        "available": True,
        "status": "ok",
        "exposure": {"银行": 0.4, "白酒": 0.35, "半导体": 0.25},
        "betas": {"银行": 1.1, "白酒": 0.7},
        "alphas": {"银行": 0.0, "白酒": 0.0},
        "t_stats": {"银行": 5.2, "白酒": 2.1},
        "significant": {"银行": True, "白酒": False},
        "correlations": {"银行": 0.8, "白酒": 0.5},
        "index_codes": {"银行": "sh000986", "白酒": "sz399997"},
        "unmapped_industries": ["半导体"],
        "window": 60,
        "sample_count": 60,
    }
    d.update(extra)
    return d


class TestStyleFactorSheet(unittest.TestCase):
    """风格与因子分析页签渲染。"""

    def _write(self, **kwargs) -> "object":
        from openpyxl import Workbook

        from src.python.report.style_factor_sheet import write_style_factor_sheet

        wb = Workbook()
        ws = wb.active
        write_style_factor_sheet(ws, **kwargs)
        return ws

    def _flat(self, ws) -> list[str]:
        return [
            str(c.value) if c.value is not None else ""
            for row in ws.iter_rows()
            for c in row
        ]

    def _pct_texts(self, ws) -> list[str]:
        """提取百分比格式单元格的展示文本（value × 100 后补 %）。"""
        texts: list[str] = []
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None and c.number_format == "0.00%":
                    texts.append(f"{float(c.value) * 100:.2f}%")
        return texts

    def _col_texts(self, ws, col: int) -> list[str]:
        """返回指定列的全部单元格文本。"""
        return [
            str(row[col - 1].value) if row[col - 1].value is not None else ""
            for row in ws.iter_rows()
        ]

    # ── 两区块同章渲染 ──

    def test_two_blocks_same_sheet(self):
        """风格表 + 因子回归两区块同章渲染。"""
        ws = self._write(style_data=_style_results(), factor_exposure=_factor_exposure())
        flat = self._flat(ws)
        self.assertTrue(any("一、基金风格表" in v for v in flat), "区块一标题应存在")
        self.assertTrue(any("二、风格因子回归" in v for v in flat), "区块二标题应存在")

    def test_style_block_rows(self):
        """风格表数据正确渲染（基金名/风格/漂移）。"""
        ws = self._write(style_data=_style_results(), factor_exposure=_factor_exposure())
        flat = self._flat(ws)
        self.assertTrue(any("易方达中小盘" in v for v in flat))
        self.assertTrue(any("大盘成长" in v for v in flat))
        self.assertTrue(any("中度" in v for v in flat))
        self.assertTrue(any("基准确立中" in v for v in flat))

    def test_factor_block_rows(self):
        """因子回归数据正确渲染（因子名/β/显著/基准对照）。"""
        ws = self._write(style_data=_style_results(), factor_exposure=_factor_exposure())
        flat = self._flat(ws)
        self.assertTrue(any("价值" in v for v in flat), "因子中文名应渲染")
        self.assertTrue(any("0.8" in v for v in flat), "价值 β 应渲染")
        self.assertTrue(any("✅ 显著" in v for v in flat), "显著标记应渲染")
        self.assertTrue(any("基准对照" in v for v in flat), "基准对照子标题应渲染")

    # ── 行业 Beta 子表 ──

    def test_industry_beta_block(self):
        """行业 Beta 子表渲染（暴露占比 + β + 显著 + 相关性）。"""
        ws = self._write(
            style_data=_style_results(),
            factor_exposure=_factor_exposure(),
            industry_beta=_industry_beta(),
        )
        flat = self._flat(ws)
        self.assertTrue(any("三、行业 Beta" in v for v in flat), "区块三标题应存在")
        self.assertTrue(any("银行" in v for v in flat))
        self.assertIn("40.00%", self._pct_texts(ws), "暴露占比 0.4 应渲染为 40.00%")
        self.assertTrue(any("sh000986" in v for v in flat), "指数代码应渲染")

    def test_industry_beta_unmapped_only_exposure(self):
        """无指数映射行业（半导体）仅显示暴露占比，β 列占位。"""
        ws = self._write(
            style_data=[],
            factor_exposure=None,
            industry_beta=_industry_beta(),
        )
        flat = self._flat(ws)
        # 半导体有暴露占比（25.00%）但无 β（— 占位）
        self.assertIn("25.00%", self._pct_texts(ws), "半导体暴露占比应渲染")
        # 半导体行 β 列（第 4 列）应为 — 占位
        for row in ws.iter_rows():
            if row[0].value == "半导体":
                self.assertEqual(row[3].value, "—", "半导体 β 应为占位")
        self.assertFalse(any("半导体" in v and "0." in v for v in flat), "半导体不应有 β 值")

    def test_industry_beta_none_hides_block(self):
        """行业 Beta None（开关关）→ 区块不渲染，前两区块正常。"""
        ws = self._write(
            style_data=_style_results(),
            factor_exposure=_factor_exposure(),
            industry_beta=None,
        )
        flat = self._flat(ws)
        self.assertTrue(any("一、基金风格表" in v for v in flat))
        self.assertTrue(any("二、风格因子回归" in v for v in flat))
        self.assertFalse(any("行业 Beta" in v for v in flat), "无数据时区块三不渲染")

    def test_industry_beta_unavailable_shows_placeholder(self):
        """行业 Beta available=False（push2 不足）→ 标题 + 数据不足占位。"""
        ws = self._write(
            style_data=[],
            factor_exposure=None,
            industry_beta=_industry_beta(available=False, status="insufficient"),
        )
        flat = self._flat(ws)
        self.assertTrue(any("三、行业 Beta" in v for v in flat), "数据不足时仍渲染标题")
        self.assertTrue(any("行业 Beta 数据不足" in v for v in flat), "数据不足占位")
        self.assertFalse(any("sh000986" in v for v in flat), "Beta 子表不渲染")

    # ── 各区块独立降级 ──

    def test_style_empty_placeholder_factor_ok(self):
        """风格表空 → 占位；因子回归正常渲染。"""
        ws = self._write(style_data=[], factor_exposure=_factor_exposure())
        flat = self._flat(ws)
        self.assertTrue(any("基金风格数据暂不可用" in v for v in flat), "风格表占位")
        self.assertTrue(any("二、风格因子回归" in v for v in flat), "因子区块不受影响")

    def test_factor_empty_placeholder_style_ok(self):
        """因子空 → 占位；风格表正常渲染。"""
        ws = self._write(style_data=_style_results(), factor_exposure=None)
        flat = self._flat(ws)
        self.assertTrue(any("因子暴露数据暂不可用" in v for v in flat), "因子区块占位")
        self.assertTrue(any("易方达中小盘" in v for v in flat), "风格表不受影响")

    def test_factor_unavailable_placeholder(self):
        """因子 available=False → 因子区块占位。"""
        ws = self._write(
            style_data=_style_results(),
            factor_exposure=_factor_exposure(available=False, status="insufficient"),
        )
        flat = self._flat(ws)
        self.assertTrue(any("因子暴露数据暂不可用" in v for v in flat))

    def test_all_empty_whole_sheet_placeholder(self):
        """三区块均无数据 → 整页占位（风格/因子占位均出现）。"""
        ws = self._write(style_data=[], factor_exposure=None, industry_beta=None)
        flat = self._flat(ws)
        self.assertTrue(any("基金风格数据暂不可用" in v for v in flat))
        self.assertTrue(any("因子暴露数据暂不可用" in v for v in flat))

    def test_drift_level_font_applied(self):
        """中度漂移行字体颜色应用（CC0000/FF8C00 系列）。"""
        ws = self._write(style_data=_style_results(), factor_exposure=None)
        # 中度漂移（FF8C00）应出现在某单元格字体颜色
        colors = {
            c.font.color.rgb if c.font and c.font.color and c.font.color.rgb else None
            for row in ws.iter_rows()
            for c in row
        }
        self.assertTrue(any("FF8C00" in str(col) for col in colors if col), "中度漂移橙色应应用")
