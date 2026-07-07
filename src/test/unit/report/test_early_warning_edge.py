"""D-7b: 智能预警模块空数据页签写入 — 边缘测试。

测试目标：
  - write_early_warning_sheet: 无行业/新闻 → 两段都显示不可用/需开启
  - write_early_warning_sheet: 有行业无新闻 → 行业段显示暂无，新闻段需开启
  - write_early_warning_sheet: 有新闻无行业 → 行业段不可用，新闻段暂无

运行：
  pytest src/test/unit/report/test_early_warning_edge.py -v
"""

from __future__ import annotations

import unittest

import openpyxl
import pytest

pytestmark = [pytest.mark.unit, pytest.mark.unit_report, pytest.mark.edge]


def _make_warnings(has_sector: bool, has_llm: bool) -> dict:
    """生成 compute_early_warnings 结果结构。"""
    return {
        "sector_alerts": [],
        "sentiment_alerts": [],
        "has_warnings": False,
        "has_sector_data": has_sector,
        "has_llm_news": has_llm,
    }


class TestWriteEarlyWarningSheetEmpty(unittest.TestCase):
    """write_early_warning_sheet 空/缺失数据场景"""

    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def _cell_texts(self) -> str:
        """返回工作表中所有单元格文本的拼接字符串（不含 None）。"""
        parts: list[str] = []
        for r in range(1, self.ws.max_row + 1):
            for c in range(1, self.ws.max_column + 1):
                v = self.ws.cell(row=r, column=c).value
                if v:
                    parts.append(str(v))
        return "\n".join(parts)

    def test_no_sector_no_news(self):
        """has_sector_data=False + has_llm_news=False → 行业不可用 + 新闻需开启。"""
        from src.python.report.early_warning import write_early_warning_sheet
        write_early_warning_sheet(self.ws, _make_warnings(False, False))
        text = self._cell_texts()
        self.assertIn("行业资金流向数据不可用", text)
        self.assertIn("开启", text)

    def test_has_sector_no_news(self):
        """has_sector_data=True + has_llm_news=False → 行业暂无 + 新闻需开启。"""
        from src.python.report.early_warning import write_early_warning_sheet
        write_early_warning_sheet(self.ws, _make_warnings(True, False))
        text = self._cell_texts()
        self.assertIn("暂无行业资金流向预警", text)
        self.assertIn("开启", text)

    def test_no_sector_has_news(self):
        """has_sector_data=False + has_llm_news=True → 行业不可用 + 新闻暂无。"""
        from src.python.report.early_warning import write_early_warning_sheet
        write_early_warning_sheet(self.ws, _make_warnings(False, True))
        text = self._cell_texts()
        self.assertIn("行业资金流向数据不可用", text)
        self.assertIn("暂无新闻情绪聚合数据", text)

    def test_has_sector_has_news_no_alerts(self):
        """两者都有但无预警 → 两段都显示暂无。"""
        from src.python.report.early_warning import write_early_warning_sheet
        write_early_warning_sheet(self.ws, _make_warnings(True, True))
        text = self._cell_texts()
        self.assertIn("暂无行业资金流向预警", text)
        self.assertIn("暂无新闻情绪聚合数据", text)

    def test_with_sector_alerts(self):
        """有行业预警数据 → 正常渲染表头。"""
        warnings = _make_warnings(True, True)
        warnings["sector_alerts"] = [{
            "sector_name": "银行", "main_net_inflow": -100_000_000,
            "main_net_inflow_pct": -5.0, "change_pct": -1.5,
            "top_stock": "工商银行",
            "matched_assets": [{"name": "招商银行", "codes": ["600036"],
                                "mv": 100000, "ratio_pct": 5.0}],
            "alert_level": "danger",
        }]
        from src.python.report.early_warning import write_early_warning_sheet
        write_early_warning_sheet(self.ws, warnings)
        text = self._cell_texts()
        self.assertIn("银行", text)
        self.assertIn("危险", text)
