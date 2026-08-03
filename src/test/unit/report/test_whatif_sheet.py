"""调仓 What-if 页签 Excel 呈现测试。

覆盖：
  - available=True → 调仓摘要（文件对比 + 变动统计 + 汇总指标 + 说明）/
    分类配置对比 / 持仓变动明细 三页签内容
  - 变动统计数字正确
  - 汇总指标 delta + 箭头（如 "2000.00 ↓"）
  - 变动明细行底色按类型标注（新增绿/清仓红/加仓黄/减仓蓝/不变灰）
  - available=False → 摘要页占位
  - whatif_data=None → 各页占位
"""

from __future__ import annotations

import unittest

import pytest

pytestmark = [pytest.mark.unit, pytest.mark.unit_report]


def _whatif_data(**extra) -> dict:
    """构造 C19 契约 whatif_data mock（新增/清仓/加仓/减仓/不变 各 1 条）。"""
    d = {
        "available": True,
        "status": "ok",
        "base_file": "before.xlsx",
        "candidate_file": "after.xlsx",
        "base": {"total_cost": 26000.0, "total_shares": 8000.0, "holding_count": 4, "hhi": 0.5},
        "candidate": {"total_cost": 24000.0, "total_shares": 6200.0, "holding_count": 4, "hhi": 0.6},
        "summary": [
            {
                "key": "total_cost",
                "label": "总成本(元)",
                "unit": "money",
                "base": 26000.0,
                "candidate": 24000.0,
                "delta": -2000.0,
                "arrow": "↓",
            },
            {
                "key": "total_shares",
                "label": "总份额",
                "unit": "shares",
                "base": 8000.0,
                "candidate": 6200.0,
                "delta": -1800.0,
                "arrow": "↓",
            },
            {
                "key": "holding_count",
                "label": "持仓品种数",
                "unit": "count",
                "base": 4,
                "candidate": 4,
                "delta": 0,
                "arrow": "→",
            },
            {
                "key": "hhi",
                "label": "持仓集中度 HHI",
                "unit": "hhi",
                "base": 0.5,
                "candidate": 0.6,
                "delta": 0.1,
                "arrow": "↑",
            },
        ],
        "categories": [
            {
                "key": "equity",
                "label": "股票",
                "base_cost": 20000.0,
                "cand_cost": 12000.0,
                "base_weight": 76.92,
                "cand_weight": 50.0,
                "delta_pct": -26.92,
            },
            {
                "key": "fixed_income",
                "label": "固收",
                "base_cost": 6000.0,
                "cand_cost": 12000.0,
                "base_weight": 23.08,
                "cand_weight": 50.0,
                "delta_pct": 26.92,
            },
        ],
        "changes": [
            {
                "code": "511880",
                "name": "货币ETF",
                "action": "新增",
                "base_shares": 0.0,
                "cand_shares": 2000.0,
                "shares_diff": 2000.0,
                "base_cost": 0.0,
                "cand_cost": 2000.0,
                "cost_diff": 2000.0,
                "base_weight": 0.0,
                "cand_weight": 8.33,
                "weight_delta_pct": 8.33,
            },
            {
                "code": "510300",
                "name": "沪深300ETF",
                "action": "清仓",
                "base_shares": 2000.0,
                "cand_shares": 0.0,
                "shares_diff": -2000.0,
                "base_cost": 6000.0,
                "cand_cost": 0.0,
                "cost_diff": -6000.0,
                "base_weight": 23.08,
                "cand_weight": 0.0,
                "weight_delta_pct": -23.08,
            },
            {
                "code": "600900",
                "name": "长江电力",
                "action": "加仓",
                "base_shares": 1000.0,
                "cand_shares": 1200.0,
                "shares_diff": 200.0,
                "base_cost": 20000.0,
                "cand_cost": 24000.0,
                "cost_diff": 4000.0,
                "base_weight": 76.92,
                "cand_weight": 100.0,
                "weight_delta_pct": 23.08,
            },
            {
                "code": "600028",
                "name": "中国石化",
                "action": "减仓",
                "base_shares": 3000.0,
                "cand_shares": 1000.0,
                "shares_diff": -2000.0,
                "base_cost": 6000.0,
                "cand_cost": 2000.0,
                "cost_diff": -4000.0,
                "base_weight": 10.0,
                "cand_weight": 5.0,
                "weight_delta_pct": -5.0,
            },
            {
                "code": "511010",
                "name": "国债ETF",
                "action": "不变",
                "base_shares": 5000.0,
                "cand_shares": 5000.0,
                "shares_diff": 0.0,
                "base_cost": 50000.0,
                "cand_cost": 50000.0,
                "cost_diff": 0.0,
                "base_weight": 50.0,
                "cand_weight": 50.0,
                "weight_delta_pct": 0.0,
            },
        ],
        "stats": {"added": 1, "removed": 1, "increased": 1, "decreased": 1, "unchanged": 1},
        "reason": "",
    }
    d.update(extra)
    return d


class TestWhatifExcelSheets(unittest.TestCase):
    """调仓 What-if 三页签 Excel 呈现测试。"""

    def _write_sheet(self, fn_name, whatif_data) -> "object":
        from openpyxl import Workbook

        from src.python.report.whatif_sheet import (
            write_whatif_backtest_sheet,
            write_whatif_category_sheet,
            write_whatif_changes_sheet,
            write_whatif_summary_sheet,
        )

        wb = Workbook()
        ws = wb.active
        {
            "write_whatif_summary_sheet": write_whatif_summary_sheet,
            "write_whatif_category_sheet": write_whatif_category_sheet,
            "write_whatif_changes_sheet": write_whatif_changes_sheet,
            "write_whatif_backtest_sheet": write_whatif_backtest_sheet,
        }[fn_name](ws, whatif_data)
        return ws

    def _flat(self, ws) -> list[str]:
        return [str(c.value) if c.value is not None else "" for row in ws.iter_rows() for c in row]

    # ── 摘要页签 ─────────────────────────────────────────

    def test_summary_sheet_full(self):
        """摘要页：标题 + 文件对比 + 变动统计 + 汇总指标（含箭头）+ 说明。"""
        ws = self._write_sheet("write_whatif_summary_sheet", _whatif_data())
        flat = self._flat(ws)
        self.assertTrue(any("调仓 What-if" in v for v in flat), "应含页标题")
        self.assertTrue(any("before.xlsx" in v and "after.xlsx" in v for v in flat), "应含文件对比")
        self.assertTrue(any("新增 1" in v and "清仓 1" in v for v in flat), "应含变动统计")
        self.assertTrue(any("总成本" in v for v in flat))
        self.assertTrue(any("26000.0" in v for v in flat), "应含基准总成本")
        self.assertTrue(any("24000.0" in v for v in flat), "应含目标总成本")
        self.assertTrue(any("-2000.0 ↓" in v for v in flat), "变化量应带箭头")
        self.assertTrue(any("口径" in v for v in flat), "应含口径说明")

    def test_summary_sheet_unavailable_placeholder(self):
        """available=False → 摘要页占位。"""
        data = _whatif_data(
            available=False,
            status="empty",
            summary=[],
            categories=[],
            changes=[],
            stats={},
            reason="基准或目标持仓为空",
        )
        ws = self._write_sheet("write_whatif_summary_sheet", data)
        flat = self._flat(ws)
        self.assertTrue(any("调仓对比数据暂不可用" in v for v in flat), "available=False 应写占位")

    def test_summary_sheet_none_placeholder(self):
        """whatif_data=None → 摘要页占位。"""
        ws = self._write_sheet("write_whatif_summary_sheet", None)
        flat = self._flat(ws)
        self.assertTrue(any("调仓对比数据暂不可用" in v for v in flat), "None 应写占位")

    # ── 分类配置页签 ─────────────────────────────────────

    def test_category_sheet_full(self):
        """分类配置页：资产大类 + 基准/目标权重 + 变化。"""
        ws = self._write_sheet("write_whatif_category_sheet", _whatif_data())
        flat = self._flat(ws)
        self.assertTrue(any("资产配置对比" in v for v in flat), "应含页标题")
        self.assertTrue(any("股票" in v for v in flat))
        self.assertTrue(any("固收" in v for v in flat))
        self.assertTrue(any("76.92" in v for v in flat), "应含基准权重")
        self.assertTrue(any("-26.92" in v for v in flat), "应含变化(百分点)")

    def test_category_sheet_none_placeholder(self):
        """whatif_data=None → 分类配置页占位。"""
        ws = self._write_sheet("write_whatif_category_sheet", None)
        flat = self._flat(ws)
        self.assertTrue(any("调仓对比数据暂不可用" in v for v in flat))

    # ── 持仓变动明细页签 ─────────────────────────────────

    def test_changes_sheet_full(self):
        """变动明细页：全部变动条目 + 动作列。"""
        ws = self._write_sheet("write_whatif_changes_sheet", _whatif_data())
        flat = self._flat(ws)
        self.assertTrue(any("持仓变动明细" in v for v in flat), "应含页标题")
        self.assertIn("新增", flat)
        self.assertIn("清仓", flat)
        self.assertIn("加仓", flat)
        self.assertIn("货币ETF", flat)
        self.assertIn("600900", flat)

    def test_changes_sheet_row_fill_by_action(self):
        """变动明细行底色：新增绿 / 清仓红 / 加仓黄 / 减仓蓝 / 不变灰。"""
        from openpyxl import Workbook

        from src.python.report.whatif_sheet import write_whatif_changes_sheet

        wb = Workbook()
        ws = wb.active
        write_whatif_changes_sheet(ws, _whatif_data())
        # 逐行动作 → 期望底色
        fill_by_row = {}
        for row in ws.iter_rows():
            action_cell = row[0]
            if action_cell.value in ("新增", "清仓", "加仓", "减仓", "不变"):
                fill = action_cell.fill
                fill_by_row[action_cell.value] = fill.start_color.rgb if fill else None
        self.assertEqual(fill_by_row["新增"], "00C6EFCE", "新增行应为绿色底")
        self.assertEqual(fill_by_row["清仓"], "00FFC7CE", "清仓行应为红色底")
        self.assertEqual(fill_by_row["加仓"], "00FFEB9C", "加仓行应为黄色底")
        self.assertEqual(fill_by_row["减仓"], "00DDEBF7", "减仓行应为蓝色底")
        self.assertEqual(fill_by_row["不变"], "00F2F2F2", "不变行应为浅灰底")

    def test_changes_sheet_none_placeholder(self):
        """whatif_data=None → 变动明细页占位。"""
        ws = self._write_sheet("write_whatif_changes_sheet", None)
        flat = self._flat(ws)
        self.assertTrue(any("调仓对比数据暂不可用" in v for v in flat))


def _bt_data(**extra) -> dict:
    """构造含可用 backtest 键的 C19 契约（Excel 呈现用，数值无需自洽）。"""
    d = _whatif_data(
        backtest={
            "available": True,
            "status": "ok",
            "effective_date": "2026-07-01",
            "reason": "",
            "metrics": [
                {
                    "key": "period_return_pct",
                    "label": "区间收益",
                    "unit": "pct",
                    "base": 24.0,
                    "candidate": 48.0,
                    "delta": 24.0,
                    "arrow": "↑",
                },
                {
                    "key": "sharpe_ratio",
                    "label": "夏普比率",
                    "unit": "ratio",
                    "base": 1.2,
                    "candidate": 0.8,
                    "delta": -0.4,
                    "arrow": "↓",
                },
            ],
            "series": {
                "labels": ["2026-07-01", "2026-07-02"],
                "base": [100.0, 124.0],
                "candidate": [100.0, 148.0],
                "base_drawdown": [0.0, 0.0],
                "candidate_drawdown": [0.0, 0.0],
            },
        }
    )
    d.update(extra)
    return d


class TestWhatifBacktestSheet(unittest.TestCase):
    """时序回测页签 Excel 呈现测试。"""

    def _write(self, whatif_data) -> "object":
        from openpyxl import Workbook

        from src.python.report.whatif_sheet import write_whatif_backtest_sheet

        wb = Workbook()
        ws = wb.active
        write_whatif_backtest_sheet(ws, whatif_data)
        return ws

    def _flat(self, ws) -> list[str]:
        return [str(c.value) if c.value is not None else "" for row in ws.iter_rows() for c in row]

    def test_backtest_sheet_full(self):
        """available=True → 生效日行 + 指标表（%单位 + 箭头）+ 序列表 + 说明。"""
        ws = self._write(_bt_data())
        flat = self._flat(ws)
        self.assertTrue(any("时序回测" in v for v in flat), "应含页标题")
        self.assertTrue(any("2026-07-01" in v for v in flat), "应含生效日")
        self.assertTrue(any("区间收益" in v for v in flat))
        self.assertTrue(any("夏普比率" in v for v in flat))
        self.assertTrue(any("24.0 ↑" in v for v in flat), "pct 变化列应带箭头")
        self.assertTrue(any("-0.4 ↓" in v for v in flat), "ratio 变化列应带箭头")
        self.assertTrue(any("2026-07-02" in v for v in flat), "应含净值序列日期")
        self.assertTrue(any("100.0" in v for v in flat), "应含归一化净值")
        self.assertTrue(any("口径" in v for v in flat), "应含口径说明")

    def test_backtest_sheet_missing_key_placeholder(self):
        """whatif_data 无 backtest 键 → 通用占位。"""
        ws = self._write(_whatif_data())
        flat = self._flat(ws)
        self.assertTrue(any("时序回测不可用" in v for v in flat), "应写占位")

    def test_backtest_sheet_none_placeholder(self):
        """whatif_data=None → 占位。"""
        ws = self._write(None)
        flat = self._flat(ws)
        self.assertTrue(any("时序回测不可用" in v for v in flat), "None 应写占位")

    def test_backtest_sheet_unavailable_reason_placeholder(self):
        """backtest available=False → reason 优先于通用文案。"""
        bt = {
            "available": False,
            "status": "unavailable",
            "reason": "生效日后数据不足",
            "effective_date": "2026-07-20",
        }
        ws = self._write(_whatif_data(backtest=bt))
        flat = self._flat(ws)
        self.assertTrue(any("生效日后数据不足" in v for v in flat), "应写回测 reason")
