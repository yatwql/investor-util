"""行情市值核算 — Excel 写入层单元测试。

测试目标：
  - _detail_to_row_values   — 行值转换
  - _num_formats            — 格式列表
  - _apply_profit_colors    — 盈亏着色
  - _apply_price_type_colors — 取价方式着色
  - _write_account_groupings — 账户分组写入
  - write_market_value_sheet — 页签写入（mock 内部函数）

通过直接导入 market_value_sheet 模块引用写入层函数。
"""

from __future__ import annotations

import unittest
from typing import Any
from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from src.python.models import Holding
from src.python.report import market_value_sheet as mvs
import pytest

pytestmark = [pytest.mark.unit, pytest.mark.unit_report]


# ═══════════════════════════════════════════════════════════
#  _detail_to_row_values
# ═══════════════════════════════════════════════════════════


class TestDetailToRowValues(unittest.TestCase):
    """测试 _detail_to_row_values 行值转换。"""

    def test_full_detail_row(self):
        """完整 DetailRow → 15 个字段的列表。"""
        d = mvs.DetailRow(
            account="证券账户", name="电池ETF", code="561910",
            price=10.5, nav_date="2026-06-26", yesterday_close=10.0,
            price_type="场内收盘价(T)", premium="--",
            shares=1000.0, market_value=10500.0, cost=1000.0,
            profit=9500.0, profit_rate=9.5, today_profit=500.0,
            source="腾讯财经", source_api="tencent",
        )
        vals = mvs._detail_to_row_values(d)
        self.assertEqual(len(vals), 15)
        self.assertEqual(vals[0], "证券账户")
        self.assertEqual(vals[1], "电池ETF")
        self.assertEqual(vals[2], "561910")
        self.assertEqual(vals[3], 10.5)
        self.assertEqual(vals[4], "2026-06-26")
        self.assertEqual(vals[5], 10.0)
        self.assertEqual(vals[6], "场内收盘价(T)")
        self.assertEqual(vals[7], "--")
        self.assertEqual(vals[8], 1000.0)
        self.assertEqual(vals[9], 10500.0)
        self.assertEqual(vals[10], 1000.0)
        self.assertEqual(vals[11], 9500.0)
        self.assertEqual(vals[12], 9.5)
        self.assertEqual(vals[13], 500.0)
        self.assertEqual(vals[14], "腾讯财经")

    def test_default_row(self):
        """默认 DetailRow → 空字符串/0 值列表。"""
        d = mvs.DetailRow()
        vals = mvs._detail_to_row_values(d)
        self.assertEqual(len(vals), 15)
        self.assertEqual(vals[0], "")
        self.assertEqual(vals[3], 0.0)
        self.assertEqual(vals[7], "--")
        self.assertIsNone(vals[12])


# ═══════════════════════════════════════════════════════════
#  _num_formats
# ═══════════════════════════════════════════════════════════


class TestNumFormats(unittest.TestCase):
    """测试 _num_formats 返回正确长度的格式列表。"""

    def test_length(self):
        """返回 15 个格式。"""
        fmts = mvs._num_formats()
        self.assertEqual(len(fmts), 15)

    def test_price_format(self):
        """第 4 个为价格格式。"""
        fmts = mvs._num_formats()
        self.assertEqual(fmts[3], '#,##0.0000')

    def test_money_format(self):
        """市值/成本/盈亏列金额格式。"""
        fmts = mvs._num_formats()
        self.assertEqual(fmts[9], '#,##0.00')
        self.assertEqual(fmts[10], '#,##0.00')
        self.assertEqual(fmts[11], '#,##0.00')

    def test_percent_format(self):
        """收益率列百分比格式。"""
        fmts = mvs._num_formats()
        self.assertEqual(fmts[12], '0.00%')

    def test_shares_format(self):
        """份额列格式。"""
        fmts = mvs._num_formats()
        self.assertEqual(fmts[8], '#,##0.00')


# ═══════════════════════════════════════════════════════════
#  _apply_profit_colors
# ═══════════════════════════════════════════════════════════


class TestApplyProfitColors(unittest.TestCase):
    """测试 _apply_profit_colors 盈亏列着色（mock ws / profit_font）。"""

    def _make_cell(self, value: Any = None) -> MagicMock:
        cell = MagicMock()
        cell.value = value
        return cell

    @patch("src.python.report.market_value_sheet.profit_font")
    def test_positive_profit(self, mock_pf):
        """正盈亏 → 调用 profit_font(正数)。"""
        mock_pf.side_effect = lambda v: f"font_for_{v}"
        ws = MagicMock()
        ws.cell.side_effect = lambda row, column: self._make_cell(500.0)
        mvs._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_any_call(500.0)
        self.assertEqual(mock_pf.call_count, 3)

    @patch("src.python.report.market_value_sheet.profit_font")
    def test_negative_profit(self, mock_pf):
        """负盈亏 → 调用 profit_font(负数)。"""
        mock_pf.side_effect = lambda v: f"font_for_{v}"
        ws = MagicMock()
        ws.cell.side_effect = lambda row, column: self._make_cell(-300.0)
        mvs._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_any_call(-300.0)

    @patch("src.python.report.market_value_sheet.profit_font")
    def test_zero_profit(self, mock_pf):
        """零盈亏 → 调用 profit_font(0)。"""
        mock_pf.side_effect = lambda v: f"font_for_{v}"
        ws = MagicMock()
        ws.cell.side_effect = lambda row, column: self._make_cell(0.0)
        mvs._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_any_call(0.0)

    @patch("src.python.report.market_value_sheet.profit_font")
    def test_non_numeric_skipped(self, mock_pf):
        """非数字值（字符串）→ 不设置字体。"""
        ws = MagicMock()
        values = {"12": "亏损", "14": "盈利"}
        def cell_side_effect(row, column):
            cell = MagicMock()
            cell.value = values.get(str(column))
            return cell
        ws.cell.side_effect = cell_side_effect
        mvs._apply_profit_colors(ws, 3, 4, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_not_called()

    @patch("src.python.report.market_value_sheet.profit_font")
    def test_none_value_skipped(self, mock_pf):
        """None 值 → 不设置字体。"""
        ws = MagicMock()
        ws.cell.return_value = self._make_cell(None)
        mvs._apply_profit_colors(ws, 3, 4, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_not_called()

    @patch("src.python.report.market_value_sheet.profit_font")
    def test_rate_col_float(self, mock_pf):
        """收益率列为 float → 调用 profit_font。"""
        mock_pf.side_effect = lambda v: f"font_for_{v}"
        ws = MagicMock()
        ws.cell.side_effect = lambda row, column: self._make_cell(0.05)
        mvs._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_called_with(0.05)

    @patch("src.python.report.market_value_sheet.profit_font")
    def test_rate_col_not_float(self, mock_pf):
        """收益率列非 float（如字符串含 %）→ 不设置字体。"""
        ws = MagicMock()
        ws.cell.side_effect = lambda row, column: self._make_cell("5.00%")
        mvs._apply_profit_colors(ws, 3, 4, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_not_called()

    @patch("src.python.report.market_value_sheet.profit_font")
    def test_multiple_rows(self, mock_pf):
        """多行数据 → 每行都着色。"""
        mock_pf.side_effect = lambda v: f"font_for_{v}"
        ws = MagicMock()
        row_values = {3: 100.0, 4: -50.0, 5: 200.0}
        def cell_side_effect(row, column):
            cell = MagicMock()
            cell.value = row_values.get(row, 0.0)
            return cell
        ws.cell.side_effect = cell_side_effect
        mvs._apply_profit_colors(ws, 3, 5, profit_col=12, rate_col=13, today_col=14)
        self.assertEqual(mock_pf.call_count, 9)

    def test_cell_font_assigned(self):
        """确保 cell.font 被赋值。"""
        ws = MagicMock()
        cell = MagicMock()
        cell.value = 100.0
        ws.cell.return_value = cell
        with patch("src.python.report.market_value_sheet.profit_font") as mock_pf:
            mock_pf.return_value = "red_font"
            mvs._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)
        self.assertEqual(cell.font, "red_font")

    def test_int_rate_value(self):
        """收益率列为 int → 不调用 profit_font（仅 float 触发）。"""
        with patch("src.python.report.market_value_sheet.profit_font") as mock_pf:
            ws = MagicMock()
            # 仅 rate 列设为 int；profit/today 列 None（非数字）
            values = {13: 5}
            ws.cell.side_effect = lambda row, column: self._make_cell(values.get(column))
            mvs._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)
            mock_pf.assert_not_called()

    def test_empty_range_no_crash(self):
        """空范围（start > end）→ 不报错。"""
        with patch("src.python.report.market_value_sheet.profit_font"):
            ws = MagicMock()
            mvs._apply_profit_colors(ws, 100, 50, profit_col=12, rate_col=13, today_col=14)


# ═══════════════════════════════════════════════════════════
#  _apply_price_type_colors
# ═══════════════════════════════════════════════════════════


class TestApplyPriceTypeColors(unittest.TestCase):
    """测试 _apply_price_type_colors 取价方式列着色（使用真实 openpyxl Worksheet）。"""

    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.test_cases = [
            ("电池ETF", "场内收盘价(T)", True),
            ("长江电力", "场内收盘价(T-1)", False),
            ("中欧医疗", "官方净值(T)", True),
            ("某混合基金", "官方净值(T-1)", False),
            ("标普500ETF(QDII)", "官方净值(T-1)", True),
            ("恒生ETF(QDII)", "场内收盘价(T-1)", False),
            ("宁德时代", "场内实时价", False),
            ("海外收益(QDII)", "官方净值(T-2)", False),
            ("--", "--", False),
        ]
        for i, (name, price_type, expected_blue) in enumerate(self.test_cases):
            row = i + 2
            self.ws.cell(row=row, column=2, value=name)
            self.ws.cell(row=row, column=7, value=price_type)

    def _assert_blue(self, row: int, msg: str = ""):
        cell = self.ws.cell(row=row, column=7)
        self.assertIsNotNone(cell.font.color, f"Row {row} font.color is None")
        self.assertEqual(str(cell.font.color.rgb), "000066CC", msg)

    def _assert_not_blue(self, row: int, msg: str = ""):
        cell = self.ws.cell(row=row, column=7)
        if cell.font.color and cell.font.color.rgb:
            self.assertNotEqual(str(cell.font.color.rgb), "000066CC", msg)

    def test_scenario(self):
        """所有场景批量验证。"""
        mvs._apply_price_type_colors(self.ws, 2, 2 + len(self.test_cases) - 1)
        errors = []
        for i, (name, price_type, expected_blue) in enumerate(self.test_cases):
            row = i + 2
            try:
                if expected_blue:
                    self._assert_blue(row, f"Row {row}: {name} / {price_type} should be blue")
                else:
                    self._assert_not_blue(row, f"Row {row}: {name} / {price_type} should NOT be blue")
            except AssertionError as e:
                errors.append(str(e))
        if errors:
            self.fail("\n".join(errors))

    def test_empty_range_no_error(self):
        """空范围（start > end）→ 不报错。"""
        mvs._apply_price_type_colors(self.ws, 100, 50)

    def test_single_row(self):
        """单行范围。"""
        self.ws.cell(row=3, column=2, value="测试")
        self.ws.cell(row=3, column=7, value="场内收盘价(T)")
        mvs._apply_price_type_colors(self.ws, 3, 3)
        self._assert_blue(3)

    def test_none_price_type_col(self):
        """取价方式列为 None → 不报错。"""
        row = 2 + len(self.test_cases)
        self.ws.cell(row=row, column=2, value="测试")
        self.ws.cell(row=row, column=7, value=None)
        mvs._apply_price_type_colors(self.ws, row, row)
        self.assertIsNone(self.ws.cell(row=row, column=7).value)


# ═══════════════════════════════════════════════════════════
#  _write_account_groupings
# ═══════════════════════════════════════════════════════════


class TestWriteAccountGroupings(unittest.TestCase):
    """测试 _write_account_groupings 账户分组写入逻辑。"""

    def setUp(self):
        self.detail_a = mvs.DetailRow(
            account="证券账户", name="电池ETF", code="561910",
            price=10.0, nav_date="2026-06-26", yesterday_close=9.5,
            price_type="场内收盘价(T)", premium="--",
            shares=100.0, market_value=1000.0, cost=100.0,
            profit=900.0, profit_rate=9.0, today_profit=50.0,
            source="腾讯财经", source_api="tencent",
        )
        self.detail_b = mvs.DetailRow(
            account="支付宝", name="中欧医疗健康混合", code="003095",
            price=1.5, nav_date="2026-06-25", yesterday_close=1.48,
            price_type="官方净值(T-1)", premium="--",
            shares=200.0, market_value=300.0, cost=400.0,
            profit=-100.0, profit_rate=-0.25, today_profit=4.0,
            source="东方财富", source_api="eastmoney",
        )

    @patch("src.python.report.market_value_sheet.write_data_row")
    @patch("src.python.report.market_value_sheet.write_subtotal_row")
    @patch("src.python.report.market_value_sheet._detail_to_row_values")
    @patch("src.python.report.market_value_sheet._num_formats")
    def test_single_account(self, mock_fmts, mock_to_row, mock_sub, mock_data):
        """单账户 → 1 个小计行。"""
        mock_fmts.return_value = [""] * 15
        mock_to_row.side_effect = lambda d: [d.account, d.name]
        ws = MagicMock()
        _, _, _, _, final_row = mvs._write_account_groupings(ws, [self.detail_a], 3)
        mock_sub.assert_called_once()
        self.assertGreater(final_row, 3)

    @patch("src.python.report.market_value_sheet.write_data_row")
    @patch("src.python.report.market_value_sheet.write_subtotal_row")
    @patch("src.python.report.market_value_sheet._detail_to_row_values")
    @patch("src.python.report.market_value_sheet._num_formats")
    def test_multiple_accounts(self, mock_fmts, mock_to_row, mock_sub, mock_data):
        """多账户 → 每个账户 1 个小计行。"""
        mock_fmts.return_value = [""] * 15
        mock_to_row.side_effect = lambda d: [d.account, d.name]
        ws = MagicMock()
        _, _, _, _, final_row = mvs._write_account_groupings(
            ws, [self.detail_a, self.detail_b], 3)
        self.assertEqual(mock_sub.call_count, 2)

    @patch("src.python.report.market_value_sheet.write_data_row")
    @patch("src.python.report.market_value_sheet.write_subtotal_row")
    @patch("src.python.report.market_value_sheet._detail_to_row_values")
    @patch("src.python.report.market_value_sheet._num_formats")
    def test_empty_details(self, mock_fmts, mock_to_row, mock_sub, mock_data):
        """空明细 → 无小计行。"""
        mock_fmts.return_value = [""] * 15
        ws = MagicMock()
        _, _, _, _, final_row = mvs._write_account_groupings(ws, [], 3)
        mock_sub.assert_not_called()
        self.assertEqual(final_row, 3)

    @patch("src.python.report.market_value_sheet.write_data_row")
    @patch("src.python.report.market_value_sheet.write_subtotal_row")
    @patch("src.python.report.market_value_sheet._detail_to_row_values")
    @patch("src.python.report.market_value_sheet._num_formats")
    def test_special_char_account(self, mock_fmts, mock_to_row, mock_sub, mock_data):
        """特殊字符账户名 → 不崩溃。"""
        mock_fmts.return_value = [""] * 15
        mock_to_row.side_effect = lambda d: [d.account, d.name]
        detail_special = mvs.DetailRow(
            account="测💹试/账户（定投）", name="基金A", code="000001",
            shares=100.0, market_value=1000.0, cost=500.0,
            profit=500.0, profit_rate=1.0, today_profit=10.0,
        )
        ws = MagicMock()
        try:
            mvs._write_account_groupings(ws, [detail_special], 3)
        except Exception as e:
            self.fail(f"特殊字符账户名崩溃: {e}")

    @patch("src.python.report.market_value_sheet.write_data_row")
    @patch("src.python.report.market_value_sheet.write_subtotal_row")
    @patch("src.python.report.market_value_sheet._detail_to_row_values")
    @patch("src.python.report.market_value_sheet._num_formats")
    def test_acc_cost_zero_rate(self, mock_fmts, mock_to_row, mock_sub, mock_data):
        """成本为 0 时收益率为 0.0 而非除零异常。"""
        mock_fmts.return_value = [""] * 15
        mock_to_row.side_effect = lambda d: [d.account, d.name]
        detail_cost_zero = mvs.DetailRow(
            account="证券", name="新股", code="688001",
            market_value=1000.0, cost=0.0, profit=1000.0,
            shares=100.0, profit_rate=None, today_profit=0.0,
        )
        ws = MagicMock()
        mvs._write_account_groupings(ws, [detail_cost_zero], 3)
        sub_call = mock_sub.call_args
        if sub_call:
            subtotal_vals = sub_call[0][3]
            self.assertEqual(subtotal_vals[11], 0.0)


# ═══════════════════════════════════════════════════════════
#  write_market_value_sheet
# ═══════════════════════════════════════════════════════════


class TestWriteMarketValueSheet(unittest.TestCase):
    """测试 write_market_value_sheet 页签写入（mock 内部函数和 Excel 写入）。"""

    def setUp(self):
        self.holdings = [
            Holding("证券账户", "电池ETF", "561910", 100.0, 1.0),
            Holding("支付宝", "中欧医疗健康混合", "003095", 200.0, 2.0),
        ]
        self.details = [
            mvs.DetailRow(
                account="证券账户", name="电池ETF", code="561910",
                price=10.0, nav_date="2026-06-26", yesterday_close=9.5,
                price_type="场内收盘价(T)", premium="--",
                shares=100.0, market_value=1000.0, cost=100.0,
                profit=900.0, profit_rate=9.0, today_profit=50.0,
                source="腾讯财经", source_api="tencent",
            ),
            mvs.DetailRow(
                account="支付宝", name="中欧医疗健康混合", code="003095",
                price=1.5, nav_date="2026-06-25", yesterday_close=1.48,
                price_type="官方净值(T-1)", premium="--",
                shares=200.0, market_value=300.0, cost=400.0,
                profit=-100.0, profit_rate=-0.25, today_profit=4.0,
                source="东方财富", source_api="eastmoney",
            ),
        ]

    @patch("src.python.report.market_value_sheet.write_total_row")
    @patch("src.python.report.market_value_sheet.write_subtotal_row")
    @patch("src.python.report.market_value_sheet.write_data_row")
    @patch("src.python.report.market_value_sheet.write_header_row")
    @patch("src.python.report.market_value_sheet.write_title_row")
    @patch("src.python.report.market_value_sheet._apply_price_type_colors")
    @patch("src.python.report.market_value_sheet._apply_profit_colors")
    @patch("src.python.report.market_value_sheet.freeze_header")
    @patch("src.python.report.market_value_sheet.auto_width")
    @patch("src.python.report.market_value_sheet._detail_to_row_values")
    @patch("src.python.report.market_value_sheet._num_formats")
    def test_basic_write(self, mock_fmts, mock_to_row,
                         mock_aw, mock_freeze, mock_color,
                         mock_pt_color, mock_tl, mock_hdr, mock_data, mock_sub, mock_total):
        """正常写入：验证汇总值正确，内部函数被调用。"""
        mock_fmts.return_value = [""] * 15
        mock_to_row.side_effect = lambda d: [
            d.account, d.name, d.code, d.price,
            d.nav_date, d.yesterday_close, d.price_type, d.premium,
            d.shares, d.market_value, d.cost, d.profit,
            d.profit_rate, d.today_profit, d.source,
        ]
        mock_tl.return_value = 2
        mock_hdr.return_value = 3
        ws = MagicMock()
        ws.title = "fixture_title"
        result = mvs.write_market_value_sheet(ws, [], "2026-06-26", details=self.details)
        grand_mv, grand_cost, grand_profit, grand_today, details = result
        self.assertAlmostEqual(grand_mv, 1300.0)
        self.assertAlmostEqual(grand_cost, 500.0)
        self.assertAlmostEqual(grand_profit, 800.0)
        self.assertAlmostEqual(grand_today, 54.0)
        self.assertEqual(len(details), 2)
        mock_tl.assert_called_once()
        mock_hdr.assert_called_once()
        self.assertEqual(mock_sub.call_count, 2)
        mock_total.assert_called_once()
        mock_color.assert_called_once()
        mock_pt_color.assert_called_once()
        mock_freeze.assert_called_once_with(ws, 2)
        mock_aw.assert_called_once_with(ws)
        self.assertEqual(ws.title, "fixture_title")

    @patch("src.python.report.market_value_sheet.write_total_row")
    @patch("src.python.report.market_value_sheet.write_subtotal_row")
    @patch("src.python.report.market_value_sheet.write_data_row")
    @patch("src.python.report.market_value_sheet.write_header_row")
    @patch("src.python.report.market_value_sheet.write_title_row")
    @patch("src.python.report.market_value_sheet._apply_price_type_colors")
    @patch("src.python.report.market_value_sheet._apply_profit_colors")
    @patch("src.python.report.market_value_sheet.freeze_header")
    @patch("src.python.report.market_value_sheet.auto_width")
    @patch("src.python.report.market_value_sheet._detail_to_row_values")
    @patch("src.python.report.market_value_sheet._num_formats")
    def test_empty_holdings(self, mock_fmts, mock_to_row,
                            mock_aw, mock_freeze, mock_color,
                            mock_pt_color, mock_tl, mock_hdr, mock_data, mock_sub, mock_total):
        """空持仓 → 总市值为 0，无小计行。"""
        mock_fmts.return_value = [""] * 15
        mock_tl.return_value = 2
        mock_hdr.return_value = 3
        ws = MagicMock()
        result = mvs.write_market_value_sheet(ws, [], "2026-06-26", details=[])
        grand_mv, grand_cost, grand_profit, grand_today, details = result
        self.assertAlmostEqual(grand_mv, 0.0)
        self.assertAlmostEqual(grand_cost, 0.0)
        self.assertAlmostEqual(grand_profit, 0.0)
        self.assertAlmostEqual(grand_today, 0.0)
        self.assertEqual(details, [])
        mock_sub.assert_not_called()
        mock_total.assert_called_once()

    @patch("src.python.report.market_value_sheet.write_total_row")
    @patch("src.python.report.market_value_sheet.write_subtotal_row")
    @patch("src.python.report.market_value_sheet.write_data_row")
    @patch("src.python.report.market_value_sheet.write_header_row")
    @patch("src.python.report.market_value_sheet.write_title_row")
    @patch("src.python.report.market_value_sheet._apply_price_type_colors")
    @patch("src.python.report.market_value_sheet._apply_profit_colors")
    @patch("src.python.report.market_value_sheet.freeze_header")
    @patch("src.python.report.market_value_sheet.auto_width")
    @patch("src.python.report.market_value_sheet._detail_to_row_values")
    @patch("src.python.report.market_value_sheet._num_formats")
    def test_subtotal_per_account(self, mock_fmts, mock_to_row,
                                   mock_aw, mock_freeze, mock_color,
                                   mock_pt_color, mock_tl, mock_hdr, mock_data, mock_sub, mock_total):
        """多个账户 → 每个账户写入小计。"""
        detail_a = self.details[0]
        detail_b = self.details[1]
        detail_c = mvs.DetailRow(
            account="证券账户", name="长江电力", code="600900",
            price=25.0, nav_date="2026-06-26", yesterday_close=24.5,
            price_type="场内收盘价(T)", premium="--",
            shares=100.0, market_value=2500.0, cost=2000.0,
            profit=500.0, profit_rate=0.25, today_profit=50.0,
            source="腾讯财经", source_api="tencent",
        )
        mock_fmts.return_value = [""] * 15
        mock_to_row.side_effect = lambda d: [
            d.account, d.name, d.code, d.price,
            d.nav_date, d.yesterday_close, d.price_type, d.premium,
            d.shares, d.market_value, d.cost, d.profit,
            d.profit_rate, d.today_profit, d.source,
        ]
        mock_tl.return_value = 2
        mock_hdr.return_value = 3
        ws = MagicMock()
        result = mvs.write_market_value_sheet(ws, [], "2026-06-26",
                                              details=[detail_a, detail_c, detail_b])
        self.assertEqual(mock_sub.call_count, 2)
        grand_mv = result[0]
        self.assertAlmostEqual(grand_mv, 3800.0)

    @patch("src.python.report.market_value_sheet.write_total_row")
    @patch("src.python.report.market_value_sheet.write_subtotal_row")
    @patch("src.python.report.market_value_sheet.write_data_row")
    @patch("src.python.report.market_value_sheet.write_header_row")
    @patch("src.python.report.market_value_sheet.write_title_row")
    @patch("src.python.report.market_value_sheet._apply_price_type_colors")
    @patch("src.python.report.market_value_sheet._apply_profit_colors")
    @patch("src.python.report.market_value_sheet.freeze_header")
    @patch("src.python.report.market_value_sheet.auto_width")
    @patch("src.python.report.market_value_sheet._detail_to_row_values")
    @patch("src.python.report.market_value_sheet._num_formats")
    def test_all_zero_price_show_warning(self, mock_fmts, mock_to_row,
                                          mock_aw, mock_freeze, mock_color,
                                          mock_pt_color, mock_tl, mock_hdr, mock_data, mock_sub, mock_total):
        """全零行情 → 写入红色警告行 + 合并单元格。"""
        mock_fmts.return_value = [""] * 15
        mock_to_row.side_effect = lambda d: [""] * 15
        mock_tl.return_value = 2
        mock_hdr.return_value = 3
        ws = MagicMock()
        ws.cell.return_value = MagicMock()
        result = mvs.write_market_value_sheet(ws, [], "2026-06-26", details=[
            mvs.DetailRow(account="证券", name="电池ETF", code="561910",
                         price=0.0, shares=100.0, market_value=0.0, cost=100.0),
        ])
        ws.merge_cells.assert_called_once()
        call_kwargs = ws.cell.call_args
        if call_kwargs:
            written_cell = ws.cell.return_value
            self.assertIsNotNone(written_cell.font)
