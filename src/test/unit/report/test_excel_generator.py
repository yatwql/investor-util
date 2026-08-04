"""Excel 报告生成单元测试 — 页签写入隔离、模块缺失降级、数据路径、LLM 用量状态。

测试目标：
  - 基本路径：外部传入明细/指数，跳过获取
  - 新闻/LLM 包含路径：页签创建 + 内容写入
  - 模块缺失降级：ImportError → add_error + 其他页签继续
  - 异常隔离：单页签 throw → add_error + 不影响其他
  - ProgressReporter 接口：info/ok/add_error 回调
  - _build_llm_usage_sheet 状态判定：缓存/成功/禁用/失败

运行：
  cd D:/codebase/zoo/investor-util
  python -m pytest src/test/test_excel_generator.py -v
"""

from __future__ import annotations

import unittest
from unittest.mock import MagicMock, patch
from collections import namedtuple

from src.python.report.progress import SilentProgressReporter
import pytest

pytestmark = [pytest.mark.unit, pytest.mark.unit_report]


# ═══════════════════════════════════════════════════════════
#  辅助函数
# ═══════════════════════════════════════════════════════════


DetailRow = namedtuple(
    "DetailRow",
    [
        "market_value",
        "cost",
        "profit",
        "today_profit",
        "code",
        "name",
        "change_pct",
        "price_type",
        "nav_date",
        "profit_rate",
        "price",
    ],
)


def _make_detail(
    mv: float = 1000,
    cost: float = 800,
    profit: float = 200,
    today_profit: float = 50,
    code: str = "600900",
    name: str = "长江电力",
) -> DetailRow:
    return DetailRow(
        market_value=mv,
        cost=cost,
        profit=profit,
        today_profit=today_profit,
        code=code,
        name=name,
        change_pct=1.5,
        price_type="tencent",
        nav_date="",
        profit_rate=25.0,
        price=28.5,
    )


# ═══════════════════════════════════════════════════════════
#  辅助：模拟各模块函数，避免真实导入依赖
# ═══════════════════════════════════════════════════════════


class _SheetMocks:
    """持有所有 mock 引用，便于断言。"""

    def __init__(self) -> None:
        self.write_summary = MagicMock()
        self.write_market_value = MagicMock(
            return_value=(
                10000.0,
                8000.0,
                2000.0,
                500.0,
                [_make_detail()],
            )
        )
        self.classify_holdings = MagicMock(return_value={})
        self.get_last_trading_day = MagicMock(return_value="2026-07-01")
        self.price_update_status = MagicMock(return_value=(1, 0, True))
        self.write_category = MagicMock()
        self.compute_penetration = MagicMock(return_value={})
        self.write_penetration = MagicMock()
        self.write_fund_performance = MagicMock()

    def start(self) -> list:
        """启动所有 patch，返回 patcher 列表以便 stop。"""
        self.patchers = [
            patch("src.python.report.summary.write_summary_sheet", self.write_summary),
            patch("src.python.report.excel_writer.create_workbook"),
            patch("src.python.report.excel_writer.save_workbook", return_value="reports/test.xlsx"),
            patch("src.python.report.market_value_sheet.write_market_value_sheet", self.write_market_value),
            patch("src.python.report.market_value.classify_holdings", self.classify_holdings),
            patch("src.python.report.market_value.get_last_trading_day", self.get_last_trading_day),
            patch("src.python.report.market_value.price_update_status", self.price_update_status),
            patch("src.python.report.category.write_category_sheet", self.write_category),
            patch("src.python.report.penetration.compute_penetration_top10", self.compute_penetration),
            patch("src.python.report.penetration_sheet.write_penetration_sheet", self.write_penetration),
            patch("src.python.report.fund_performance.write_fund_performance_sheet", self.write_fund_performance),
            patch("src.python.report.excel_content_sheets.get_report_sheet_name", side_effect=lambda k: k),
            patch("src.python.report.excel_news_warning.get_report_sheet_name", side_effect=lambda k: k, create=True),
            patch("src.python.report.excel_news_warning.get_llm_module_name", MagicMock()),
        ]
        for p in self.patchers:
            p.start()
        # 保存 create_workbook mock 引用
        self.mock_create_wb = self.patchers[1].new
        self.mock_wb = MagicMock()
        self.mock_ws = MagicMock()
        self.mock_wb.active = self.mock_ws
        self.mock_wb.create_sheet.return_value = MagicMock()
        self.mock_create_wb.return_value = self.mock_wb
        return self.patchers

    def stop(self) -> None:
        for p in self.patchers:
            p.stop()


class TestGenerateExcelReport(unittest.TestCase):
    """generate_excel_report 主流程。"""

    def setUp(self) -> None:
        self.progress = SilentProgressReporter()
        self.holdings = []
        self.sheets = _SheetMocks()
        self.sheets.start()
        # mock 指数获取
        self.idx_patcher = patch("src.python.fetcher.index.fetch_indices", return_value={})
        self.mock_fetch_idx = self.idx_patcher.start()
        self.us_idx_patcher = patch("src.python.fetcher.index.fetch_us_indices", return_value={})
        self.mock_fetch_us_idx = self.us_idx_patcher.start()

    def tearDown(self) -> None:
        self.sheets.stop()
        self.idx_patcher.stop()
        self.us_idx_patcher.stop()

    # ── 基本路径 ──

    @pytest.mark.smoke
    def test_basic_generation(self) -> None:
        """基本路径：所有模块正常，外部传入明细+指数。"""
        from src.python.report.excel_generator import generate_excel_report

        details = [_make_detail()]
        a_idx = {"sh000001": {"name": "上证指数", "price": 3120, "change_pct": 0.5}}

        generate_excel_report(
            self.holdings,
            details=details,
            a_indices=a_idx,
            us_indices={},
            progress=self.progress,
        )

        self.assertEqual(len(self.progress.get_errors()), 0)

    def test_generation_without_external_data(self) -> None:
        """无外部传入数据 → 内部获取指数。"""
        from src.python.report.excel_generator import generate_excel_report

        generate_excel_report(self.holdings, progress=self.progress)

        self.assertEqual(len(self.progress.get_errors()), 0)
        self.mock_fetch_idx.assert_called_once()
        self.mock_fetch_us_idx.assert_called_once()

    # ── 新闻路径 ──

    def test_with_news(self) -> None:
        """include_news=True → 新闻页签创建。"""
        from src.python.report.excel_generator import generate_excel_report

        with patch("src.python.report.news_correlation.write_news_sheet") as mock_news:
            with patch("src.python.report.news_correlation.build_news_data", return_value=([], {})):
                generate_excel_report(
                    self.holdings,
                    include_news=True,
                    progress=self.progress,
                )

        self.assertEqual(len(self.progress.get_errors()), 0)

    def test_with_news_external_data(self) -> None:
        """include_news + 外部新闻数据 → 复用。"""
        from src.python.report.excel_generator import generate_excel_report

        news_data = [{"title": "新闻1", "intro": "简介", "matched_keywords": ["test"]}]
        news_llm_meta = {"llm_enabled": False}

        with patch("src.python.report.news_correlation.write_news_sheet") as mock_news:
            generate_excel_report(
                self.holdings,
                include_news=True,
                news_data=news_data,
                news_llm_meta=news_llm_meta,
                progress=self.progress,
            )

        self.assertEqual(len(self.progress.get_errors()), 0)

    # ── LLM 路径 ──

    def test_with_llm(self) -> None:
        """enable_llm=True → LLM 内容写入。"""
        from src.python.report.excel_generator import generate_excel_report

        llm_content = ("<p>宏</p>", "<p>策略</p>", "<p>体检</p>", "<p>穿透</p>")

        with patch("src.python.report.llm_content.write_llm_sheets") as mock_llm:
            mock_llm.return_value = ("<p>宏</p>", "<p>策略</p>", "<p>体检</p>", "<p>穿透</p>")
            with patch("src.python.llm.session.get_session_usage", return_value={"call_count": 0, "per_module": {}}):
                generate_excel_report(
                    self.holdings,
                    include_llm=True,
                    llm_content=llm_content,
                    progress=self.progress,
                )

        mock_llm.assert_called_once()
        self.assertLessEqual(len(self.progress.get_errors()), 1)

    def test_with_llm_and_session_usage(self) -> None:
        """include_llm + 有会话统计 → 写入用量页签 + 汇总页追加。"""
        from src.python.report.excel_generator import generate_excel_report

        llm_content = ("<p>宏</p>", "<p>策略</p>", "<p>体检</p>", "<p>穿透</p>")
        session_usage = {
            "call_count": 2,
            "input_tokens": 1000,
            "output_tokens": 500,
            "cache_hit_tokens": 200,
            "models": ["deepseek-v4-flash"],
            "per_module": {
                "global_macro": {
                    "input_tokens": 500,
                    "output_tokens": 200,
                    "model": "deepseek-v4-flash",
                    "cached": False,
                    "thinking": False,
                    "endpoint": "",
                    "cache_hit_tokens": 0,
                    "cost": 0.0,
                },
                "expert_review": {
                    "input_tokens": 500,
                    "output_tokens": 300,
                    "model": "deepseek-v4-flash",
                    "cached": False,
                    "thinking": True,
                    "endpoint": "",
                    "cache_hit_tokens": 0,
                    "cost": 0.0,
                },
            },
        }

        with patch("src.python.report.llm_content.write_llm_sheets") as mock_llm:
            mock_llm.return_value = ("<p>宏</p>", "<p>策略</p>", "<p>体检</p>", "<p>穿透</p>")
            with patch("src.python.llm.session.get_session_usage", return_value=session_usage):
                with patch("src.python.llm.session.format_session_usage") as mock_fmt:
                    mock_fmt.return_value = {
                        "has_usage": True,
                        "call_count": 2,
                        "total_tokens": 1500,
                        "cost_display": "¥0.002",
                        "per_module": session_usage["per_module"],
                    }
                    with patch("src.python.report.summary.write_llm_usage_sheet"):
                        generate_excel_report(
                            self.holdings,
                            include_llm=True,
                            llm_content=llm_content,
                            progress=self.progress,
                        )

        mock_llm.assert_called_once()

    # ── 模块缺失降级 ──

    @pytest.mark.smoke
    def test_summary_module_missing(self) -> None:
        """汇总模块缺失 → add_error + 其他模块继续。"""
        from src.python.report.excel_generator import generate_excel_report

        with patch("src.python.report.summary.write_summary_sheet", None):
            generate_excel_report(
                self.holdings,
                progress=self.progress,
            )

        errors = self.progress.get_errors()
        self.assertTrue(
            any("summary" in e.lower() or "汇总" in e for e in errors), f"预期 summary 错误，得到: {errors}"
        )

    def test_market_value_module_missing(self) -> None:
        """行情市值模块缺失 → add_error + 后续模块继续。"""
        from src.python.report.excel_generator import generate_excel_report

        with patch("src.python.report.market_value_sheet.write_market_value_sheet", None):
            generate_excel_report(
                self.holdings,
                progress=self.progress,
            )

        errors = self.progress.get_errors()
        self.assertTrue(
            any("market_value_sheet" in e.lower() or "行情市值" in e for e in errors),
            f"预期 market_value 错误，得到: {errors}",
        )

    # ── 页签写入异常隔离 ──

    @pytest.mark.smoke
    def test_sheet_exception_isolation(self) -> None:
        """某个页签抛出异常 → add_error + 不影响其他页签写入。"""
        from src.python.report.excel_generator import generate_excel_report

        broken_sheet = MagicMock(side_effect=ValueError("写入失败"))
        # 替换 summary 的 mock 为会抛异常的版本
        with patch("src.python.report.summary.write_summary_sheet", broken_sheet):
            generate_excel_report(
                self.holdings,
                progress=self.progress,
            )

        errors = self.progress.get_errors()
        self.assertTrue(any("生成失败" in e for e in errors), f"预期 sheet 写入错误记录，得到: {errors}")
        # 确认不暴露原始异常堆栈
        for e in errors:
            self.assertNotIn("ValueError", e, f"错误信息不应包含原始异常类型: {e}")

    def test_sheet_exception_others_still_called(self):
        """某页签失败 → 其他页签仍被调用（业务语义验证）。"""
        from src.python.report.excel_generator import generate_excel_report

        mocks = _SheetMocks()
        patchers = mocks.start()
        try:
            # 让 summary 抛出异常
            mocks.write_summary.side_effect = ValueError("写入失败")
            # 同时让穿透模块成功返回有效数据
            mocks.compute_penetration.return_value = {
                "top10": [
                    {"rank": 1, "name": "茅台", "mv": 10000.0, "ratio_pct": 50.0, "sources": [], "codes": ["600519"]}
                ],
                "summary": {
                    "total_mv": 20000.0,
                    "total_funds": 0,
                    "total_stocks": 1,
                    "fund_breakdown": "",
                    "merged_count": 1,
                    "top10_coverage_pct": "50.0%",
                    "unknown_mv": 0,
                    "failed_funds": 0,
                },
            }
            details = [_make_detail()]
            generate_excel_report(
                self.holdings,
                details=details,
                a_indices={},
                us_indices={},
                progress=self.progress,
            )
            errors = self.progress.get_errors()
            self.assertTrue(any("生成失败" in e for e in errors), f"预期 summary 写入错误，得到: {errors}")
            # 业务语义：穿透模块仍应被调用
            mocks.compute_penetration.assert_called_once()
            mocks.write_penetration.assert_called_once()
            # 市值模块仍应被调用
            mocks.write_market_value.assert_called_once()
        finally:
            for p in patchers:
                p.stop()

    # ── ProgressReporter 默认值 ──

    @pytest.mark.smoke
    def test_default_progress_reporter(self) -> None:
        """progress=None → 使用 SilentProgressReporter（不抛异常）。"""
        from src.python.report.excel_generator import generate_excel_report

        try:
            generate_excel_report(self.holdings)
        except Exception as e:
            self.fail(f"progress=None 时抛出异常: {e}")

    def test_timer_records(self) -> None:
        """模块耗时应有记录（_Timer 上下文管理器）。"""
        from src.python.report.excel_generator import generate_excel_report
        from src.python.report.progress import _timing_records

        _timing_records.clear()
        details = [_make_detail()]

        generate_excel_report(
            self.holdings,
            details=details,
            a_indices={},
            us_indices={},
            progress=self.progress,
        )

        self.assertGreater(len(_timing_records), 0)


# ═══════════════════════════════════════════════════════════
#  ProgressReporter 单元测试（补充 test_progress.py 未覆盖）
# ═══════════════════════════════════════════════════════════


class TestProgressReporterCallSheet(unittest.TestCase):
    """ProgressReporter.call_sheet 的完整行为。"""

    def setUp(self) -> None:
        self.prog = SilentProgressReporter()

    def test_call_sheet_success(self) -> None:
        """fn 正常 → 返回 True。"""
        fn = MagicMock(return_value=42)
        result = self.prog.call_sheet("测试", fn, 1, 2, key="val")
        self.assertTrue(result)
        fn.assert_called_once_with(1, 2, key="val")

    def test_call_sheet_fn_none(self) -> None:
        """fn=None → 返回 False + add_error。"""
        result = self.prog.call_sheet("缺失模块", None)
        self.assertFalse(result)
        errors = self.prog.get_errors()
        self.assertTrue(any("缺失" in e for e in errors))

    def test_call_sheet_exception(self) -> None:
        """fn 抛出异常 → 返回 False + add_error（友好提示，不暴露堆栈）。"""

        def _broken(*a, **kw):
            raise RuntimeError("写入失败")

        result = self.prog.call_sheet("损坏模块", _broken)
        self.assertFalse(result)
        errors = self.prog.get_errors()
        self.assertTrue(any("生成失败" in e for e in errors), f"预期友好错误提示，得到: {errors}")
        for e in errors:
            self.assertNotIn("RuntimeError", e, f"错误信息不应包含原始异常类型: {e}")


# ═══════════════════════════════════════════════════════════
#  _build_llm_usage_sheet — LLM 用量页签状态判定
# ═══════════════════════════════════════════════════════════


class TestBuildLlmUsageSheet(unittest.TestCase):
    """测试 _build_llm_usage_sheet 中缓存/成功/禁用/失败的状态判定。"""

    def setUp(self):
        self.wb = MagicMock()
        self.prog = SilentProgressReporter()
        self._name_map = {
            "global_macro": "全球政经局势",
            "expert_review": "智囊团深度复盘",
            "health_check": "持仓体检报告",
            "penetration_deep": "穿透深度分析",
        }

    def _run(self, raw_session: dict, formatted: dict, module_failure: dict | None = None) -> MagicMock:
        """执行 build_llm_usage_sheet 并返回 write_llm_usage_sheet 的 mock。"""
        if module_failure is None:
            module_failure = {}
        from contextlib import ExitStack

        with ExitStack() as stack:
            mock_write = stack.enter_context(patch("src.python.report.summary.write_llm_usage_sheet"))
            stack.enter_context(patch("src.python.llm.get_session_usage", return_value=raw_session))
            stack.enter_context(patch("src.python.llm.format_session_usage", return_value=formatted))
            stack.enter_context(patch("src.python.llm.prompts.LLM_MODULE_FAILURE", module_failure))
            stack.enter_context(patch("src.python.core.registry.get_llm_module_names", return_value=self._name_map))
            from src.python.report.excel_llm_usage import build_llm_usage_sheet

            build_llm_usage_sheet(self.wb, self.prog)
        return mock_write

    def test_cache_hit_all_modules(self):
        """全部模块缓存命中 → 各模块状态均为 'cached'、标签为 '缓存'。"""
        formatted = {
            "has_usage": True,
            "call_count": 0,
            "per_module": {
                "global_macro": {
                    "model": "deepseek-v4-flash",
                    "cached": True,
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "cache_hit_tokens": 500,
                    "cost": 0.0,
                    "thinking": False,
                    "endpoint": "",
                },
                "expert_review": {
                    "model": "claude-sonnet-4",
                    "cached": True,
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "cache_hit_tokens": 1200,
                    "cost": 0.0,
                    "thinking": True,
                    "endpoint": "",
                },
                "health_check": {
                    "model": "gpt-4o",
                    "cached": True,
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "cache_hit_tokens": 800,
                    "cost": 0.0,
                    "thinking": False,
                    "endpoint": "",
                },
                "penetration_deep": {
                    "model": "deepseek-v4-flash",
                    "cached": True,
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "cache_hit_tokens": 600,
                    "cost": 0.0,
                    "thinking": False,
                    "endpoint": "",
                },
            },
        }
        mock_write = self._run({}, formatted)
        mock_write.assert_called_once()
        excel_module_info = mock_write.call_args[0][2]

        by_key = {e["key"]: e for e in excel_module_info}
        for mk in ["global_macro", "expert_review", "health_check", "penetration_deep"]:
            with self.subTest(module=mk):
                self.assertEqual(by_key[mk]["status"], "cached")
                self.assertEqual(by_key[mk]["status_label"], "缓存")
                self.assertTrue(by_key[mk]["cached"])
        self.assertTrue(by_key["expert_review"]["thinking"])

    def test_mixed_cache_and_success(self):
        """混合缓存和真实调用 → 各自正确的状态。"""
        formatted = {
            "has_usage": True,
            "call_count": 2,
            "per_module": {
                "global_macro": {
                    "model": "deepseek-v4-flash",
                    "cached": True,
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "cache_hit_tokens": 500,
                    "cost": 0.0,
                    "thinking": False,
                    "endpoint": "",
                },
                "expert_review": {
                    "model": "claude-sonnet-4",
                    "cached": False,
                    "input_tokens": 1500,
                    "output_tokens": 800,
                    "cache_hit_tokens": 0,
                    "cost": 0.005,
                    "thinking": True,
                    "endpoint": "https://api.test.com",
                },
            },
        }
        mock_write = self._run({}, formatted)
        mock_write.assert_called_once()
        excel_module_info = mock_write.call_args[0][2]
        by_key = {e["key"]: e for e in excel_module_info}

        self.assertEqual(by_key["global_macro"]["status"], "cached")
        self.assertEqual(by_key["global_macro"]["total_tokens"], 0)
        self.assertEqual(by_key["expert_review"]["status"], "success")
        self.assertEqual(by_key["expert_review"]["status_label"], "成功")
        self.assertEqual(by_key["expert_review"]["total_tokens"], 2300)
        self.assertEqual(by_key["expert_review"]["input_tokens"], 1500)
        self.assertEqual(by_key["expert_review"]["output_tokens"], 800)
        self.assertTrue(by_key["expert_review"]["thinking"])
        self.assertEqual(by_key["expert_review"]["endpoint"], "https://api.test.com")

    def test_disabled_module(self):
        """禁用模块 → excel_module_info 含已禁用状态。"""
        from src.python.llm import FAIL_REASON_DISABLED

        formatted = {"has_usage": True, "per_module": {}}
        mock_write = self._run(
            {},
            formatted,
            module_failure={"global_macro": FAIL_REASON_DISABLED},
        )
        mock_write.assert_called_once()
        excel_module_info = mock_write.call_args[0][2]
        by_key = {e["key"]: e for e in excel_module_info}

        self.assertIn("global_macro", by_key)
        self.assertEqual(by_key["global_macro"]["status"], "disabled")
        self.assertEqual(by_key["global_macro"]["status_label"], "已禁用")
        self.assertEqual(by_key["global_macro"]["model"], "")

    def test_failed_module(self):
        """失败模块 → excel_module_info 含失败描述。"""
        from src.python.llm import FAIL_REASON_API_ERROR

        formatted = {"has_usage": True, "per_module": {}}
        mock_write = self._run(
            {},
            formatted,
            module_failure={"health_check": FAIL_REASON_API_ERROR},
        )
        mock_write.assert_called_once()
        excel_module_info = mock_write.call_args[0][2]
        by_key = {e["key"]: e for e in excel_module_info}

        self.assertIn("health_check", by_key)
        self.assertEqual(by_key["health_check"]["status"], "failed")
        self.assertEqual(by_key["health_check"]["status_label"], "LLM API 调用失败")

    def test_disabled_overrides_per_module(self):
        """禁用标记优先于 per_module 数据（即使有缓存数据也不显示）。"""
        from src.python.llm import FAIL_REASON_DISABLED

        formatted = {
            "has_usage": True,
            "per_module": {
                "global_macro": {
                    "model": "deepseek-v4-flash",
                    "cached": True,
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "cache_hit_tokens": 500,
                    "cost": 0.0,
                    "thinking": False,
                    "endpoint": "",
                },
            },
        }
        mock_write = self._run(
            {},
            formatted,
            module_failure={"global_macro": FAIL_REASON_DISABLED},
        )
        excel_module_info = mock_write.call_args[0][2]
        gm = next(e for e in excel_module_info if e["key"] == "global_macro")
        # 即使有缓存数据，禁用标记优先
        self.assertEqual(gm["status"], "disabled")
        self.assertEqual(gm["status_label"], "已禁用")

    def test_unknown_modules_skipped(self):
        """无 per_module 且无失败原因 → excel_module_info 为空，不调用 write_llm_usage_sheet。"""
        formatted = {"has_usage": True, "per_module": {}}
        mock_write = self._run({}, formatted)
        # 当 excel_module_info 为空时 build_llm_usage_sheet 直接 return，不调用 write_llm_usage_sheet
        mock_write.assert_not_called()

    def test_no_usage_returns_early(self):
        """format_session_usage 无 has_usage → 不调用 write_llm_usage_sheet。"""
        formatted = {"has_usage": False, "per_module": {}}
        mock_write = self._run({}, formatted)
        mock_write.assert_not_called()

    def test_raw_session_per_module_fallback(self):
        """raw_session 有 per_module 数据但 formatted 无 per_module → 仍应正确写入。

        回归验证：format_session_usage 在 call_count=0 且 per_module 有数据时返回
        has_usage=True（含 per_module），但如果时序/状态导致返回 {"has_usage": False}
        （不含 per_module），build_llm_usage_sheet 应通过 raw_session 拿到数据。
        """
        raw_session = {
            "per_module": {
                "global_macro": {
                    "model": "deepseek-v4-flash",
                    "cached": True,
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "cache_hit_tokens": 500,
                    "cost": 0.0,
                    "thinking": False,
                    "endpoint": "",
                },
            },
        }
        # format_session_usage 返回无 per_module 的"空"结果（模拟边界情况）
        formatted = {"has_usage": False}
        mock_write = self._run(raw_session, formatted)
        # 必须仍被调用，而非因 excel_module_info 为空而跳过
        mock_write.assert_called_once()
        excel_module_info = mock_write.call_args[0][2]
        self.assertEqual(len(excel_module_info), 1)
        self.assertEqual(excel_module_info[0]["key"], "global_macro")
        self.assertEqual(excel_module_info[0]["status"], "cached")

    def test_all_states_mixed(self):
        """禁用、失败、缓存、成功混合 → 各模块正确渲染且数量正确。"""
        from src.python.llm import FAIL_REASON_DISABLED, FAIL_REASON_TIMEOUT

        formatted = {
            "has_usage": True,
            "call_count": 1,
            "per_module": {
                "global_macro": {
                    "model": "deepseek-v4-flash",
                    "cached": False,
                    "input_tokens": 500,
                    "output_tokens": 300,
                    "cache_hit_tokens": 0,
                    "cost": 0.002,
                    "thinking": False,
                    "endpoint": "",
                },
                "expert_review": {
                    "model": "claude-sonnet-4",
                    "cached": True,
                    "input_tokens": 0,
                    "output_tokens": 0,
                    "cache_hit_tokens": 1000,
                    "cost": 0.0,
                    "thinking": True,
                    "endpoint": "",
                },
            },
        }
        mock_write = self._run(
            {},
            formatted,
            module_failure={
                "health_check": FAIL_REASON_DISABLED,
                "penetration_deep": FAIL_REASON_TIMEOUT,
            },
        )
        mock_write.assert_called_once()
        excel_module_info = mock_write.call_args[0][2]
        # 4 个模块都应出现在表格中
        by_key = {e["key"]: e for e in excel_module_info}

        self.assertEqual(by_key["global_macro"]["status"], "success")
        self.assertEqual(by_key["global_macro"]["status_label"], "成功")
        self.assertEqual(by_key["expert_review"]["status"], "cached")
        self.assertEqual(by_key["expert_review"]["status_label"], "缓存")
        self.assertEqual(by_key["health_check"]["status"], "disabled")
        self.assertEqual(by_key["health_check"]["status_label"], "已禁用")
        self.assertEqual(by_key["penetration_deep"]["status"], "failed")
        self.assertIn("请求超时", by_key["penetration_deep"]["status_label"])
        self.assertEqual(len(excel_module_info), 4)


class TestCreateSheets(unittest.TestCase):
    """create_sheets 页签创建与标题设置测试。"""

    _CUSTOM_ORDER = [
        {"key": "fund_performance", "name": "基金业绩分析", "number": 1, "type": "always"},
        {"key": "summary", "name": "投资分析汇总", "number": 2, "type": "always"},
        {"key": "market_value", "name": "市值核算明细表", "number": 3, "type": "always"},
    ]

    def _make_wb(self):
        """创建一个空 Workbook。"""
        from openpyxl import Workbook

        wb = Workbook()
        # 删除默认 Sheet
        wb.remove(wb.active)
        return wb

    def test_default_order_uses_default_titles(self):
        """默认 section_order → 标题使用连续重新编号（非注册序号）。"""
        from src.python.core.registry import _REPORT_SECTION_DEFAULT
        from src.python.report.excel_sheet_factory import create_sheets

        wb = self._make_wb()
        # always(6) + history(1，合并章) + evolution(1) = 8 个页签，连续编号 1-8（组合演进为独立 evolution 类型）
        sheets = create_sheets(
            wb, _REPORT_SECTION_DEFAULT, enable_fund_deep_analysis=False, enable_news=False, enable_llm=False
        )
        self.assertEqual(len(sheets), 8)
        expected_titles = {
            "summary": "1.投资分析汇总",
            "market_value": "2.市值核算明细表",
            "category": "3.持仓分类表",
            "penetration": "4.资产穿透TOP10",
            "fund_performance": "5.基金业绩分析",
            "portfolio_history_drawdown": "6.组合历史走势与回撤",
            "portfolio_evolution": "7.组合演进",
            "data_source_status": "8.数据源可用性矩阵",
        }
        for key, title in expected_titles.items():
            self.assertIn(key, sheets, f"{key} should be created")
            self.assertEqual(sheets[key].title, title, f"{key} title mismatch")
        # 回归：旧「组合历史走势」「历史回撤分析」独立 sheet 不再生成（已物理合并）
        self.assertNotIn("portfolio_history", sheets)
        self.assertNotIn("drawdown_analysis", sheets)

    def test_custom_order_uses_custom_titles(self):
        """自定义 section_order → 标题使用配置序号。"""
        from src.python.report.excel_sheet_factory import create_sheets

        wb = self._make_wb()
        sheets = create_sheets(
            wb, self._CUSTOM_ORDER, enable_fund_deep_analysis=False, enable_news=False, enable_llm=False
        )
        self.assertEqual(len(sheets), 3)
        self.assertEqual(sheets["fund_performance"].title, "1.基金业绩分析")
        self.assertEqual(sheets["summary"].title, "2.投资分析汇总")
        self.assertEqual(sheets["market_value"].title, "3.市值核算明细表")

    def test_visibility_filtering(self):
        """可见性过滤 → 只创建匹配 type 的页签。"""
        from src.python.core.registry import _REPORT_SECTION_DEFAULT
        from src.python.report.excel_sheet_factory import create_sheets

        wb = self._make_wb()
        # board 层启用 news + data 层 news 可用 → 新闻版块页签应出现
        sheets = create_sheets(
            wb,
            _REPORT_SECTION_DEFAULT,
            enable_fund_deep_analysis=False,
            enable_news=True,
            enable_llm=False,
            data_availability={"news_data_available": True},
        )
        news_keys = {s["key"] for s in _REPORT_SECTION_DEFAULT if s["type"] == "news"}
        # always(6) + history(1，合并章) + evolution(1) + news(1) = 9
        self.assertEqual(len(sheets), 9)
        for key in news_keys:
            self.assertIn(key, sheets)

    def test_evolution_disabled_hides_sheet(self):
        """enable_portfolio_evolution=False → 组合演进页签不创建。"""
        from src.python.core.registry import _REPORT_SECTION_DEFAULT
        from src.python.report.excel_sheet_factory import create_sheets

        wb = self._make_wb()
        # always(6) + history(1，合并章) = 7（evolution 关闭，无组合演进页签）
        sheets = create_sheets(
            wb,
            _REPORT_SECTION_DEFAULT,
            enable_fund_deep_analysis=False,
            enable_news=False,
            enable_llm=False,
            enable_portfolio_evolution=False,
        )
        self.assertEqual(len(sheets), 7)
        self.assertNotIn("portfolio_evolution", sheets)
        # 其他 always 页签不受影响
        self.assertIn("summary", sheets)
        self.assertIn("data_source_status", sheets)

    def test_evolution_enabled_creates_sheet(self):
        """enable_portfolio_evolution=True（默认）→ 组合演进页签创建。"""
        from src.python.core.registry import _REPORT_SECTION_DEFAULT
        from src.python.report.excel_sheet_factory import create_sheets

        wb = self._make_wb()
        sheets = create_sheets(
            wb, _REPORT_SECTION_DEFAULT, enable_fund_deep_analysis=False, enable_news=False, enable_llm=False
        )
        self.assertEqual(len(sheets), 8)
        self.assertIn("portfolio_evolution", sheets)


if __name__ == "__main__":
    unittest.main()
