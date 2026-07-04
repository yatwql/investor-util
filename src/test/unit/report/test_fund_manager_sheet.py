"""report/fund_manager_sheet.py 单元测试。

测试目标：
  - write_fund_manager_sheet：写入基本结构和内容正确
  - _change_label：变更标签正确
  - 预警级别颜色映射

运行：
  pytest src/test/ -m "unit_report" -k "fund_manager_sheet" -v
"""

from __future__ import annotations

import unittest
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from src.python.report.fund_manager_sheet import write_fund_manager_sheet

pytestmark = [pytest.mark.unit, pytest.mark.unit_report]


class TestWriteFundManagerSheet(unittest.TestCase):
    """write_fund_manager_sheet：Excel 写入"""

    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.manager_data = [
            {
                "name": "易方达中小盘混合", "code": "110011",
                "current_manager": "张坤", "tenure_days": 5000,
                "changed_1m": False, "changed_3m": False, "changed_6m": False,
                "alert_level": "正常", "is_first_check": False,
            },
            {
                "name": "某主动基金", "code": "007844",
                "current_manager": "刘彦春", "tenure_days": 2000,
                "changed_1m": True, "changed_3m": True, "changed_6m": True,
                "alert_level": "紧急", "is_first_check": False,
            },
            {
                "name": "新基金", "code": "999999",
                "current_manager": "王经理", "tenure_days": 30,
                "changed_1m": False, "changed_3m": False, "changed_6m": False,
                "alert_level": "首检", "is_first_check": True,
            },
        ]

    def test_write_title_exists(self):
        """标题行写入。"""
        write_fund_manager_sheet(self.ws, self.manager_data)
        self.assertIsNotNone(self.ws.cell(row=1, column=1).value)
        self.assertIn("基金经理变更监控", str(self.ws.cell(row=1, column=1).value))

    def test_write_headers(self):
        """表头写入。"""
        write_fund_manager_sheet(self.ws, self.manager_data)
        headers = [self.ws.cell(row=2, column=c).value for c in range(1, 9)]
        self.assertIn("基金名称", headers)
        self.assertIn("预警级别", headers)

    def test_write_data_rows(self):
        """数据行写入。"""
        write_fund_manager_sheet(self.ws, self.manager_data)
        # 3 条数据 → 3 行
        for i, item in enumerate(self.manager_data, start=3):
            self.assertEqual(self.ws.cell(row=i, column=1).value, item["name"])
            self.assertEqual(self.ws.cell(row=i, column=8).value, item["alert_level"])

    def test_empty_data_writes_title_and_headers(self):
        """空数据列表写入标题和表头（无数据行）。"""
        write_fund_manager_sheet(self.ws, [])
        self.assertIsNotNone(self.ws.cell(row=1, column=1).value)
        self.assertIsNotNone(self.ws.cell(row=2, column=1).value)
        # 第 3 行应为空（无数据）
        self.assertIsNone(self.ws.cell(row=3, column=1).value)

    def test_first_check_change_labels(self):
        """首检行变更列显示"—"。"""
        first_check_data = [
            {
                "name": "首检基金", "code": "999999",
                "current_manager": "王经理", "tenure_days": 30,
                "changed_1m": False, "changed_3m": False, "changed_6m": False,
                "alert_level": "首检", "is_first_check": True,
            },
        ]
        write_fund_manager_sheet(self.ws, first_check_data)
        # 第 5/6/7 列（1月内/3月内/6月内变更）都应显示"—"
        for col in (5, 6, 7):
            val = str(self.ws.cell(row=3, column=col).value)
            self.assertEqual(val, "—")
