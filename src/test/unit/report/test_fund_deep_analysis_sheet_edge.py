"""基金深度分析 Excel 写入模块空数据占位边缘测试。

测试目标：
  - fund_manager_sheet：manager_data 为空 → 写占位
  - position_relationship_sheet：重合度区块基金数 < 2 → 写占位（STATUS_MESSAGES）
  - fund_concentration_sheet：concentration_data 为空 → 写占位
  - style_factor_sheet：风格表 style_data 为空 → 写占位（风格与因子分析章区块一）

运行：
  pytest src/test/unit/report/test_fund_deep_analysis_sheet_edge.py -v
"""

from __future__ import annotations

import unittest

import openpyxl
import pytest

from src.python.report.data_status import STATUS_MESSAGES

pytestmark = [pytest.mark.unit, pytest.mark.unit_report, pytest.mark.edge]


class TestFundManagerSheetEmpty(unittest.TestCase):
    """fund_manager_sheet 空数据占位"""

    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_empty_manager_data_writes_placeholder(self):
        """manager_data=[] → 第4行含占位文本。"""
        from src.python.report.fund_manager_sheet import write_fund_manager_sheet
        write_fund_manager_sheet(self.ws, [])
        placeholder = self.ws.cell(row=4, column=1).value
        self.assertIsNotNone(placeholder)
        self.assertIn("暂不可用", str(placeholder))

    def test_empty_manager_data_status_message(self):
        """占位文本来自 STATUS_MESSAGES。"""
        from src.python.report.fund_manager_sheet import write_fund_manager_sheet
        write_fund_manager_sheet(self.ws, [])
        self.assertEqual(
            self.ws.cell(row=4, column=1).value,
            STATUS_MESSAGES["manager_unavailable"],
        )


class TestOverlapBlockEmpty(unittest.TestCase):
    """持仓关系矩阵·重合度区块 空/不足数据占位"""

    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_single_fund_writes_placeholder(self):
        """只有 1 只基金 → 重合度区块写占位。"""
        from src.python.report.position_relationship_sheet import write_position_relationship_sheet
        overlap_result = {"funds": ["110011"], "matrix": [], "pairs": []}
        write_position_relationship_sheet(self.ws, overlap_result)
        placeholder = self.ws.cell(row=3, column=1).value
        self.assertIsNotNone(placeholder)
        self.assertIn("无法计算", str(placeholder))

    def test_no_funds_writes_placeholder(self):
        """0 只基金 → 重合度区块写占位。"""
        from src.python.report.position_relationship_sheet import write_position_relationship_sheet
        overlap_result = {"funds": [], "matrix": [], "pairs": []}
        write_position_relationship_sheet(self.ws, overlap_result)
        placeholder = self.ws.cell(row=3, column=1).value
        self.assertEqual(placeholder, STATUS_MESSAGES["overlap_unavailable"])


class TestFundConcentrationSheetEmpty(unittest.TestCase):
    """fund_concentration_sheet 空数据占位"""

    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_empty_data_writes_placeholder(self):
        """concentration_data=[] → 第4行含占位文本。"""
        from src.python.report.fund_concentration_sheet import write_concentration_sheet
        write_concentration_sheet(self.ws, [])
        placeholder = self.ws.cell(row=4, column=1).value
        self.assertIsNotNone(placeholder)
        self.assertEqual(placeholder, STATUS_MESSAGES["concentration_unavailable"])


class TestFundStyleSheetEmpty(unittest.TestCase):
    """style_factor_sheet 风格表空数据占位（风格与因子分析章区块一）"""

    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_empty_style_data_writes_placeholder(self):
        """style_data=[] → 风格表区块写占位，因子区块正常。"""
        from src.python.report.style_factor_sheet import write_style_factor_sheet
        write_style_factor_sheet(self.ws, style_data=[], factor_exposure=None)
        flat = [str(c.value) for row in self.ws.iter_rows() for c in row if c.value is not None]
        self.assertIn(STATUS_MESSAGES["style_unavailable"], flat)
        # 因子区块独立降级：占位文本同时存在（一章三区块互不影响）
        self.assertIn(STATUS_MESSAGES["factor_exposure_unavailable"], flat)


if __name__ == "__main__":
    unittest.main()
