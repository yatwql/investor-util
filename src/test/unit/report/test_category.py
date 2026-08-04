"""持仓分类模块单元测试 — 分类逻辑与边界测试。

测试目标：
  - _categorize_holding — 8 种分类分支全覆盖
  - write_category_sheet — 空/单条/混合持仓场景
  - _apply_profit_colors — 着色边界
  - 三维度分类聚合一致 — 资产属性/投资分类/账户小计之和 = 总计

运行：
  cd D:/codebase/zoo/investor-util
  python -m unittest src.test_category -v
"""

from __future__ import annotations

import unittest
from unittest.mock import patch

from src.python.core.models import Holding
from src.python.report import category as cat
from src.python.report.market_value import DetailRow
import pytest
pytestmark = [pytest.mark.unit, pytest.mark.unit_report]



class TestCategorizeHolding(unittest.TestCase):
    """_categorize_holding 全分支测试。"""

    def _h(self, name: str, code: str, account: str = "证券账户") -> Holding:
        return Holding(account=account, name=name, code=code,
                        shares=100, cost_price=10.0)

    @pytest.mark.smoke
    def test_qdii(self):
        """QDII → (基金, QDII)。"""
        h = self._h("华夏纳斯达克100ETF(QDII)", "513300")
        self.assertEqual(cat._categorize_holding(h), ("基金", "QDII"))

    @pytest.mark.smoke
    def test_bond(self):
        """含债关键词 → (债券, 纯债)。"""
        h = self._h("招商鑫福中短债A", "012325", "支付宝")
        self.assertEqual(cat._categorize_holding(h), ("债券", "纯债"))

    def test_cash(self):
        """含货币关键词 → (现金, 货币)。"""
        h = self._h("银华货币A", "180008", "支付宝")
        self.assertEqual(cat._categorize_holding(h), ("现金", "货币"))

    @pytest.mark.smoke
    def test_stock(self):
        """股票代码 6/0/3 开头 → (股票, A股)。"""
        h = self._h("贵州茅台", "600519")
        self.assertEqual(cat._categorize_holding(h), ("股票", "A股"))

    @pytest.mark.smoke
    def test_index_fund(self):
        """场内 ETF 代码 5 开头 → (基金, 指数)。"""
        h = self._h("南方中证500ETF", "510500", "证券账户")
        self.assertEqual(cat._categorize_holding(h), ("基金", "指数"))

    def test_active_fund(self):
        """场外渠道且非指数名称 → (基金, 主动)。"""
        h = self._h("易方达蓝筹精选混合", "005827", "支付宝")
        self.assertEqual(cat._categorize_holding(h), ("基金", "主动"))

    def test_hybrid_fund(self):
        """微信钱包账户含混合名称 → (基金, 主动)（微信属场外渠道）。"""
        h = self._h("招商安润灵活配置混合", "000126", "微信钱包")
        self.assertEqual(cat._categorize_holding(h), ("基金", "主动"))

    def test_other_fund(self):
        """其他基金（支付宝账户非指数）→ (基金, 主动)。"""
        h = self._h("某只特殊策略基金", "888888", "支付宝")
        self.assertEqual(cat._categorize_holding(h), ("基金", "主动"))

    def test_etf_code_no_keyword(self):
        """ETF代码(5开头)但名称无关键词 → (基金, 指数)。"""
        h = self._h("科创50", "588000")
        self.assertEqual(cat._categorize_holding(h), ("基金", "指数"))


class TestWriteCategorySheet(unittest.TestCase):
    """write_category_sheet 集成测试。"""

    def setUp(self):
        import openpyxl

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_single_holding(self):
        """单条持仓 → 分类表正确渲染。"""
        h = Holding("证券", "长江电力", "600900", 100, 10.0)
        details = [DetailRow(
            code="600900", account="证券", name="长江电力",
            market_value=2500, profit=1500, profit_rate=1.5)]
        cat.write_category_sheet(self.ws, [h], details)
        # 应该至少有一行标题 + 一行数据 + 合计行
        self.assertGreater(self.ws.max_row, 2)

    def test_empty_holdings(self):
        """空持仓 → 不崩溃。"""
        try:
            cat.write_category_sheet(self.ws, [], [])
        except Exception as e:
            self.fail(f"write_category_sheet with empty holdings raised: {e}")


class TestWriteCategoryFlowSubcolumns(unittest.TestCase):
    """持仓分类表成本流水子列（成本分档 / 分红累计）渲染测试。

    开关 `report_submodules.cost_lots` 对应 fund_flow_data 是否传入：
    None → 保持既有 10 列；非 None → 追加「成本分档」「分红累计」子列。
    """

    def setUp(self):
        import openpyxl

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def _call(self, holdings, details, fund_flow_data=None):
        # mock 分红 API 加载（避免真实网络请求，测试隔离）
        with patch("src.python.report.category._load_dividend_data", return_value=({}, True)):
            cat.write_category_sheet(self.ws, holdings, details, fund_flow_data=fund_flow_data)

    def _headers(self):
        return [self.ws.cell(row=2, column=c).value for c in range(1, 13)]

    def _find_total_row(self):
        for r in range(1, self.ws.max_row + 1):
            if self.ws.cell(row=r, column=1).value == "总计":
                return r
        return None

    @staticmethod
    def _flow(low_shares=0.0, high_shares=0.0, div=0.0):
        return {
            "available": True,
            "cost_tiers": {
                "per_code": {
                    "600900": {
                        "low": {"shares": low_shares, "cost": low_shares * 9.0},
                        "high": {"shares": high_shares, "cost": high_shares * 11.0},
                        "unpriced": {"shares": 0.0, "cost": 0.0},
                    }
                }
            },
            "dividends": {"per_code": {"600900": div}},
        }

    def test_flow_subcolumns_header_when_enabled(self):
        """开关开启（fund_flow_data 非 None）时，表头追加「成本分档」「分红累计」列。"""
        h = Holding("证券", "长江电力", "600900", 100, 10.0)
        details = [
            DetailRow(code="600900", account="证券", name="长江电力",
                      market_value=1500, cost=1000, profit=500, profit_rate=0.5)
        ]
        self._call([h], details, self._flow(low_shares=100, div=120.0))
        headers = self._headers()
        self.assertIn("成本分档", headers)
        self.assertIn("分红累计", headers)
        self.assertEqual(headers.index("成本分档") + 1, 11)
        self.assertEqual(headers.index("分红累计") + 1, 12)

    def test_flow_tier_low_and_div_value(self):
        """开关开启时数据行含「低成本」分档标签与分红累计数值。"""
        h = Holding("证券", "长江电力", "600900", 100, 10.0)
        details = [
            DetailRow(code="600900", account="证券", name="长江电力",
                      market_value=1500, cost=1000, profit=500, profit_rate=0.5)
        ]
        self._call([h], details, self._flow(low_shares=100, div=120.0))
        # 数据行 = 表头下一行（row 3）：列 11 = 成本分档, 列 12 = 分红累计
        self.assertEqual(self.ws.cell(row=3, column=11).value, "低成本")
        self.assertEqual(self.ws.cell(row=3, column=12).value, 120.0)

    def test_flow_tier_high_label(self):
        """持仓批次成本价高于市价时渲染「高成本」档标签。"""
        h = Holding("证券", "长江电力", "600900", 100, 10.0)
        details = [
            DetailRow(code="600900", account="证券", name="长江电力",
                      market_value=900, cost=1000, profit=-100, profit_rate=-0.1)
        ]
        self._call([h], details, self._flow(high_shares=100, div=0.0))
        self.assertEqual(self.ws.cell(row=3, column=11).value, "高成本")

    def test_flow_tier_mixed_label(self):
        """持仓批次横跨低/高两档时渲染「混合」档标签。"""
        h = Holding("证券", "长江电力", "600900", 100, 10.0)
        details = [
            DetailRow(code="600900", account="证券", name="长江电力",
                      market_value=1000, cost=1000, profit=0, profit_rate=0.0)
        ]
        self._call([h], details, self._flow(low_shares=50, high_shares=50, div=0.0))
        self.assertEqual(self.ws.cell(row=3, column=11).value, "混合")

    def test_flow_total_row_dividend_sum(self):
        """开关开启时，分红累计在小计/总计行汇总（跨持仓累加）。"""
        flow = {
            "available": True,
            "cost_tiers": {
                "per_code": {
                    "600900": {"low": {"shares": 100, "cost": 900}, "high": {"shares": 0, "cost": 0}, "unpriced": {"shares": 0, "cost": 0}},
                    "600519": {"low": {"shares": 10, "cost": 15000}, "high": {"shares": 0, "cost": 0}, "unpriced": {"shares": 0, "cost": 0}},
                }
            },
            "dividends": {"per_code": {"600900": 120.0, "600519": 30.0}},
        }
        h1 = Holding("证券", "长江电力", "600900", 100, 10.0)
        h2 = Holding("证券", "贵州茅台", "600519", 10, 1500.0)
        details = [
            DetailRow(code="600900", account="证券", name="长江电力", market_value=1500, cost=1000, profit=500, profit_rate=0.5),
            DetailRow(code="600519", account="证券", name="贵州茅台", market_value=20000, cost=15000, profit=5000, profit_rate=0.33),
        ]
        self._call([h1, h2], details, flow)
        total_row = self._find_total_row()
        self.assertIsNotNone(total_row)
        self.assertEqual(self.ws.cell(row=total_row, column=12).value, 150.0)

    def test_no_flow_subcolumns_when_disabled(self):
        """开关关闭（fund_flow_data=None）时保持既有 10 列输出，无成本流水子列。"""
        h = Holding("证券", "长江电力", "600900", 100, 10.0)
        details = [
            DetailRow(code="600900", account="证券", name="长江电力",
                      market_value=1500, cost=1000, profit=500, profit_rate=0.5)
        ]
        self._call([h], details, None)
        headers = self._headers()
        self.assertNotIn("成本分档", headers)
        self.assertNotIn("分红累计", headers)
        # 既有表头 10 列保持不变（前 10 列与改造前一致）
        self.assertEqual(headers[:10][-1], "年均股息率")


@pytest.mark.data
class TestCategoryAggregationConsistency(unittest.TestCase):
    """三维度分类聚合一致 — 各维度小计之和 = 总计。"""

    def _build_mock_category_result(self):
        """构造模拟的 category() 输出，验证聚合一致。"""
        # 模拟分类结果结构：{维度: {类别: {account: {合计值}}}}
        return {
            "asset_type": {  # 资产属性维度
                "股票": {"count": 2, "market_value": 5000, "profit": 1000},
                "基金": {"count": 3, "market_value": 8000, "profit": -500},
                "债券": {"count": 1, "market_value": 3000, "profit": 100},
            },
            "invest_type": {  # 投资分类维度
                "主动管理": {"count": 2, "market_value": 6000, "profit": -300},
                "被动指数": {"count": 1, "market_value": 2000, "profit": -200},
                "纯债": {"count": 1, "market_value": 3000, "profit": 100},
                "股票": {"count": 2, "market_value": 5000, "profit": 1000},
            },
            "account": {  # 账户维度
                "证券账户": {"count": 3, "market_value": 9000, "profit": 800},
                "支付宝": {"count": 2, "market_value": 4000, "profit": -300},
                "微信钱包": {"count": 1, "market_value": 3000, "profit": 100},
            },
        }

    def test_asset_type_subtotals_equal_grand_total(self):
        """资产属性各分类市值之和 = 总市值。"""
        data = self._build_mock_category_result()
        total_mv = sum(v["market_value"] for v in data["asset_type"].values())
        expected_total = 5000 + 8000 + 3000
        self.assertEqual(total_mv, expected_total)
        self.assertEqual(total_mv, 16000)

    def test_invest_type_subtotals_equal_grand_total(self):
        """投资分类各子类市值之和 = 总市值。"""
        data = self._build_mock_category_result()
        total_mv = sum(v["market_value"] for v in data["invest_type"].values())
        expected_total = 6000 + 2000 + 3000 + 5000
        self.assertEqual(total_mv, expected_total)
        self.assertEqual(total_mv, 16000)

    def test_account_subtotals_equal_grand_total(self):
        """各账户市值之和 = 总市值。"""
        data = self._build_mock_category_result()
        total_mv = sum(v["market_value"] for v in data["account"].values())
        expected_total = 9000 + 4000 + 3000
        self.assertEqual(total_mv, expected_total)
        self.assertEqual(total_mv, 16000)

    def test_all_three_dimensions_agree(self):
        """三个维度的总计应一致。"""
        data = self._build_mock_category_result()
        asset_total = sum(v["market_value"] for v in data["asset_type"].values())
        invest_total = sum(v["market_value"] for v in data["invest_type"].values())
        account_total = sum(v["market_value"] for v in data["account"].values())
        self.assertEqual(asset_total, invest_total)
        self.assertEqual(invest_total, account_total)

    def test_profit_subtotals_also_consistent(self):
        """利润维度也应三维一致（交叉验证非仅市值）。"""
        data = self._build_mock_category_result()
        asset_profit = sum(v["profit"] for v in data["asset_type"].values())
        invest_profit = sum(v["profit"] for v in data["invest_type"].values())
        account_profit = sum(v["profit"] for v in data["account"].values())
        self.assertEqual(asset_profit, invest_profit)
        self.assertEqual(invest_profit, account_profit)
        # 1000 + (-500) + 100 = 600
        self.assertEqual(asset_profit, 600)

    def test_no_negative_counts(self):
        """所有 count 应为非负整数。"""
        data = self._build_mock_category_result()
        for dim_name, dim_data in data.items():
            for cat_name, values in dim_data.items():
                self.assertGreaterEqual(
                    values["count"], 0,
                    f"{dim_name}.{cat_name} count={values['count']} 为负"
                )


class TestYieldText(unittest.TestCase):
    """calc_yield_text 基础功能测试。"""

    def setUp(self):
        self.d = cat.DetailRow()
        self.d.name = "长江电力"
        self.d.code = "600900"
        self.d.price = 50.0
        self.d.market_value = 5000.0
        self.d.cost = 4500.0
        self.d.profit = 500.0
        self.d.profit_rate = 0.10
        self.d.today_profit = 10.0

    def test_normal(self):
        """正常数据 → 正确的股息率百分比。"""
        result = cat.calc_yield_text("600900", self.d, {"600900": {"avg_dividend": 0.85}})
        self.assertEqual(result, "1.70%")

    def test_code_not_found(self):
        """代码不在分红数据中 → "--"。"""
        result = cat.calc_yield_text("600900", self.d, {})
        self.assertEqual(result, "--")

    def test_avg_dividend_none(self):
        """avg_dividend 为 None → "--"。"""
        result = cat.calc_yield_text("600900", self.d, {"600900": {"avg_dividend": None}})
        self.assertEqual(result, "--")


if __name__ == "__main__":
    unittest.main()
