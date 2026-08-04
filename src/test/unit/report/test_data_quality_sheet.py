"""数据质量仪表盘页签写入单元测试（「数据源可用性矩阵」升级：源健康 + 品种覆盖 + 可信度）。

测试目标：
  - build_coverage_block 规范化品种覆盖区块（有数据 / 无数据 / 降级）
  - write_data_quality_sheet 写入标题「数据质量仪表盘」+ 源健康 + 品种覆盖 + 可信度区块
  - position_status 为空 → 品种覆盖区块写降级占位
  - data_freshness 为空 → 可信度区块写降级占位
  - 异常品种行以醒目状态标注；跳变品种可信度区块可见

运行：
  python -m pytest src/test/unit/report/test_data_quality_sheet.py -v
"""

from __future__ import annotations

import unittest

import openpyxl
import pytest

from unittest.mock import patch

from src.python.report.data_quality_sheet import (
    _COVERAGE_PLACEHOLDER,
    _FRESHNESS_PLACEHOLDER,
    build_coverage_block,
    write_data_quality_sheet,
)

pytestmark = [pytest.mark.unit, pytest.mark.unit_report]


# ── 辅助构造 ──────────────────────────────────────────────


def _matrix_row(name: str, status: str = "ok", detail: str = "1 正常") -> dict:
    """构造 build_data_source_matrix 输出格式的单行。"""
    return {
        "key": "price",
        "name": name,
        "status": status,
        "detail": detail,
        "total": 1,
        "ok": 1 if status == "ok" else 0,
        "degraded": 1 if status == "degraded" else 0,
        "failed": 1 if status == "failed" else 0,
        "sample_failures": [] if status != "failed" else ["price_x: unreachable"],
        "degraded_list": [] if status != "degraded" else ["price_x: timeout"],
    }


def _coverage_item(code: str, name: str, status: str) -> dict:
    """构造品种覆盖 items 单行。"""
    return {
        "code": code,
        "name": name,
        "account": "账户A",
        "status": status,
        "status_label": {"ok": "正常", "nav_missing": "净值缺失", "possibly_delisted": "可能退市"}.get(status, status),
        "reason": {"ok": "", "nav_missing": "未获取到净值", "possibly_delisted": "无有效行情"}.get(status, ""),
    }


def _coverage_status(items: list[dict]) -> dict:
    """构造 position_status C19 契约 dict。"""
    abnormal = [i for i in items if i["status"] != "ok"]
    return {
        "available": True,
        "items": items,
        "abnormal_count": len(abnormal),
        "summary": f"{len(items)} 个品种，{len(abnormal)} 个数据异常",
    }


def _freshness_item(code: str, name: str, freshness: str, change_pct: float = 0.0, jump: bool = False) -> dict:
    """构造可信度摘要 items 单行。"""
    label_map = {"fresh": "实时", "cached": "缓存（T-1）", "stale": "过期", "degraded": "降级"}
    jump_label = f"疑似数据错误（单日 {change_pct:+.2f}%）" if jump else ""
    return {
        "code": code,
        "name": name,
        "account": "账户A",
        "freshness": freshness,
        "freshness_label": label_map.get(freshness, freshness),
        "jump": jump,
        "jump_label": jump_label,
        "change_pct": change_pct,
    }


def _freshness_status(items: list[dict]) -> dict:
    """构造 data_freshness C19 契约 dict。"""
    abnormal = [i for i in items if i["freshness"] in ("stale", "degraded") or i["jump"]]
    return {
        "available": True,
        "items": items,
        "abnormal_count": len(abnormal),
        "summary": f"{len(items)} 个品种，{len(abnormal)} 个数据异常",
    }


class TestBuildCoverageBlock(unittest.TestCase):
    """build_coverage_block 规范化测试。"""

    def setUp(self):
        self.ws = openpyxl.Workbook().active

    def test_has_data_with_abnormal(self):
        """有 items 且含异常 → has_data=True，abnormal_count 正确。"""
        block = build_coverage_block(
            _coverage_status(
                [_coverage_item("600900", "长江电力", "ok"), _coverage_item("99999", "坏码", "nav_missing")]
            )
        )
        self.assertTrue(block["has_data"])
        self.assertEqual(block["abnormal_count"], 1)
        self.assertEqual(len(block["items"]), 2)

    def test_none_position_status(self):
        """position_status=None → has_data=False（未获取行情数据）。"""
        block = build_coverage_block(None)
        self.assertFalse(block["has_data"])
        self.assertEqual(block["items"], [])

    def test_unavailable_status(self):
        """available=False → has_data=False（数据不可用）。"""
        block = build_coverage_block({"available": False, "items": [], "abnormal_count": 0, "summary": ""})
        self.assertFalse(block["has_data"])


class TestWriteDataQualitySheet(unittest.TestCase):
    """write_data_quality_sheet Excel 写入测试。"""

    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_title_is_dashboard(self):
        """页签标题为「数据质量仪表盘」。"""
        write_data_quality_sheet(
            self.ws, [_matrix_row("行情数据")], _coverage_status([_coverage_item("600900", "长江电力", "ok")])
        )
        self.assertIn("数据质量仪表盘", str(self.ws.cell(row=1, column=1).value))

    def test_source_health_block_kept(self):
        """源健康区块保留矩阵数据源行与列头。"""
        write_data_quality_sheet(self.ws, [_matrix_row("行情数据", "degraded", "1 降级")], None)
        cells = [
            str(self.ws.cell(row=r, column=c).value or "") for r in range(1, self.ws.max_row + 1) for c in range(1, 6)
        ]
        joined = "|".join(cells)
        self.assertIn("源健康", joined)
        self.assertIn("行情数据", joined)
        self.assertIn("降级", joined)

    def test_position_coverage_block_renders(self):
        """品种覆盖区块列出品种状态行。"""
        write_data_quality_sheet(
            self.ws,
            [_matrix_row("行情数据")],
            _coverage_status([_coverage_item("600900", "长江电力", "nav_missing")]),
        )
        cells = [
            str(self.ws.cell(row=r, column=c).value or "") for r in range(1, self.ws.max_row + 1) for c in range(1, 6)
        ]
        joined = "|".join(cells)
        self.assertIn("品种覆盖", joined)
        self.assertIn("600900", joined)
        self.assertIn("长江电力", joined)
        self.assertIn("净值缺失", joined)

    def test_coverage_placeholder_when_no_data(self):
        """position_status=None → 品种覆盖区块写占位文本。"""
        write_data_quality_sheet(self.ws, [_matrix_row("行情数据")], None)
        cells = [
            str(self.ws.cell(row=r, column=c).value or "") for r in range(1, self.ws.max_row + 1) for c in range(1, 6)
        ]
        self.assertIn(_COVERAGE_PLACEHOLDER, "|".join(cells))

    def test_abnormal_rows_highlighted(self):
        """异常品种行标注状态文案（非正常状态可见）。"""
        write_data_quality_sheet(
            self.ws,
            [_matrix_row("行情数据")],
            _coverage_status(
                [_coverage_item("110011", "易方达", "ok"), _coverage_item("88888", "退市股", "possibly_delisted")]
            ),
        )
        cells = [
            str(self.ws.cell(row=r, column=c).value or "") for r in range(1, self.ws.max_row + 1) for c in range(1, 6)
        ]
        joined = "|".join(cells)
        self.assertIn("可能退市", joined)
        self.assertIn("88888", joined)

    def test_empty_matrix_still_writes_coverage(self):
        """源健康矩阵为空时，品种覆盖区块仍可写入。"""
        end_row = write_data_quality_sheet(
            self.ws,
            [],
            _coverage_status([_coverage_item("600900", "长江电力", "ok")]),
        )
        cells = [str(self.ws.cell(row=r, column=c).value or "") for r in range(1, end_row + 1) for c in range(1, 6)]
        self.assertIn("品种覆盖", "|".join(cells))

    def test_freshness_block_renders(self):
        """可信度区块列出新鲜度与跳变信息。"""
        write_data_quality_sheet(
            self.ws,
            [_matrix_row("行情数据")],
            _coverage_status([_coverage_item("600900", "长江电力", "ok")]),
            _freshness_status([_freshness_item("600900", "长江电力", "fresh", 0.5)]),
        )
        cells = [
            str(self.ws.cell(row=r, column=c).value or "") for r in range(1, self.ws.max_row + 1) for c in range(1, 6)
        ]
        joined = "|".join(cells)
        self.assertIn("可信度", joined)
        self.assertIn("新鲜度", joined)
        self.assertIn("600900", joined)

    def test_jump_highlighted_in_freshness_block(self):
        """跳变品种在可信度区块标注疑似数据错误。"""
        write_data_quality_sheet(
            self.ws,
            [_matrix_row("行情数据")],
            _coverage_status([_coverage_item("005827", "易方达", "ok")]),
            _freshness_status([_freshness_item("005827", "易方达", "fresh", 25.0, jump=True)]),
        )
        cells = [
            str(self.ws.cell(row=r, column=c).value or "") for r in range(1, self.ws.max_row + 1) for c in range(1, 6)
        ]
        joined = "|".join(cells)
        self.assertIn("疑似数据错误", joined)
        self.assertIn("+25.00%", joined)

    def test_freshness_placeholder_when_no_data(self):
        """data_freshness=None → 可信度区块写占位文本。"""
        end_row = write_data_quality_sheet(
            self.ws,
            [_matrix_row("行情数据")],
            _coverage_status([_coverage_item("600900", "长江电力", "ok")]),
            None,
        )
        cells = [str(self.ws.cell(row=r, column=c).value or "") for r in range(1, end_row + 1) for c in range(1, 6)]
        self.assertIn(_FRESHNESS_PLACEHOLDER, "|".join(cells))


class TestLegacySourceMatrixStyle(unittest.TestCase):
    """旧样式回归测试 — 开关关闭时「数据源可用性矩阵」保持原样。"""

    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_legacy_title_and_matrix_rows(self):
        """旧样式标题为「数据源可用性矩阵」，矩阵行保留。"""
        from src.python.report.excel_generator import _write_data_source_matrix_sheet
        from src.python.report.progress import SilentProgressReporter

        with patch(
            "src.python.report.data_source_matrix.build_data_source_matrix",
            return_value=[_matrix_row("行情数据", "degraded", "1 降级")],
        ):
            _write_data_source_matrix_sheet(self.ws, SilentProgressReporter())
        cells = [
            str(self.ws.cell(row=r, column=c).value or "") for r in range(1, self.ws.max_row + 1) for c in range(1, 6)
        ]
        joined = "|".join(cells)
        self.assertIn("数据源可用性矩阵", joined)
        self.assertIn("行情数据", joined)
        self.assertIn("降级", joined)
        # 旧样式不含「品种覆盖」区块（开关关闭时无新增内容）
        self.assertNotIn("品种覆盖", joined)
