"""持仓关系矩阵页签·相关性区块 Excel 呈现测试。

覆盖（position_relationship_sheet 的 _write_correlation_block）：
  - available=True → 写入下三角矩阵 + 配对明细 + 说明区
  - 配对按 |r| 降序、显著标记、r 颜色字体
  - available=False（数据不足）→ 占位文本
  - 相关性数据 None → 相关性区块占位
  - 上三角留空、对角线=1.00、重叠样本不足格=N/A
"""

from __future__ import annotations

import unittest

import pytest

pytestmark = [pytest.mark.unit, pytest.mark.unit_report]


def _correlation_data(**extra) -> dict:
    """构造持仓关系矩阵·相关性区块数据契约 mock（2 品种，强负相关）。"""
    d = {
        "available": True,
        "status": "ok",
        "window": 60,
        "sample_count": 60,
        "codes": ["a", "b"],
        "names": {"a": "资产A", "b": "资产B"},
        "matrix": [[1.0, None], [-0.87, 1.0]],
        "p_values": [[None, None], [0.0001, None]],
        "pairs": [
            {
                "code_a": "b",
                "name_a": "资产B",
                "code_b": "a",
                "name_b": "资产A",
                "pearson": -0.87,
                "p_value": 0.0001,
                "significant": True,
                "samples": 60,
            }
        ],
        "insufficient_codes": [],
        "note": "",
    }
    d.update(extra)
    return d


class TestExcelCorrelationSheet(unittest.TestCase):
    """持仓关系矩阵页签·相关性区块 Excel 呈现测试。"""

    def _write(self, correlation_data) -> "object":
        from openpyxl import Workbook

        from src.python.report.position_relationship_sheet import write_position_relationship_sheet

        wb = Workbook()
        ws = wb.active
        write_position_relationship_sheet(ws, overlap_result=None, correlation_data=correlation_data)
        return ws

    def _all_text(self, ws) -> list[list[str]]:
        return [[str(c.value) if c.value is not None else "" for c in row] for row in ws.iter_rows()]

    def _flat(self, ws) -> list[str]:
        return [v for row in self._all_text(ws) for v in row]

    def test_matrix_pairs_and_notes_rendered(self):
        """available=True → 矩阵 + 配对明细 + 说明区齐全。"""
        ws = self._write(_correlation_data())
        titles = [r[0] for r in self._all_text(ws)]
        self.assertTrue(any("相关性矩阵" in t or "持仓相关性" in t for t in titles), f"应含相关性标题，实际: {titles}")
        self.assertTrue(any("配对明细" in t for t in titles), f"应含配对明细标题，实际: {titles}")
        self.assertTrue(any("说明" in t for t in titles), f"应含说明标题，实际: {titles}")
        # 标题顺序：矩阵 → 配对 → 说明
        corr_idx = next(i for i, t in enumerate(titles) if "相关性" in t)
        pairs_idx = next(i for i, t in enumerate(titles) if "配对明细" in t)
        notes_idx = next(i for i, t in enumerate(titles) if "说明" in t)
        self.assertLess(corr_idx, pairs_idx)
        self.assertLess(pairs_idx, notes_idx)
        # 配对数据完整
        flat = self._flat(ws)
        self.assertTrue(any("资产A" in v for v in flat))
        self.assertTrue(any("资产B" in v for v in flat))
        self.assertIn("-0.87", flat)
        self.assertIn("0.0001", flat)
        self.assertTrue(any("显著" in v for v in flat))
        # 说明区包含窗口与样本
        self.assertTrue(any("计算窗口" in v for v in flat))

    def test_matrix_lower_triangle_and_diagonal(self):
        """下三角有值、对角线 1.0、上三角留空。"""
        ws = self._write(_correlation_data())
        grid = self._all_text(ws)
        # 矩阵数据行（r[0] 非空即矩阵/标题行，标题不以此前缀开头）
        data_rows = [r for r in grid if r and r[0]]
        a_row = next(r for r in data_rows if r[0].startswith("资产A"))
        b_row = next(r for r in data_rows if r[0].startswith("资产B"))
        self.assertIn("1.0", a_row)  # 对角线（raw float，number_format 显示 1.00）
        self.assertIn("", a_row)  # 上三角留空
        self.assertIn("-0.87", b_row)  # 下三角 r

    def test_na_cell_for_insufficient_overlap(self):
        """重叠样本不足品种 → 相关性格为 N/A。"""
        data = _correlation_data()
        data["matrix"] = [[1.0, None], [None, 1.0]]
        data["insufficient_codes"] = ["a", "b"]
        ws = self._write(data)
        flat = self._flat(ws)
        self.assertIn("N/A", flat)
        # 说明区提示数据不足品种
        self.assertTrue(any("不足" in v and "a" in v for v in flat))

    def test_unavailable_placeholder(self):
        """available=False（数据不足/源故障）→ 占位文本。"""
        data = _correlation_data(available=False, status="insufficient", codes=[], matrix=[], pairs=[])
        ws = self._write(data)
        flat = self._flat(ws)
        self.assertTrue(any("持仓相关性数据暂不可用" in v for v in flat), "available=False 应写占位")

    def test_none_placeholder(self):
        """correlation_data=None → 整页占位。"""
        ws = self._write(None)
        flat = self._flat(ws)
        self.assertTrue(any("持仓相关性数据暂不可用" in v for v in flat), "None 应写整页占位")

    def test_pairs_sorted_by_abs_r(self):
        """配对明细按 |r| 降序（与 HTML 一致）。"""
        data = _correlation_data()
        data["codes"] = ["a", "b", "c"]
        data["names"] = {"a": "资产A", "b": "资产B", "c": "资产C"}
        data["matrix"] = [
            [1.0, None, None],
            [-0.2, 1.0, None],
            [0.9, 0.1, 1.0],
        ]
        data["pairs"] = [
            {
                "code_a": "c",
                "name_a": "资产C",
                "code_b": "a",
                "name_b": "资产A",
                "pearson": 0.9,
                "p_value": 0.001,
                "significant": True,
                "samples": 60,
            },
            {
                "code_a": "b",
                "name_a": "资产B",
                "code_b": "a",
                "name_b": "资产A",
                "pearson": -0.2,
                "p_value": 0.1,
                "significant": False,
                "samples": 60,
            },
            {
                "code_a": "c",
                "name_a": "资产C",
                "code_b": "b",
                "name_b": "资产B",
                "pearson": 0.1,
                "p_value": 0.2,
                "significant": False,
                "samples": 60,
            },
        ]
        ws = self._write(data)
        rows = self._all_text(ws)
        # 只取「配对明细」表内的 r 列（第 4 列，index 3），排除矩阵格干扰
        pairs_idx = next(i for i, r in enumerate(rows) if r and "配对明细" in r[0])
        pearson_vals: list[float] = []
        for r in rows[pairs_idx + 2 :]:  # +1 表头行，+1 跳到首条数据
            if not r or not r[0] or r[0] == "说明":
                break  # 数据行结束（空行或说明标题）
            pearson_vals.append(float(r[3]))
        self.assertEqual(pearson_vals, [0.9, -0.2, 0.1])
        self.assertEqual(pearson_vals, sorted(pearson_vals, key=abs, reverse=True), "配对应严格按 |r| 降序")
        # 显著标记出现在 0.9 那对
        flat_pairs = [v for r in rows[pairs_idx:] for v in r]
        self.assertTrue(any("显著" in v for v in flat_pairs))


def _overlap_result(**extra) -> dict:
    """构造持仓关系矩阵·重合度区块结构（2 只基金，部分重合 50%）。"""
    d = {
        "fund_names": {"a": "基金A", "b": "基金B"},
        "funds": ["a", "b"],
        "matrix": [
            [1.0, 0.5],
            [0.5, 1.0],
        ],
        "pairs": [
            {
                "fund_a": "a",
                "fund_b": "b",
                "name_a": "基金A",
                "name_b": "基金B",
                "code_a": "a",
                "code_b": "b",
                "common_count": 2,
                "jaccard": 0.5,
                "common_stocks": [
                    {"name": "贵州茅台", "code": "600519"},
                    {"name": "五粮液", "code": "000858"},
                ],
            }
        ],
    }
    d.update(extra)
    return d


class TestExcelMergedRelationshipSheet(unittest.TestCase):
    """持仓关系矩阵页签·一章两区块（重合度 + 相关性）Excel 呈现测试。"""

    def _write(self, overlap_result, correlation_data) -> "object":
        from openpyxl import Workbook

        from src.python.report.position_relationship_sheet import write_position_relationship_sheet

        wb = Workbook()
        ws = wb.active
        write_position_relationship_sheet(ws, overlap_result=overlap_result, correlation_data=correlation_data)
        return ws

    def _flat(self, ws) -> list[str]:
        return [str(c.value) if c.value is not None else "" for row in ws.iter_rows() for c in row]

    def test_both_blocks_render_in_one_sheet(self):
        """重合度 + 相关性同时提供 → 一章两区块同页呈现（章节标题带序号 7）。"""
        ws = self._write(_overlap_result(), _correlation_data())
        flat = self._flat(ws)
        self.assertTrue(any("7. 持仓关系矩阵" in v for v in flat), f"应含章节标题（序号 7），实际: {flat[:3]}")
        self.assertTrue(any("一、持仓重合度矩阵" in v for v in flat), "应含重合度区块标题")
        self.assertTrue(any("二、持仓相关性矩阵" in v for v in flat), "应含相关性区块标题")
        self.assertTrue(any("基金A" in v for v in flat), "重合度区块应含基金名")
        self.assertTrue(any("资产A" in v for v in flat), "相关性区块应含资产名")
        # 区块顺序：重合度在上、相关性在下
        overlap_idx = next(i for i, v in enumerate(flat) if "一、持仓重合度矩阵" in v)
        corr_idx = next(i for i, v in enumerate(flat) if "二、持仓相关性矩阵" in v)
        self.assertLess(overlap_idx, corr_idx, "重合度区块应位于相关性区块之前")

    def test_overlap_placeholder_when_correlation_only(self):
        """仅提供相关性数据 → 重合度区块写占位（相关度矩阵照常呈现）。"""
        ws = self._write(None, _correlation_data())
        flat = self._flat(ws)
        self.assertTrue(any("无法计算重合度" in v for v in flat), "重合度区块应写占位")
        self.assertTrue(any("资产B" in v for v in flat), "相关性矩阵应正常呈现")

    def test_correlation_placeholder_when_overlap_only(self):
        """仅提供重合度数据 → 相关性区块写占位（重合度矩阵照常呈现）。"""
        ws = self._write(_overlap_result(), None)
        flat = self._flat(ws)
        self.assertTrue(any("基金A" in v for v in flat), "重合度矩阵应正常呈现")
        self.assertTrue(any("持仓相关性数据暂不可用" in v for v in flat), "相关性区块应写占位")

    def test_overlap_jaccard_value_rendered(self):
        """重合度区块 Jaccard 系数值以百分比呈现。"""
        ws = self._write(_overlap_result(), _correlation_data())
        flat = self._flat(ws)
        self.assertTrue(any("50.00%" in v for v in flat), "重合度 Jaccard 0.5 应呈现为 50.00%")
        self.assertTrue(any("100.00%" in v for v in flat), "对角线 1.0 应呈现为 100.00%")
