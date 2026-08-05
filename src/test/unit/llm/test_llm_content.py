"""LLM 内容 Excel 写入模块单元测试。

测试目标：
  - _strip_html — HTML 标签剥离
  - _write_content_sheet — 段落分行、wrap_text、行高、列宽、footer 提取
  - write_llm_sheets — 多页签写入 + 禁用模块处理

运行：
  pytest src/test/unit/llm/test_llm_content.py -v
"""

from __future__ import annotations

import unittest
from math import ceil

from openpyxl import Workbook

from src.python.report.llm_content import (
    _ROW_HEIGHT_MIN,
    _ROW_HEIGHT_PER_LINE,
    _CHARS_PER_LINE,
    _FACT_CHECK_DETAIL_FONT,
    _FACT_CHECK_WARN_FONT,
    _calc_row_height,
    _get_module_key_map,
    _get_placeholder,
    _split_html_blocks,
    _strip_html,
    _extract_footer_text,
    _write_content_sheet,
    write_llm_sheets,
)
from src.python.report.styles import CONTENT_FONT, GREEN_FONT, TITLE_FILL, TITLE_FONT
import pytest

pytestmark = [pytest.mark.unit, pytest.mark.unit_llm, pytest.mark.llm]


# ═══════════════════════════════════════════════════════════
#  _calc_row_height
# ═══════════════════════════════════════════════════════════


class TestCalcRowHeight(unittest.TestCase):
    """测试行高估算函数。"""

    def test_empty(self) -> None:
        self.assertEqual(_calc_row_height(""), _ROW_HEIGHT_MIN)

    def test_short_text(self) -> None:
        """短文本 → 最小行高。"""
        text = "短文本"
        self.assertEqual(_calc_row_height(text), _ROW_HEIGHT_MIN)

    def test_long_text(self) -> None:
        """超长文本 → 行高按行数倍乘。"""
        text = "中" * 120  # 120 中文字符 ≈ 3 行
        expected = ceil(120 / _CHARS_PER_LINE) * _ROW_HEIGHT_PER_LINE
        self.assertEqual(_calc_row_height(text), expected)

    def test_boundary(self) -> None:
        """正好 _CHARS_PER_LINE 个字符 → 最小行高。"""
        text = "中" * _CHARS_PER_LINE
        self.assertEqual(_calc_row_height(text), _ROW_HEIGHT_MIN)

    def test_multiline_uses_newline_count(self) -> None:
        """含 \\n 的多行块 → 行高按换行数而非字符宽度。"""
        # "A\nB\nC\nD\nE" 5 行短文本，按字符宽度仅算 1 行 → 取 5 行
        text = "A\nB\nC\nD\nE"
        expected = 5 * _ROW_HEIGHT_PER_LINE
        self.assertEqual(_calc_row_height(text), expected)

    def test_multiline_takes_max_of_both(self) -> None:
        """多行 + 长文本 → 取换行数与字符宽度的较大者。"""
        text = "中" * 120 + "\n" + "中" * 120  # 2 行，每行 3 个宽度行
        self.assertEqual(_calc_row_height(text), 6 * _ROW_HEIGHT_PER_LINE)


# ═══════════════════════════════════════════════════════════
#  _strip_html
# ═══════════════════════════════════════════════════════════


class TestStripHtml(unittest.TestCase):
    """测试 HTML 标签剥离。"""

    def test_plain_text(self) -> None:
        self.assertEqual(_strip_html("hello"), "hello")

    def test_strip_p_tag(self) -> None:
        self.assertEqual(_strip_html("<p>内容</p>"), "内容")

    def test_strip_strong(self) -> None:
        self.assertEqual(_strip_html("<strong>粗体</strong>"), "粗体")

    def test_strip_nested(self) -> None:
        html = "<p>这是 <strong>粗体</strong> 和 <em>斜体</em></p>"
        self.assertEqual(_strip_html(html), "这是 粗体 和 斜体")

    def test_html_entity_unchanged(self) -> None:
        """HTML 实体（如 &amp;）保留原样——已知行为。"""
        self.assertEqual(_strip_html("A &amp; B"), "A &amp; B")

    def test_none_input(self) -> None:
        self.assertEqual(_strip_html(None), "")

    def test_empty_string(self) -> None:
        self.assertEqual(_strip_html(""), "")


# ═══════════════════════════════════════════════════════════
#  _write_content_sheet
# ═══════════════════════════════════════════════════════════


class TestWriteContentSheet(unittest.TestCase):
    """测试单个 LLM 分析章节的写入行为。"""

    def setUp(self):
        self.wb = Workbook()

    def _write_and_get_sheet(self, content=None):
        """写入内容并返回 worksheet 对象。"""
        ws = self.wb.create_sheet()
        _write_content_sheet(ws, "测试页签", content)
        return ws

    def test_title_row(self):
        """第 1 行标题 — 合并 A1:B1，使用 TITLE_FONT 和 TITLE_FILL。"""
        ws = self._write_and_get_sheet(content="一段落")
        cell = ws.cell(row=1, column=1)
        self.assertEqual(cell.value, "测试页签")
        self.assertEqual(cell.font.name, TITLE_FONT.name)
        self.assertEqual(cell.fill.fgColor.rgb, TITLE_FILL.fgColor.rgb)

    def test_single_paragraph(self):
        """单段落 → 写入第 2 行，wrap_text=True。"""
        ws = self._write_and_get_sheet(content="单段落内容")
        cell = ws.cell(row=2, column=1)
        self.assertEqual(cell.value, "单段落内容")
        self.assertTrue(cell.alignment.wrap_text)

    def test_multi_paragraphs(self):
        """三个段落 → 各一段（空行间距），无内容。"""
        content = "<p>第一段</p>\n\n<p>第二段</p>\n\n<p>第三段</p>"
        ws = self._write_and_get_sheet(content=content)

        self.assertEqual(ws.cell(row=2, column=1).value, "第一段")
        self.assertEqual(ws.cell(row=4, column=1).value, "第二段")
        self.assertEqual(ws.cell(row=6, column=1).value, "第三段")
        # 空行无值
        self.assertIsNone(ws.cell(row=3, column=1).value)
        self.assertIsNone(ws.cell(row=5, column=1).value)

    def test_content_none_placeholder(self):
        """content=None → 写入占位符。"""
        ws = self._write_and_get_sheet(content=None)
        cell = ws.cell(row=2, column=1)
        self.assertIn("LLM API Key", str(cell.value))

    def test_wrap_text_enabled(self):
        """每行内容单元格启用文本换行。"""
        ws = self._write_and_get_sheet(content="段落一\n\n段落二")
        for r in (2, 4):
            cell = ws.cell(row=r, column=1)
            self.assertTrue(cell.alignment.wrap_text, f"Row {r} wrap_text should be True")

    def test_paragraph_font(self):
        """段落使用 CONTENT_FONT。"""
        ws = self._write_and_get_sheet(content="测试内容")
        cell = ws.cell(row=2, column=1)
        self.assertEqual(cell.font.size, CONTENT_FONT.size)
        self.assertEqual(cell.font.color.rgb, CONTENT_FONT.color.rgb)

    def test_row_height_short(self):
        """短段落行高 = 最小行高。"""
        ws = self._write_and_get_sheet(content="短")
        self.assertGreaterEqual(ws.row_dimensions[2].height, _ROW_HEIGHT_MIN)

    def test_column_width_fixed(self):
        """A 列宽度固定。"""
        ws = self._write_and_get_sheet(content="测试内容")
        # 默认 openpyxl 宽度约 8.43，此处应被设为固定值 80
        width = ws.column_dimensions["A"].width
        self.assertIsNotNone(width)
        self.assertEqual(width, 80)

    def test_freeze_panes(self):
        """冻结首行。"""
        ws = self._write_and_get_sheet(content="测试")
        self.assertEqual(ws.freeze_panes, "A2")


# ═══════════════════════════════════════════════════════════
#  _get_module_key_map
# ═══════════════════════════════════════════════════════════


class TestGetModuleKeyMap(unittest.TestCase):
    """测试 _get_module_key_map 在自定义 section_order 下的行为。"""

    _CUSTOM_ORDER = [
        {"key": "global_macro", "name": "全球局势", "number": 1},
        {"key": "expert_review", "name": "专家复盘", "number": 2},
        {"key": "news_correlation", "name": "新闻关联", "number": 3},
        {"key": "health_check", "name": "持仓体检", "number": 4},
        {"key": "penetration_deep", "name": "穿透分析", "number": 5},
        {"key": "llm_usage", "name": "LLM 用量", "number": 6},
    ]

    def test_custom_order_maps_correctly(self):
        """自定义 section_order → 映射使用配置序号，排除 news_correlation/llm_usage。"""
        result = _get_module_key_map(self._CUSTOM_ORDER)
        expected = {
            "1.全球局势": "global_macro",
            "2.专家复盘": "expert_review",
            "4.持仓体检": "health_check",
            "5.穿透分析": "penetration_deep",
        }
        self.assertEqual(result, expected)

    def test_skips_news_correlation_and_llm_usage(self):
        """news_correlation 和 llm_usage 被排除。"""
        result = _get_module_key_map(self._CUSTOM_ORDER)
        self.assertNotIn("3.新闻关联", result)
        self.assertNotIn("6.LLM 用量", result)

    def test_none_falls_back_to_default(self):
        """section_order=None → 使用默认全局顺序（排除 news_correlation/llm_usage）。"""
        result = _get_module_key_map(None)
        self.assertIn("global_macro", result.values())
        self.assertNotIn("news_correlation", result.values())
        self.assertNotIn("llm_usage", result.values())
        # 默认模块排除 2 个 = total - 2
        from src.python.core.registry import _REPORT_SECTION_DEFAULT

        self.assertEqual(len(result), len(_REPORT_SECTION_DEFAULT) - 2)

    def test_empty_list_falls_back_to_default(self):
        """空列表（falsy）→ 回退到默认全局顺序。"""
        result = _get_module_key_map([])
        # 空列表 falsy → section_order or get_report_section_order() → 默认值
        self.assertIn("summary", result.values())
        self.assertGreater(len(result), 4)


# ═══════════════════════════════════════════════════════════
#  _get_placeholder — failure reason lookup
# ═══════════════════════════════════════════════════════════


class TestGetPlaceholder(unittest.TestCase):
    """测试 _get_placeholder 在自定义 section_order 下的占位符查找。"""

    _CUSTOM_ORDER = [
        {"key": "global_macro", "name": "全球局势", "number": 1, "type": "llm"},
        {"key": "expert_review", "name": "专家复盘", "number": 2, "type": "llm"},
    ]

    def test_custom_order_finds_placeholder(self):
        """自定义 section_order → 占位符按配置序号查找。"""
        from src.python.llm.prompts import LLM_MODULE_FAILURE, FAIL_REASON_NOT_CONFIGURED

        LLM_MODULE_FAILURE.clear()
        LLM_MODULE_FAILURE["global_macro"] = FAIL_REASON_NOT_CONFIGURED
        try:
            text = _get_placeholder("1.全球局势", self._CUSTOM_ORDER)
            self.assertIn("LLM 未配置", text)
        finally:
            LLM_MODULE_FAILURE.clear()

    def test_unknown_title_returns_generic(self):
        """未知标题 → 通用占位符文本。"""
        text = _get_placeholder("99.不存在的页签", self._CUSTOM_ORDER)
        self.assertIn("LLM API Key", text)

    def test_none_falls_back_to_default_order(self):
        """section_order=None → 使用默认顺序查找。"""
        text = _get_placeholder("12.全球政经局势", None)
        self.assertIn("LLM API Key", text)  # 默认状态未设置失败原因


# ═══════════════════════════════════════════════════════════
#  write_llm_sheets — custom section_order
# ═══════════════════════════════════════════════════════════


class TestWriteLlmSheetsCustomOrder(unittest.TestCase):
    """测试 write_llm_sheets 在自定义 section_order 下的标题行。"""

    _CUSTOM_ORDER = [
        {"key": "fund_performance", "name": "基金业绩", "number": 1, "type": "always"},
        {"key": "global_macro", "name": "全球局势", "number": 2, "type": "llm"},
        {"key": "expert_review", "name": "专家复盘", "number": 3, "type": "llm"},
        {"key": "health_check", "name": "持仓体检", "number": 4, "type": "llm"},
        {"key": "penetration_deep", "name": "穿透分析", "number": 5, "type": "llm"},
        {"key": "summary", "name": "投资汇总", "number": 6, "type": "always"},
    ]
    _LLM_KEYS = ["global_macro", "expert_review", "health_check", "penetration_deep"]
    _CUSTOM_TITLES = ["2.全球局势", "3.专家复盘", "4.持仓体检", "5.穿透分析"]

    def setUp(self):
        self.wb = Workbook()
        self.sheets = {}
        for key, title in zip(self._LLM_KEYS, self._CUSTOM_TITLES):
            ws = self.wb.create_sheet(title=title)
            self.sheets[key] = ws

    def test_title_row_uses_custom_numbers(self):
        """自定义 section_order → 标题行使用配置序号。"""
        write_llm_sheets(
            self.sheets, ("<p>宏观</p>", "<p>复盘</p>", "<p>体检</p>", "<p>穿透</p>"), section_order=self._CUSTOM_ORDER
        )
        for key, expected in zip(self._LLM_KEYS, self._CUSTOM_TITLES):
            cell = self.sheets[key].cell(row=1, column=1)
            self.assertEqual(cell.value, expected, f"{key} 标题行应为 {expected!r}")

    def test_content_none_uses_custom_order_placeholder(self):
        """自定义 section_order + content=None → 占位符正确。"""
        write_llm_sheets(self.sheets, (None, None, None, None), section_order=self._CUSTOM_ORDER)
        for ws in self.sheets.values():
            cell = ws.cell(row=2, column=1)
            self.assertIn("LLM API Key", str(cell.value or ""))


# ═══════════════════════════════════════════════════════════
#  write_llm_sheets
# ═══════════════════════════════════════════════════════════


class TestWriteLlmSheets(unittest.TestCase):
    """测试 write_llm_sheets 集成行为。"""

    _LLM_KEYS = ["global_macro", "expert_review", "health_check", "penetration_deep"]
    _LLM_TITLES = ["12.全球政经局势", "13.智囊团深度复盘", "14.持仓体检报告", "15.穿透深度分析"]

    def setUp(self):
        self.wb = Workbook()
        self.sheets = {}
        for key, title in zip(self._LLM_KEYS, self._LLM_TITLES):
            ws = self.wb.create_sheet(title=title)
            self.sheets[key] = ws

    def test_sheet_titles(self):
        """四个 sheet 标题正确。"""
        write_llm_sheets(self.sheets, ("<p>宏观</p>", "<p>复盘</p>", None, None))
        sheet_names = [ws.title for ws in self.wb.worksheets]
        for title in self._LLM_TITLES:
            self.assertIn(title, sheet_names)

    def test_return_text_quad(self):
        """返回 (text7, text8, text9, textA) 纯文本四元组。"""
        text7, text8, text9, textA = write_llm_sheets(
            self.sheets, ("<p>全球政经局势</p>", "<p>复盘内容</p>", "<p>持仓体检报告</p>", "<p>穿透深度分析</p>")
        )
        self.assertEqual(text7, "全球政经局势")
        self.assertEqual(text8, "复盘内容")
        self.assertEqual(text9, "持仓体检报告")
        self.assertEqual(textA, "穿透深度分析")

    def test_content_none(self):
        """content=(None, None, None, None) → 占位符，不崩溃。"""
        text7, text8, text9, textA = write_llm_sheets(self.sheets, (None, None, None, None))
        self.assertEqual(text7, "")
        self.assertEqual(text8, "")
        self.assertEqual(text9, "")
        self.assertEqual(textA, "")
        ws_list = [self.sheets[k] for k in self._LLM_KEYS]
        self.assertIn("LLM API Key", str(ws_list[0].cell(row=2, column=1).value or ""))
        self.assertIn("LLM API Key", str(ws_list[1].cell(row=2, column=1).value or ""))


class TestLlmModuleFailureReset(unittest.TestCase):
    """LLM_MODULE_FAILURE 跨测试残留由 conftest autouse fixture 自动清理。

    write_llm_sheets() 读取模块级全局 LLM_MODULE_FAILURE 判断模块是否被禁用。
    若某测试设置 LLM_MODULE_FAILURE[key]=FAIL_REASON_DISABLED 后未清理，
    后续测试的页签会被跳过不写入，占位符断言失败。
    conftest.py 的 _auto_reset_llm_module_failure autouse fixture 负责清除
    已污染的状态，本用例验证该清理逻辑。
    """

    def test_autouse_fixture_clears_polluted_state(self):
        """模拟上一测试残留的禁用状态，验证 autouse fixture 的清理逻辑。"""
        from src.python.llm.prompts import FAIL_REASON_DISABLED, LLM_MODULE_FAILURE

        LLM_MODULE_FAILURE["global_macro"] = FAIL_REASON_DISABLED
        LLM_MODULE_FAILURE["health_check"] = FAIL_REASON_DISABLED
        self.assertNotEqual(LLM_MODULE_FAILURE, {})

        # 复现 autouse fixture 的执行时机（__wrapped__ 取 fixture 底层函数）
        from src.test.conftest import _auto_reset_llm_module_failure

        _auto_reset_llm_module_failure.__wrapped__()

        self.assertEqual(LLM_MODULE_FAILURE, {})


class TestWriteLlmSheetsDisabled(unittest.TestCase):
    """测试 write_llm_sheets 在模块禁用时的行为。"""

    _LLM_KEYS = ["global_macro", "expert_review", "health_check", "penetration_deep"]
    _LLM_TITLES = ["12.全球政经局势", "13.智囊团深度复盘", "14.持仓体检报告", "15.穿透深度分析"]

    def setUp(self):
        self.wb = Workbook()
        self.sheets = {}
        for key, title in zip(self._LLM_KEYS, self._LLM_TITLES):
            ws = self.wb.create_sheet(title=title)
            self.sheets[key] = ws

    def test_disabled_global_macro_skips_sheet(self):
        """global_macro 禁用 → 不写入该页签，其他页签正常。"""
        from src.python.llm.prompts import LLM_MODULE_FAILURE, FAIL_REASON_DISABLED

        # 清除可能残留的状态
        LLM_MODULE_FAILURE.clear()
        LLM_MODULE_FAILURE["global_macro"] = FAIL_REASON_DISABLED
        try:
            text7, text8, text9, textA = write_llm_sheets(
                self.sheets, ("<p>宏观</p>", "<p>复盘</p>", "<p>体检</p>", "<p>穿透</p>")
            )
            # 禁用模块返回空文本
            self.assertEqual(text7, "")
            self.assertNotEqual(text8, "")
            # global_macro 页签存在但未写入，其他页签正常写入
            sheet_titles = [ws.title for ws in self.wb.worksheets]
            self.assertIn("12.全球政经局势", sheet_titles)  # 页签仍存在（预创建）
            self.assertIn("13.智囊团深度复盘", sheet_titles)
            self.assertIn("14.持仓体检报告", sheet_titles)
            self.assertIn("15.穿透深度分析", sheet_titles)
        finally:
            LLM_MODULE_FAILURE.clear()

    def test_all_disabled_no_sheets_created(self):
        """所有 4 个模块都禁用 → sheets 字典中的页签不写入内容（不修改）。"""
        from src.python.llm.prompts import LLM_MODULE_FAILURE, FAIL_REASON_DISABLED

        LLM_MODULE_FAILURE.clear()
        LLM_MODULE_FAILURE.update(
            global_macro=FAIL_REASON_DISABLED,
            expert_review=FAIL_REASON_DISABLED,
            health_check=FAIL_REASON_DISABLED,
            penetration_deep=FAIL_REASON_DISABLED,
        )
        try:
            text7, text8, text9, textA = write_llm_sheets(self.sheets, (None, None, None, None))
            # 全部返回空文本
            self.assertEqual(text7, "")
            self.assertEqual(text8, "")
            self.assertEqual(text9, "")
            self.assertEqual(textA, "")
            # 页签仍存在（预创建传入），但禁用模块不写入内容
            sheet_titles = [ws.title for ws in self.wb.worksheets]
            for title in self._LLM_TITLES:
                self.assertIn(title, sheet_titles)
        finally:
            LLM_MODULE_FAILURE.clear()


# ═══════════════════════════════════════════════════════════
#  _extract_footer_text
# ═══════════════════════════════════════════════════════════


class TestExtractFooterText(unittest.TestCase):
    """测试从 HTML 中提取底部 LLM 标识行。"""

    def test_none_input(self) -> None:
        self.assertEqual(_extract_footer_text(None), "")

    def test_empty_input(self) -> None:
        self.assertEqual(_extract_footer_text(""), "")

    def test_no_footer(self) -> None:
        """无 footer 标签 → 空字符串。"""
        html = "<p>纯内容</p>\n\n<p>更多内容</p>"
        self.assertEqual(_extract_footer_text(html), "")

    def test_extract_token_footer(self) -> None:
        """提取包含模型和 Token 用量的 footer。"""
        html = (
            "<p>全球政经局势</p>\n\n"
            '<p style="color:#888;font-size:12px">'
            "模型：gpt-4o | Token 用量：输入 1,234 / 输出 567 = 1,801"
            "</p>"
        )
        self.assertEqual(
            _extract_footer_text(html),
            "模型：gpt-4o | Token 用量：输入 1,234 / 输出 567 = 1,801",
        )

    def test_extract_footer_with_cost(self) -> None:
        """提取含估算费用和 Extended Thinking 的 footer。"""
        html = (
            "<p>内容</p>\n\n"
            '<p style="color:#888;font-size:12px">'
            "模型：claude-sonnet-4 | Token 用量：输入 5,000 / 输出 1,000 = 6,000 | "
            "估算费用：$0.015 | Extended Thinking"
            "</p>"
        )
        result = _extract_footer_text(html)
        self.assertIn("模型：claude-sonnet-4", result)
        self.assertIn("Token 用量：输入 5,000 / 输出 1,000 = 6,000", result)
        self.assertIn("估算费用：$0.015", result)
        self.assertIn("Extended Thinking", result)

    def test_extract_cache_footer(self) -> None:
        """提取缓存 footer。"""
        html = '<p>内容</p>\n\n<p style="color:#888;font-size:12px">本次使用LLM缓存，未直接使用LLM服务能力</p>'
        self.assertEqual(
            _extract_footer_text(html),
            "本次使用LLM缓存，未直接使用LLM服务能力",
        )

    def test_extract_cache_footer_with_thinking(self) -> None:
        """提取含 Extended Thinking 的缓存 footer。"""
        html = (
            "<p>内容</p>\n\n"
            '<p style="color:#888;font-size:12px">'
            "本次使用LLM缓存（原始模型：claude-sonnet-4） | Extended Thinking"
            "</p>"
        )
        result = _extract_footer_text(html)
        self.assertIn("本次使用LLM缓存", result)
        self.assertIn("Extended Thinking", result)

    def test_extract_last_of_multiple_footers(self) -> None:
        """多个 <p style='color:#888;font-size:12px'> 标签 → 提取最后一条（缓存提示）。"""
        html = (
            "<p>正文内容</p>\n\n"
            '<p style="color:#888;font-size:12px">'
            "模型：DeepSeek-V4-Flash | Token 用量：输入 3,200 / 输出 1,321 = 4,521 | "
            "估算费用：¥0.0063"
            "</p>\n\n"
            '<p style="color:#888;font-size:12px">'
            "本次使用LLM缓存（原始模型：DeepSeek-V4-Flash）"
            "</p>"
        )
        result = _extract_footer_text(html)
        self.assertIn("本次使用LLM缓存", result)
        self.assertNotIn("Token 用量", result)


# ═══════════════════════════════════════════════════════════
#  _write_content_sheet — unified footer
# ═══════════════════════════════════════════════════════════


class TestWriteContentSheetUnifiedFooter(unittest.TestCase):
    """测试 _write_content_sheet 中统一 footer 的行为。"""

    def setUp(self):
        self.wb = Workbook()

    def test_footer_extracted_from_html(self) -> None:
        """HTML 含 footer → 最后一行为 footer 纯文本。"""
        html = (
            "<p>全球政经局势</p>\n\n"
            '<p style="color:#888;font-size:12px">'
            "模型：gpt-4o | Token 用量：输入 1,234 / 输出 567 = 1,801 | "
            "估算费用：$0.005 | Extended Thinking"
            "</p>"
        )
        ws = self.wb.create_sheet()
        _write_content_sheet(ws, "测试", html)
        footer_cell = ws.cell(row=4, column=1)
        self.assertIn("模型：gpt-4o", str(footer_cell.value or ""))
        self.assertIn("Token 用量", str(footer_cell.value or ""))
        self.assertIn("估算费用：$0.005", str(footer_cell.value or ""))
        self.assertIn("Extended Thinking", str(footer_cell.value or ""))

    def test_footer_not_repeated_in_content(self) -> None:
        """footer 被排除在正文之外，避免重复。"""
        html = (
            "<p>内容段落</p>\n\n"
            '<p style="color:#888;font-size:12px">'
            "模型：gpt-4o | Token 用量：输入 100 / 输出 50 = 150"
            "</p>"
        )
        ws = self.wb.create_sheet()
        _write_content_sheet(ws, "测试", html)
        content_cell = ws.cell(row=2, column=1)
        self.assertEqual(content_cell.value, "内容段落")
        self.assertNotIn("模型", str(content_cell.value or ""))

    def test_cached_with_embedded_footer_uses_it(self) -> None:
        """缓存内容含嵌入式 footer → footer 正确展示。"""
        html = '<p>缓存内容</p>\n\n<p style="color:#888;font-size:12px">本次使用LLM缓存，未直接使用LLM服务能力</p>'
        ws = self.wb.create_sheet()
        _write_content_sheet(ws, "测试", html)
        footer_cell = ws.cell(row=4, column=1)
        self.assertIn("LLM缓存", str(footer_cell.value or ""))
        # 不应再有额外行
        self.assertIsNone(ws.cell(row=5, column=1).value)

    def test_dual_footer_cache_line_shown(self) -> None:
        """模型/Token footer + 缓存 footer 并存 → 缓存 footer 为末尾标识行。"""
        html = (
            "<p>正文段落</p>\n\n"
            '<p style="color:#888;font-size:12px">'
            "模型：DeepSeek-V4-Flash | Token 用量：输入 100 / 输出 50 = 150"
            "</p>\n\n"
            '<p style="color:#888;font-size:12px">'
            "本次使用LLM缓存（原始模型：DeepSeek-V4-Flash）"
            "</p>"
        )
        ws = self.wb.create_sheet()
        _write_content_sheet(ws, "测试", html)
        # 模型/Token 行在正文中（不会被排除，因为不是最后一段）
        content_cell_after_token = ws.cell(row=4, column=1)
        self.assertIn("模型：DeepSeek-V4-Flash", str(content_cell_after_token.value or ""))
        # 缓存提示为末尾 footer
        footer_cell = ws.cell(row=6, column=1)
        self.assertIn("本次使用LLM缓存", str(footer_cell.value or ""))
        self.assertNotIn("Token 用量", str(footer_cell.value or ""))


# ═══════════════════════════════════════════════════════════
#  _split_html_blocks
# ═══════════════════════════════════════════════════════════


class TestSplitHtmlBlocks(unittest.TestCase):
    """测试按块级 HTML 元素分段。"""

    def test_blank_input(self) -> None:
        self.assertEqual(_split_html_blocks(""), [])

    def test_no_block_tags(self) -> None:
        """无块级标签 → 按 \\n\\n 分段（兼容纯文本输入）。"""
        self.assertEqual(_split_html_blocks("段落一\n\n段落二"), ["段落一", "段落二"])

    def test_no_newline_between_p(self) -> None:
        """markdown_to_html 无换行拼接 <p> 标签 → 正确切分为独立块。"""
        blocks = _split_html_blocks("<p>段一</p><p>段二</p><p>段三</p>")
        self.assertEqual(blocks, ["<p>段一</p>", "<p>段二</p>", "<p>段三</p>"])

    def test_p_newline_p(self) -> None:
        """<p> 之间含 \\n\\n → 仍按块切分。"""
        blocks = _split_html_blocks("<p>第一段</p>\n\n<p>第二段</p>")
        self.assertEqual(blocks, ["<p>第一段</p>", "<p>第二段</p>"])

    def test_orphan_span_merged_into_prev_block(self) -> None:
        """游离 <span> 附属行并入前一块（校验摘要尾部场景）。"""
        html = (
            '<p style="color:#a40;font-size:12px">[标签]事实校验：1/2 项通过，1 项提示</p>'
            '<span style="color:#888;font-size:11px">已修正明细: 5%→6%（句段）</span>'
        )
        blocks = _split_html_blocks(html)
        self.assertEqual(len(blocks), 1)
        self.assertIn("事实校验：1/2 项通过", blocks[0])
        self.assertIn("已修正明细", blocks[0])

    def test_orphan_span_without_newline_gets_separator(self) -> None:
        """游离 <span> 与前块间无换行 → 并入时补 \\n，避免剥离后粘连。"""
        html = '<p>正文段落</p><span style="color:#888;font-size:11px">已修正明细: 5%→6%</span>'
        blocks = _split_html_blocks(html)
        self.assertEqual(len(blocks), 1)
        stripped = _strip_html(blocks[0])
        self.assertEqual(stripped, "正文段落\n已修正明细: 5%→6%")

    def test_hr_separates_blocks(self) -> None:
        """<hr> 作为独立分块点。"""
        blocks = _split_html_blocks("<p>段一</p><hr><p>段二</p>")
        self.assertEqual(blocks, ["<p>段一</p>", "<hr>", "<p>段二</p>"])

    def test_ul_list_kept_as_single_block(self) -> None:
        """<ul><li>…</li></ul> 整体一块，剥离后保留列表项多行。"""
        blocks = _split_html_blocks("<ul><li>项一</li><li>项二</li></ul>")
        self.assertEqual(len(blocks), 1)
        stripped = _strip_html(blocks[0])
        self.assertIn("项一", stripped)
        self.assertIn("\n", stripped)


# ═══════════════════════════════════════════════════════════
#  _write_content_sheet — fact check summary
# ═══════════════════════════════════════════════════════════


class TestWriteContentSheetFactCheck(unittest.TestCase):
    """测试事实校验摘要块在 Excel 中的呈现（footer 去重 + 明细行灰色）。"""

    _FOOTER = "模型：claude-sonnet-4 | Token 用量：输入 100 / 输出 50 = 150"

    def setUp(self):
        self.wb = Workbook()

    def _write(self, html):
        ws = self.wb.create_sheet()
        _write_content_sheet(ws, "测试页签", html)
        return ws

    def test_footer_not_repeated_after_summary_appended(self) -> None:
        """orchestrator 以单 \\n 追加摘要后 → footer 只在末尾灰字出现一次。"""
        html = (
            "<p>段落一</p><p>段落二</p>"
            f'<p style="color:#888;font-size:12px">{self._FOOTER}</p>'
            '<p style="color:#a40;font-size:12px">[智囊团深度复盘]事实校验：3/5 项通过，2 项提示（自动修正 1 处数值）'
            "\n⚠ [智囊团深度复盘]持仓占比 25% 与真实值 20% 不符</p>"
        )
        ws = self._write(html)
        # 摘要在 起（段落一 、段落二 ）
        self.assertEqual(ws.cell(row=2, column=1).value, "段落一")
        self.assertEqual(ws.cell(row=4, column=1).value, "段落二")
        # footer 只在末尾一次，正文各单元格不含 footer 文本
        for r in (2, 4, 6, 7):
            value = str(ws.cell(row=r, column=1).value or "")
            self.assertNotIn("模型：claude-sonnet-4", value, f"Row {r} 不应含 footer")
        last_nonempty = [r for r in range(1, 30) if ws.cell(row=r, column=1).value]
        footer_row = last_nonempty[-1]
        self.assertIn("模型：claude-sonnet-4", str(ws.cell(row=footer_row, column=1).value))

    def test_warn_summary_first_line_amber_detail_gray(self) -> None:
        """告警摘要 → 首行琥珀，明细行灰色小字。"""
        html = (
            '<p style="color:#a40;font-size:12px">[智囊团深度复盘]事实校验：3/5 项通过，2 项提示（自动修正 1 处数值）'
            "\n⚠ [智囊团深度复盘]持仓占比 25% 与真实值 20% 不符"
            "\n⚠ [智囊团深度复盘]某基金 60% 与真实值 55% 不符</p>"
        )
        ws = self._write(html)
        first = ws.cell(row=2, column=1)
        detail = ws.cell(row=3, column=1)
        self.assertIn("事实校验：3/5 项通过", str(first.value))
        self.assertEqual(first.font.color.rgb, _FACT_CHECK_WARN_FONT.color.rgb)
        self.assertIn("持仓占比 25%", str(detail.value))
        self.assertEqual(detail.font.color.rgb, _FACT_CHECK_DETAIL_FONT.color.rgb)
        # 第二明细行继续灰色
        self.assertEqual(ws.cell(row=4, column=1).font.color.rgb, _FACT_CHECK_DETAIL_FONT.color.rgb)

    def test_pass_summary_green_corrections_gray(self) -> None:
        """通过摘要（含自动修正）→ 首行绿色，已修正明细灰色。"""
        html = (
            '<p style="color:#4a4;font-size:12px">[智囊团深度复盘]✓ 事实校验通过：21/21 项检查全部通过（自动修正 1 处数值）</p>'
            '\n<span style="color:#888;font-size:11px">已修正明细: 5.0%→30.3%（某某句段）</span>'
        )
        ws = self._write(html)
        first = ws.cell(row=2, column=1)
        detail = ws.cell(row=3, column=1)
        self.assertIn("✓ 事实校验通过", str(first.value))
        self.assertEqual(first.font.color.rgb, GREEN_FONT.color.rgb)
        self.assertIn("已修正明细: 5.0%→30.3%", str(detail.value))
        self.assertEqual(detail.font.color.rgb, _FACT_CHECK_DETAIL_FONT.color.rgb)

    def test_no_newline_p_tags_split_into_rows(self) -> None:
        """markdown_to_html 无换行多段 → 每段独立单元格（修复整模块坍缩）。"""
        html = "<p>段一</p><p>段二</p><p>段三</p>"
        ws = self._write(html)
        self.assertEqual(ws.cell(row=2, column=1).value, "段一")
        self.assertEqual(ws.cell(row=4, column=1).value, "段二")
        self.assertEqual(ws.cell(row=6, column=1).value, "段三")

    def test_ul_list_row_height_sufficient(self) -> None:
        """<ul> 列表块剥离后多行 → 行高按换行数，避免内容截断。"""
        html = "<ul><li>项一</li><li>项二</li><li>项三</li></ul>"
        ws = self._write(html)
        cell = ws.cell(row=2, column=1)
        self.assertIn("项一", str(cell.value))
        self.assertIn("项二", str(cell.value))
        # 3 行文本 → 行高 >= 3 * _ROW_HEIGHT_PER_LINE
        self.assertGreaterEqual(ws.row_dimensions[2].height, 3 * _ROW_HEIGHT_PER_LINE)


if __name__ == "__main__":
    unittest.main()
