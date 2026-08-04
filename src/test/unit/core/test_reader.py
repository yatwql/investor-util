"""持仓读取模块单元测试。

测试目标：
  - list_xlsx_files — 目录遍历/排序/临时文件跳过
  - get_xlsx_info — xlsx 元信息读取（账号数量、数据行数）
  - read_holdings — 正常解析/多账户/表头校验/空行跳过/数值校验/格式异常
  - _safe_str / _safe_float / _match_header — 辅助函数

运行：
  cd D:/codebase/zoo/investor-util
  python -m unittest src.test_reader -v
"""

from __future__ import annotations

import os
import tempfile
import time
import zipfile
import unittest
from typing import Any, List
from unittest.mock import MagicMock, patch

from openpyxl.utils.exceptions import InvalidFileException

from src.python.core import reader
from src.python.core.models import Holding
import pytest

pytestmark = [pytest.mark.unit, pytest.mark.unit_core]


# ═════════════════════════════════════════════════════════════
#  辅助函数测试
# ═════════════════════════════════════════════════════════════


class TestSafeStr(unittest.TestCase):
    """测试 _safe_str 转换逻辑。"""

    @pytest.mark.smoke
    def test_none_returns_empty(self):
        """None -> 空字符串。"""
        self.assertEqual(reader._safe_str(None), "")

    @pytest.mark.smoke
    def test_string_stripped(self):
        """字符串 -> 去除前后空格。"""
        self.assertEqual(reader._safe_str("  贵州茅台  "), "贵州茅台")

    @pytest.mark.smoke
    def test_number_to_string(self):
        """数字 -> 字符串。"""
        self.assertEqual(reader._safe_str(513300), "513300")

    def test_float_to_string(self):
        """浮点数 -> 字符串。"""
        self.assertEqual(reader._safe_str(100.5), "100.5")

    @pytest.mark.smoke
    def test_empty_string(self):
        """空字符串 -> 空字符串。"""
        self.assertEqual(reader._safe_str(""), "")

    def test_whitespace_only(self):
        """纯空格 -> 空字符串。"""
        self.assertEqual(reader._safe_str("   "), "")


class TestSafeFloat(unittest.TestCase):
    """测试 _safe_float 转换逻辑。"""

    def test_normal_float(self):
        """正常浮点数 -> float。"""
        self.assertEqual(reader._safe_float(100.5, "份额", "测试表", 2), 100.5)

    def test_integer(self):
        """整数 -> float。"""
        self.assertEqual(reader._safe_float(100, "份额", "测试表", 2), 100.0)

    def test_string_numeric(self):
        """数字字符串 -> float。"""
        self.assertEqual(reader._safe_float("100.5", "份额", "测试表", 2), 100.5)

    def test_none_returns_none(self):
        """None -> 返回 None 并记录警告。"""
        with self.assertLogs("invest", level="WARNING") as log:
            result = reader._safe_float(None, "份额", "测试表", 2)
        self.assertIsNone(result)
        self.assertIn("第 2 行 '份额' 为空", log.output[0])

    def test_invalid_string_returns_none(self):
        """不可解析的字符串 -> 返回 None 并记录警告。"""
        with self.assertLogs("invest", level="WARNING") as log:
            result = reader._safe_float("abc", "份额", "测试表", 2)
        self.assertIsNone(result)
        self.assertIn("第 2 行 '份额' 无法解析 (abc)", log.output[0])

    def test_boolean_true(self):
        """布尔值 True -> 1.0。"""
        self.assertEqual(reader._safe_float(True, "份额", "测试表", 2), 1.0)

    def test_boolean_false(self):
        """布尔值 False -> 0.0。"""
        self.assertEqual(reader._safe_float(False, "份额", "测试表", 2), 0.0)


class TestMatchHeader(unittest.TestCase):
    """测试 _match_header 表头匹配逻辑。"""

    _expected = ["名称", "代码", "持仓份额", "每份成本"]

    def test_exact_match(self):
        """完全匹配 -> True。"""
        self.assertTrue(reader._match_header(self._expected, self._expected))

    def test_extra_columns(self):
        """比预期多列 -> True。"""
        actual = ["名称", "代码", "持仓份额", "每份成本", "备注", "市值"]
        self.assertTrue(reader._match_header(actual, self._expected))

    def test_fewer_columns(self):
        """比预期少列 -> False。"""
        actual = ["名称", "代码"]
        self.assertFalse(reader._match_header(actual, self._expected))

    def test_first_column_mismatch(self):
        """首列不匹配 -> False。"""
        actual = ["股票名称", "代码", "持仓份额", "每份成本"]
        self.assertFalse(reader._match_header(actual, self._expected))

    def test_middle_column_mismatch(self):
        """中间列不匹配 -> False。"""
        actual = ["名称", "证券代码", "持仓份额", "每份成本"]
        self.assertFalse(reader._match_header(actual, self._expected))

    def test_empty_actual(self):
        """空表头 -> False。"""
        self.assertFalse(reader._match_header([], self._expected))

    def test_whitespace_in_header(self):
        """表头含前后空格（实际不会被 _safe_str 传入，但防御性验证）。"""
        actual = ["  名称  ", "代码", "持仓份额", "每份成本"]
        self.assertFalse(reader._match_header(actual, self._expected))


# ═════════════════════════════════════════════════════════════
#  list_xlsx_files 测试
# ═════════════════════════════════════════════════════════════


class TestListXlsxFiles(unittest.TestCase):
    """测试 list_xlsx_files 目录遍历与排序。"""

    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.dir_path = self.tmpdir.name

    def tearDown(self):
        self.tmpdir.cleanup()

    def test_directory_not_exist(self):
        """不存在的目录 -> 返回空列表。"""
        result = reader.list_xlsx_files("D:/nonexistent_dir_12345_test")
        self.assertEqual(result, [])

    def test_directory_empty(self):
        """空目录 -> 返回空列表。"""
        result = reader.list_xlsx_files(self.dir_path)
        self.assertEqual(result, [])

    def test_only_xlsx_files(self):
        """目录中只有 .xlsx 文件 -> 返回这些文件。"""
        f1 = os.path.join(self.dir_path, "a.xlsx")
        f2 = os.path.join(self.dir_path, "b.xlsx")
        open(f1, "w").close()
        open(f2, "w").close()
        result = reader.list_xlsx_files(self.dir_path)
        self.assertEqual(len(result), 2)
        self.assertIn(f1, result)
        self.assertIn(f2, result)

    def test_skip_non_xlsx(self):
        """非 .xlsx 文件被过滤。"""
        open(os.path.join(self.dir_path, "a.xlsx"), "w").close()
        open(os.path.join(self.dir_path, "b.csv"), "w").close()
        open(os.path.join(self.dir_path, "c.txt"), "w").close()
        result = reader.list_xlsx_files(self.dir_path)
        self.assertEqual(len(result), 1)
        self.assertTrue(result[0].endswith("a.xlsx"))

    def test_skip_temp_files(self):
        """跳过 ~$ 开头的临时文件。"""
        open(os.path.join(self.dir_path, "~$工作簿1.xlsx"), "w").close()
        open(os.path.join(self.dir_path, "工作簿1.xlsx"), "w").close()
        result = reader.list_xlsx_files(self.dir_path)
        self.assertEqual(len(result), 1)
        self.assertTrue(result[0].endswith("工作簿1.xlsx"))

    def test_all_temp_files_skipped(self):
        """全是 ~$ 临时文件 -> 返回空列表。"""
        open(os.path.join(self.dir_path, "~$a.xlsx"), "w").close()
        open(os.path.join(self.dir_path, "~$b.xlsx"), "w").close()
        result = reader.list_xlsx_files(self.dir_path)
        self.assertEqual(result, [])

    def test_case_insensitive_extension(self):
        """不区分 .xlsx 大小写。"""
        open(os.path.join(self.dir_path, "a.XLSX"), "w").close()
        open(os.path.join(self.dir_path, "b.Xlsx"), "w").close()
        open(os.path.join(self.dir_path, "c.xlsX"), "w").close()
        result = reader.list_xlsx_files(self.dir_path)
        self.assertEqual(len(result), 3)

    def test_sorted_by_mtime_desc(self):
        """按修改时间降序排列。"""
        f1 = os.path.join(self.dir_path, "older.xlsx")
        f2 = os.path.join(self.dir_path, "newer.xlsx")
        open(f1, "w").close()
        open(f2, "w").close()
        now = time.time()
        os.utime(f1, (now, now - 100))
        os.utime(f2, (now, now))
        result = reader.list_xlsx_files(self.dir_path)
        self.assertEqual(result[0], f2)
        self.assertEqual(result[1], f1)

    def test_absolute_paths_returned(self):
        """返回绝对路径。"""
        open(os.path.join(self.dir_path, "a.xlsx"), "w").close()
        result = reader.list_xlsx_files(self.dir_path)
        self.assertTrue(os.path.isabs(result[0]))

    def test_mixed_files_with_xlsx(self):
        """混合各种类型的文件，只保留 .xlsx（不含 ~$）。"""
        for name in ["a.xlsx", "b.xlsx", "c.csv", "d.txt", "~$e.xlsx", "f.xlsx"]:
            open(os.path.join(self.dir_path, name), "w").close()
        result = reader.list_xlsx_files(self.dir_path)
        basenames = {os.path.basename(p) for p in result}
        self.assertEqual(basenames, {"a.xlsx", "b.xlsx", "f.xlsx"})


# ═════════════════════════════════════════════════════════════
#  get_xlsx_info 测试
# ═════════════════════════════════════════════════════════════


class TestGetXlsxInfo(unittest.TestCase):
    """测试 get_xlsx_info 元信息读取。"""

    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_single_sheet(self, mock_load):
        """单个工作表 -> 返回正确元信息。"""
        mock_wb = MagicMock()
        mock_wb.sheetnames = ["证券账户"]
        mock_ws = MagicMock()
        mock_ws.max_row = 10
        mock_wb.__getitem__.return_value = mock_ws
        mock_load.return_value = mock_wb

        result = reader.get_xlsx_info("dummy.xlsx")
        self.assertEqual(result["sheet_names"], ["证券账户"])
        self.assertEqual(result["accounts"], 1)
        self.assertEqual(result["total_rows"], 9)

    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_multiple_sheets(self, mock_load):
        """多个工作表 -> 汇总所有账号行数。"""
        mock_wb = MagicMock()
        mock_wb.sheetnames = ["账户A", "账户B", "账户C"]

        def getitem(name):
            ws = MagicMock()
            ws.max_row = {"账户A": 10, "账户B": 5, "账户C": 2}[name]
            return ws

        mock_wb.__getitem__.side_effect = getitem
        mock_load.return_value = mock_wb

        result = reader.get_xlsx_info("dummy.xlsx")
        self.assertEqual(result["accounts"], 3)
        self.assertEqual(result["total_rows"], 14)

    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_sheet_with_only_header(self, mock_load):
        """只有标题行无数据 -> total_rows 为 0。"""
        mock_wb = MagicMock()
        mock_wb.sheetnames = ["空白表"]
        mock_ws = MagicMock()
        mock_ws.max_row = 1
        mock_wb.__getitem__.return_value = mock_ws
        mock_load.return_value = mock_wb

        result = reader.get_xlsx_info("dummy.xlsx")
        self.assertEqual(result["total_rows"], 0)

    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_sheet_with_none_max_row(self, mock_load):
        """max_row 为 None -> 按 1 处理，total_rows 为 0。"""
        mock_wb = MagicMock()
        mock_wb.sheetnames = ["空表"]
        mock_ws = MagicMock()
        mock_ws.max_row = None
        mock_wb.__getitem__.return_value = mock_ws
        mock_load.return_value = mock_wb

        result = reader.get_xlsx_info("dummy.xlsx")
        self.assertEqual(result["total_rows"], 0)

    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_invalid_file_returns_error_dict(self, mock_load):
        """无法读取的文件 -> 返回含 error 字段的字典。"""
        mock_load.side_effect = OSError("无法打开文件")
        result = reader.get_xlsx_info("dummy.xlsx")
        self.assertIn("error", result)
        self.assertEqual(result["error"], "无法打开文件")

    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_workbook_closed(self, mock_load):
        """get_xlsx_info 结束后应当 close workbook。"""
        mock_wb = MagicMock()
        mock_wb.sheetnames = ["证券账户"]
        mock_ws = MagicMock()
        mock_ws.max_row = 5
        mock_wb.__getitem__.return_value = mock_ws
        mock_load.return_value = mock_wb

        reader.get_xlsx_info("dummy.xlsx")
        mock_wb.close.assert_called_once()


# ═════════════════════════════════════════════════════════════
#  _parse_workbook 测试（直接测试内部解析逻辑）
# ═════════════════════════════════════════════════════════════


class TestParseWorkbook(unittest.TestCase):
    """测试 _parse_workbook 的工作表解析逻辑。"""

    def _make_cell(self, value):
        """创建 mock cell，支持 .value 属性。"""
        cell = MagicMock()
        cell.value = value
        return cell

    def _make_worksheet(self, header: List[str], data_rows: List[List[Any]], max_row: int | None = None):
        """创建 mock worksheet，模拟 iter_rows 行为。"""
        ws = MagicMock()
        ws.max_row = max_row if max_row is not None else (len(data_rows) + 1)
        header_cells = [self._make_cell(v) for v in header]

        def iter_rows_side_effect(min_row=1, max_row=None, values_only=False):
            if min_row == 1 and not values_only:
                return iter([header_cells])
            elif min_row >= 2 and values_only:
                return iter(data_rows)
            return iter([])

        ws.iter_rows.side_effect = iter_rows_side_effect
        return ws

    def _make_workbook(self, sheets: dict):
        """创建 mock workbook。

        sheets: {name: (header, data_rows, max_row?)}
        """
        wb = MagicMock()
        wb.sheetnames = list(sheets.keys())

        sheet_map = {}
        for name, params in sheets.items():
            header = params[0]
            data = params[1]
            max_row = params[2] if len(params) > 2 else None
            sheet_map[name] = self._make_worksheet(header, data, max_row)

        wb.__getitem__.side_effect = sheet_map.get
        return wb

    def test_single_sheet_valid(self):
        """单个工作表中有效数据行 -> 正确解析。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [
            ["长江电力", "600900", 200, 50.0],
            ["贵州茅台", "600519", 10, 2000.0],
        ]
        wb = self._make_workbook({"证券账户": (header, data)})
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 2)
        self.assertEqual(holdings[0].account, "证券账户")
        self.assertEqual(holdings[0].name, "长江电力")
        self.assertEqual(holdings[0].code, "600900")
        self.assertAlmostEqual(holdings[0].shares, 200.0)
        self.assertAlmostEqual(holdings[0].cost_price, 50.0)
        self.assertEqual(holdings[1].name, "贵州茅台")

    def test_multiple_accounts(self):
        """多个账户 -> 分别解析并合并。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        sheets = {
            "证券账户": (header, [["长江电力", "600900", 200, 50.0]]),
            "支付宝": (header, [["某混合基金", "001234", 1000, 1.0]]),
        }
        wb = self._make_workbook(sheets)
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 2)
        accounts = {h.account for h in holdings}
        self.assertEqual(accounts, {"证券账户", "支付宝"})

    def test_empty_sheet_skipped(self):
        """max_row < 2 的空表 -> 跳过。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [["长江电力", "600900", 200, 50.0]]
        wb = self._make_workbook(
            {
                "证券账户": (header, data),
                "空表": (header, [], 1),
            }
        )
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 1)

    def test_sheet_with_none_max_row_skipped(self):
        """max_row 为 None 的工作表 -> 跳过。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [["长江电力", "600900", 200, 50.0]]
        wb = self._make_workbook(
            {
                "证券账户": (header, data),
                "空表": (header, [], None),
            }
        )
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 1)

    def test_header_mismatch_skipped(self):
        """表头不匹配 -> 整表跳过。"""
        header_ok = ["名称", "代码", "持仓份额", "每份成本"]
        header_bad = ["股票名称", "代码", "持仓份额", "每份成本"]
        sheets = {
            "证券账户": (header_ok, [["长江电力", "600900", 200, 50.0]]),
            "支付宝": (header_bad, [["某基金", "001234", 1000, 1.0]]),
        }
        wb = self._make_workbook(sheets)
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 1)
        self.assertEqual(holdings[0].account, "证券账户")

    def test_header_fewer_columns_skipped(self):
        """表头列数不足 -> 跳过。"""
        header_ok = ["名称", "代码", "持仓份额", "每份成本"]
        header_short = ["名称", "代码"]
        sheets = {
            "证券账户": (header_ok, [["长江电力", "600900", 200, 50.0]]),
            "短表": (header_short, [["某基金", "001234"]]),
        }
        wb = self._make_workbook(sheets)
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 1)

    def test_extra_columns_allowed(self):
        """多列（超出预期）-> 仍可解析。"""
        header = ["名称", "代码", "持仓份额", "每份成本", "备注"]
        data = [["长江电力", "600900", 200, 50.0, "测试备注"]]
        wb = self._make_workbook({"证券账户": (header, data)})
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 1)

    def test_skip_empty_data_rows(self):
        """全空数据行 -> 跳过。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [
            ["长江电力", "600900", 200, 50.0],
            [None, None, None, None],
            ["贵州茅台", "600519", 10, 2000.0],
        ]
        wb = self._make_workbook({"证券账户": (header, data)})
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 2)

    def test_skip_row_with_empty_name(self):
        """名称为空的行 -> 跳过。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [
            ["长江电力", "600900", 200, 50.0],
            ["", "000001", 100, 10.0],
            [None, "000002", 100, 10.0],
        ]
        wb = self._make_workbook({"证券账户": (header, data)})
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 1)

    def test_skip_invalid_numeric_values(self):
        """数值格式无效 -> 跳过并记录警告。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [
            ["长江电力", "600900", "not_a_number", 50.0],
            ["贵州茅台", "600519", 10, "invalid"],
        ]
        wb = self._make_workbook({"证券账户": (header, data)})
        with self.assertLogs("invest", level="WARNING") as log:
            holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 0)
        self.assertTrue(any("份额" in msg for msg in log.output))
        self.assertTrue(any("成本" in msg for msg in log.output))

    def test_skip_zero_negative_values(self):
        """份额 <= 0 或成本 < 0 -> 跳过并记录警告；成本为 0 时视为有效（零成本获赠/继承）。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [
            ["长江电力", "600900", 0, 50.0],
            ["贵州茅台", "600519", -10, 2000.0],
            ["某基金", "001234", 1000, 0],
            ["另一基金", "005678", 500, -1.0],
        ]
        wb = self._make_workbook({"证券账户": (header, data)})
        with self.assertLogs("invest", level="WARNING") as log:
            holdings = reader._parse_workbook(wb)
        # cost_price=0 的行（某基金）应被保留，其余 3 行跳过
        self.assertEqual(len(holdings), 1)
        self.assertEqual(holdings[0].name, "某基金")
        self.assertEqual(holdings[0].cost_price, 0.0)
        self.assertTrue(any("无效数值" in msg for msg in log.output))

    def test_mixed_valid_and_invalid(self):
        """混合有效行与无效行 -> 只返回有效行。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [
            ["长江电力", "600900", 200, 50.0],
            ["无效份额", "000001", "abc", 10.0],
            ["贵州茅台", "600519", 10, 2000.0],
            [None, "000002", 100, 10.0],
        ]
        wb = self._make_workbook({"证券账户": (header, data)})
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 2)
        self.assertEqual(holdings[0].name, "长江电力")
        self.assertEqual(holdings[1].name, "贵州茅台")

    def test_whitespace_account_name_stripped(self):
        """账户名称去除空格。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [["长江电力", "600900", 200, 50.0]]
        wb = self._make_workbook({"  证券账户  ": (header, data)})
        holdings = reader._parse_workbook(wb)
        self.assertEqual(holdings[0].account, "证券账户")

    def test_float_string_conversion(self):
        """字符串格式的数字应能正确转换为 float。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [["长江电力", "600900", "200.5", "50.0"]]
        wb = self._make_workbook({"证券账户": (header, data)})
        holdings = reader._parse_workbook(wb)
        self.assertEqual(len(holdings), 1)
        self.assertAlmostEqual(holdings[0].shares, 200.5)
        self.assertAlmostEqual(holdings[0].cost_price, 50.0)

    def test_all_sheets_empty_returns_empty(self):
        """所有表都为空 -> 返回空列表。"""
        wb = self._make_workbook({})
        holdings = reader._parse_workbook(wb)
        self.assertEqual(holdings, [])

    def test_sheet_name_stripped_from_account(self):
        """账号从 sheet_name 获取并 strip。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [["长江电力", "600900", 200, 50.0]]
        wb = self._make_workbook({" 证券账户 ": (header, data)})
        holdings = reader._parse_workbook(wb)
        self.assertEqual(holdings[0].account, "证券账户")

    def test_code_as_numeric_parsed(self):
        """代码为数值类型时转字符串。"""
        header = ["名称", "代码", "持仓份额", "每份成本"]
        data = [["长江电力", 600900, 200, 50.0]]
        wb = self._make_workbook({"证券账户": (header, data)})
        holdings = reader._parse_workbook(wb)
        self.assertEqual(holdings[0].code, "600900")


# ═════════════════════════════════════════════════════════════
#  read_holdings 测试（文件级入口 + 异常处理）
# ═════════════════════════════════════════════════════════════


class TestReadHoldings(unittest.TestCase):
    """测试 read_holdings 文件打开与异常处理。"""

    def test_file_not_found(self):
        """文件不存在 -> FileNotFoundError。"""
        with self.assertRaises(FileNotFoundError):
            reader.read_holdings("D:/nonexistent_file_for_test.xlsx")

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_invalid_file_format(self, mock_load, mock_exists):
        """InvalidFileException -> ValueError。"""
        mock_exists.return_value = True
        mock_load.side_effect = InvalidFileException("格式错误")

        with self.assertRaises(ValueError) as ctx:
            reader.read_holdings("bad.xlsx")
        self.assertIn("格式错误", str(ctx.exception))

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_bad_zip_file(self, mock_load, mock_exists):
        """BadZipFile -> ValueError。"""
        mock_exists.return_value = True
        mock_load.side_effect = zipfile.BadZipFile("不是 zip 文件")

        with self.assertRaises(ValueError) as ctx:
            reader.read_holdings("bad.xlsx")
        self.assertIn("格式错误", str(ctx.exception))

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_empty_workbook(self, mock_load, mock_exists):
        """空工作簿 -> 返回空列表。"""
        mock_exists.return_value = True
        mock_wb = MagicMock()
        mock_wb.sheetnames = []
        mock_load.return_value = mock_wb

        holdings = reader.read_holdings("empty.xlsx")
        self.assertEqual(holdings, [])

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_workbook_closed_after_read(self, mock_load, mock_exists):
        """文件读取完后 close 被调用。"""
        mock_exists.return_value = True
        mock_wb = MagicMock()
        mock_wb.sheetnames = []
        mock_load.return_value = mock_wb

        reader.read_holdings("dummy.xlsx")
        mock_wb.close.assert_called_once()

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_logging_on_empty_result(self, mock_load, mock_exists):
        """无持仓记录时记录 WARNING 日志。"""
        mock_exists.return_value = True
        mock_wb = MagicMock()
        mock_wb.sheetnames = ["空表"]
        mock_ws = MagicMock()
        mock_ws.max_row = 1
        mock_wb.__getitem__.return_value = mock_ws
        mock_load.return_value = mock_wb

        with self.assertLogs("invest", level="WARNING") as log:
            holdings = reader.read_holdings("empty.xlsx")
        self.assertEqual(holdings, [])
        self.assertTrue(any("未读取到任何持仓记录" in msg for msg in log.output))

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_file_not_exists_before_load(self, mock_load, mock_exists):
        """os.path.exists 在 load_workbook 之前被调用。"""
        mock_exists.return_value = False

        with self.assertRaises(FileNotFoundError):
            reader.read_holdings("missing.xlsx")
        mock_load.assert_not_called()

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_logging_info_on_start(self, mock_load, mock_exists):
        """读取开始时记录 INFO 日志。"""
        mock_exists.return_value = True
        mock_wb = MagicMock()
        mock_wb.sheetnames = []
        mock_load.return_value = mock_wb

        with self.assertLogs("invest", level="INFO") as log:
            reader.read_holdings("dummy.xlsx")
        self.assertTrue(any("正在读取持仓文件" in msg for msg in log.output))


class TestParseFlowSheets(unittest.TestCase):
    """交易流水 / 分红流水页签解析测试（持仓文件格式扩展）。

    覆盖：正常解析、表头不匹配跳过、空表跳过、非法行容错（日期/操作/数值）、
    费用列可选、操作归一化、无流水页签回退、文件级入口、向后兼容。
    """

    def _make_cell(self, value):
        cell = MagicMock()
        cell.value = value
        return cell

    def _make_worksheet(self, title, header, data_rows, max_row=None):
        ws = MagicMock()
        ws.title = title
        ws.max_row = max_row if max_row is not None else (len(data_rows) + 1)
        header_cells = [self._make_cell(v) for v in header]

        def iter_rows_side_effect(min_row=1, max_row=None, values_only=False):
            if min_row == 1 and not values_only:
                return iter([header_cells])
            elif min_row >= 2 and values_only:
                return iter(data_rows)
            return iter([])

        ws.iter_rows.side_effect = iter_rows_side_effect
        return ws

    # ── _parse_trade_sheet ─────────────────────────────────

    def test_trade_sheet_valid(self):
        """含费用列的交易流水 -> 正确解析并归一化操作。"""
        ws = self._make_worksheet(
            "交易流水",
            ["日期", "代码", "操作", "份额", "价格", "费用"],
            [
                ["2026-01-05", "600900", "买入", 200, 25.0, 5.0],
                ["2026-02-10", "600519", "卖出", 10, 2100.0, 10.0],
            ],
        )
        records = reader._parse_trade_sheet(ws)
        self.assertEqual(len(records), 2)
        self.assertEqual(records[0].date, "2026-01-05")
        self.assertEqual(records[0].code, "600900")
        self.assertEqual(records[0].action, "buy")
        self.assertAlmostEqual(records[0].shares, 200.0)
        self.assertAlmostEqual(records[0].price, 25.0)
        self.assertAlmostEqual(records[0].fee, 5.0)
        self.assertEqual(records[1].action, "sell")

    def test_trade_sheet_without_fee_column(self):
        """无费用列 -> fee 默认 0。"""
        ws = self._make_worksheet(
            "交易流水",
            ["日期", "代码", "操作", "份额", "价格"],
            [["2026-01-05", "600900", "买入", 200, 25.0]],
        )
        records = reader._parse_trade_sheet(ws)
        self.assertEqual(len(records), 1)
        self.assertAlmostEqual(records[0].fee, 0.0)

    def test_trade_header_mismatch_skipped(self):
        """表头不匹配 -> 返回空列表。"""
        ws = self._make_worksheet(
            "交易流水",
            ["日期", "代码", "方向", "数量", "单价"],
            [["2026-01-05", "600900", "买入", 200, 25.0]],
        )
        self.assertEqual(reader._parse_trade_sheet(ws), [])

    def test_trade_empty_sheet(self):
        """空表（max_row<2）-> 空列表。"""
        ws = self._make_worksheet("交易流水", ["日期", "代码", "操作", "份额", "价格"], [], 1)
        self.assertEqual(reader._parse_trade_sheet(ws), [])

    def test_trade_invalid_date_skipped(self):
        """非法日期行跳过，其余行保留。"""
        ws = self._make_worksheet(
            "交易流水",
            ["日期", "代码", "操作", "份额", "价格"],
            [
                ["昨天", "600900", "买入", 200, 25.0],
                ["2026-02-10", "600519", "买入", 10, 2000.0],
            ],
        )
        records = reader._parse_trade_sheet(ws)
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].code, "600519")

    def test_trade_invalid_action_skipped(self):
        """非法操作方向行跳过。"""
        ws = self._make_worksheet(
            "交易流水",
            ["日期", "代码", "操作", "份额", "价格"],
            [
                ["2026-01-05", "600900", "转仓", 200, 25.0],
                ["2026-02-10", "600519", "买入", 10, 2000.0],
            ],
        )
        records = reader._parse_trade_sheet(ws)
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].code, "600519")

    def test_trade_invalid_numbers_skipped(self):
        """非法份额/价格/负值行跳过。"""
        ws = self._make_worksheet(
            "交易流水",
            ["日期", "代码", "操作", "份额", "价格"],
            [
                ["2026-01-05", "600900", "买入", "abc", 25.0],
                ["2026-01-06", "600900", "买入", 200, -5.0],
                ["2026-01-07", "600900", "买入", 200, 25.0],
            ],
        )
        records = reader._parse_trade_sheet(ws)
        self.assertEqual(len(records), 1)
        self.assertAlmostEqual(records[0].price, 25.0)

    def test_trade_action_normalization(self):
        """买卖中英文操作归一化为 buy/sell。"""
        for raw, expected in [
            ("买入", "buy"),
            ("买", "buy"),
            ("buy", "buy"),
            ("申购", "buy"),
            ("卖出", "sell"),
            ("卖", "sell"),
            ("sell", "sell"),
            ("赎回", "sell"),
        ]:
            ws = self._make_worksheet(
                "交易流水",
                ["日期", "代码", "操作", "份额", "价格"],
                [["2026-01-05", "600900", raw, 200, 25.0]],
            )
            records = reader._parse_trade_sheet(ws)
            self.assertEqual(records[0].action, expected, f"操作 {raw} 应归一化为 {expected}")

    # ── _parse_dividend_sheet ──────────────────────────────

    def test_dividend_sheet_valid(self):
        """分红流水正常解析。"""
        ws = self._make_worksheet(
            "分红流水",
            ["日期", "代码", "每份分红"],
            [
                ["2026-06-01", "600900", 0.35],
                ["2026-06-01", "600519", 0.0],
            ],
        )
        records = reader._parse_dividend_sheet(ws)
        self.assertEqual(len(records), 2)
        self.assertEqual(records[0].date, "2026-06-01")
        self.assertEqual(records[0].code, "600900")
        self.assertAlmostEqual(records[0].amount, 0.35)

    def test_dividend_header_mismatch(self):
        """分红表头不匹配 -> 空列表。"""
        ws = self._make_worksheet(
            "分红流水",
            ["日期", "代码", "分红"],
            [["2026-06-01", "600900", 0.35]],
        )
        self.assertEqual(reader._parse_dividend_sheet(ws), [])

    def test_dividend_invalid_amount_skipped(self):
        """非法/负分红金额行跳过。"""
        ws = self._make_worksheet(
            "分红流水",
            ["日期", "代码", "每份分红"],
            [
                ["2026-06-01", "600900", -1.0],
                ["2026-06-01", "600519", 2.0],
            ],
        )
        records = reader._parse_dividend_sheet(ws)
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].code, "600519")

    # ── read_flow_sheets 文件级入口 ────────────────────────

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_read_flow_sheets_with_sheets(self, mock_load, mock_exists):
        """含流水页签的工作簿 -> 解析两个页签。"""
        mock_exists.return_value = True
        wb = MagicMock()
        wb.sheetnames = ["交易流水", "分红流水"]
        wb.__getitem__.side_effect = {
            "交易流水": self._make_worksheet(
                "交易流水",
                ["日期", "代码", "操作", "份额", "价格"],
                [["2026-01-05", "600900", "买入", 200, 25.0]],
            ),
            "分红流水": self._make_worksheet(
                "分红流水",
                ["日期", "代码", "每份分红"],
                [["2026-06-01", "600900", 0.35]],
            ),
        }.get
        mock_load.return_value = wb

        transactions, dividends = reader.read_flow_sheets("flows.xlsx")
        self.assertEqual(len(transactions), 1)
        self.assertEqual(transactions[0].code, "600900")
        self.assertEqual(len(dividends), 1)
        self.assertEqual(dividends[0].amount, 0.35)
        wb.close.assert_called_once()

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_read_flow_sheets_missing_sheets(self, mock_load, mock_exists):
        """无流水页签 -> 返回空列表（向后兼容）。"""
        mock_exists.return_value = True
        wb = MagicMock()
        wb.sheetnames = ["证券账户"]
        mock_load.return_value = wb

        transactions, dividends = reader.read_flow_sheets("plain.xlsx")
        self.assertEqual(transactions, [])
        self.assertEqual(dividends, [])

    def test_read_flow_sheets_file_not_found(self):
        """文件不存在 -> FileNotFoundError。"""
        with self.assertRaises(FileNotFoundError):
            reader.read_flow_sheets("D:/nonexistent_flows_for_test.xlsx")

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_read_holdings_with_flows(self, mock_load, mock_exists):
        """read_holdings_with_flows 组合主表 + 流水。"""
        mock_exists.return_value = True
        wb = MagicMock()
        wb.sheetnames = ["证券账户", "交易流水", "分红流水"]
        wb.__getitem__.side_effect = {
            "证券账户": self._make_worksheet(
                "证券账户",
                ["名称", "代码", "持仓份额", "每份成本"],
                [["长江电力", "600900", 200, 50.0]],
            ),
            "交易流水": self._make_worksheet(
                "交易流水",
                ["日期", "代码", "操作", "份额", "价格"],
                [["2026-01-05", "600900", "买入", 200, 25.0]],
            ),
            "分红流水": self._make_worksheet(
                "分红流水",
                ["日期", "代码", "每份分红"],
                [["2026-06-01", "600900", 0.35]],
            ),
        }.get
        mock_load.return_value = wb

        result = reader.read_holdings_with_flows("full.xlsx")
        self.assertEqual(len(result.holdings), 1)
        self.assertEqual(result.holdings[0].name, "长江电力")
        self.assertEqual(len(result.transactions), 1)
        self.assertEqual(len(result.dividends), 1)

    def test_valid_date_formats(self):
        """日期格式判定：支持 - 与 / 分隔。"""
        self.assertTrue(reader._valid_date("2026-01-05"))
        self.assertTrue(reader._valid_date("2026/1/5"))
        self.assertFalse(reader._valid_date("20260105"))
        self.assertFalse(reader._valid_date("昨天"))
        self.assertFalse(reader._valid_date(""))

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_multi_account_with_flows(self, mock_load, mock_exists):
        """多账户主表 + 流水页签组合解析（流水为文件级单页签，不干扰多账户主表）。"""
        mock_exists.return_value = True
        wb = MagicMock()
        wb.sheetnames = ["账户A", "账户B", "交易流水", "分红流水"]
        wb.__getitem__.side_effect = {
            "账户A": self._make_worksheet(
                "账户A",
                ["名称", "代码", "持仓份额", "每份成本"],
                [["长江电力", "600900", 200, 50.0]],
            ),
            "账户B": self._make_worksheet(
                "账户B",
                ["名称", "代码", "持仓份额", "每份成本"],
                [["贵州茅台", "600519", 10, 1800.0]],
            ),
            "交易流水": self._make_worksheet(
                "交易流水",
                ["日期", "代码", "操作", "份额", "价格"],
                [["2026-01-05", "600900", "买入", 200, 25.0]],
            ),
            "分红流水": self._make_worksheet(
                "分红流水",
                ["日期", "代码", "每份分红"],
                [["2026-06-01", "600900", 0.35]],
            ),
        }.get
        mock_load.return_value = wb

        result = reader.read_holdings_with_flows("multi.xlsx")
        self.assertEqual(len(result.holdings), 2)
        accounts = {h.account for h in result.holdings}
        self.assertEqual(accounts, {"账户A", "账户B"})
        self.assertEqual(len(result.transactions), 1)
        self.assertEqual(len(result.dividends), 1)

    @patch("src.python.core.reader.os.path.exists")
    @patch("src.python.core.reader.openpyxl.load_workbook")
    def test_backward_compat_no_flows_field_equal(self, mock_load, mock_exists):
        """旧格式文件（无流水）：read_holdings_with_flows 的 holdings 与 read_holdings 逐字段一致。"""
        mock_exists.return_value = True
        wb = MagicMock()
        wb.sheetnames = ["证券账户"]
        wb.__getitem__.side_effect = {
            "证券账户": self._make_worksheet(
                "证券账户",
                ["名称", "代码", "持仓份额", "每份成本"],
                [["长江电力", "600900", 200, 50.0]],
            ),
        }.get
        mock_load.return_value = wb

        holdings_plain = reader.read_holdings("plain.xlsx")
        result = reader.read_holdings_with_flows("plain.xlsx")
        self.assertEqual(result.transactions, [])
        self.assertEqual(result.dividends, [])
        self.assertEqual(len(result.holdings), len(holdings_plain))
        for new, old in zip(result.holdings, holdings_plain):
            self.assertEqual(
                (new.account, new.name, new.code, new.shares, new.cost_price, new.data_status),
                (old.account, old.name, old.code, old.shares, old.cost_price, old.data_status),
            )

    def test_normalize_action_unknown(self):
        """未知操作返回 None。"""
        self.assertIsNone(reader._normalize_action("转仓"))
        self.assertIsNone(reader._normalize_action(""))
        self.assertEqual(reader._normalize_action("BUY"), "buy")


if __name__ == "__main__":
    unittest.main()
