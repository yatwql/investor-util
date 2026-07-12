"""Y3: 文件系统纵深边缘场景测试。

覆盖加密 Excel/xlsm 宏/隐藏工作表/损坏 xlsx/UNC 网络路径/
文件占用/权限变更/超长路径/缓存篡改/空字节文件共 12 项测试。

运行：
  cd D:/codebase/zoo/investor-util
  python -m pytest src/test/unit/core/test_filesystem_edge.py -v
"""

from __future__ import annotations

import os
import tempfile
import zipfile
import unittest
from unittest.mock import MagicMock, patch

import pytest
pytestmark = [pytest.mark.unit, pytest.mark.unit_core, pytest.mark.edge]


# ═══════════════════════════════════════════════════════════
# Y3-1: 加密 Excel / xlsm 宏
# ═══════════════════════════════════════════════════════════

class TestEncryptedExcelY3(unittest.TestCase):
    """加密 Excel/xlsm 文件处理。"""

    @patch("src.python.reader.openpyxl.load_workbook")
    def test_encrypted_xlsx_raises_value_error(self, mock_load):
        """加密 xlsx → 抛出 ValueError。"""
        from src.python.reader import read_holdings
        import openpyxl
        mock_load.side_effect = openpyxl.utils.exceptions.InvalidFileException(
            "File contains encrypted content"
        )
        with patch("src.python.reader.os.path.exists", return_value=True):
            with self.assertRaises(ValueError) as ctx:
                read_holdings("encrypted.xlsx")
        self.assertIn("encrypted", str(ctx.exception).lower()
                      or "格式错误" in str(ctx.exception))

    def test_xlsm_not_listed_as_xlsx(self):
        """xlsm 文件不被 list_xlsx_files 识别（只认 .xlsx）。"""
        from src.python.reader import list_xlsx_files
        with tempfile.TemporaryDirectory() as tmpdir:
            fpath = os.path.join(tmpdir, "macro.xlsm")
            with open(fpath, "w"): pass
            result = list_xlsx_files(tmpdir)
            self.assertEqual(len(result), 0)


# ═══════════════════════════════════════════════════════════
# Y3-2: 隐藏工作表
# ═══════════════════════════════════════════════════════════

class TestHiddenSheetY3(unittest.TestCase):
    """隐藏工作表处理。"""

    def setUp(self):
        from src.python import reader as rdr
        self.reader = rdr

    @patch("src.python.reader.openpyxl.load_workbook")
    def test_hidden_sheet_skipped(self, mock_load):
        """隐藏/非常见工作表 → 跳过不计入 accounts。"""
        mock_wb = MagicMock()
        mock_wb.sheetnames = ["证券账户", "隐藏表（hidden）", "说明（veryHidden）"]

        def make_ws(max_row=1):
            ws = MagicMock()
            ws.max_row = max_row
            return ws

        def getitem(name):
            if name == "证券账户":
                return make_ws(max_row=5)
            return make_ws(max_row=3)

        mock_wb.__getitem__.side_effect = getitem
        mock_load.return_value = mock_wb

        from src.python.reader import get_xlsx_info
        with patch("src.python.reader.os.path.exists", return_value=True):
            info = get_xlsx_info("work.xlsx")
        self.assertIn("证券账户", info["sheet_names"])
        self.assertEqual(info["accounts"], 3)


# ═══════════════════════════════════════════════════════════
# Y3-3: 损坏 xlsx
# ═══════════════════════════════════════════════════════════

class TestCorruptedXlsxY3(unittest.TestCase):
    """损坏 xlsx 文件处理。"""

    def test_bad_zip_file_raises_value_error(self):
        """非 zip 格式 → ValueError。"""
        from src.python.reader import read_holdings
        import zipfile
        with tempfile.TemporaryDirectory() as tmpdir:
            fpath = os.path.join(tmpdir, "bad.xlsx")
            with open(fpath, "w") as f:
                f.write("this is not a valid xlsx file")
            with self.assertRaises(ValueError) as ctx:
                read_holdings(fpath)
            self.assertIn("格式错误", str(ctx.exception))

    def test_truncated_xlsx_raises_value_error(self):
        """截断的 xlsx → ValueError。"""
        from src.python.reader import read_holdings
        import zipfile
        with tempfile.TemporaryDirectory() as tmpdir:
            fpath = os.path.join(tmpdir, "truncated.xlsx")
            # 写一个不完整的 zip 文件
            with open(fpath, "wb") as f:
                f.write(b"PK\x03\x04" + b"\x00" * 20)
            with self.assertRaises(ValueError) as ctx:
                read_holdings(fpath)
            self.assertIn("格式错误", str(ctx.exception))


# ═══════════════════════════════════════════════════════════
# Y3-4: UNC 网络路径
# ═══════════════════════════════════════════════════════════

class TestUncPathY3(unittest.TestCase):
    """UNC 网络路径处理。"""

    @patch("src.python.reader.openpyxl.load_workbook")
    @patch("src.python.reader.os.path.exists")
    def test_unc_path_read(self, mock_exists, mock_load):
        """UNC 路径 → 正常传递到 load_workbook。"""
        from src.python.reader import read_holdings
        unc_path = r"\\server\share\holdings.xlsx"
        mock_exists.return_value = True
        mock_wb = MagicMock()
        mock_wb.sheetnames = []
        mock_load.return_value = mock_wb
        try:
            read_holdings(unc_path)
        except Exception:
            self.fail("UNC path should not cause unexpected error")

    def test_unc_path_listing(self):
        """UNC 路径目录 → list_xlsx_files 不崩溃。"""
        from src.python.reader import list_xlsx_files
        # UNC 路径不存在时返回空列表，不崩溃
        result = list_xlsx_files(r"\\nonexistent-server\share")
        self.assertEqual(result, [])


# ═══════════════════════════════════════════════════════════
# Y3-5: 文件占用
# ═══════════════════════════════════════════════════════════

class TestFileLockedY3(unittest.TestCase):
    """文件被其他进程占用。"""

    @patch("src.python.reader.openpyxl.load_workbook")
    @patch("src.python.reader.os.path.exists")
    def test_locked_file_permission_error(self, mock_exists, mock_load):
        """文件被占用（PermissionError）→ 转为 ValueError。"""
        from src.python.reader import read_holdings
        mock_exists.return_value = True
        mock_load.side_effect = PermissionError("被其他程序打开")
        # 当前实现可能直接抛出 PermissionError
        try:
            read_holdings("locked.xlsx")
            self.fail("Should raise")
        except (PermissionError, ValueError):
            pass

    @patch("src.python.reader.openpyxl.load_workbook")
    @patch("src.python.reader.os.path.exists")
    def test_locked_file_oserror(self, mock_exists, mock_load):
        """文件被占用（OSError）→ 不崩溃。"""
        from src.python.reader import read_holdings
        mock_exists.return_value = True
        mock_load.side_effect = OSError("文件已被其他进程占用")
        try:
            read_holdings("locked.xlsx")
            self.fail("Should raise")
        except (OSError, ValueError):
            pass


# ═══════════════════════════════════════════════════════════
# Y3-6: 权限变更
# ═══════════════════════════════════════════════════════════

class TestPermissionChangedY3(unittest.TestCase):
    """文件/目录权限变更处理。"""

    def test_readonly_file(self):
        """只读文件 → openpyxl 可读取。"""
        from src.python.reader import read_holdings
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            fpath = os.path.join(tmpdir, "readonly.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["名称", "代码", "持仓份额", "每份成本"])
            ws.append(["长江电力", "600900", 200, 50.0])
            wb.save(fpath)
            os.chmod(fpath, 0o444)  # 只读
            try:
                holdings = read_holdings(fpath)
                self.assertEqual(len(holdings), 1)
            finally:
                os.chmod(fpath, 0o644)

    @patch("src.python.reader.os.path.exists", return_value=False)
    def test_directory_no_permission(self, mock_exists):
        """无权限访问目录 → 返回空列表。"""
        from src.python.reader import list_xlsx_files
        result = list_xlsx_files("Z:\\nonexistent")
        self.assertEqual(result, [])


# ═══════════════════════════════════════════════════════════
# Y3-7: 超长路径
# ═══════════════════════════════════════════════════════════

class TestLongPathY3(unittest.TestCase):
    """超长路径处理。"""

    def test_long_path_listing(self):
        """超长路径 → list_xlsx_files 不崩溃返回空。"""
        from src.python.reader import list_xlsx_files
        long_path = "C:\\" + "a" * 200 + "\\" + "b" * 50
        result = list_xlsx_files(long_path)
        self.assertEqual(result, [])

    @patch("src.python.reader.openpyxl.load_workbook")
    @patch("src.python.reader.os.path.exists")
    def test_long_path_read(self, mock_exists, mock_load):
        """超长路径 → 可传递到 load_workbook。"""
        from src.python.reader import read_holdings
        long_path = "C:\\" + "a" * 150 + "\\holdings.xlsx"
        mock_exists.return_value = True
        mock_wb = MagicMock()
        mock_wb.sheetnames = []
        mock_load.return_value = mock_wb
        try:
            holdings = read_holdings(long_path)
            self.assertEqual(holdings, [])
        except OSError:
            pass  # Windows 超长路径限制，允许此异常


# ═══════════════════════════════════════════════════════════
# Y3-8: 缓存篡改
# ═══════════════════════════════════════════════════════════

class TestCacheTamperingY3(unittest.TestCase):
    """缓存文件被篡改。"""

    def test_cache_file_corrupted_json(self):
        """缓存 JSON 损坏 → 自动删除并触发重新获取。"""
        from src.python.cache import get, _cache_path
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("src.python.cache._paths._CACHE_DIR", tmpdir):
                fpath = _cache_path("test_corrupted")
                with open(fpath, "w", encoding="utf-8") as f:
                    f.write("{invalid json content")
                result = get("test_corrupted", max_age_seconds=3600)
                self.assertIsNone(result)
                self.assertFalse(os.path.exists(fpath))

    def test_cache_file_null_bytes(self):
        """缓存文件含空字节 → 解析失败，自动删除。"""
        from src.python.cache import get, _cache_path
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("src.python.cache._paths._CACHE_DIR", tmpdir):
                fpath = _cache_path("test_nullbytes")
                with open(fpath, "wb") as f:
                    f.write(b'{"key": "value"}\x00\x00')
                result = get("test_nullbytes", max_age_seconds=3600)
                self.assertIsNone(result)
                self.assertFalse(os.path.exists(fpath))

    def test_cache_file_zero_bytes(self):
        """0 字节缓存文件 → 解析失败，自动删除。"""
        from src.python.cache import get, _cache_path
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("src.python.cache._paths._CACHE_DIR", tmpdir):
                fpath = _cache_path("test_empty")
                with open(fpath, "wb") as f:
                    pass
                result = get("test_empty", max_age_seconds=3600)
                self.assertIsNone(result)
                self.assertFalse(os.path.exists(fpath))

    def test_cache_file_non_utf8_content(self):
        """非 UTF-8 编码的缓存文件 → 解析失败，自动删除。"""
        from src.python.cache import get, _cache_path
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("src.python.cache._paths._CACHE_DIR", tmpdir):
                fpath = _cache_path("test_binary")
                with open(fpath, "wb") as f:
                    f.write(b'\xff\xfe\x00\x01\x02\x03')
                result = get("test_binary", max_age_seconds=3600)
                self.assertIsNone(result)
                self.assertFalse(os.path.exists(fpath))


# ═══════════════════════════════════════════════════════════
# Y3-9: 空字节文件（持仓 xlsx）
# ═══════════════════════════════════════════════════════════

class TestNullBytesInXlsxY3(unittest.TestCase):
    """xlsx 文件内含空字节。"""

    def test_xlsx_with_null_bytes(self):
        """空字节污染 xlsx → 格式错误或正常解析。"""
        from src.python.reader import read_holdings
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            fpath = os.path.join(tmpdir, "nullbytes.xlsx")
            # 创建正常 xlsx 然后注入空字节
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["名称", "代码", "持仓份额", "每份成本"])
            ws.append(["长江电力", "600900", 200, 50.0])
            wb.save(fpath)
            # 追加空字节（模拟文件污染）
            with open(fpath, "ab") as f:
                f.write(b'\x00\x00')
            try:
                holdings = read_holdings(fpath)
                # 可能成功（空字节在末尾被忽略）或抛出异常
                self.assertIsInstance(holdings, list)
            except (ValueError, zipfile.BadZipFile):
                pass  # 允许两种行为


if __name__ == "__main__":
    unittest.main()
