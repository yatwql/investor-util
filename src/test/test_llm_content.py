"""LLM 内容 Excel 写入模块单元测试。

测试目标：
  - _strip_html — HTML 标签剥离
  - _write_content_sheet — 段落分行、wrap_text、行高、列宽、缓存提示
  - write_llm_sheets — 双页签写入 + 缓存标记独立

运行：
  cd D:/codebase/zoo/investor-util
  python -m unittest src.test_llm_content -v
"""

from __future__ import annotations

import unittest
from math import ceil

from openpyxl import Workbook

from src.python.report.llm_content import (
    _ROW_HEIGHT_MIN,
    _ROW_HEIGHT_PER_LINE,
    _CHARS_PER_LINE,
    _calc_row_height,
    _strip_html,
    _write_content_sheet,
    write_llm_sheets,
)
from src.python.report.styles import CONTENT_FONT, TITLE_FILL, TITLE_FONT


# ═══════════════════════════════════════════════════════════
#  _calc_row_height
# ═══════════════════════════════════════════════════════════


class TestCalcRowHeight(unittest.TestCase):
    """测试行高估算函数。"""

    def test_empty(self) -> None:
        self.assertEqual(_calc_row_height(""), _ROW_HEIGHT_MIN)

    def test_short_text(self) -> None:
        """短文本 → 最小行高。"""
        text = "短文本"
        self.assertEqual(_calc_row_height(text), _ROW_HEIGHT_MIN)

    def test_long_text(self) -> None:
        """超长文本 → 行高按行数倍乘。"""
        text = "中" * 120  # 120 中文字符 ≈ 3 行
        expected = ceil(120 / _CHARS_PER_LINE) * _ROW_HEIGHT_PER_LINE
        self.assertEqual(_calc_row_height(text), expected)

    def test_boundary(self) -> None:
        """正好 _CHARS_PER_LINE 个字符 → 最小行高。"""
        text = "中" * _CHARS_PER_LINE
        self.assertEqual(_calc_row_height(text), _ROW_HEIGHT_MIN)


# ═══════════════════════════════════════════════════════════
#  _strip_html
# ═══════════════════════════════════════════════════════════


class TestStripHtml(unittest.TestCase):
    """测试 HTML 标签剥离。"""

    def test_plain_text(self) -> None:
        self.assertEqual(_strip_html("hello"), "hello")

    def test_strip_p_tag(self) -> None:
        self.assertEqual(_strip_html("<p>内容</p>"), "内容")

    def test_strip_strong(self) -> None:
        self.assertEqual(_strip_html("<strong>粗体</strong>"), "粗体")

    def test_strip_nested(self) -> None:
        html = "<p>这是 <strong>粗体</strong> 和 <em>斜体</em></p>"
        self.assertEqual(_strip_html(html), "这是 粗体 和 斜体")

    def test_html_entity_unchanged(self) -> None:
        """HTML 实体（如 &amp;）保留原样——已知行为。"""
        self.assertEqual(_strip_html("A &amp; B"), "A &amp; B")

    def test_none_input(self) -> None:
        self.assertEqual(_strip_html(None), "")

    def test_empty_string(self) -> None:
        self.assertEqual(_strip_html(""), "")


# ═══════════════════════════════════════════════════════════
#  _write_content_sheet
# ═══════════════════════════════════════════════════════════


class TestWriteContentSheet(unittest.TestCase):
    """测试单个 LLM 内容页签的写入行为。"""

    def setUp(self):
        self.wb = Workbook()

    def _write_and_get_sheet(self, content=None, from_cache=False):
        """写入内容并返回 worksheet 对象。"""
        ws = self.wb.create_sheet()
        _write_content_sheet(ws, "测试页签", content, from_cache=from_cache)
        return ws

    def test_title_row(self):
        """第 1 行标题 — 合并 A1:B1，使用 TITLE_FONT 和 TITLE_FILL。"""
        ws = self._write_and_get_sheet(content="一段落")
        cell = ws.cell(row=1, column=1)
        self.assertEqual(cell.value, "测试页签")
        self.assertEqual(cell.font.name, TITLE_FONT.name)
        self.assertEqual(cell.fill.fgColor.rgb, TITLE_FILL.fgColor.rgb)

    def test_single_paragraph(self):
        """单段落 → 写入第 2 行，wrap_text=True。"""
        ws = self._write_and_get_sheet(content="单段落内容")
        cell = ws.cell(row=2, column=1)
        self.assertEqual(cell.value, "单段落内容")
        self.assertTrue(cell.alignment.wrap_text)

    def test_multi_paragraphs(self):
        """三个段落 → A2/A4/A6 各一段（A3/A5 空行间距），A7 无内容。"""
        content = "<p>第一段</p>\n\n<p>第二段</p>\n\n<p>第三段</p>"
        ws = self._write_and_get_sheet(content=content)

        self.assertEqual(ws.cell(row=2, column=1).value, "第一段")
        self.assertEqual(ws.cell(row=4, column=1).value, "第二段")
        self.assertEqual(ws.cell(row=6, column=1).value, "第三段")
        # 空行无值
        self.assertIsNone(ws.cell(row=3, column=1).value)
        self.assertIsNone(ws.cell(row=5, column=1).value)

    def test_content_none_placeholder(self):
        """content=None → A2 写入占位符，无缓存提示行。"""
        ws = self._write_and_get_sheet(content=None)
        cell = ws.cell(row=2, column=1)
        self.assertIn("LLM API Key", str(cell.value))

    def test_wrap_text_enabled(self):
        """每行内容单元格启用文本换行。"""
        ws = self._write_and_get_sheet(content="段落一\n\n段落二")
        for r in (2, 4):
            cell = ws.cell(row=r, column=1)
            self.assertTrue(cell.alignment.wrap_text,
                            f"Row {r} wrap_text should be True")

    def test_paragraph_font(self):
        """段落使用 CONTENT_FONT。"""
        ws = self._write_and_get_sheet(content="测试内容")
        cell = ws.cell(row=2, column=1)
        self.assertEqual(cell.font.size, CONTENT_FONT.size)
        self.assertEqual(cell.font.color.rgb, CONTENT_FONT.color.rgb)

    def test_from_cache_hint_row(self):
        """from_cache=True → 段落后追加灰色缓存提示行。"""
        ws = self._write_and_get_sheet(content="一段落话", from_cache=True)
        # 第 2 行 = 段落, 第 3 行 = 空行, 第 4 行 = 缓存提示
        hint_cell = ws.cell(row=4, column=1)
        self.assertIn("LLM缓存", str(hint_cell.value))
        self.assertTrue(hint_cell.font.italic)

    def test_from_cache_false_no_hint(self):
        """from_cache=False → 无缓存提示行。"""
        ws = self._write_and_get_sheet(content="一段落话", from_cache=False)
        # 第 2 行 = 段落, 第 3 行 = 空行, 第 4 行应为空
        self.assertIsNone(ws.cell(row=4, column=1).value)

    def test_content_none_no_cache_hint(self):
        """content=None 即使 from_cache=True 也不追加提示行。"""
        ws = self._write_and_get_sheet(content=None, from_cache=True)
        cell = ws.cell(row=3, column=1)
        self.assertIsNone(cell.value)

    def test_row_height_short(self):
        """短段落行高 = 最小行高。"""
        ws = self._write_and_get_sheet(content="短")
        self.assertGreaterEqual(
            ws.row_dimensions[2].height, _ROW_HEIGHT_MIN)

    def test_column_width_fixed(self):
        """A 列宽度固定。"""
        ws = self._write_and_get_sheet(content="测试内容")
        # 默认 openpyxl 宽度约 8.43，此处应被设为固定值 80
        width = ws.column_dimensions["A"].width
        self.assertIsNotNone(width)
        self.assertEqual(width, 80)

    def test_freeze_panes(self):
        """冻结首行。"""
        ws = self._write_and_get_sheet(content="测试")
        self.assertEqual(ws.freeze_panes, "A2")


# ═══════════════════════════════════════════════════════════
#  write_llm_sheets
# ═══════════════════════════════════════════════════════════


class TestWriteLlmSheets(unittest.TestCase):
    """测试 write_llm_sheets 集成行为。"""

    def setUp(self):
        self.wb = Workbook()

    def _get_llm_sheets(self):
        """返回 write_llm_sheets 创建的四个 sheet（跳过 Workbook 默认 sheet）。"""
        return self.wb.worksheets[-4:]

    def test_sheet_titles(self):
        """四个 sheet 标题正确。"""
        write_llm_sheets(self.wb, ("<p>宏观</p>", "<p>复盘</p>", None, None))
        sheet_names = [ws.title for ws in self.wb.worksheets]
        self.assertIn("7. 全球政经局势", sheet_names)
        self.assertIn("8. 智囊团深度复盘", sheet_names)
        self.assertIn("9. 持仓体检报告", sheet_names)
        self.assertIn("10. 穿透深度分析", sheet_names)

    def test_return_text_quad(self):
        """返回 (text7, text8, text9, textA) 纯文本四元组。"""
        text7, text8, text9, textA = write_llm_sheets(
            self.wb, ("<p>宏观分析</p>", "<p>复盘内容</p>", "<p>体检结果</p>", "<p>穿透分析</p>"))
        self.assertEqual(text7, "宏观分析")
        self.assertEqual(text8, "复盘内容")
        self.assertEqual(text9, "体检结果")
        self.assertEqual(textA, "穿透分析")

    def test_cache_macro_only(self):
        """macro 缓存 + expert 未缓存 → sheet7 有提示行，sheet8 无。"""
        write_llm_sheets(
            self.wb,
            ("<p>宏观内容</p>", "<p>复盘内容</p>", None, None),
            llm_cached=(True, False, False, False),
        )
        ws7, ws8, ws9, wsA = self._get_llm_sheets()

        # sheet7：第 4 行应有缓存提示
        self.assertIn("LLM缓存", str(ws7.cell(row=4, column=1).value or ""))
        # sheet8：第 4 行应无提示
        self.assertIsNone(ws8.cell(row=4, column=1).value)

    def test_cache_both(self):
        """macro + expert 都来自缓存 → 两个 sheet 都有提示。"""
        write_llm_sheets(
            self.wb,
            ("<p>宏观</p>", "<p>复盘</p>", None, None),
            llm_cached=(True, True, False, False),
        )
        ws7, ws8, ws9, wsA = self._get_llm_sheets()
        self.assertIn("LLM缓存", str(ws7.cell(row=4, column=1).value or ""))
        self.assertIn("LLM缓存", str(ws8.cell(row=4, column=1).value or ""))

    def test_cache_off_default(self):
        """默认 llm_cached=(False, False, False, False) → 无缓存提示。"""
        write_llm_sheets(self.wb, ("<p>宏观</p>", "<p>复盘</p>", None, None))
        ws7, ws8, ws9, wsA = self._get_llm_sheets()
        self.assertIsNone(ws7.cell(row=4, column=1).value)
        self.assertIsNone(ws8.cell(row=4, column=1).value)

    def test_content_none(self):
        """content=(None, None, None, None) → 占位符，不崩溃。"""
        text7, text8, text9, textA = write_llm_sheets(self.wb, (None, None, None, None))
        self.assertEqual(text7, "")
        self.assertEqual(text8, "")
        self.assertEqual(text9, "")
        self.assertEqual(textA, "")
        ws7, ws8, ws9, wsA = self._get_llm_sheets()
        self.assertIn("LLM API Key", str(ws7.cell(row=2, column=1).value or ""))
        self.assertIn("LLM API Key", str(ws8.cell(row=2, column=1).value or ""))


if __name__ == "__main__":
    unittest.main()
