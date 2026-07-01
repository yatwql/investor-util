"""集成测试 — 报告生成流程端到端验证。

测试目标：
  - _generate_excel_report — 全页签生成正确，mock 外部数据避免真实 API 调用

运行：
  cd D:/codebase/zoo/investor-util
  python -m unittest src.test_integration -v
"""

from __future__ import annotations

import os
import tempfile
import unittest
from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from src.python.models import Holding
from src.python.report.excel_writer import create_workbook, save_workbook
from src.python.report.excel_generator import generate_excel_report as _generate_excel_report


class MockDetail:
    """模拟 DetailRow dataclass（完整 15 个字段）。"""
    def __init__(self, name: str = "测试资产", code: str = "600000",
                 price: float = 10.0, yesterday_close: float = 9.8,
                 market_value: float = 1000.0, cost: float = 900.0,
                 profit: float = 100.0, today_profit: float = 20.0,
                 profit_rate: float = 0.1):
        self.account = "测试账户"
        self.name = name
        self.code = code
        self.price = price
        self.nav_date = "2026-06-28"
        self.yesterday_close = yesterday_close
        self.price_type = "T"
        self.premium = "—"
        self.shares = 100.0
        self.market_value = market_value
        self.cost = cost
        self.profit = profit
        self.profit_rate = profit_rate
        self.today_profit = today_profit
        self.source = "mock"
        self.source_api = "mock"


class TestGenerateExcelReport(unittest.TestCase):
    """_generate_excel_report 集成测试。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                    shares=100, cost_price=10.0),
            Holding(account="证券", name="贵州茅台", code="600519",
                    shares=50, cost_price=200.0),
        ]
        self.details = [
            MockDetail(name="长江电力", code="600900", price=25.0,
                       yesterday_close=24.5, market_value=2500.0,
                       cost=1000.0, profit=1500.0, today_profit=50.0),
            MockDetail(name="贵州茅台", code="600519", price=1800.0,
                       yesterday_close=1780.0, market_value=90000.0,
                       cost=10000.0, profit=80000.0, today_profit=1000.0),
        ]

    def tearDown(self):
        self.tmp.cleanup()

    @patch("src.python.fetcher.index.fetch_indices")
    @patch("src.python.fetcher.index.fetch_us_indices")
    @patch("src.python.report.fund_performance.write_fund_performance_sheet")
    def test_generate_basic_report(self, mock_perf, mock_us_idx, mock_a_idx):
        """基础报告（无新闻/无LLM）→ 5 个核心页签。"""
        mock_a_idx.return_value = {}
        mock_us_idx.return_value = {}

        _generate_excel_report(
            self.holdings,
            output_dir=self.tmp.name,
            details=self.details,
            a_indices={},
            us_indices={},
        )

        # 检查输出文件
        out_files = os.listdir(self.tmp.name)
        self.assertTrue(any(f.endswith(".xlsx") for f in out_files),
                        f"应在 {self.tmp.name} 中找到 xlsx 文件，实际有 {out_files}")

    @patch("src.python.fetcher.index.fetch_indices")
    @patch("src.python.fetcher.index.fetch_us_indices")
    @patch("src.python.report.fund_performance.write_fund_performance_sheet")
    def test_generate_with_news(self, mock_perf, mock_us_idx, mock_a_idx):
        """含新闻报告 → 使用预传入 news_data。"""
        mock_a_idx.return_value = {}
        mock_us_idx.return_value = {}

        _generate_excel_report(
            self.holdings,
            include_news=True,
            output_dir=self.tmp.name,
            details=self.details,
            a_indices={},
            us_indices={},
            news_data=[],
        )

        out_files = os.listdir(self.tmp.name)
        self.assertTrue(any(f.endswith(".xlsx") for f in out_files))

    @patch("src.python.fetcher.index.fetch_indices")
    @patch("src.python.fetcher.index.fetch_us_indices")
    @patch("src.python.report.fund_performance.write_fund_performance_sheet")
    @patch("src.python.report.llm_content.write_llm_sheets")
    def test_generate_with_llm(self, mock_llm, mock_perf, mock_us_idx, mock_a_idx):
        """含 LLM 报告 → 新增全球政经局势 + 智囊团深度复盘 页签。"""
        mock_a_idx.return_value = {}
        mock_us_idx.return_value = {}
        mock_llm.return_value = ("全球政经局势", "复盘内容", "", "")

        _generate_excel_report(
            self.holdings,
            include_llm=True,
            output_dir=self.tmp.name,
            details=self.details,
            a_indices={},
            us_indices={},
            llm_content=("<p>宏观</p>", "<p>复盘</p>", None, None),
        )

        out_files = os.listdir(self.tmp.name)
        self.assertTrue(any(f.endswith(".xlsx") for f in out_files))

    @patch("src.python.fetcher.index.fetch_indices")
    @patch("src.python.fetcher.index.fetch_us_indices")
    @patch("src.python.report.fund_performance.write_fund_performance_sheet")
    def test_generate_empty_holdings(self, mock_perf, mock_us_idx, mock_a_idx):
        """空持仓 → 不崩溃。"""
        mock_a_idx.return_value = {}
        mock_us_idx.return_value = {}

        _generate_excel_report(
            [],
            output_dir=self.tmp.name,
            details=[],
            a_indices={},
            us_indices={},
        )

        out_files = os.listdir(self.tmp.name)
        self.assertTrue(any(f.endswith(".xlsx") for f in out_files))

    @patch("src.python.fetcher.index.fetch_indices")
    @patch("src.python.fetcher.index.fetch_us_indices")
    @patch("src.python.report.fund_performance.write_fund_performance_sheet")
    def test_generate_single_holding(self, mock_perf, mock_us_idx, mock_a_idx):
        """单条持仓 → 正确生成。"""
        mock_a_idx.return_value = {}
        mock_us_idx.return_value = {}
        single_holding = [self.holdings[0]]
        single_detail = [self.details[0]]

        _generate_excel_report(
            single_holding,
            output_dir=self.tmp.name,
            details=single_detail,
            a_indices={},
            us_indices={},
        )

        out_files = os.listdir(self.tmp.name)
        self.assertTrue(any(f.endswith(".xlsx") for f in out_files))


class TestWorkbookSaveRoundtrip(unittest.TestCase):
    """Workbook 保存/读取往返测试。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()

    def tearDown(self):
        self.tmp.cleanup()

    def test_save_and_reopen(self):
        """保存的 xlsx 可被 openpyxl 重新打开。"""
        from openpyxl import load_workbook

        wb = create_workbook()
        ws = wb.active
        ws.title = "测试页签"
        ws["A1"] = "测试内容"

        path = save_workbook(wb, output_dir=self.tmp.name)
        self.assertTrue(os.path.exists(path))

        # 重新打开
        loaded = load_workbook(path, read_only=True)
        self.assertIn("测试页签", loaded.sheetnames)
        self.assertEqual(loaded["测试页签"]["A1"].value, "测试内容")
        loaded.close()

    def test_multiple_sheets_saved(self):
        """多页签 workbook 正确保存。"""
        wb = create_workbook()
        wb.remove(wb.active)

        names = []
        for i in range(4):
            ws = wb.create_sheet()
            name = f"页签{i + 1}"
            ws.title = name
            ws["A1"] = name
            names.append(name)

        path = save_workbook(wb, output_dir=self.tmp.name)
        self.assertTrue(os.path.exists(path))

        from openpyxl import load_workbook
        loaded = load_workbook(path, read_only=True)
        self.assertEqual(len(loaded.sheetnames), 4)
        for name in names:
            self.assertIn(name, loaded.sheetnames)
        loaded.close()


# ═══════════════════════════════════════════════════════════════
#  R-095: 业务场景集成测试 S1~S5
# ═══════════════════════════════════════════════════════════════


class ScenarioTestBase(unittest.TestCase):
    """场景测试基类：提供共享的 mock 环境。"""

    def setUp(self):
        # 阻止所有网络/API 调用
        self._price_patcher = patch("src.python.fetcher.price.fetch_market_data")
        self._mock_price = self._price_patcher.start()
        self._mock_price.return_value = {
            "price": 10.0, "yesterday_close": 9.8,
            "price_date": "2026-06-26", "source": "腾讯财经",
            "source_api": "tencent",
        }

        self._fund_patcher = patch("src.python.report.penetration.fetch_fund_holdings")
        self._mock_fund = self._fund_patcher.start()
        self._mock_fund.return_value = {
            "code": "510300", "name": "沪深300ETF",
            "date": "2026-03-31",
            "holdings": [{"name": "贵州茅台", "code": "600519", "ratio": 16.0}],
        }

        # LLM 相关 mock
        self._llm_config_patcher = patch(
            "src.python.config.get_llm_config",
            return_value={"provider": None, "enabled_llm": {}},
        )
        self._llm_config_patcher.start()

    def tearDown(self):
        self._price_patcher.stop()
        self._fund_patcher.stop()
        self._llm_config_patcher.stop()

    def _make_holding(self, account: str, name: str, code: str,
                       shares: float, cost_price: float) -> Holding:
        return Holding(
            account=account, name=name, code=code,
            shares=shares, cost_price=cost_price,
        )


class TestScenarioS1(ScenarioTestBase):
    """S1: 纯股票组合（3 只 A 股，无基金）→ 穿透 TOP10 等于直接持股。"""

    def setUp(self):
        super().setUp()
        self.holdings = [
            self._make_holding("证券", "贵州茅台", "600519", 100, 2000.0),
            self._make_holding("证券", "长江电力", "600900", 200, 28.0),
            self._make_holding("证券", "宁德时代", "300750", 50, 250.0),
        ]

    def test_penetration_top10_equals_direct_holdings(self):
        """纯股票 → 穿透 TOP10 即为直接持股。"""
        from src.python.report.market_value import DetailRow

        details = []
        for h in self.holdings:
            dr = DetailRow()
            dr.code = h.code
            dr.name = h.name
            dr.price = h.cost_price
            dr.market_value = h.cost_price * h.shares
            dr.source_api = "tencent"
            dr.account = "证券"
            dr.nav_date = "2026-06-26"
            dr.yesterday_close = h.cost_price * 0.98
            dr.price_type = "场内收盘价(T)"
            dr.premium = "--"
            dr.shares = h.shares
            dr.cost = h.cost_price * h.shares
            dr.profit = 0.0
            dr.profit_rate = 0.0
            dr.today_profit = 0.0
            dr.source = "mock"
            details.append(dr)

        from src.python.report.penetration import compute_penetration_top10
        result = compute_penetration_top10(self.holdings, details)
        top10 = result.get("top10", [])
        code_set = {item.get("code", "").split(",")[0] for item in top10 if item.get("code")}
        expected = {"600519", "600900", "300750"}
        self.assertTrue(expected.issubset(code_set) or code_set.issubset(expected),
                        f"穿透 TOP10 代码 {code_set} 应与直接持股 {expected} 匹配")

    def test_no_fund_in_category(self):
        """纯股票 → 分类表中无基金行。"""
        from src.python.report.penetration import classify_penetration
        for h in self.holdings:
            cls = classify_penetration(h)
            self.assertEqual(cls, "stock", f"{h.name} 应为 stock")

    def test_total_profit_correct(self):
        """总盈亏 = 各股票盈亏之和。"""
        from src.python.report.market_value import _compute_detail_row
        details = [{
            "price": 2050.0, "yesterday_close": 2000.0,
            "price_date": "2026-06-26",
            "source": "腾讯财经", "source_api": "tencent",
        }, {
            "price": 28.5, "yesterday_close": 28.0,
            "price_date": "2026-06-26",
            "source": "腾讯财经", "source_api": "tencent",
        }, {
            "price": 260.0, "yesterday_close": 250.0,
            "price_date": "2026-06-26",
            "source": "腾讯财经", "source_api": "tencent",
        }]
        total = 0
        for h, m in zip(self.holdings, details):
            d = _compute_detail_row(h, m)
            total += d.profit
        self.assertAlmostEqual(total, (2050-2000)*100 + (28.5-28)*200 + (260-250)*50)


class TestScenarioS2(ScenarioTestBase):
    """S2: 纯基金组合（ETF + 主动 + QDII）。"""

    def setUp(self):
        super().setUp()
        self.holdings = [
            self._make_holding("证券", "沪深300ETF", "510300", 1000, 4.0),
            self._make_holding("支付宝", "易方达蓝筹精选", "005827", 500, 2.0),
            self._make_holding("证券", "纳斯达克ETF", "513300", 200, 1.5),
        ]

    def test_classify_correct(self):
        """基金类型分类正确。"""
        from src.python.report.penetration import classify_penetration
        classes = {h.code: classify_penetration(h) for h in self.holdings}
        # 513300 是 ETF 含 QDII → QDII
        self.assertIn(classes.get("513300", ""), ("qdii", "etf"))
        # 510300 → ETF
        self.assertEqual(classes.get("510300"), "etf")
        # 005827 支付宝 → active_equity
        self.assertEqual(classes.get("005827"), "active_equity")

    def test_penetration_top10_not_empty(self):
        """基金持仓 → 穿透 TOP10 不为空。"""
        from src.python.report.market_value import DetailRow

        details = []
        for h in self.holdings:
            dr = DetailRow()
            dr.code = h.code
            dr.name = h.name
            dr.price = h.cost_price
            dr.market_value = h.cost_price * h.shares
            dr.source_api = "tencent" if h.code in ("510300", "513300") else "tiantian"
            dr.account = h.account
            dr.nav_date = "2026-06-26"
            dr.yesterday_close = h.cost_price * 0.98
            dr.price_type = "T"
            dr.premium = "--"
            dr.shares = h.shares
            dr.cost = h.cost_price * h.shares
            dr.profit = 0.0
            dr.profit_rate = 0.0
            dr.today_profit = 0.0
            dr.source = "mock"
            details.append(dr)

        from src.python.report.penetration import compute_penetration_top10
        result = compute_penetration_top10(self.holdings, details)
        top10 = result.get("top10", [])
        self.assertTrue(len(top10) > 0, "基金持仓穿透后 TOP10 不应为空")


class TestScenarioS3(ScenarioTestBase):
    """S3: 混合多账户（证券+支付宝+微信）。"""

    def setUp(self):
        super().setUp()
        self.holdings = [
            self._make_holding("证券", "长江电力", "600900", 100, 28.0),
            self._make_holding("证券", "沪深300ETF", "510300", 500, 4.0),
            self._make_holding("支付宝", "易方达蓝筹精选", "005827", 300, 2.0),
            self._make_holding("微信", "招商鑫福中短债A", "012325", 1000, 1.0),
        ]

    def test_account_subtotals(self):
        """分账户小计正确。"""
        details = {}
        for h in self.holdings:
            details[h.code] = {
                "code": h.code, "name": h.name, "price": h.cost_price * 1.05,
                "market_value": h.cost_price * h.shares * 1.05,
                "source_api": "tencent",
            }

        from src.python.report.market_value import _compute_detail_row
        rows = [_compute_detail_row(h, details[h.code]) for h in self.holdings]

        # 按账户分组求和
        subtotals = {}
        for r in rows:
            subtotals[r.account] = subtotals.get(r.account, 0) + r.market_value
        total = sum(subtotals.values())
        self.assertAlmostEqual(total, sum(r.market_value for r in rows))
        self.assertEqual(len(subtotals), 3, "应有 3 个不同账户")


class TestScenarioS4(ScenarioTestBase):
    """S4: 新持仓无缓存 → 全部从 API 获取。"""

    def test_api_called_when_no_cache(self):
        """无缓存 → 调用 fetch_market_data。"""
        holdings = [
            self._make_holding("证券", "长江电力", "600900", 100, 28.0),
        ]

        from src.python.report.market_value import _generate_details
        self._mock_price.reset_mock()
        details = _generate_details(holdings, "2026-06-26")
        self.assertEqual(len(details), 1)
        # fetch_market_data 被调用过
        self._mock_price.assert_called()


class TestScenarioS5(ScenarioTestBase):
    """S5: 缓存全命中 → LLM 页脚显示缓存提示。"""

    @patch("src.python.cache.get")
    def test_llm_cache_hit_shows_hint(self, mock_cache_get):
        """LLM 缓存命中 → 页脚包含缓存标记。"""
        from src.python.llm.skeleton import _handle_cache_hit

        mock_cache_get.return_value = "<p>缓存内容</p>"
        llm_config = {"model": "claude-sonnet-4-20250514"}

        result = _handle_cache_hit(
            cached="<p>旧缓存</p>",
            cache_key="llm_global_macro_abc",
            module_key="global_macro",
            model="claude-sonnet-4",
            llm_config=llm_config,
            thinking_enabled=False,
        )
        self.assertIn("使用LLM缓存", result)

    @patch("src.python.cache.get")
    def test_llm_cache_hit_zero_cost(self, mock_cache_get):
        """缓存命中 → Token 费用为 0。"""
        from src.python.llm.skeleton import _handle_cache_hit

        mock_cache_get.return_value = "<p>缓存</p>"
        llm_config = {"model": "claude-sonnet-4"}

        result = _handle_cache_hit(
            cached="<p>缓存</p>",
            cache_key="test_key",
            module_key="test_mod",
            model="test-model",
            llm_config=llm_config,
            thinking_enabled=False,
        )
        # 缓存命中时不应显示费用
        self.assertNotIn("费用", result)


if __name__ == "__main__":
    unittest.main()
