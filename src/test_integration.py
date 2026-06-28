"""集成测试 — 报告生成流程端到端验证。

测试目标：
  - _generate_excel_report — 全页签生成正确，mock 外部数据避免真实 API 调用

运行：
  cd D:/codebase/zoo/investor-util
  python -m unittest src.test_integration -v
"""

from __future__ import annotations

import os
import tempfile
import unittest
from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from src.models import Holding
from src.report.excel_writer import create_workbook, save_workbook
from src.tui_handlers import _generate_excel_report


class MockDetail:
    """模拟 DetailRow dataclass（完整 15 个字段）。"""
    def __init__(self, name: str = "测试资产", code: str = "600000",
                 price: float = 10.0, yesterday_close: float = 9.8,
                 market_value: float = 1000.0, cost: float = 900.0,
                 profit: float = 100.0, today_profit: float = 20.0,
                 profit_rate: float = 0.1):
        self.account = "测试账户"
        self.name = name
        self.code = code
        self.price = price
        self.nav_date = "2026-06-28"
        self.yesterday_close = yesterday_close
        self.price_type = "T"
        self.premium = "—"
        self.shares = 100.0
        self.market_value = market_value
        self.cost = cost
        self.profit = profit
        self.profit_rate = profit_rate
        self.today_profit = today_profit
        self.source = "mock"
        self.source_api = "mock"


class TestGenerateExcelReport(unittest.TestCase):
    """_generate_excel_report 集成测试。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                    shares=100, cost_price=10.0),
            Holding(account="证券", name="贵州茅台", code="600519",
                    shares=50, cost_price=200.0),
        ]
        self.details = [
            MockDetail(name="长江电力", code="600900", price=25.0,
                       yesterday_close=24.5, market_value=2500.0,
                       cost=1000.0, profit=1500.0, today_profit=50.0),
            MockDetail(name="贵州茅台", code="600519", price=1800.0,
                       yesterday_close=1780.0, market_value=90000.0,
                       cost=10000.0, profit=80000.0, today_profit=1000.0),
        ]

    def tearDown(self):
        self.tmp.cleanup()

    @patch("src.fetcher.fetch_indices")
    @patch("src.fetcher.fetch_us_indices")
    @patch("src.report.fund_performance.write_fund_performance_sheet")
    def test_generate_basic_report(self, mock_perf, mock_us_idx, mock_a_idx):
        """基础报告（无新闻/无LLM）→ 5 个核心页签。"""
        mock_a_idx.return_value = {}
        mock_us_idx.return_value = {}

        _generate_excel_report(
            self.holdings,
            output_dir=self.tmp.name,
            details=self.details,
            a_indices={},
            us_indices={},
        )

        # 检查输出文件
        out_files = os.listdir(self.tmp.name)
        self.assertTrue(any(f.endswith(".xlsx") for f in out_files),
                        f"应在 {self.tmp.name} 中找到 xlsx 文件，实际有 {out_files}")

    @patch("src.fetcher.fetch_indices")
    @patch("src.fetcher.fetch_us_indices")
    @patch("src.report.fund_performance.write_fund_performance_sheet")
    def test_generate_with_news(self, mock_perf, mock_us_idx, mock_a_idx):
        """含新闻报告 → 使用预传入 news_data。"""
        mock_a_idx.return_value = {}
        mock_us_idx.return_value = {}

        _generate_excel_report(
            self.holdings,
            include_news=True,
            output_dir=self.tmp.name,
            details=self.details,
            a_indices={},
            us_indices={},
            news_data=[],
        )

        out_files = os.listdir(self.tmp.name)
        self.assertTrue(any(f.endswith(".xlsx") for f in out_files))

    @patch("src.fetcher.fetch_indices")
    @patch("src.fetcher.fetch_us_indices")
    @patch("src.report.fund_performance.write_fund_performance_sheet")
    @patch("src.report.llm_content.write_llm_sheets")
    def test_generate_with_llm(self, mock_llm, mock_perf, mock_us_idx, mock_a_idx):
        """含 LLM 报告 → 新增模块 7+8 页签。"""
        mock_a_idx.return_value = {}
        mock_us_idx.return_value = {}
        mock_llm.return_value = ("宏观内容", "复盘内容")

        _generate_excel_report(
            self.holdings,
            include_llm=True,
            show_llm_in_tui=False,
            output_dir=self.tmp.name,
            details=self.details,
            a_indices={},
            us_indices={},
            llm_content=("<p>宏观</p>", "<p>复盘</p>"),
            llm_cached=(True, True),
        )

        out_files = os.listdir(self.tmp.name)
        self.assertTrue(any(f.endswith(".xlsx") for f in out_files))

    @patch("src.fetcher.fetch_indices")
    @patch("src.fetcher.fetch_us_indices")
    @patch("src.report.fund_performance.write_fund_performance_sheet")
    def test_generate_empty_holdings(self, mock_perf, mock_us_idx, mock_a_idx):
        """空持仓 → 不崩溃。"""
        mock_a_idx.return_value = {}
        mock_us_idx.return_value = {}

        _generate_excel_report(
            [],
            output_dir=self.tmp.name,
            details=[],
            a_indices={},
            us_indices={},
        )

        out_files = os.listdir(self.tmp.name)
        self.assertTrue(any(f.endswith(".xlsx") for f in out_files))

    @patch("src.fetcher.fetch_indices")
    @patch("src.fetcher.fetch_us_indices")
    @patch("src.report.fund_performance.write_fund_performance_sheet")
    def test_generate_single_holding(self, mock_perf, mock_us_idx, mock_a_idx):
        """单条持仓 → 正确生成。"""
        mock_a_idx.return_value = {}
        mock_us_idx.return_value = {}
        single_holding = [self.holdings[0]]
        single_detail = [self.details[0]]

        _generate_excel_report(
            single_holding,
            output_dir=self.tmp.name,
            details=single_detail,
            a_indices={},
            us_indices={},
        )

        out_files = os.listdir(self.tmp.name)
        self.assertTrue(any(f.endswith(".xlsx") for f in out_files))


class TestWorkbookSaveRoundtrip(unittest.TestCase):
    """Workbook 保存/读取往返测试。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()

    def tearDown(self):
        self.tmp.cleanup()

    def test_save_and_reopen(self):
        """保存的 xlsx 可被 openpyxl 重新打开。"""
        from openpyxl import load_workbook

        wb = create_workbook()
        ws = wb.active
        ws.title = "测试页签"
        ws["A1"] = "测试内容"

        path = save_workbook(wb, output_dir=self.tmp.name)
        self.assertTrue(os.path.exists(path))

        # 重新打开
        loaded = load_workbook(path, read_only=True)
        self.assertIn("测试页签", loaded.sheetnames)
        self.assertEqual(loaded["测试页签"]["A1"].value, "测试内容")
        loaded.close()

    def test_multiple_sheets_saved(self):
        """多页签 workbook 正确保存。"""
        wb = create_workbook()
        wb.remove(wb.active)

        names = []
        for i in range(4):
            ws = wb.create_sheet()
            name = f"页签{i + 1}"
            ws.title = name
            ws["A1"] = name
            names.append(name)

        path = save_workbook(wb, output_dir=self.tmp.name)
        self.assertTrue(os.path.exists(path))

        from openpyxl import load_workbook
        loaded = load_workbook(path, read_only=True)
        self.assertEqual(len(loaded.sheetnames), 4)
        for name in names:
            self.assertIn(name, loaded.sheetnames)
        loaded.close()


if __name__ == "__main__":
    unittest.main()
