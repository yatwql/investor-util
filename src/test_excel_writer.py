"""Excel 输出引擎单元测试 — 异常场景与边界测试。

测试目标：
  - _ensure_reports_dir — 目录创建/写入权限异常
  - save_workbook — 存档权限异常降级
  - create_workbook/write_title_row/write_header_row/write_data_row
  - write_subtotal_row/write_total_row
  - auto_width / freeze_header

运行：
  cd D:/codebase/zoo/investor-util
  python -m unittest src.test_excel_writer -v
"""

from __future__ import annotations

import os
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from src.report import excel_writer as ew
from src.report.styles import FMT_MONEY, FMT_PERCENT


class TestCreateWorkbook(unittest.TestCase):
    """Workbook 创建测试。"""

    def test_create_workbook(self):
        """create_workbook 返回合法 Workbook。"""
        wb = ew.create_workbook()
        self.assertIsInstance(wb, Workbook)

    def test_create_workbook_has_active_sheet(self):
        """新 Workbook 默认有一个活动页签。"""
        wb = ew.create_workbook()
        self.assertIsNotNone(wb.active)


class TestEnsureReportsDir(unittest.TestCase):
    """_ensure_reports_dir 异常场景测试。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()

    def tearDown(self):
        self.tmp.cleanup()

    def test_creates_directory(self):
        """目录不存在 → 创建目录和子目录。"""
        sub = os.path.join(self.tmp.name, "reports")
        ew._ensure_reports_dir(sub)
        # 目录应存在（内含 YYYYMMDD 子目录）
        self.assertTrue(os.path.exists(sub))

    def test_no_permission_raises(self):
        """无写入权限 → 抛出 PermissionError。"""
        # 创建一个只读目录
        readonly_dir = os.path.join(self.tmp.name, "readonly")
        os.makedirs(readonly_dir, exist_ok=True)
        with patch("builtins.open", side_effect=PermissionError("denied")):
            with self.assertRaises(PermissionError):
                ew._ensure_reports_dir(readonly_dir)


class TestSaveWorkbook(unittest.TestCase):
    """save_workbook 异常场景测试。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.wb = ew.create_workbook()
        ws = self.wb.active
        ws.title = "测试"
        ws.cell(row=1, column=1, value="test")

    def tearDown(self):
        self.tmp.cleanup()

    def test_save_success(self):
        """正常保存 → 返回最新文件路径。"""
        path = ew.save_workbook(self.wb, output_dir=self.tmp.name)
        self.assertTrue(os.path.exists(path))
        self.assertTrue(path.endswith(".xlsx"))

    def test_archive_failure_still_returns_latest(self):
        """存档权限异常 → 降级，最新版仍保存成功。"""
        with patch("openpyxl.Workbook.save") as mock_save:
            # 第一次调用（最新版）成功，第二次（存档）失败
            mock_save.side_effect = [None, PermissionError("denied")]
            # 不应抛异常
            path = ew.save_workbook(self.wb, output_dir=self.tmp.name)
            self.assertIsNotNone(path)


class TestWriteRows(unittest.TestCase):
    """各类行写入函数测试。"""

    def setUp(self):
        self.wb = ew.create_workbook()
        self.ws = self.wb.active

    def test_write_title_row(self):
        """标题行写入并返回下一行号。"""
        row = ew.write_title_row(self.ws, 1, "测试标题", 5)
        self.assertEqual(row, 2)
        self.assertEqual(self.ws.cell(row=1, column=1).value, "测试标题")

    def test_write_header_row(self):
        """表头行写入。"""
        headers = ["名称", "代码", "价格"]
        row = ew.write_header_row(self.ws, 1, headers)
        self.assertEqual(row, 2)
        self.assertEqual(self.ws.cell(row=1, column=1).value, "名称")

    def test_write_data_row(self):
        """数据行写入。"""
        vals = ["长江电力", "600900", 26.65]
        row = ew.write_data_row(self.ws, 1, vals)
        self.assertEqual(row, 2)

    def test_write_data_row_with_fmts(self):
        """带格式的数据行写入。"""
        fmts = ["", "", FMT_MONEY]
        ew.write_data_row(self.ws, 1, [1, 2, 1234.56], fmts)
        cell = self.ws.cell(row=1, column=3)
        self.assertEqual(cell.value, 1234.56)

    def test_write_subtotal_row(self):
        """小计行写入。"""
        ew.write_title_row(self.ws, 1, "测试", 8)
        ew.write_header_row(self.ws, 2, ["A", "B", "C", "D", "E", "F", "G", "H"])
        vals = ["", "", "", "", 100.0, 80.0, 20.0, 0.25, 5.0]
        row = ew.write_subtotal_row(self.ws, 3, "测试小计", vals, 8)
        self.assertIsNotNone(row)

    def test_write_total_row(self):
        """总计行写入。"""
        ew.write_title_row(self.ws, 1, "测试", 8)
        ew.write_header_row(self.ws, 2, ["A", "B", "C", "D", "E", "F", "G", "H"])
        vals = ["", "", "", "", 1000.0, 800.0, 200.0, 0.25, 50.0]
        row = ew.write_total_row(self.ws, 3, "总计", vals, 8)
        self.assertIsNotNone(row)


class TestAutoWidth(unittest.TestCase):
    """auto_width 边界测试。"""

    def setUp(self):
        self.wb = ew.create_workbook()
        self.ws = self.wb.active

    def test_auto_width_basic(self):
        """基本列宽自适应。"""
        self.ws.cell(row=1, column=1, value="名称")
        self.ws.cell(row=2, column=1, value="长江电力")
        ew.auto_width(self.ws)
        col_letter = chr(64 + 1)  # A
        width = self.ws.column_dimensions[col_letter].width
        self.assertIsNotNone(width)

    def test_auto_width_cjk(self):
        """CJK 字符占用双倍宽度。"""
        self.ws.cell(row=1, column=1, value="长江电力投资有限公司")
        ew.auto_width(self.ws)
        width = self.ws.column_dimensions["A"].width
        self.assertGreater(width, 10)

    def test_auto_width_with_min_max(self):
        """min_width / max_width 参数。"""
        self.ws.cell(row=1, column=1, value="短")
        ew.auto_width(self.ws, min_width=10, max_width=40)
        width = self.ws.column_dimensions["A"].width
        self.assertGreaterEqual(width, 10)

    def test_auto_width_empty(self):
        """空工作表 → 不崩溃。"""
        try:
            ew.auto_width(self.ws)
        except Exception as e:
            self.fail(f"auto_width on empty sheet raised: {e}")


class TestFreezeHeader(unittest.TestCase):
    """freeze_header 测试。"""

    def setUp(self):
        self.wb = ew.create_workbook()
        self.ws = self.wb.active

    def test_freeze_header(self):
        """冻结首行。"""
        ew.freeze_header(self.ws, 2)
        self.assertIsNotNone(self.ws.freeze_panes)

    def test_freeze_header_empty(self):
        """空 sheet 冻结 → 不崩溃。"""
        try:
            ew.freeze_header(self.ws, 2)
        except Exception as e:
            self.fail(f"freeze_header raised: {e}")


if __name__ == "__main__":
    unittest.main()
