"""市值核算模块单元测试。

测试目标：
  - _is_qdii / _is_etf  — 基金类型识别
  - _date_within_days   — 日期范围判断
  - classify_holdings   — 持仓分类逻辑
  - price_update_status — 价格更新状态检测
  - is_market_open      — A 股交易时段判断
  - get_last_trading_day / get_prev_trading_day — 交易日计算
  - _determine_price_type — 取价方式标签生成
  - _generate_details   — 明细行生成（mock API）
  - _detail_to_row_values — 行值转换
  - _num_formats        — 格式列表
  - _apply_profit_colors — 盈亏着色
  - _apply_price_type_colors — 取价方式着色
  - write_market_value_sheet — 页签写入（mock 内部函数）

运行：
  cd D:/codebase/zoo/investor-util
  python -m unittest src.test_market_value -v
"""

from __future__ import annotations

import unittest
from datetime import datetime, timedelta
from typing import Any
from unittest.mock import MagicMock, call, patch

from openpyxl import Workbook

from src.models import Holding
from src.report import market_value as mv
from src.report.styles import BLUE_FONT


# ═══════════════════════════════════════════════════════════
#  _is_qdii
# ═══════════════════════════════════════════════════════════


class TestIsQdii(unittest.TestCase):
    """测试 _is_qdii 名称含 QDII 判断。"""

    def test_qdii_in_name(self):
        """名称含 QDII → True。"""
        self.assertTrue(mv._is_qdii("华夏纳斯达克100ETF(QDII)"))

    def test_qdii_lowercase(self):
        """名称含小写 qdii → True（大小写不敏感）。"""
        self.assertTrue(mv._is_qdii("华夏纳斯达克100ETF(qdii)"))

    def test_qdii_mixed_case(self):
        """名称含混合大小写 QdIi → True。"""
        self.assertTrue(mv._is_qdii("测试(QdIi)"))

    def test_non_qdii(self):
        """不含 QDII → False。"""
        self.assertFalse(mv._is_qdii("电池ETF"))

    def test_empty_string(self):
        """空字符串 → False。"""
        self.assertFalse(mv._is_qdii(""))

    def test_no_market_value_keyword(self):
        """含有其他相似关键词但不含 QDII → False。"""
        self.assertFalse(mv._is_qdii("QD股票基金"))


# ═══════════════════════════════════════════════════════════
#  _is_etf
# ═══════════════════════════════════════════════════════════


class TestIsEtf(unittest.TestCase):
    """测试 _is_etf 名称含 ETF 判断。"""

    def test_etf_in_name(self):
        """名称含 ETF → True。"""
        self.assertTrue(mv._is_etf("电池ETF"))

    def test_etf_lowercase(self):
        """名称含小写 etf → True。"""
        self.assertTrue(mv._is_etf("电池etf"))

    def test_etf_mixed_case(self):
        """名称含混合大小写 Etf → True。"""
        self.assertTrue(mv._is_etf("电池Etf"))

    def test_non_etf(self):
        """不含 ETF → False。"""
        self.assertFalse(mv._is_etf("长江电力"))

    def test_empty_string(self):
        """空字符串 → False。"""
        self.assertFalse(mv._is_etf(""))


# ═══════════════════════════════════════════════════════════
#  _date_within_days
# ═══════════════════════════════════════════════════════════


class TestDateWithinDays(unittest.TestCase):
    """测试 _date_within_days 日期范围判断。"""

    def test_same_day(self):
        """同一天 → True（0 天差）。"""
        self.assertTrue(mv._date_within_days("2026-06-26", "2026-06-26", 3))

    def test_within_days(self):
        """在范围内（1 天差）→ True。"""
        self.assertTrue(mv._date_within_days("2026-06-25", "2026-06-26", 3))

    def test_exactly_max_days(self):
        """恰好 max_days 天差 → True。"""
        self.assertTrue(mv._date_within_days("2026-06-23", "2026-06-26", 3))

    def test_beyond_max_days(self):
        """超过 max_days 天差 → False。"""
        self.assertFalse(mv._date_within_days("2026-06-22", "2026-06-26", 3))

    def test_future_date(self):
        """未来日期（负数天差）→ False。"""
        self.assertFalse(mv._date_within_days("2026-06-28", "2026-06-26", 3))

    def test_invalid_date_str(self):
        """无效日期字符串 → False。"""
        self.assertFalse(mv._date_within_days("not-a-date", "2026-06-26", 3))

    def test_empty_date_str(self):
        """空日期字符串 → False。"""
        self.assertFalse(mv._date_within_days("", "2026-06-26", 3))

    def test_none_date_str(self):
        """None 日期 → False（TypeError 分支）。"""
        self.assertFalse(mv._date_within_days(None, "2026-06-26", 3))

    def test_invalid_today_str(self):
        """无效 today_str → False。"""
        self.assertFalse(mv._date_within_days("2026-06-26", "", 3))

    def test_max_days_zero(self):
        """max_days=0，仅当天匹配 → True。"""
        self.assertTrue(mv._date_within_days("2026-06-26", "2026-06-26", 0))

    def test_max_days_zero_beyond(self):
        """max_days=0，差 1 天 → False。"""
        self.assertFalse(mv._date_within_days("2026-06-25", "2026-06-26", 0))


# ═══════════════════════════════════════════════════════════
#  classify_holdings
# ═══════════════════════════════════════════════════════════


class TestClassifyHoldings(unittest.TestCase):
    """测试 classify_holdings 按类型分类持仓。"""

    def _h(self, name: str, code: str = "", account: str = "证券账户") -> Holding:
        return Holding(
            account=account, name=name, code=code,
            shares=1.0, cost_price=1.0,
        )

    # ── QDII ─────────────────────────────────────────────

    def test_qdii_in_name(self):
        """名称含 QDII → QDII。"""
        h = self._h("华夏纳斯达克100ETF(QDII)", "513300")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["QDII"]), 1)
        self.assertEqual(result["QDII"][0], h)

    def test_qdii_lowercase(self):
        """名称含小写 qdii → QDII（大小写不敏感）。"""
        h = self._h("易方达标普500(Qdii)", "161125")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["QDII"]), 1)

    # ── 场外渠道 ─────────────────────────────────────────

    def test_fund_account(self):
        """基金账户 → 国内场外。"""
        h = self._h("某混合基金", "002943", "基金账户")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["国内场外"]), 1)

    def test_alipay_account(self):
        """支付宝账户 → 国内场外。"""
        h = self._h("中欧医疗健康混合", "003095", "支付宝")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["国内场外"]), 1)

    def test_wechat_account(self):
        """微信账户 → 国内场外。"""
        h = self._h("易方达蓝筹精选", "005827", "微信理财")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["国内场外"]), 1)

    def test_bank_account(self):
        """银行账户 → 国内场外。"""
        h = self._h("某稳健增长", "001234", "银行")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["国内场外"]), 1)

    # ── 场内 ETF ─────────────────────────────────────────

    def test_etf_in_name(self):
        """名称含 ETF → 场内ETF。"""
        h = self._h("电池ETF", "561910")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["场内ETF"]), 1)

    def test_code_starts_with_5(self):
        """代码 5 开头 → 场内ETF。"""
        h = self._h("黄金ETF", "518880")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["场内ETF"]), 1)

    def test_code_starts_with_1(self):
        """代码 1 开头 → 场内ETF。"""
        h = self._h("某转债", "110059")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["场内ETF"]), 1)

    # ── 场内股票 ─────────────────────────────────────────

    def test_stock_code_6(self):
        """代码 6 开头 → 场内股票。"""
        h = self._h("长江电力", "600900")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["场内股票"]), 1)

    def test_stock_code_0(self):
        """代码 0 开头 → 场内股票。"""
        h = self._h("平安银行", "000001")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["场内股票"]), 1)

    def test_stock_code_3(self):
        """代码 3 开头 → 场内股票。"""
        h = self._h("宁德时代", "300750")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["场内股票"]), 1)

    # ── 兜底 ─────────────────────────────────────────────

    def test_other_code_falls_back(self):
        """其余（代码非 6/0/3/5/1）→ 国内场外。"""
        h = self._h("某基金", "888888")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["国内场外"]), 1)

    # ── 优先级：QDII > 场外渠道 > ETF > 股票 > 兜底 ────

    def test_qdii_priority_over_fund_account(self):
        """QDII 优先级高于场外渠道。"""
        h = self._h("标普500ETF(QDII)", "161125", "支付宝")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["QDII"]), 1)

    def test_account_priority_over_code(self):
        """场外渠道账户优先级高于代码匹配。"""
        h = self._h("某基金", "600900", "支付宝")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["国内场外"]), 1)

    def test_qdii_priority_over_etf(self):
        """QDII 优先级高于 ETF 名称匹配。"""
        h = self._h("恒生ETF(QDII)", "159920")
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["QDII"]), 1)

    # ── 边缘情况 ─────────────────────────────────────────

    def test_empty_holdings(self):
        """空列表 → 所有分类为空。"""
        result = mv.classify_holdings([])
        for cat in result.values():
            self.assertEqual(len(cat), 0)

    def test_whitespace_stripped(self):
        """持仓名称/代码/账户的空格被清理。"""
        h = Holding(account="  证券账户  ", name="  电池ETF  ", code="  561910  ",
                    shares=1.0, cost_price=1.0)
        result = mv.classify_holdings([h])
        self.assertEqual(len(result["场内ETF"]), 1)

    def test_mixed_holdings(self):
        """多种类型混合分类正确。"""
        holdings = [
            self._h("电池ETF", "561910"),
            self._h("长江电力", "600900"),
            self._h("华夏纳斯达克100ETF(QDII)", "513300"),
            self._h("中欧医疗健康混合", "003095", "支付宝"),
        ]
        result = mv.classify_holdings(holdings)
        self.assertEqual(len(result["QDII"]), 1)
        self.assertEqual(len(result["场内ETF"]), 1)
        self.assertEqual(len(result["场内股票"]), 1)
        self.assertEqual(len(result["国内场外"]), 1)


# ═══════════════════════════════════════════════════════════
#  price_update_status
# ═══════════════════════════════════════════════════════════


class TestPriceUpdateStatus(unittest.TestCase):
    """测试 price_update_status 价格更新状态检测。"""

    def _row(self, source_api: str, nav_date: str, name: str = "") -> mv.DetailRow:
        return mv.DetailRow(
            source_api=source_api, nav_date=nav_date, name=name,
        )

    # ── Tencent（场内）────────────────────────────────────

    def test_tencent_updated(self):
        """tencent + nav_date == trading_day → 已更新。"""
        d = self._row("tencent", "2026-06-26")
        updated, total, all_ok = mv.price_update_status([d], "2026-06-26")
        self.assertEqual(updated, 1)
        self.assertEqual(total, 1)
        self.assertTrue(all_ok)

    def test_tencent_not_updated(self):
        """tencent + nav_date != trading_day → 未更新。"""
        d = self._row("tencent", "2026-06-25")
        updated, total, all_ok = mv.price_update_status([d], "2026-06-26")
        self.assertEqual(updated, 0)
        self.assertEqual(total, 1)
        self.assertFalse(all_ok)

    def test_tencent_empty_nav_date(self):
        """tencent + 空 nav_date → 未更新。"""
        d = self._row("tencent", "")
        updated, _, _ = mv.price_update_status([d], "2026-06-26")
        self.assertEqual(updated, 0)

    # ── EastMoney + QDII ────────────────────────────────

    def test_eastmoney_qdii_within_3_days(self):
        """eastmoney + QDII + nav_date 在 3 天内 → 已更新。"""
        d = self._row("eastmoney", "2026-06-24", name="标普500(QDII)")
        updated, _, _ = mv.price_update_status([d], "2026-06-26")
        self.assertEqual(updated, 1)

    def test_eastmoney_qdii_exactly_3_days(self):
        """eastmoney + QDII + nav_date 恰好 3 天 → 已更新。"""
        d = self._row("eastmoney", "2026-06-23", name="标普500(QDII)")
        updated, _, _ = mv.price_update_status([d], "2026-06-26")
        self.assertEqual(updated, 1)

    def test_eastmoney_qdii_beyond_3_days(self):
        """eastmoney + QDII + nav_date 超出 3 天 → 未更新。"""
        d = self._row("eastmoney", "2026-06-22", name="标普500(QDII)")
        updated, _, _ = mv.price_update_status([d], "2026-06-26")
        self.assertEqual(updated, 0)

    def test_eastmoney_qdii_empty_nav_date(self):
        """eastmoney + QDII + 空 nav_date → 未更新（nav_date 为 falsy）。"""
        d = self._row("eastmoney", "", name="标普500(QDII)")
        updated, _, _ = mv.price_update_status([d], "2026-06-26")
        self.assertEqual(updated, 0)

    # ── EastMoney + 非 QDII（国内场外）────────────────────

    def test_eastmoney_domestic_equal_trading_day(self):
        """eastmoney + 非 QDII + nav_date == trading_day → 已更新。"""
        d = self._row("eastmoney", "2026-06-26", name="中欧医疗健康混合")
        updated, _, _ = mv.price_update_status([d], "2026-06-26")
        self.assertEqual(updated, 1)

    def test_eastmoney_domestic_equal_prev_trading_day(self):
        """eastmoney + 非 QDII + nav_date == prev_trading_day → 已更新。"""
        # 周三交易日，周二为前一日
        d = self._row("eastmoney", "2026-06-23", name="中欧医疗健康混合")
        updated, _, _ = mv.price_update_status([d], "2026-06-24")
        self.assertEqual(updated, 1)

    def test_eastmoney_domestic_neither(self):
        """eastmoney + 非 QDII + nav_date 不是 T 也不是 T-1 → 未更新。"""
        d = self._row("eastmoney", "2026-06-22", name="中欧医疗健康混合")
        updated, _, _ = mv.price_update_status([d], "2026-06-26")
        self.assertEqual(updated, 0)

    # ── 混合场景 ───────────────────────────────────────────

    def test_mixed_status(self):
        """部分更新 → all_ok 为 False。"""
        details = [
            self._row("tencent", "2026-06-26"),        # 已更新
            self._row("tencent", "2026-06-25"),          # 未更新
            self._row("eastmoney", "2026-06-26", name="某基金"),  # 已更新
        ]
        updated, total, all_ok = mv.price_update_status(details, "2026-06-26")
        self.assertEqual(updated, 2)
        self.assertEqual(total, 3)
        self.assertFalse(all_ok)

    def test_all_updated(self):
        """全部已更新 → all_ok 为 True。"""
        details = [
            self._row("tencent", "2026-06-26"),
            self._row("eastmoney", "2026-06-26", name="某基金"),
        ]
        _, _, all_ok = mv.price_update_status(details, "2026-06-26")
        self.assertTrue(all_ok)

    def test_all_not_updated(self):
        """全部未更新 → all_ok 为 False。"""
        details = [
            self._row("tencent", "2026-06-25"),
            self._row("tencent", "2026-06-24"),
        ]
        _, _, all_ok = mv.price_update_status(details, "2026-06-26")
        self.assertFalse(all_ok)

    def test_empty_list(self):
        """空列表 → (0, 0, True)。"""
        updated, total, all_ok = mv.price_update_status([], "2026-06-26")
        self.assertEqual(updated, 0)
        self.assertEqual(total, 0)
        self.assertTrue(all_ok)

    def test_unknown_source_api_ignored(self):
        """未知 source_api → 不计入已更新。"""
        d = self._row("unknown_api", "2026-06-26")
        updated, total, _ = mv.price_update_status([d], "2026-06-26")
        self.assertEqual(updated, 0)
        self.assertEqual(total, 1)


# ═══════════════════════════════════════════════════════════
#  is_market_open
# ═══════════════════════════════════════════════════════════


class TestIsMarketOpen(unittest.TestCase):
    """测试 is_market_open A 股交易时段判断（mock datetime.now）。"""

    @patch("src.report.market_value.datetime")
    def test_weekend_saturday(self, mock_dt):
        """周六 → False。"""
        mock_dt.now.return_value = datetime(2026, 6, 27, 10, 0, 0)  # Saturday
        self.assertFalse(mv.is_market_open())

    @patch("src.report.market_value.datetime")
    def test_weekend_sunday(self, mock_dt):
        """周日 → False。"""
        mock_dt.now.return_value = datetime(2026, 6, 28, 10, 0, 0)  # Sunday
        self.assertFalse(mv.is_market_open())

    @patch("src.report.market_value.datetime")
    def test_before_open(self, mock_dt):
        """周一 9:00（开盘前）→ False。"""
        mock_dt.now.return_value = datetime(2026, 6, 22, 9, 0, 0)  # Mon
        self.assertFalse(mv.is_market_open())

    @patch("src.report.market_value.datetime")
    def test_morning_session(self, mock_dt):
        """周一 10:00（上午交易时段）→ True。"""
        mock_dt.now.return_value = datetime(2026, 6, 22, 10, 0, 0)
        self.assertTrue(mv.is_market_open())

    @patch("src.report.market_value.datetime")
    def test_morning_open_boundary(self, mock_dt):
        """周一 9:30（开盘边界）→ True。"""
        mock_dt.now.return_value = datetime(2026, 6, 22, 9, 30, 0)
        self.assertTrue(mv.is_market_open())

    @patch("src.report.market_value.datetime")
    def test_morning_close_boundary(self, mock_dt):
        """周一 11:30（午休边界）→ True。"""
        mock_dt.now.return_value = datetime(2026, 6, 22, 11, 30, 0)
        self.assertTrue(mv.is_market_open())

    @patch("src.report.market_value.datetime")
    def test_lunch_break(self, mock_dt):
        """周一 12:00（午休）→ False。"""
        mock_dt.now.return_value = datetime(2026, 6, 22, 12, 0, 0)
        self.assertFalse(mv.is_market_open())

    @patch("src.report.market_value.datetime")
    def test_afternoon_session(self, mock_dt):
        """周一 14:00（下午交易时段）→ True。"""
        mock_dt.now.return_value = datetime(2026, 6, 22, 14, 0, 0)
        self.assertTrue(mv.is_market_open())

    @patch("src.report.market_value.datetime")
    def test_afternoon_open_boundary(self, mock_dt):
        """周一 13:00（下午开盘边界）→ True。"""
        mock_dt.now.return_value = datetime(2026, 6, 22, 13, 0, 0)
        self.assertTrue(mv.is_market_open())

    @patch("src.report.market_value.datetime")
    def test_afternoon_close_boundary(self, mock_dt):
        """周一 15:00（收盘边界）→ True。"""
        mock_dt.now.return_value = datetime(2026, 6, 22, 15, 0, 0)
        self.assertTrue(mv.is_market_open())

    @patch("src.report.market_value.datetime")
    def test_after_close(self, mock_dt):
        """周一 15:30（收盘后）→ False。"""
        mock_dt.now.return_value = datetime(2026, 6, 22, 15, 30, 0)
        self.assertFalse(mv.is_market_open())


# ═══════════════════════════════════════════════════════════
#  get_last_trading_day
# ═══════════════════════════════════════════════════════════


class TestGetLastTradingDay(unittest.TestCase):
    """测试 get_last_trading_day 最近交易日计算（mock datetime.now）。"""

    @patch("src.report.market_value.datetime")
    def test_saturday(self, mock_dt):
        """周六 → 上周五。"""
        mock_dt.now.return_value = datetime(2026, 6, 27, 10, 0, 0)
        self.assertEqual(mv.get_last_trading_day(), "2026-06-26")

    @patch("src.report.market_value.datetime")
    def test_sunday(self, mock_dt):
        """周日 → 上周五。"""
        mock_dt.now.return_value = datetime(2026, 6, 28, 10, 0, 0)
        self.assertEqual(mv.get_last_trading_day(), "2026-06-26")

    @patch("src.report.market_value.datetime")
    def test_monday(self, mock_dt):
        """周一 → 当天。"""
        mock_dt.now.return_value = datetime(2026, 6, 22, 10, 0, 0)
        self.assertEqual(mv.get_last_trading_day(), "2026-06-22")

    @patch("src.report.market_value.datetime")
    def test_wednesday(self, mock_dt):
        """周三 → 当天。"""
        mock_dt.now.return_value = datetime(2026, 6, 24, 10, 0, 0)
        self.assertEqual(mv.get_last_trading_day(), "2026-06-24")

    @patch("src.report.market_value.datetime")
    def test_friday(self, mock_dt):
        """周五 → 当天。"""
        mock_dt.now.return_value = datetime(2026, 6, 26, 10, 0, 0)
        self.assertEqual(mv.get_last_trading_day(), "2026-06-26")


# ═══════════════════════════════════════════════════════════
#  get_prev_trading_day
# ═══════════════════════════════════════════════════════════


class TestGetPrevTradingDay(unittest.TestCase):
    """测试 get_prev_trading_day 前一交易日计算。"""

    def test_monday_to_friday(self):
        """周一 → 上周五（减 3 天）。"""
        self.assertEqual(mv.get_prev_trading_day("2026-06-22"), "2026-06-19")

    def test_tuesday_to_monday(self):
        """周二 → 周一（减 1 天）。"""
        self.assertEqual(mv.get_prev_trading_day("2026-06-23"), "2026-06-22")

    def test_wednesday_to_tuesday(self):
        """周三 → 周二。"""
        self.assertEqual(mv.get_prev_trading_day("2026-06-24"), "2026-06-23")

    def test_thursday_to_wednesday(self):
        """周四 → 周三。"""
        self.assertEqual(mv.get_prev_trading_day("2026-06-25"), "2026-06-24")

    def test_friday_to_thursday(self):
        """周五 → 周四。"""
        self.assertEqual(mv.get_prev_trading_day("2026-06-26"), "2026-06-25")

    def test_saturday_to_friday(self):
        """周六 → 周五。"""
        self.assertEqual(mv.get_prev_trading_day("2026-06-27"), "2026-06-26")

    def test_sunday_to_saturday(self):
        """周日 → 周六。"""
        self.assertEqual(mv.get_prev_trading_day("2026-06-28"), "2026-06-27")

    def test_empty_string_calls_get_last_trading_day(self):
        """空字符串 → 调用 get_last_trading_day。"""
        with patch("src.report.market_value.get_last_trading_day") as mock_ltd:
            mock_ltd.return_value = "2026-06-26"
            result = mv.get_prev_trading_day("")
            self.assertEqual(result, "2026-06-25")
            mock_ltd.assert_called_once()

    def test_invalid_date(self):
        """无效日期字符串 → 返回空字符串。"""
        self.assertEqual(mv.get_prev_trading_day("not-a-date"), "")

    def test_none_date(self):
        """None 作为日期 → falsy 判断触发，回退到 get_last_trading_day（不会进入异常分支）。"""
        with patch("src.report.market_value.get_last_trading_day") as mock_ltd:
            mock_ltd.return_value = "2026-06-26"
            result = mv.get_prev_trading_day(None)
            self.assertEqual(result, "2026-06-25")


# ═══════════════════════════════════════════════════════════
#  _determine_price_type
# ═══════════════════════════════════════════════════════════


class TestDeterminePriceType(unittest.TestCase):
    """测试 _determine_price_type 取价方式标签生成。

    需要 mock is_market_open 和 get_prev_trading_day。
    """

    def setUp(self):
        self.td = "2026-06-26"   # Friday
        self.prev = "2026-06-25"  # Thursday

    # ── Tencent ───────────────────────────────────────────

    @patch("src.report.market_value.is_market_open", return_value=True)
    def test_tencent_intraday(self, _):
        """tencent + 交易时段 → 场内实时价。"""
        result = mv._determine_price_type("tencent", self.td, self.td)
        self.assertEqual(result, "场内实时价")

    @patch("src.report.market_value.is_market_open", return_value=False)
    def test_tencent_closed_no_nav_date(self, _):
        """tencent + 已收市 + 无净值日期 → 场内收盘价(--)。"""
        with patch("src.report.market_value.get_prev_trading_day",
                   return_value=self.prev):
            result = mv._determine_price_type("tencent", "", self.td)
            self.assertEqual(result, "场内收盘价(--)")

    @patch("src.report.market_value.is_market_open", return_value=False)
    def test_tencent_closed_nav_today(self, _):
        """tencent + 已收市 + nav_date == T → 场内收盘价(T)。"""
        result = mv._determine_price_type("tencent", self.td, self.td)
        self.assertEqual(result, "场内收盘价(T)")

    @patch("src.report.market_value.is_market_open", return_value=False)
    def test_tencent_closed_nav_prev(self, _):
        """tencent + 已收市 + nav_date == T-1 → 场内收盘价(T-1)。"""
        with patch("src.report.market_value.get_prev_trading_day",
                   return_value=self.prev):
            result = mv._determine_price_type("tencent", self.prev, self.td)
            self.assertEqual(result, "场内收盘价(T-1)")

    @patch("src.report.market_value.is_market_open", return_value=False)
    def test_tencent_closed_nav_other(self, _):
        """tencent + 已收市 + nav_date 为其他日期 → 场内收盘价(date)。"""
        result = mv._determine_price_type("tencent", "2026-06-20", self.td)
        self.assertEqual(result, "场内收盘价(2026-06-20)")

    # ── EastMoney（场外）──────────────────────────────────

    def test_eastmoney_no_nav_date(self):
        """eastmoney + 空 nav_date → 官方净值(--)。"""
        result = mv._determine_price_type("eastmoney", "", self.td)
        self.assertEqual(result, "官方净值(--)")

    def test_eastmoney_nav_today(self):
        """eastmoney + nav_date == T → 官方净值(T)。"""
        result = mv._determine_price_type("eastmoney", self.td, self.td)
        self.assertEqual(result, "官方净值(T)")

    def test_eastmoney_nav_prev(self):
        """eastmoney + nav_date == T-1 → 官方净值(T-1)。"""
        with patch("src.report.market_value.get_prev_trading_day",
                   return_value=self.prev):
            result = mv._determine_price_type("eastmoney", self.prev, self.td)
            self.assertEqual(result, "官方净值(T-1)")

    def test_eastmoney_nav_2_days_ago(self):
        """eastmoney + nav_date 2 天前 → 官方净值(T-2)。"""
        result = mv._determine_price_type("eastmoney", "2026-06-24", self.td)
        self.assertEqual(result, "官方净值(T-2)")

    def test_eastmoney_nav_5_days_ago(self):
        """eastmoney + nav_date 5 天前 → 官方净值(T-5)。"""
        result = mv._determine_price_type("eastmoney", "2026-06-21", self.td)
        self.assertEqual(result, "官方净值(T-5)")

    def test_eastmoney_nav_6_days_ago(self):
        """eastmoney + nav_date 6 天前 → 官方净值(date)。"""
        result = mv._determine_price_type("eastmoney", "2026-06-20", self.td)
        self.assertEqual(result, "官方净值(2026-06-20)")

    def test_eastmoney_nav_invalid_format(self):
        """eastmoney + 无效日期格式 → 官方净值(原字符串)（ValueError 分支）。"""
        result = mv._determine_price_type("eastmoney", "invalid-date", self.td)
        self.assertEqual(result, "官方净值(invalid-date)")

    def test_eastmoney_nav_future(self):
        """eastmoney + 未来日期（days_diff < 0）→ 官方净值(T)。"""
        result = mv._determine_price_type("eastmoney", "2026-06-30", self.td)
        self.assertEqual(result, "官方净值(T)")


# ═══════════════════════════════════════════════════════════
#  _generate_details
# ═══════════════════════════════════════════════════════════


class TestGenerateDetails(unittest.TestCase):
    """测试 _generate_details 明细行生成（mock API 调用）。"""

    def setUp(self):
        self.tencent_mock_data = {
            "name": "电池ETF", "code": "561910",
            "price": 10.5, "yesterday_close": 10.0,
            "price_date": "2026-06-26",
            "source_api": "tencent", "source": "腾讯财经",
        }
        self.eastmoney_mock_data = {
            "name": "中欧医疗健康混合", "code": "003095",
            "price": 1.5, "yesterday_close": 1.48,
            "price_date": "2026-06-26",
            "source_api": "eastmoney", "source": "东方财富",
        }

    @patch("src.report.market_value.fetch_market_data")
    @patch("src.report.market_value.get_last_trading_day")
    @patch("src.report.market_value.is_market_open")
    def test_tencent_asset(self, mock_open, mock_ltd, mock_fetch):
        """Tencent 场内资产：各字段正确赋值，today_profit 按公式计算。"""
        mock_open.return_value = False
        mock_ltd.return_value = "2026-06-26"
        mock_fetch.return_value = self.tencent_mock_data

        h = Holding("证券账户", "电池ETF", "561910", 1000.0, 1.0)
        details = mv._generate_details([h], "2026-06-26")
        self.assertEqual(len(details), 1)

        d = details[0]
        self.assertEqual(d.account, "证券账户")
        self.assertEqual(d.name, "电池ETF")
        self.assertEqual(d.code, "561910")
        self.assertEqual(d.price, 10.5)
        self.assertEqual(d.nav_date, "2026-06-26")
        self.assertEqual(d.yesterday_close, 10.0)
        self.assertEqual(d.source_api, "tencent")
        self.assertEqual(d.source, "腾讯财经")
        # tencent + 已收市 + nav_date==T → "场内收盘价(T)"
        self.assertEqual(d.price_type, "场内收盘价(T)")
        self.assertEqual(d.premium, "--")
        self.assertEqual(d.shares, 1000.0)
        self.assertEqual(d.market_value, 10500.0)       # 10.5 * 1000
        self.assertEqual(d.cost, 1000.0)                 # 1.0 * 1000
        self.assertEqual(d.profit, 9500.0)               # 10500 - 1000
        self.assertAlmostEqual(d.profit_rate, 9.5)       # 9500 / 1000
        # today_profit = (10.5 - 10.0) * 1000
        self.assertEqual(d.today_profit, 500.0)

    @patch("src.report.market_value.fetch_market_data")
    @patch("src.report.market_value.get_last_trading_day")
    @patch("src.report.market_value.is_market_open")
    def test_eastmoney_asset(self, mock_open, mock_ltd, mock_fetch):
        """EastMoney 场外资产：字段正确赋值。"""
        mock_open.return_value = False
        mock_ltd.return_value = "2026-06-26"
        mock_fetch.return_value = self.eastmoney_mock_data

        h = Holding("支付宝", "中欧医疗健康混合", "003095", 500.0, 2.0)
        details = mv._generate_details([h], "2026-06-26")
        self.assertEqual(len(details), 1)

        d = details[0]
        self.assertEqual(d.account, "支付宝")
        self.assertEqual(d.source_api, "eastmoney")
        # eastmoney + nav_date==T → "官方净值(T)"
        self.assertEqual(d.price_type, "官方净值(T)")
        self.assertEqual(d.cost, 1000.0)                 # 2.0 * 500
        self.assertEqual(d.market_value, 750.0)           # 1.5 * 500
        self.assertEqual(d.profit, -250.0)                # 750 - 1000
        # nav_date == T → today_profit = (1.5 - 1.48) * 500
        self.assertEqual(d.today_profit, 10.0)

    @patch("src.report.market_value.fetch_market_data")
    @patch("src.report.market_value.get_last_trading_day")
    def test_eastmoney_stale_nav(self, mock_ltd, mock_fetch):
        """East Money + nav_date 过期 → today_profit 为 0。"""
        mock_ltd.return_value = "2026-06-26"
        mock_fetch.return_value = {
            "name": "中欧医疗健康混合", "code": "003095",
            "price": 1.5, "yesterday_close": 1.48,
            "price_date": "2026-06-20",          # 6 天前，过期数据
            "source_api": "eastmoney", "source": "东方财富",
        }

        h = Holding("支付宝", "中欧医疗健康混合", "003095", 100.0, 1.0)
        details = mv._generate_details([h], "2026-06-26")
        self.assertEqual(details[0].today_profit, 0.0)
        # price_type 也是 T-6 对应的格式
        self.assertEqual(details[0].price_type, "官方净值(2026-06-20)")

    @patch("src.report.market_value.fetch_market_data")
    def test_fetch_returns_none(self, mock_fetch):
        """API 返回 None → 默认值填充，日志告警。"""
        mock_fetch.return_value = None

        h = Holding("证券账户", "电池ETF", "561910", 1000.0, 1.0)
        details = mv._generate_details([h], "2026-06-26")
        self.assertEqual(len(details), 1)

        d = details[0]
        self.assertEqual(d.price, 0.0)
        self.assertEqual(d.yesterday_close, 0.0)
        self.assertEqual(d.nav_date, "")
        self.assertEqual(d.source, "--")
        self.assertEqual(d.source_api, "")
        self.assertEqual(d.price_type, "--")
        self.assertEqual(d.today_profit, 0.0)

    @patch("src.report.market_value.fetch_market_data")
    @patch("src.report.market_value.get_last_trading_day")
    @patch("src.report.market_value.is_market_open")
    def test_multiple_holdings(self, mock_open, mock_ltd, mock_fetch):
        """多个持仓 → 返回多条明细。"""
        mock_open.return_value = False
        mock_ltd.return_value = "2026-06-26"

        def side_effect(code, name=""):
            if code == "561910":
                return self.tencent_mock_data
            elif code == "003095":
                return self.eastmoney_mock_data
            return None
        mock_fetch.side_effect = side_effect

        holdings = [
            Holding("证券账户", "电池ETF", "561910", 100.0, 1.0),
            Holding("支付宝", "中欧医疗健康混合", "003095", 200.0, 1.0),
        ]
        details = mv._generate_details(holdings, "2026-06-26")
        self.assertEqual(len(details), 2)
        self.assertEqual(details[0].source_api, "tencent")
        self.assertEqual(details[1].source_api, "eastmoney")

    @patch("src.report.market_value.fetch_market_data")
    def test_empty_holdings(self, mock_fetch):
        """空持仓列表 → 返回空列表。"""
        details = mv._generate_details([], "2026-06-26")
        self.assertEqual(details, [])

    @patch("src.report.market_value.fetch_market_data")
    @patch("src.report.market_value.get_last_trading_day")
    @patch("src.report.market_value.is_market_open")
    def test_today_str_none(self, mock_open, mock_ltd, mock_fetch):
        """today_str 为空（默认取当天）→ 不报错。"""
        mock_open.return_value = False
        mock_ltd.return_value = "2026-06-26"
        mock_fetch.return_value = self.tencent_mock_data

        h = Holding("证券账户", "电池ETF", "561910", 100.0, 1.0)
        # 传入空字符串，函数内部回退到 datetime.now()
        with patch("src.report.market_value.datetime") as mock_dt:
            mock_dt.now.return_value = datetime(2026, 6, 26, 15, 30, 0)
            details = mv._generate_details([h], "")
        self.assertEqual(len(details), 1)

    @patch("src.report.market_value.fetch_market_data")
    @patch("src.report.market_value.get_last_trading_day")
    @patch("src.report.market_value.is_market_open")
    def test_account_strip(self, mock_open, mock_ltd, mock_fetch):
        """账户名前后空格被清理。"""
        mock_open.return_value = False
        mock_ltd.return_value = "2026-06-26"
        mock_fetch.return_value = self.tencent_mock_data

        h = Holding("  证券账户  ", "电池ETF", "561910", 100.0, 1.0)
        details = mv._generate_details([h], "2026-06-26")
        self.assertEqual(details[0].account, "证券账户")


# ═══════════════════════════════════════════════════════════
#  _detail_to_row_values
# ═══════════════════════════════════════════════════════════


class TestDetailToRowValues(unittest.TestCase):
    """测试 _detail_to_row_values 行值转换。"""

    def test_full_detail_row(self):
        """完整 DetailRow → 15 个字段的列表。"""
        d = mv.DetailRow(
            account="证券账户", name="电池ETF", code="561910",
            price=10.5, nav_date="2026-06-26", yesterday_close=10.0,
            price_type="场内收盘价(T)", premium="--",
            shares=1000.0, market_value=10500.0, cost=1000.0,
            profit=9500.0, profit_rate=9.5, today_profit=500.0,
            source="腾讯财经", source_api="tencent",
        )
        vals = mv._detail_to_row_values(d)
        self.assertEqual(len(vals), 15)
        self.assertEqual(vals[0], "证券账户")
        self.assertEqual(vals[1], "电池ETF")
        self.assertEqual(vals[2], "561910")
        self.assertEqual(vals[3], 10.5)
        self.assertEqual(vals[4], "2026-06-26")
        self.assertEqual(vals[5], 10.0)
        self.assertEqual(vals[6], "场内收盘价(T)")
        self.assertEqual(vals[7], "--")
        self.assertEqual(vals[8], 1000.0)
        self.assertEqual(vals[9], 10500.0)
        self.assertEqual(vals[10], 1000.0)
        self.assertEqual(vals[11], 9500.0)
        self.assertEqual(vals[12], 9.5)
        self.assertEqual(vals[13], 500.0)
        self.assertEqual(vals[14], "腾讯财经")

    def test_default_row(self):
        """默认 DetailRow → 空字符串/0 值列表。"""
        d = mv.DetailRow()
        vals = mv._detail_to_row_values(d)
        self.assertEqual(len(vals), 15)
        self.assertEqual(vals[0], "")
        self.assertEqual(vals[3], 0.0)
        self.assertEqual(vals[7], "--")
        self.assertEqual(vals[12], 0.0)


# ═══════════════════════════════════════════════════════════
#  _num_formats
# ═══════════════════════════════════════════════════════════


class TestNumFormats(unittest.TestCase):
    """测试 _num_formats 返回正确长度的格式列表。"""

    def test_length(self):
        """返回 15 个格式。"""
        fmts = mv._num_formats()
        self.assertEqual(len(fmts), 15)

    def test_price_format(self):
        """第 4 个为价格格式。"""
        fmts = mv._num_formats()
        self.assertEqual(fmts[3], '#,##0.0000')

    def test_money_format(self):
        """市值/成本/盈亏列金额格式。"""
        fmts = mv._num_formats()
        self.assertEqual(fmts[9], '#,##0.00')
        self.assertEqual(fmts[10], '#,##0.00')
        self.assertEqual(fmts[11], '#,##0.00')

    def test_percent_format(self):
        """收益率列百分比格式。"""
        fmts = mv._num_formats()
        self.assertEqual(fmts[12], '0.00%')

    def test_shares_format(self):
        """份额列格式。"""
        fmts = mv._num_formats()
        self.assertEqual(fmts[8], '#,##0.00')


# ═══════════════════════════════════════════════════════════
#  _apply_profit_colors
# ═══════════════════════════════════════════════════════════


class TestApplyProfitColors(unittest.TestCase):
    """测试 _apply_profit_colors 盈亏列着色（mock ws / profit_font）。"""

    def _make_cell(self, value: Any = None) -> MagicMock:
        cell = MagicMock()
        cell.value = value
        return cell

    @patch("src.report.market_value.profit_font")
    def test_positive_profit(self, mock_pf):
        """正盈亏 → 调用 profit_font(正数)。"""
        mock_pf.side_effect = lambda v: f"font_for_{v}"
        ws = MagicMock()
        ws.cell.side_effect = lambda row, column: self._make_cell(500.0)

        mv._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)

        # 验证 profit 列和 today_profit 列调用了 profit_font(500.0)
        mock_pf.assert_any_call(500.0)
        # 3 行 × 2 列（profit col + today col）+ 1 列 rate = 3 次 profit_font 调用
        self.assertEqual(mock_pf.call_count, 3)

    @patch("src.report.market_value.profit_font")
    def test_negative_profit(self, mock_pf):
        """负盈亏 → 调用 profit_font(负数)。"""
        mock_pf.side_effect = lambda v: f"font_for_{v}"
        ws = MagicMock()
        ws.cell.side_effect = lambda row, column: self._make_cell(-300.0)

        mv._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_any_call(-300.0)

    @patch("src.report.market_value.profit_font")
    def test_zero_profit(self, mock_pf):
        """零盈亏 → 调用 profit_font(0)。"""
        mock_pf.side_effect = lambda v: f"font_for_{v}"
        ws = MagicMock()
        ws.cell.side_effect = lambda row, column: self._make_cell(0.0)

        mv._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_any_call(0.0)

    @patch("src.report.market_value.profit_font")
    def test_non_numeric_skipped(self, mock_pf):
        """非数字值（字符串）→ 不设置字体。"""
        ws = MagicMock()
        # profit col 返回字符串，today col 返回字符串
        values = {"12": "亏损", "14": "盈利"}
        def cell_side_effect(row, column):
            cell = MagicMock()
            cell.value = values.get(str(column))
            return cell
        ws.cell.side_effect = cell_side_effect

        mv._apply_profit_colors(ws, 3, 4, profit_col=12, rate_col=13, today_col=14)
        # profit_font 从未被调用（cell.value 不是 int/float）
        mock_pf.assert_not_called()

    @patch("src.report.market_value.profit_font")
    def test_none_value_skipped(self, mock_pf):
        """None 值 → 不设置字体。"""
        ws = MagicMock()
        ws.cell.return_value = self._make_cell(None)

        mv._apply_profit_colors(ws, 3, 4, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_not_called()

    @patch("src.report.market_value.profit_font")
    def test_rate_col_float(self, mock_pf):
        """收益率列为 float → 调用 profit_font。"""
        mock_pf.side_effect = lambda v: f"font_for_{v}"
        ws = MagicMock()
        ws.cell.side_effect = lambda row, column: self._make_cell(0.05)

        mv._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_called_with(0.05)

    @patch("src.report.market_value.profit_font")
    def test_rate_col_not_float(self, mock_pf):
        """收益率列非 float（如字符串含 %）→ 不设置字体。"""
        ws = MagicMock()
        ws.cell.side_effect = lambda row, column: self._make_cell("5.00%")

        mv._apply_profit_colors(ws, 3, 4, profit_col=12, rate_col=13, today_col=14)
        mock_pf.assert_not_called()

    @patch("src.report.market_value.profit_font")
    def test_multiple_rows(self, mock_pf):
        """多行数据 → 每行都着色。"""
        mock_pf.side_effect = lambda v: f"font_for_{v}"

        ws = MagicMock()
        row_values = {3: 100.0, 4: -50.0, 5: 200.0}
        def cell_side_effect(row, column):
            cell = MagicMock()
            cell.value = row_values.get(row, 0.0)
            return cell
        ws.cell.side_effect = cell_side_effect

        mv._apply_profit_colors(ws, 3, 5, profit_col=12, rate_col=13, today_col=14)
        # 3 行 × 3 列 = 9 次 profit_font 调用
        self.assertEqual(mock_pf.call_count, 9)

    def test_cell_font_assigned(self):
        """确保 cell.font 被赋值。"""
        ws = MagicMock()
        cell = MagicMock()
        cell.value = 100.0
        ws.cell.return_value = cell

        with patch("src.report.market_value.profit_font") as mock_pf:
            mock_pf.return_value = "red_font"
            mv._apply_profit_colors(ws, 3, 3, profit_col=12, rate_col=13, today_col=14)

        self.assertEqual(cell.font, "red_font")


# ═══════════════════════════════════════════════════════════
#  _apply_price_type_colors
# ═══════════════════════════════════════════════════════════


class TestApplyPriceTypeColors(unittest.TestCase):
    """测试 _apply_price_type_colors 取价方式列着色（使用真实 openpyxl Worksheet）。"""

    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        # 数据行（行 2 起，行 1 保留给标题）
        # 每行数据：(名称, 取价方式)
        self.test_cases = [
            ("电池ETF", "场内收盘价(T)", True),       # 场内收盘价(T) → 蓝
            ("长江电力", "场内收盘价(T-1)", False),    # 非 QDII + T-1 → 不蓝
            ("中欧医疗", "官方净值(T)", True),          # 官方净值(T) → 蓝
            ("某混合基金", "官方净值(T-1)", False),     # 非 QDII + T-1 → 不蓝
            ("标普500ETF(QDII)", "官方净值(T-1)", True),  # QDII + T-1 → 蓝
            ("恒生ETF(QDII)", "场内收盘价(T-1)", False),  # QDII + 场内 T-1 → 不蓝
            ("宁德时代", "场内实时价", False),          # 实时价 → 不蓝
            ("海外收益(QDII)", "官方净值(T-2)", False),  # QDII + 过期 → 不蓝
            ("--", "--", False),                       # 占位符 → 不蓝
        ]
        for i, (name, price_type, expected_blue) in enumerate(self.test_cases):
            row = i + 2
            self.ws.cell(row=row, column=2, value=name)
            self.ws.cell(row=row, column=7, value=price_type)

    def _assert_blue(self, row: int, msg: str = ""):
        cell = self.ws.cell(row=row, column=7)
        self.assertIsNotNone(cell.font.color, f"Row {row} font.color is None")
        # openpyxl 写入时 0066CC 存储为 ARGB 格式 000066CC
        self.assertEqual(str(cell.font.color.rgb), "000066CC", msg)

    def _assert_not_blue(self, row: int, msg: str = ""):
        cell = self.ws.cell(row=row, column=7)
        if cell.font.color and cell.font.color.rgb:
            self.assertNotEqual(str(cell.font.color.rgb), "000066CC", msg)

    def test_scenario(self):
        """所有场景批量验证。"""
        mv._apply_price_type_colors(self.ws, 2, 2 + len(self.test_cases) - 1)

        errors = []
        for i, (name, price_type, expected_blue) in enumerate(self.test_cases):
            row = i + 2
            try:
                if expected_blue:
                    self._assert_blue(row, f"Row {row}: {name} / {price_type} should be blue")
                else:
                    self._assert_not_blue(row, f"Row {row}: {name} / {price_type} should NOT be blue")
            except AssertionError as e:
                errors.append(str(e))
        if errors:
            self.fail("\n".join(errors))

    def test_empty_range_no_error(self):
        """空范围（start > end）→ 不报错。"""
        mv._apply_price_type_colors(self.ws, 100, 50)  # 无异常即为成功

    def test_single_row(self):
        """单行范围。"""
        self.ws.cell(row=3, column=2, value="测试")
        self.ws.cell(row=3, column=7, value="场内收盘价(T)")
        mv._apply_price_type_colors(self.ws, 3, 3)
        self._assert_blue(3)

    def test_none_price_type_col(self):
        """取价方式列为 None → 不报错。"""
        row = 2 + len(self.test_cases)  # 新行
        self.ws.cell(row=row, column=2, value="测试")
        self.ws.cell(row=row, column=7, value=None)  # 明确设为 None
        mv._apply_price_type_colors(self.ws, row, row)  # 不抛异常即通过
        self.assertIsNone(self.ws.cell(row=row, column=7).value)


# ═══════════════════════════════════════════════════════════
#  write_market_value_sheet
# ═══════════════════════════════════════════════════════════


class TestWriteMarketValueSheet(unittest.TestCase):
    """测试 write_market_value_sheet 页签写入（mock 内部函数和 Excel 写入）。"""

    def setUp(self):
        self.holdings = [
            Holding("证券账户", "电池ETF", "561910", 100.0, 1.0),
            Holding("支付宝", "中欧医疗健康混合", "003095", 200.0, 2.0),
        ]
        self.details = [
            mv.DetailRow(
                account="证券账户", name="电池ETF", code="561910",
                price=10.0, nav_date="2026-06-26", yesterday_close=9.5,
                price_type="场内收盘价(T)", premium="--",
                shares=100.0, market_value=1000.0, cost=100.0,
                profit=900.0, profit_rate=9.0, today_profit=50.0,
                source="腾讯财经", source_api="tencent",
            ),
            mv.DetailRow(
                account="支付宝", name="中欧医疗健康混合", code="003095",
                price=1.5, nav_date="2026-06-25", yesterday_close=1.48,
                price_type="官方净值(T-1)", premium="--",
                shares=200.0, market_value=300.0, cost=400.0,
                profit=-100.0, profit_rate=-0.25, today_profit=4.0,
                source="东方财富", source_api="eastmoney",
            ),
        ]

    @patch("src.report.market_value.write_total_row")
    @patch("src.report.market_value.write_subtotal_row")
    @patch("src.report.market_value.write_data_row")
    @patch("src.report.market_value.write_header_row")
    @patch("src.report.market_value.write_title_row")
    @patch("src.report.market_value._apply_price_type_colors")
    @patch("src.report.market_value._apply_profit_colors")
    @patch("src.report.market_value.freeze_header")
    @patch("src.report.market_value.auto_width")
    @patch("src.report.market_value._generate_details")
    @patch("src.report.market_value._detail_to_row_values")
    @patch("src.report.market_value._num_formats")
    def test_basic_write(self, mock_fmts, mock_to_row, mock_gen,
                         mock_aw, mock_freeze, mock_color,
                         mock_pt_color, mock_tl, mock_hdr, mock_data, mock_sub, mock_total):
        """正常写入：验证汇总值正确，内部函数被调用。"""
        mock_gen.return_value = self.details
        mock_fmts.return_value = [""] * 15
        mock_to_row.side_effect = lambda d: [
            d.account, d.name, d.code, d.price,
            d.nav_date, d.yesterday_close, d.price_type, d.premium,
            d.shares, d.market_value, d.cost, d.profit,
            d.profit_rate, d.today_profit, d.source,
        ]
        # write_title_row 返回下一行号
        mock_tl.return_value = 2
        # write_header_row 返回下一行号
        mock_hdr.return_value = 3

        ws = MagicMock()
        result = mv.write_market_value_sheet(ws, self.holdings, "2026-06-26")

        grand_mv, grand_cost, grand_profit, grand_today, details = result
        # 验证汇总值
        self.assertAlmostEqual(grand_mv, 1300.0)    # 1000 + 300
        self.assertAlmostEqual(grand_cost, 500.0)   # 100 + 400
        self.assertAlmostEqual(grand_profit, 800.0)  # 900 + (-100)
        self.assertAlmostEqual(grand_today, 54.0)    # 50 + 4
        self.assertEqual(len(details), 2)

        # 验证 _generate_details 被调用
        mock_gen.assert_called_once_with(self.holdings, "2026-06-26")

        # 验证标题和表头行
        mock_tl.assert_called_once()
        mock_hdr.assert_called_once()

        # 验证小计和总计
        self.assertEqual(mock_sub.call_count, 2)  # 两个账户各一个小计
        mock_total.assert_called_once()

        # 验证着色
        mock_color.assert_called_once()
        mock_pt_color.assert_called_once()

        # 验证冻结和列宽
        mock_freeze.assert_called_once_with(ws, 2)
        mock_aw.assert_called_once_with(ws)

        # 验证工作表标题
        self.assertEqual(ws.title, "市值核算")

    @patch("src.report.market_value.write_total_row")
    @patch("src.report.market_value.write_subtotal_row")
    @patch("src.report.market_value.write_data_row")
    @patch("src.report.market_value.write_header_row")
    @patch("src.report.market_value.write_title_row")
    @patch("src.report.market_value._apply_price_type_colors")
    @patch("src.report.market_value._apply_profit_colors")
    @patch("src.report.market_value.freeze_header")
    @patch("src.report.market_value.auto_width")
    @patch("src.report.market_value._generate_details")
    @patch("src.report.market_value._detail_to_row_values")
    @patch("src.report.market_value._num_formats")
    def test_empty_holdings(self, mock_fmts, mock_to_row, mock_gen,
                            mock_aw, mock_freeze, mock_color,
                            mock_pt_color, mock_tl, mock_hdr, mock_data, mock_sub, mock_total):
        """空持仓 → 总市值为 0，无小计行。"""
        mock_gen.return_value = []
        mock_fmts.return_value = [""] * 15
        mock_tl.return_value = 2
        mock_hdr.return_value = 3

        ws = MagicMock()
        result = mv.write_market_value_sheet(ws, [], "2026-06-26")

        grand_mv, grand_cost, grand_profit, grand_today, details = result
        self.assertAlmostEqual(grand_mv, 0.0)
        self.assertAlmostEqual(grand_cost, 0.0)
        self.assertAlmostEqual(grand_profit, 0.0)
        self.assertAlmostEqual(grand_today, 0.0)
        self.assertEqual(details, [])

        # 无小计行（没有账户分组）
        mock_sub.assert_not_called()

        # 总计行仍写入
        mock_total.assert_called_once()

    @patch("src.report.market_value.write_total_row")
    @patch("src.report.market_value.write_subtotal_row")
    @patch("src.report.market_value.write_data_row")
    @patch("src.report.market_value.write_header_row")
    @patch("src.report.market_value.write_title_row")
    @patch("src.report.market_value._apply_price_type_colors")
    @patch("src.report.market_value._apply_profit_colors")
    @patch("src.report.market_value.freeze_header")
    @patch("src.report.market_value.auto_width")
    @patch("src.report.market_value._generate_details")
    @patch("src.report.market_value._detail_to_row_values")
    @patch("src.report.market_value._num_formats")
    def test_subtotal_per_account(self, mock_fmts, mock_to_row, mock_gen,
                                   mock_aw, mock_freeze, mock_color,
                                   mock_pt_color, mock_tl, mock_hdr, mock_data, mock_sub, mock_total):
        """多个账户 → 每个账户写入小计。"""
        detail_a = self.details[0]  # 证券账户
        detail_b = self.details[1]  # 支付宝
        # 增加第二条证券账户记录
        detail_c = mv.DetailRow(
            account="证券账户", name="长江电力", code="600900",
            price=25.0, nav_date="2026-06-26", yesterday_close=24.5,
            price_type="场内收盘价(T)", premium="--",
            shares=100.0, market_value=2500.0, cost=2000.0,
            profit=500.0, profit_rate=0.25, today_profit=50.0,
            source="腾讯财经", source_api="tencent",
        )
        mock_gen.return_value = [detail_a, detail_c, detail_b]
        mock_fmts.return_value = [""] * 15
        mock_to_row.side_effect = lambda d: [
            d.account, d.name, d.code, d.price,
            d.nav_date, d.yesterday_close, d.price_type, d.premium,
            d.shares, d.market_value, d.cost, d.profit,
            d.profit_rate, d.today_profit, d.source,
        ]
        mock_tl.return_value = 2
        mock_hdr.return_value = 3

        ws = MagicMock()
        result = mv.write_market_value_sheet(ws, [], "2026-06-26")

        # 两个账户 → 两个小计
        self.assertEqual(mock_sub.call_count, 2)

        grand_mv = result[0]
        # 证券账户: 1000 + 2500 = 3500; 支付宝: 300
        self.assertAlmostEqual(grand_mv, 3800.0)


if __name__ == "__main__":
    unittest.main()
