"""财经新闻关联模块单元测试 — 异常场景与边界测试。

测试目标：
  - build_news_data — 空持仓/API 失败降级（mock API）
  - write_news_sheet — 空数据/正常数据/LLM 分析列渲染
  - LLM 增强路径 — 配置开启时调用 enhance_news_correlation

运行：
  cd D:/codebase/zoo/investor-util
  python -m unittest src.test_news_correlation -v
"""

from __future__ import annotations

import unittest
from unittest.mock import MagicMock, patch

from src.models import Holding
from src.report import news_correlation as nc


class TestBuildNewsData(unittest.TestCase):
    """build_news_data 异常场景测试（mock 网络层 + LLM 配置）。"""

    @patch("src.config.get_llm_config")
    @patch("src.providers.news_aggregator.aggregate_news")
    def test_empty_holdings_returns_empty_list(
        self, mock_aggregate: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """空持仓 → aggregate_news 不会被调或返回空列表。"""
        mock_aggregate.return_value = []
        mock_llm_cfg.return_value = None
        result, meta = nc.build_news_data([])
        self.assertEqual(result, [])
        self.assertFalse(meta.get("llm_enabled"))

    @patch("src.config.get_llm_config")
    @patch("src.providers.news_aggregator.aggregate_news")
    def test_api_failure_returns_empty(
        self, mock_aggregate: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """API 失败 → 返回空列表。"""
        mock_aggregate.return_value = []
        mock_llm_cfg.return_value = None
        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)
        self.assertEqual(result, [])
        self.assertFalse(meta.get("llm_enabled"))

    @patch("src.config.get_llm_config")
    @patch("src.providers.news_aggregator.aggregate_news")
    def test_api_returns_data(
        self, mock_aggregate: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """API 返回数据 → 正确传递。"""
        mock_aggregate.return_value = [
            {"title": "新闻标题", "intro": "简介", "url": "http://example.com",
             "ctime": "2026-06-27", "media_name": "新浪财经", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": False}
        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)
        self.assertEqual(len(result), 1)
        self.assertIn("matched_keywords", result[0])
        self.assertFalse(meta.get("llm_enabled"))


class TestBuildNewsDataWithLLM(unittest.TestCase):
    """build_news_data 开启 LLM 增强时的测试。"""

    @patch("src.config.get_llm_config")
    @patch("src.llm_client.enhance_news_correlation")
    @patch("src.providers.news_aggregator.aggregate_news")
    def test_llm_enabled_calls_enhance(
        self, mock_aggregate: MagicMock, mock_enhance: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """llm_news_analysis=true → 调用 enhance_news_correlation。"""
        mock_aggregate.return_value = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": True}
        enriched = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪",
             "matched_keywords": ["长江电力"],
             "llm_analysis": "[高] 直接涉及电力政策"},
        ]
        mock_enhance.return_value = (enriched, False, {"total_tokens": 300})

        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)

        mock_enhance.assert_called_once()
        self.assertTrue(meta.get("llm_enabled"))
        self.assertFalse(meta.get("llm_cached"))
        self.assertEqual(meta.get("token_usage", {}).get("total_tokens"), 300)
        self.assertEqual(len(result), 1)
        self.assertIn("llm_analysis", result[0])

    @patch("src.config.get_llm_config")
    @patch("src.llm_client.enhance_news_correlation")
    @patch("src.providers.news_aggregator.aggregate_news")
    def test_llm_cache_hit(
        self, mock_aggregate: MagicMock, mock_enhance: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """LLM 缓存命中 → meta 中 llm_cached 为 True。"""
        mock_aggregate.return_value = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": True}
        enriched = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪",
             "matched_keywords": ["长江电力"],
             "llm_analysis": "[高] 直接涉及电力政策"},
        ]
        mock_enhance.return_value = (enriched, True, {})  # cached

        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)

        self.assertTrue(meta.get("llm_cached"))

    @patch("src.config.get_llm_config")
    @patch("src.llm_client.enhance_news_correlation")
    @patch("src.providers.news_aggregator.aggregate_news")
    def test_llm_enhance_failure_falls_back_gracefully(
        self, mock_aggregate: MagicMock, mock_enhance: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """LLM 增强失败 → 返回原始新闻数据 + llm_enabled=True 但无分析。"""
        mock_aggregate.return_value = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": True}
        mock_enhance.side_effect = Exception("LLM 服务异常")

        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)

        self.assertEqual(len(result), 1)
        self.assertNotIn("llm_analysis", result[0])
        self.assertTrue(meta.get("llm_enabled"))
        # 失败后 token_usage 为空
        self.assertEqual(meta.get("token_usage", {}), {})

    @patch("src.config.get_llm_config")
    @patch("src.providers.news_aggregator.aggregate_news")
    def test_llm_disabled_returns_no_llm_meta(
        self, mock_aggregate: MagicMock, mock_llm_cfg: MagicMock,
    ) -> None:
        """llm_news_analysis=false → llm_enabled=False。"""
        mock_aggregate.return_value = [
            {"title": "新闻", "intro": "简介", "url": "http://ex.com",
             "ctime": "2026-06-28", "media_name": "新浪", "matched_keywords": ["长江电力"]},
        ]
        mock_llm_cfg.return_value = {"llm_news_analysis": False}

        holdings = [
            Holding(account="证券", name="长江电力", code="600900",
                     shares=100, cost_price=10.0)
        ]
        result, meta = nc.build_news_data(holdings)

        self.assertEqual(len(result), 1)
        self.assertNotIn("llm_analysis", result[0])
        self.assertFalse(meta.get("llm_enabled"))


class TestWriteNewsSheet(unittest.TestCase):
    """write_news_sheet 边界测试。"""

    def setUp(self) -> None:
        from openpyxl import Workbook
        self.wb = Workbook()
        self.ws = self.wb.active

    def test_empty_data(self) -> None:
        """空数据 → 写入"暂无关联新闻"占位。"""
        nc.write_news_sheet(self.ws, [])
        any_text = False
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and "暂无" in str(cell.value):
                    any_text = True
        self.assertTrue(any_text)

    def test_with_data(self) -> None:
        """有数据 → 正确写入标题、序号、关键词（6 列，无 LLM）。"""
        data = [
            {"title": "新闻A", "intro": "简介A", "url": "http://a.com",
             "ctime": "2026-06-27", "media_name": "新浪", "matched_keywords": ["茅台"]},
            {"title": "新闻B", "intro": "简介B", "url": "http://b.com",
             "ctime": "2026-06-26", "media_name": "新浪", "matched_keywords": ["五粮液", "白酒"]},
        ]
        nc.write_news_sheet(self.ws, data)
        headers = [self.ws.cell(row=2, column=c).value for c in range(1, 7)]
        self.assertIn("序号", headers)
        self.assertEqual(self.ws.cell(row=3, column=1).value, 1)

    def test_partial_missing_fields(self) -> None:
        """数据缺少可选字段 → 不崩溃。"""
        data = [
            {"title": "仅标题", "matched_keywords": ["茅台"]},
        ]
        try:
            nc.write_news_sheet(self.ws, data)
        except Exception as e:
            self.fail(f"write_news_sheet with partial data raised: {e}")

    def test_with_llm_analysis_column(self) -> None:
        """数据含 llm_analysis → 自动增加第 7 列。"""
        data = [
            {"title": "新闻A", "intro": "简介A", "url": "http://a.com",
             "ctime": "2026-06-27", "media_name": "新浪",
             "matched_keywords": ["茅台"], "llm_analysis": "[高] 白酒利好"},
        ]
        nc.write_news_sheet(self.ws, data)
        col7 = self.ws.cell(row=2, column=7).value
        self.assertEqual(col7, "LLM 关联分析")
        self.assertEqual(self.ws.cell(row=3, column=7).value, "[高] 白酒利好")

    def test_llm_enabled_with_token_usage(self) -> None:
        """LLM 启用 + 非缓存 → 底部显示 token 消耗。"""
        data = [
            {"title": "新闻A", "intro": "简介A", "url": "http://a.com",
             "ctime": "2026-06-27", "media_name": "新浪",
             "matched_keywords": ["茅台"], "llm_analysis": "[高] 白酒利好"},
        ]
        llm_meta = {
            "llm_enabled": True,
            "llm_cached": False,
            "token_usage": {"input_tokens": 2000, "output_tokens": 500, "total_tokens": 2500},
        }
        nc.write_news_sheet(self.ws, data, llm_meta=llm_meta)

        found_token = False
        for row in self.ws.iter_rows():
            for cell in row:
                val = str(cell.value) if cell.value else ""
                if "Token" in val and "2,500" in val:
                    found_token = True
        self.assertTrue(found_token, "Token 消耗行应出现在页签底部")

    def test_llm_cache_hit_footnote(self) -> None:
        """LLM 缓存命中 → 底部写"使用了LLM缓存"。"""
        data = [{"title": "新闻A", "intro": "简介", "url": "http://a.com",
                 "ctime": "2026-06-27", "media_name": "新浪",
                 "matched_keywords": ["茅台"], "llm_analysis": "[高] 利好"}]
        llm_meta = {"llm_enabled": True, "llm_cached": True, "token_usage": {}}
        nc.write_news_sheet(self.ws, data, llm_meta=llm_meta)

        found = False
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and "LLM缓存" in str(cell.value):
                    found = True
        self.assertTrue(found, "缓存命中时应显示'使用了LLM缓存'")

    def test_llm_disabled_footnote(self) -> None:
        """LLM 未启用 → 底部写"未依赖于LLM服务"。"""
        data = [{"title": "新闻A", "intro": "简介", "url": "http://a.com",
                 "ctime": "2026-06-27", "media_name": "新浪", "matched_keywords": ["茅台"]}]
        llm_meta = {"llm_enabled": False, "llm_cached": False, "token_usage": {}}
        nc.write_news_sheet(self.ws, data, llm_meta=llm_meta)

        found = False
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and "未依赖于LLM" in str(cell.value):
                    found = True
        self.assertTrue(found, "LLM 未启用时应显示'未依赖于LLM服务'")

    def test_llm_analysis_is_empty_when_all_irrelevant(self) -> None:
        """所有新闻 LLM 判定为无关 → 不显示 LLM 分析列。"""
        data = [
            {"title": "无关新闻", "intro": "无关内容", "url": "http://irr.com",
             "ctime": "2026-06-28", "media_name": "新浪",
             "matched_keywords": [], "llm_analysis": ""},
        ]
        nc.write_news_sheet(self.ws, data)
        col7 = self.ws.cell(row=2, column=7).value
        self.assertIsNone(col7)


if __name__ == "__main__":
    unittest.main()
