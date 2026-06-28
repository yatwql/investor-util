"""LLM 内容输出模块 — 报告第 7、8 页。

由调用方预生成 LLM 内容后传入本模块写入 Excel 页签。
"""

from __future__ import annotations

import logging
import re
from math import ceil

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from src.report.excel_writer import (
    freeze_header,
    write_title_row,
)
from src.report.styles import CONTENT_FONT

logger = logging.getLogger("invest")

_CONTENT_NCOLS = 2
_COL_WIDTH = 80          # 固定列宽（字符宽度）
_CHARS_PER_LINE = 40     # 每行约 40 中文字符
_ROW_HEIGHT_MIN = 30     # 最小行高 (pt)
_ROW_HEIGHT_PER_LINE = 15  # 每行增加高度 (pt)
_CACHE_HINT_TEXT = "本次使用LLM缓存，未直接使用LLM服务能力"


def _strip_html(text: str) -> str:
    """移除文本中的 HTML 标签，保留纯文本内容。"""
    if not text:
        return ""
    return re.sub(r"<[^>]+>", "", text).strip()


def _calc_row_height(text: str) -> int:
    """根据文本长度估算行高。

    按中文字符宽度估算：列宽 _COL_WIDTH 约容纳 _CHARS_PER_LINE 个中文。
    """
    if not text:
        return _ROW_HEIGHT_MIN
    lines = ceil(len(text) / _CHARS_PER_LINE)
    return max(lines * _ROW_HEIGHT_PER_LINE, _ROW_HEIGHT_MIN)


_CONTENT_ALIGN = Alignment(
    wrap_text=True,
    vertical="top",
    horizontal="left",
)

_CACHE_HINT_FONT = Font(
    color="999999",
    size=9,
    italic=True,
)


def _write_content_sheet(
    ws: Worksheet,
    title: str,
    content: str | None,
    from_cache: bool = False,
) -> None:
    """写入一个 LLM 内容页签。

    结构：
      - 第 1 行：标题（合并 A1:B1，居中标题样式）
      - 第 2 行起：每段落占一行 + 一行空行间距
      - 若 from_cache 为 True，末尾追加灰色缓存提示行

    Args:
        ws: 目标工作表
        title: 页签标题行文本
        content: LLM 返回的 HTML 文本（已剥离标签），为 None 时写入占位符
        from_cache: 是否来自缓存（为 True 时追加缓存提示）
    """
    ws.title = title

    # 标题行
    row = write_title_row(ws, 1, title, _CONTENT_NCOLS)

    # 固定列宽
    ws.column_dimensions["A"].width = _COL_WIDTH

    if content:
        text = _strip_html(content)
        # 按双换行分段，过滤空段
        paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
        for para in paragraphs:
            cell = ws.cell(row=row, column=1, value=para)
            cell.font = CONTENT_FONT
            cell.alignment = _CONTENT_ALIGN
            ws.row_dimensions[row].height = _calc_row_height(para)
            row += 1
            # 段落间空行
            row += 1
    else:
        placeholder = (
            "本节内容待生成 — 请配置 LLM API Key（data/config/llm_key.json）"
        )
        cell = ws.cell(row=row, column=1, value=placeholder)
        cell.font = CONTENT_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = _ROW_HEIGHT_MIN
        row += 1

    # 缓存来源提示（仅在 content 中尚未包含时追加，避免重复）
    if from_cache and content and _CACHE_HINT_TEXT not in content:
        cell = ws.cell(row=row, column=1, value=_CACHE_HINT_TEXT)
        cell.font = _CACHE_HINT_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20

    # 冻结标题行
    freeze_header(ws, 1)


def write_llm_sheets(
    wb: Any,
    llm_content: tuple[str | None, str | None],
    llm_cached: tuple[bool, bool] = (False, False),
) -> tuple[str, str]:
    """写入 LLM 内容页签（模块 7 & 8）。

    调用方必须预先生成 llm_content，本函数仅负责写入 Excel。

    Args:
        wb: 工作簿
        llm_content: (macro_html, expert_html) 预生成内容，各可能为 None
        llm_cached: (macro_cached, expert_cached) 缓存标记，
            分别对应两个页签的缓存状态

    Returns:
        (macro_text, expert_text) 纯文本二元组，供 TUI 展示
    """
    ws7 = wb.create_sheet()
    ws7.title = "全球政经局势"
    ws8 = wb.create_sheet()
    ws8.title = "智囊团深度复盘"

    content7, content8 = llm_content
    macro_cached, expert_cached = llm_cached

    _write_content_sheet(ws7, "全球政经局势", content7, from_cache=macro_cached)
    _write_content_sheet(ws8, "智囊团深度复盘", content8, from_cache=expert_cached)

    logger.info("LLM 内容页签写入完成")

    # 返回纯文本内容，供 TUI 展示
    return _strip_html(content7) if content7 else "", _strip_html(content8) if content8 else ""
