"""财经新闻热点与持仓关联分析模块 — 报告增补页签。

从新浪财经/东方财富/财联社三个源获取最新财经新闻，
与持仓名称/代码以及穿透 TOP10 资产名称进行关键词匹配，
按关联度排序输出 TOP N，在 Excel 中以单独页签呈现。

输出列：
  序号 | 新闻标题 | 摘要 | 来源 | 发布时间 | 关联关键词

可选 LLM 增强：为每条新闻做关联度判定，增加"LLM 关联分析"列。
"""

from __future__ import annotations

import logging
import re
from typing import Any, List, Optional

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models import Holding
from src.report.excel_writer import (
    auto_width,
    freeze_header,
    write_data_row,
    write_header_row,
    write_title_row,
)

logger = logging.getLogger("invest")

_NCOLS = 6
_BASE_HEADERS = ["序号", "新闻标题", "摘要", "来源", "发布时间", "关联关键词"]

# ── 关键词富化 ────────────────────────────────────────────────


def _build_keyword_lookup(
    holdings: List[Holding],
    penetrated_assets: Optional[List[dict]] = None,
    industry_data: Optional[dict[str, dict]] = None,
) -> dict[str, dict]:
    """构建关键词→来源正向查找表。

    对每条持仓的名称片段和代码、穿透资产的名称和代码建立索引，
    可选加入行业分类和概念板块关键词（type="concept" 类型），
    用于将 matched_keywords 还原为具体的持仓/穿透/概念/行业标签。

    Args:
        holdings: 持仓列表
        penetrated_assets: 穿透 TOP10 资产列表
        industry_data: 行业/概念数据字典 {code: {industry, concepts, ...}}
            industry_data 由 fetcher.batch_fetch_industry_data 返回，
            每项含 industry（行业名称）和 concepts（概念板块名称列表）。

    Returns:
        {keyword: {"type": "holding"|"penetration"|"concept", "name": str, "code": str}}
        同一关键词若同时匹配持仓和穿透，持仓优先（先到先得）。
    """
    lookup: dict[str, dict] = {}

    _suffixes = [
        "ETF", "联接", "A", "C", "(QDII)", "基金", "混合",
        "指数", "开放", "式", "发起", "LOF",
    ]
    for h in holdings:
        code = h.code.strip()
        name = h.name.strip()

        if code and code not in lookup:
            lookup[code] = {"type": "holding", "name": name, "code": code}

        clean = name
        for suffix in _suffixes:
            clean = clean.replace(suffix, "")
        terms = re.findall(r"[一-鿿]{2,}", clean)
        # 从多字词中额外提取双字窗口（如"长江电力"→"长江""电力"）
        _extra: set[str] = set()
        for t in terms:
            if len(t) > 2:
                for i in range(len(t) - 1):
                    _extra.add(t[i:i + 2])
        terms = list(set(terms) | _extra)
        for t in terms:
            if t not in lookup:
                lookup[t] = {"type": "holding", "name": name, "code": code}

        if "ETF" in name:
            core = name.replace("ETF", "").strip()
            core_terms = re.findall(r"[一-鿿]{2,}", core)
            for t in core_terms:
                if t not in lookup:
                    lookup[t] = {"type": "holding", "name": name, "code": code}

    if penetrated_assets:
        for asset in penetrated_assets:
            asset_name = (asset.get("name") or "").strip()
            asset_codes = asset.get("codes") or []

            for ac in asset_codes:
                ac_stripped = ac.strip()
                if ac_stripped and ac_stripped not in lookup:
                    lookup[ac_stripped] = {
                        "type": "penetration", "name": asset_name, "code": ac_stripped,
                    }

            if asset_name:
                clean_name = asset_name
                for suffix in _suffixes:
                    clean_name = clean_name.replace(suffix, "")
                terms = re.findall(r"[一-鿿]{2,}", clean_name)
                _extra_pen: set[str] = set()
                for t in terms:
                    if len(t) > 2:
                        for i in range(len(t) - 1):
                            _extra_pen.add(t[i:i + 2])
                terms = list(set(terms) | _extra_pen)
                for t in terms:
                    if t not in lookup:
                        lookup[t] = {
                            "type": "penetration", "name": asset_name, "code": t,
                        }

        # ── 3) 从行业分类/概念板块数据提取 ──
    if industry_data:
        for code, idata in industry_data.items():
            # 行业名称（如 "电力设备"）
            industry_name = (idata.get("industry") or "").strip()
            if industry_name and industry_name not in lookup:
                lookup[industry_name] = {
                    "type": "concept", "name": industry_name,
                    "code": code, "source": "industry",
                }
            # 概念板块名称（如 "CPO光模块"）
            for cname in idata.get("concepts", []):
                cname = cname.strip()
                if cname and cname not in lookup:
                    lookup[cname] = {
                        "type": "concept", "name": cname,
                        "code": code, "source": "concept",
                    }

    # ── 4) 为持仓/穿透条目附加行业/概念信息 ───
    # 只有 code 为真实证券代码（纯数字或含字母）的条目才查找 industry_data
    if industry_data:
        for _key, _entry in lookup.items():
            if _entry.get("type") in ("holding", "penetration"):
                _code = _entry.get("code", "")
                if _code in industry_data:
                    _idata = industry_data[_code]
                    if _idata.get("industry"):
                        _entry["industry"] = _idata["industry"]
                    if _idata.get("concepts"):
                        _entry["concepts_list"] = _idata["concepts"]

    return lookup


def _format_industry_tags(entry: dict) -> str:
    """为持仓/穿透条目生成行业/概念标签后缀。

    Args:
        entry: lookup 条目（可能含 industry / concepts_list 字段）

    Returns:
        标签字符串，如 " [电力·水电]"；无信息时返回 ""
    """
    tags: list[str] = []
    if entry.get("industry"):
        tags.append(entry["industry"])
    concepts_list = entry.get("concepts_list", [])
    if concepts_list:
        tags.extend(concepts_list[:2])  # 最多取前 2 个概念
    if tags:
        return f" [{' · '.join(tags)}]"
    return ""


def _enrich_keywords_for_item(
    item: dict[str, Any],
    keyword_lookup: dict[str, dict],
) -> list[dict]:
    """将单条新闻的 matched_keywords 富化为 enriched_keywords。

    返回列表：
        [{"display": "长江电力(600900)", "type": "holding"}, ...]

    去重规则：同一持仓只出现一次，同一穿透资产同理。
    """
    matched = item.get("matched_keywords", [])
    if not matched:
        return []

    seen: set[tuple[str, str]] = set()
    enriched: list[dict] = []

    for kw in matched:
        entry = keyword_lookup.get(kw)
        if entry and entry["type"] == "holding":
            dedup_key = entry["code"]
            if ("holding", dedup_key) not in seen:
                seen.add(("holding", dedup_key))
                enriched.append({
                    "display": f"{entry['name']}({entry['code']})"
                               f"{_format_industry_tags(entry)}",
                    "type": "holding",
                })
        elif entry and entry["type"] == "penetration":
            dedup_key = entry.get("code", entry["name"])
            if ("penetration", dedup_key) not in seen:
                seen.add(("penetration", dedup_key))
                enriched.append({
                    "display": f"{entry['name']}[穿透]"
                               f"{_format_industry_tags(entry)}",
                    "type": "penetration",
                })
        elif entry and entry["type"] == "concept":
            dedup_key = entry.get("name", kw)
            if ("concept", dedup_key) not in seen:
                seen.add(("concept", dedup_key))
                enriched.append({
                    "display": f"{entry['name']}[概念]",
                    "type": "concept",
                })
        else:
            if ("industry", kw) not in seen:
                seen.add(("industry", kw))
                enriched.append({
                    "display": kw,
                    "type": "industry",
                })

    type_order = {"holding": 0, "penetration": 1, "concept": 2, "industry": 3}
    enriched.sort(key=lambda x: type_order.get(x["type"], 99))

    return enriched


def _format_enriched_keywords(enriched: list[dict]) -> str:
    """将 enriched_keywords 列表格式化为单行显示字符串。"""
    if not enriched:
        return ""
    return ", ".join(item["display"] for item in enriched)


# ── 核心函数 ──────────────────────────────────────────────────


def build_news_data(
    holdings: List[Holding],
    top_n: int = 100,
    penetrated_assets: Optional[List[dict]] = None,
) -> tuple[list[dict[str, Any]], dict]:
    """获取新闻数据并与持仓关联。

    从多个财经新闻源获取最新新闻，
    与持仓名称/代码及穿透 TOP10 资产进行关键词匹配，
    按关联度排序返回 TOP N。

    若 llm_settings.json 中 llm_news_analysis 为 true，自动启用 LLM 二次分析，
    对新闻逐条判定关联度并给出原因分析，结果写入 llm_analysis 字段。

    Args:
        holdings: 持仓列表
        top_n: 最多返回的关联新闻条数
        penetrated_assets: 穿透 TOP10 资产列表（可选），
            每项含 name 和 codes 字段。传入后新闻关键词
            会额外覆盖穿透到的底层资产。

    Returns:
        (news_data, meta)
        news_data: [{title, intro, url, ctime, media_name, matched_keywords, llm_analysis?}, ...]
        meta: {
            "token_usage": {...},        # LLM token 消耗（启用时）
            "llm_cached": bool,          # LLM 结果是否来自缓存
            "llm_enabled": bool,         # 是否启用了 LLM 分析
        }
        获取失败时 news_data 为 []。
    """
    from src.providers.news_aggregator import aggregate_news
    from src.providers.news_keywords import build_holding_keywords

    keywords = build_holding_keywords(holdings, penetrated_assets=penetrated_assets)
    logger.info("新闻关联关键词（含穿透）: %s", keywords)

    # ── 行业/概念关键词扩展 ──────────────────────────────────────
    _industry_data: dict[str, dict] = {}
    try:
        # 收集所有唯一代码（持仓 + 穿透资产）
        _all_codes: set[str] = set()
        for h in holdings:
            if h.code and h.code.strip():
                _all_codes.add(h.code.strip())
        if penetrated_assets:
            for _asset in penetrated_assets:
                for _ac in (_asset.get("codes") or []):
                    if _ac and _ac.strip():
                        _all_codes.add(_ac.strip())

        if _all_codes:
            from src.fetcher import batch_fetch_industry_data as _batch_industry
            _industry_data = _batch_industry(list(_all_codes))
            if _industry_data:
                # 将行业名称和概念名称追加为关键词，提高匹配率
                _extra_kw: list[str] = []
                for _idata in _industry_data.values():
                    if _idata.get("industry"):
                        _extra_kw.append(_idata["industry"])
                    for _cname in _idata.get("concepts", []):
                        if _cname.strip():
                            _extra_kw.append(_cname.strip())
                if _extra_kw:
                    _all_kw = list(set(keywords + _extra_kw))
                    _all_kw.sort(key=lambda x: (-len(x), x))
                    keywords = _all_kw
                    logger.info("行业/概念关键词扩展: 新增 %d 个 → 共 %d 个",
                                len(_extra_kw), len(keywords))
    except Exception as e:
        logger.warning("行业/概念数据获取失败（非关键错误，继续）: %s", e)
        _industry_data = {}

    news_items = aggregate_news(keywords, top_n=top_n)

    # 提取成功访问的数据源列表（用于报告底部脚注）
    _active_sources: list[str] = []
    if news_items:
        seen_labels: list[str] = []
        for item in news_items:
            label = item.get("_source", "")
            if label and label not in seen_labels:
                seen_labels.append(label)
                _active_sources.append(label)

    # 初始化元数据
    meta: dict = {
        "token_usage": {},
        "llm_cached": False,
        "llm_enabled": False,
        "active_sources": _active_sources,
    }

    if news_items:
        logger.info("新闻关联完成: 获取 %d 条, 匹配 %d 条",
                    len(news_items), sum(1 for n in news_items if n.get("matched_keywords")))
    else:
        logger.warning("新闻获取失败")
        return news_items, meta

    # ── LLM 增强（可选） ──────────────────────────────────────
    from src.config import get_llm_config
    _llm_config = get_llm_config()
    if _llm_config and _llm_config.get("llm_news_analysis", False):
        # 检查 API Key 是否实际配置 — 未配置时降级为传统分析
        _api_key = (_llm_config.get("api_key") or "").strip()
        if not _api_key:
            logger.warning("llm_news_analysis 已开启但未配置 api_key，降级为传统关键词匹配分析")
        else:
            meta["llm_enabled"] = True
            try:
                from src.llm_client import enhance_news_correlation
                news_items, _cached, _token_usage = enhance_news_correlation(
                    news_items, holdings, penetrated_assets=penetrated_assets,
                    industry_data=_industry_data,
                )
                meta["llm_cached"] = _cached
                meta["token_usage"] = _token_usage
                if _cached:
                    logger.info("LLM 新闻关联分析（缓存）: 富化 %d 条",
                                sum(1 for n in news_items if n.get("llm_analysis")))
            except Exception as e:
                logger.warning("LLM 新闻关联分析出错: %s", e)

    # ── 关键词富化（标注每个关键词的来源） ─────────────────
    _lookup = _build_keyword_lookup(holdings, penetrated_assets, industry_data=_industry_data)
    for item in news_items:
        enriched = _enrich_keywords_for_item(item, _lookup)
        if enriched:
            item["enriched_keywords"] = enriched
        else:
            item["enriched_keywords"] = []

    return news_items, meta


def write_news_sheet(
    ws: Worksheet,
    news_data: List[dict[str, Any]],
    llm_meta: Optional[dict] = None,
) -> None:
    """写入财经新闻热点与持仓关联分析页签。

    这是 Excel 的增补页签（仅在 N 选项时生成）。
    若有 LLM 分析数据，自动增加 "LLM 关联分析" 列。

    Args:
        ws: 目标工作表
        news_data: build_news_data() 返回的数据
        llm_meta: LLM 元数据，含 token_usage / llm_cached / llm_enabled
    """
    ws.title = "财经新闻热点"

    # 检测是否有 LLM 分析数据（按 item 中的 llm_analysis 字段）
    has_llm = any(
        isinstance(item, dict) and item.get("llm_analysis")
        for item in news_data
    )
    ncols = _NCOLS + (1 if has_llm else 0)
    headers = _BASE_HEADERS + (["LLM 关联分析"] if has_llm else [])

    row = write_title_row(ws, 1, "财经新闻热点与持仓关联分析", ncols)
    row = write_header_row(ws, row, headers)

    if not news_data:
        write_data_row(ws, row, ["暂无关联新闻"])
        logger.info("新闻关联分析：无数据")
        freeze_header(ws, 2)
        auto_width(ws)
        return

    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for idx, item in enumerate(news_data, 1):
        enriched = item.get("enriched_keywords", [])
        if enriched:
            keywords_str = _format_enriched_keywords(enriched)
        else:
            keywords_str = ", ".join(item.get("matched_keywords", []))

        vals: list = [
            idx,
            item.get("title", ""),
            item.get("intro", ""),
            item.get("media_name", ""),
            item.get("ctime", ""),
            keywords_str,
        ]
        if has_llm:
            vals.append(item.get("llm_analysis", ""))
        write_data_row(ws, row, vals)
        ws.cell(row=row, column=2).alignment = wrap_left
        ws.cell(row=row, column=3).alignment = wrap_left
        row += 1

    # 底部说明
    row += 1

    if llm_meta and llm_meta.get("llm_enabled"):
        # LLM 已启用
        if llm_meta.get("llm_cached"):
            note_parts = [
                f"共获取 {len(news_data)} 条关联新闻。"
                "本次使用LLM缓存，未直接使用LLM服务能力"
            ]
        else:
            note_parts = [f"共获取 {len(news_data)} 条关联新闻"]
            if has_llm:
                note_parts.append("（含 LLM 智能关联分析）")
            note_parts.append("，关键词匹配基于持仓名称和代码")
    else:
        # LLM 未启用（或配置关闭）
        note_parts = [
            f"共获取 {len(news_data)} 条关联新闻。"
            "本次未使用LLM服务能力增强支持，使用传统关键字匹配技术"
        ]

    write_data_row(ws, row, ["".join(note_parts)])

    # Token 用量行（LLM 启用且非缓存命中时）
    if llm_meta and llm_meta.get("llm_enabled") and not llm_meta.get("llm_cached"):
        token_usage = llm_meta.get("token_usage") or {}
        if token_usage.get("total_tokens", 0) > 0:
            row += 1
            token_note = (
                f"Token 用量："
                f"输入 {token_usage.get('input_tokens', 0):,} / "
                f"输出 {token_usage.get('output_tokens', 0):,} = "
                f"{token_usage.get('total_tokens', 0):,}"
            )
            write_data_row(ws, row, [token_note])

    freeze_header(ws, 2)
    auto_width(ws)
    # 覆盖新闻标题和摘录列宽（auto_width 的 max_width=30 偏窄）
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    if has_llm:
        llm_col = _NCOLS + 1
        ws.column_dimensions[get_column_letter(llm_col)].width = 30
    llm_info = f"，LLM 分析 {sum(1 for n in news_data if n.get('llm_analysis'))} 条" if has_llm else ""
    logger.info("新闻关联分析页签写入完成%s，共 %d 条", llm_info, len(news_data))
